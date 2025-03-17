import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import matplotlib.pyplot as plt
import seaborn as sns
import datetime
from io import BytesIO
import os
import time
import calendar
from sklearn.linear_model import LinearRegression, HuberRegressor
from sklearn.ensemble import IsolationForest, RandomForestRegressor
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
from scipy import stats
from scipy.spatial.distance import cdist
import warnings
import re

# Suppress specific warnings
warnings.filterwarnings('ignore', category=FutureWarning)
warnings.filterwarnings('ignore', category=UserWarning)

# Configuración inicial de la página
st.set_page_config(
    page_title="Smart Fuel Analytics 3.0",
    page_icon="⛽",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Definición de temas de color y estilos principales
THEMES = {
    'modern': {
        'primary': '#4361EE',
        'secondary': '#3A0CA3',
        'accent': '#F72585',
        'success': '#4CC9F0',
        'info': '#4895EF',
        'warning': '#F8961E',
        'danger': '#F94144',
        'light': '#F8F9FA',
        'dark': '#212529',
        'background': '#F8F9FA',
        'card': '#FFFFFF',
        'text': '#212529',
        'subtext': '#6C757D',
        'border': '#DEE2E6',
        'heat_1': '#F94144',
        'heat_2': '#F8961E',
        'heat_3': '#F9C74F',
        'heat_4': '#90BE6D',
        'heat_5': '#43AA8B',
        'gradient_start': 'rgba(67, 97, 238, 0.8)',
        'gradient_end': 'rgba(58, 12, 163, 0.6)',
    },
    'dark': {
        'primary': '#BB86FC',
        'secondary': '#3700B3',
        'accent': '#03DAC6',
        'success': '#00E676',
        'info': '#2196F3',
        'warning': '#FB8C00',
        'danger': '#CF6679',
        'light': '#BBBBBB',
        'dark': '#121212',
        'background': '#121212',
        'card': '#1E1E1E',
        'text': '#E1E1E1',
        'subtext': '#BBBBBB',
        'border': '#333333',
        'heat_1': '#CF6679',
        'heat_2': '#FB8C00',
        'heat_3': '#FFAB00',
        'heat_4': '#64DD17',
        'heat_5': '#00B0FF',
        'gradient_start': 'rgba(187, 134, 252, 0.8)',
        'gradient_end': 'rgba(55, 0, 179, 0.6)',
    },
    'corporate': {
        'primary': '#0A66C2',
        'secondary': '#2867B2',
        'accent': '#E68523',
        'success': '#0AB94E',
        'info': '#2493DF',
        'warning': '#F5B700',
        'danger': '#E11D48',
        'light': '#F9FAFB',
        'dark': '#111827',
        'background': '#F9FAFB',
        'card': '#FFFFFF',
        'text': '#1F2937',
        'subtext': '#6B7280',
        'border': '#E5E7EB',
        'heat_1': '#E11D48',
        'heat_2': '#F59E0B',
        'heat_3': '#F5B700',
        'heat_4': '#10B981',
        'heat_5': '#0AB94E',
        'gradient_start': 'rgba(10, 102, 194, 0.8)',
        'gradient_end': 'rgba(40, 103, 178, 0.6)',
    }
}

# Selección del tema predeterminado
DEFAULT_THEME = 'modern'
COLORS = THEMES[DEFAULT_THEME]

# Agregamos manejo de sesión para tema
if 'theme' not in st.session_state:
    st.session_state['theme'] = DEFAULT_THEME
    
COLORS = THEMES[st.session_state['theme']]

# Estilos CSS personalizados avanzados
def load_css(colors):
    return f"""
<style>

     @keyframes fadeIn {{
        from {{ opacity: 0; transform: translateY(10px); }}
        to {{ opacity: 1; transform: translateY(0); }}
    }}
    
    @keyframes pulse {{
        0% {{ transform: scale(1); }}  <!-- ERROR: Debe ser }} al final -->
        50% {{ transform: scale(1.05); }}
        100% {{ transform: scale(1); }}
    }}



    /* Estilos generales */
    html {{
        font-family: 'SF Pro Display', 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen-Sans, Ubuntu, Cantarell, 'Helvetica Neue', sans-serif;
    }}
    
    .reportview-container .main .block-container {{
        padding-top: 1rem;
        padding-right: 1rem;
        padding-left: 1rem;
        padding-bottom: 1rem;
        max-width: 1200px;
    }}
    
    .reportview-container .main {{
        color: {colors['text']};
        background-color: {colors['background']};
    }}
    

    css = css.replace("{colors['text']}", colors['text'])
    css = css.replace("{colors['background']}", colors['background'])
    
    /* Scrollbar personalizado */
    ::-webkit-scrollbar {{
        width: 10px;
        height: 10px;
    }}
    
    ::-webkit-scrollbar-track {{
        background: {colors['background']};
        border-radius: 5px;
    }}
    
    ::-webkit-scrollbar-thumb {{
        background-color: {colors['primary']};
        border-radius: 5px;
        border: 2px solid {colors['background']};
    }}
    
    ::-webkit-scrollbar-thumb:hover {{
        background-color: {colors['secondary']};
    }}
    
    /* Títulos */
    h1, h2, h3, h4, h5, h6 {{
        font-family: 'SF Pro Display', 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
        font-weight: 600;
    }}
    
    .title {{
        font-size: 2.75rem;
        background: linear-gradient(90deg, {colors['primary']} 0%, {colors['secondary']} 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        margin-bottom: 1.5rem;
        text-align: center;
        font-weight: 800;
        letter-spacing: -0.5px;
        line-height: 1.2;
    }}
    
    .subtitle {{
        font-size: 1.8rem;
        margin-top: 2rem;
        margin-bottom: 1.5rem;
        color: {colors['secondary']};
        padding-bottom: 0.75rem;
        position: relative;
        font-weight: 700;
    }}
    
    .subtitle::after {{
        content: '';
        position: absolute;
        left: 0;
        bottom: 0;
        height: 4px;
        width: 80px;
        background: linear-gradient(90deg, {colors['primary']} 0%, {colors['secondary']} 100%);
        border-radius: 2px;
    }}
    
    .section-title {{
        font-size: 1.5rem;
        font-weight: 600;
        color: {colors['secondary']};
        margin-top: 1.75rem;
        margin-bottom: 1.25rem;
        display: flex;
        align-items: center;
    }}
    
    .section-title::before {{
        content: '';
        display: inline-block;
        width: 18px;
        height: 18px;
        background-color: {colors['primary']};
        margin-right: 10px;
        border-radius: 4px;
    }}
    
    /* Cards modernas */
    .card {{
        background-color: {colors['card']};
        border-radius: 16px;
        box-shadow: 0 10px 25px rgba(0, 0, 0, 0.05);
        padding: 1.75rem;
        margin-bottom: 1.75rem;
        transition: all 0.3s cubic-bezier(0.25, 0.8, 0.25, 1);
        border: 1px solid {colors['border']};
        overflow: hidden;
        position: relative;
    }}
    
    .card::before {{
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        width: 100%;
        height: 4px;
        background: linear-gradient(90deg, {colors['primary']} 0%, {colors['secondary']} 100%);
    }}
    
    .card:hover {{
        transform: translateY(-5px);
        box-shadow: 0 15px 30px rgba(0, 0, 0, 0.1);
    }}
    
    /* Métricas modernas */
    .metric-container {{
        display: flex;
        flex-direction: column;
        align-items: center;
        justify-content: center;
        text-align: center;
        padding: 1.5rem;
        border-radius: 16px;
        background-color: {colors['card']};
        box-shadow: 0 4px 15px rgba(0, 0, 0, 0.05);
        height: 100%;
        transition: all 0.3s cubic-bezier(0.25, 0.8, 0.25, 1);
        border: 1px solid {colors['border']};
        position: relative;
        overflow: hidden;
    }}
    
    .metric-container::before {{
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        width: 4px;
        height: 100%;
        background: linear-gradient(180deg, {colors['primary']} 0%, {colors['secondary']} 100%);
    }}
    
    .metric-container:hover {{
        transform: translateY(-5px);
        box-shadow: 0 10px 20px rgba(0, 0, 0, 0.08);
    }}
    
    .metric-value {{
        font-size: 2.75rem;
        font-weight: 700;
        background: linear-gradient(90deg, {colors['primary']} 0%, {colors['secondary']} 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        margin-bottom: 0.5rem;
        line-height: 1.1;
    }}
    
    .metric-label {{
        font-size: 1rem;
        color: {colors['subtext']};
        margin-bottom: 0.5rem;
        font-weight: 500;
    }}
    
    .metric-delta {{
        font-size: 0.9rem;
        display: flex;
        align-items: center;
        justify-content: center;
        font-weight: 500;
    }}
    
    .delta-up {{
        color: {colors['success']};
    }}
    
    .delta-down {{
        color: {colors['danger']};
    }}
    
    /* Alertas modernas */
    .alert {{
        border-radius: 12px;
        padding: 1.25rem;
        margin-bottom: 1.25rem;
        border-left: 5px solid;
        display: flex;
        align-items: flex-start;
        box-shadow: 0 4px 15px rgba(0, 0, 0, 0.05);
    }}
    
    .alert-icon {{
        font-size: 1.5rem;
        margin-right: 1rem;
    }}
    
    .alert-content {{
        flex: 1;
    }}
    
    .alert-danger {{
        background-color: rgba({int(colors['danger'][1:3], 16)}, {int(colors['danger'][3:5], 16)}, {int(colors['danger'][5:7], 16)}, 0.1);
        border-left-color: {colors['danger']};
    }}
    
    .alert-warning {{
        background-color: rgba({int(colors['warning'][1:3], 16)}, {int(colors['warning'][3:5], 16)}, {int(colors['warning'][5:7], 16)}, 0.1);
        border-left-color: {colors['warning']};
    }}
    
    .alert-info {{
        background-color: rgba({int(colors['info'][1:3], 16)}, {int(colors['info'][3:5], 16)}, {int(colors['info'][5:7], 16)}, 0.1);
        border-left-color: {colors['info']};
    }}
    
    .alert-success {{
        background-color: rgba({int(colors['success'][1:3], 16)}, {int(colors['success'][3:5], 16)}, {int(colors['success'][5:7], 16)}, 0.1);
        border-left-color: {colors['success']};
    }}
    
    /* Botones modernos */
    .primary-button {{
        background: linear-gradient(90deg, {colors['primary']} 0%, {colors['secondary']} 100%);
        color: white;
        border: none;
        border-radius: 12px;
        padding: 0.85rem 1.75rem;
        font-size: 1rem;
        font-weight: 600;
        cursor: pointer;
        transition: all 0.3s ease;
        display: inline-block;
        text-align: center;
        text-decoration: none;
        margin: 0.75rem 0;
        box-shadow: 0 4px 15px rgba(0, 0, 0, 0.1);
    }}
    
    .primary-button:hover {{
        transform: translateY(-2px);
        box-shadow: 0 7px 20px rgba(0, 0, 0, 0.2);
    }}
    
    /* Tablas modernas */
    .table-container {{
        border-radius: 12px;
        overflow: hidden;
        box-shadow: 0 4px 15px rgba(0, 0, 0, 0.05);
        margin-bottom: 1.5rem;
    }}
    
    .dataframe {{
        width: 100%;
        border-collapse: collapse;
    }}
    
    .dataframe th {{
        background: linear-gradient(90deg, {colors['primary']} 0%, {colors['secondary']} 100%);
        color: white;
        text-align: left;
        padding: 14px 20px;
        font-weight: 600;
        font-size: 0.9rem;
        letter-spacing: 0.5px;
        text-transform: uppercase;
    }}
    
    .dataframe td {{
        padding: 12px 20px;
        border-bottom: 1px solid {colors['border']};
        font-size: 0.95rem;
    }}
    
    .dataframe tr:nth-child(even) {{
        background-color: rgba(0, 0, 0, 0.02);
    }}
    
    .dataframe tr:hover {{
        background-color: rgba(0, 0, 0, 0.05);
    }}
    
    /* Heat Map Styling */
    .heat-cell {{
        padding: 10px;
        border-radius: 8px;
        font-weight: 500;
        text-align: center;
    }}
    
    .heat-1 {{ background-color: {colors['heat_1']}; color: white; }}
    .heat-2 {{ background-color: {colors['heat_2']}; color: white; }}
    .heat-3 {{ background-color: {colors['heat_3']}; color: black; }}
    .heat-4 {{ background-color: {colors['heat_4']}; color: black; }}
    .heat-5 {{ background-color: {colors['heat_5']}; color: black; }}
    
    /* Animaciones y efectos visuales */
    @keyframes fadeIn {{
        from {{ opacity: 0; transform: translateY(10px); }}
        to {{ opacity: 1; transform: translateY(0); }}
    }}
    
    .fade-in {{
        animation: fadeIn 0.6s ease-out;
    }}
    
   @keyframes pulse {{
    0% {{ transform: scale(1); }}
    50% {{ transform: scale(1.05); }}
    100% {{ transform: scale(1); }}
}}
    
    .pulse {{
        animation: pulse 2s infinite;
    }}
    
    /* Tarjetas de dashboard */
    .dashboard-card {{
        background: white;
        border-radius: 16px;
        padding: 1.5rem;
        box-shadow: 0 8px 20px rgba(0, 0, 0, 0.1);
        transition: all 0.3s ease;
        height: 100%;
        display: flex;
        flex-direction: column;
        position: relative;
        overflow: hidden;
    }}
    
    .dashboard-card::after {{
        content: '';
        position: absolute;
        bottom: 0;
        left: 0;
        width: 100%;
        height: 4px;
        background: linear-gradient(90deg, {colors['primary']} 0%, {colors['secondary']} 100%);
        transform: scaleX(0);
        transform-origin: left;
        transition: transform 0.3s ease;
    }}
    
    .dashboard-card:hover {{
        transform: translateY(-5px);
        box-shadow: 0 15px 30px rgba(0, 0, 0, 0.15);
    }}
    
    .dashboard-card:hover::after {{
        transform: scaleX(1);
    }}
    
    .dashboard-card-header {{
        margin-bottom: 1rem;
        padding-bottom: 1rem;
        border-bottom: 1px solid {colors['border']};
        display: flex;
        justify-content: space-between;
        align-items: center;
    }}
    
    .dashboard-card-title {{
        font-size: 1.2rem;
        font-weight: 600;
        color: {colors['secondary']};
        margin: 0;
    }}
    
    .dashboard-card-icon {{
        font-size: 1.5rem;
        color: {colors['primary']};
    }}
    
    .dashboard-card-content {{
        flex: 1;
        display: flex;
        flex-direction: column;
        justify-content: center;
    }}
    
    .dashboard-card-footer {{
        margin-top: 1rem;
        padding-top: 1rem;
        border-top: 1px solid {colors['border']};
        display: flex;
        justify-content: space-between;
        align-items: center;
        font-size: 0.85rem;
        color: {colors['subtext']};
    }}
    
    /* Personalización de plotly */
    .js-plotly-plot .plotly .modebar {{
        opacity: 0.2;
    }}
    
    .js-plotly-plot .plotly .modebar:hover {{
        opacity: 1;
    }}
    
    /* Badges */
    .badge {{
        display: inline-block;
        padding: 0.35em 0.65em;
        font-size: 0.75em;
        font-weight: 700;
        line-height: 1;
        text-align: center;
        white-space: nowrap;
        vertical-align: baseline;
        border-radius: 0.375rem;
    }}
    
    .badge-primary {{ background-color: {colors['primary']}; color: white; }}
    .badge-secondary {{ background-color: {colors['secondary']}; color: white; }}
    .badge-success {{ background-color: {colors['success']}; color: white; }}
    .badge-danger {{ background-color: {colors['danger']}; color: white; }}
    .badge-warning {{ background-color: {colors['warning']}; color: black; }}
    .badge-info {{ background-color: {colors['info']}; color: white; }}
    
    /* Progress bar */
    .progress {{
        height: 0.75rem;
        background-color: rgba(0, 0, 0, 0.1);
        border-radius: 0.375rem;
        margin: 0.5rem 0;
        overflow: hidden;
    }}
    
    .progress-bar {{
        height: 100%;
        border-radius: 0.375rem;
        transition: width 0.6s ease;
    }}
    
    .progress-bar-success {{ background-color: {colors['success']}; }}
    .progress-bar-warning {{ background-color: {colors['warning']}; }}
    .progress-bar-danger {{ background-color: {colors['danger']}; }}
    
    /* Tooltips */
    .tooltip {{
        position: relative;
        display: inline-block;
        cursor: pointer;
    }}
    
    .tooltip .tooltiptext {{
        visibility: hidden;
        width: 200px;
        background-color: {colors['dark']};
        color: white;
        text-align: center;
        border-radius: 6px;
        padding: 10px;
        position: absolute;
        z-index: 1;
        bottom: 125%;
        left: 50%;
        transform: translateX(-50%);
        opacity: 0;
        transition: opacity 0.3s;
        box-shadow: 0 5px 15px rgba(0, 0, 0, 0.2);
        font-size: 0.85rem;
    }}
    
    .tooltip:hover .tooltiptext {{
        visibility: visible;
        opacity: 1;
    }}
    
    /* Banner elegante */
    .banner {{
        background: linear-gradient(135deg, {colors['gradient_start']}, {colors['gradient_end']}),
                    url('https://images.unsplash.com/photo-1553708881-112abc53fe54?ixlib=rb-1.2.1&auto=format&fit=crop&w=1350&q=80');
        background-size: cover;
        background-position: center;
        background-blend-mode: overlay;
        color: white;
        padding: 3rem 2rem;
        margin-bottom: 2rem;
        border-radius: 16px;
        box-shadow: 0 10px 30px rgba(0, 0, 0, 0.1);
    }}
    
    .banner-content {{
        max-width: 600px;
    }}
    
    .banner-title {{
        font-size: 2.5rem;
        font-weight: 800;
        margin-bottom: 1rem;
        text-shadow: 0 2px 4px rgba(0, 0, 0, 0.2);
    }}
    
    .banner-subtitle {{
        font-size: 1.2rem;
        font-weight: 400;
        margin-bottom: 1.5rem;
        opacity: 0.9;
    }}
    
    /* Mini stats */
    .mini-stat {{
        display: flex;
        align-items: center;
        margin-bottom: 1rem;
    }}
    
    .mini-stat-icon {{
        width: 40px;
        height: 40px;
        border-radius: 12px;
        background: linear-gradient(135deg, {colors['primary']}, {colors['secondary']});
        display: flex;
        align-items: center;
        justify-content: center;
        margin-right: 1rem;
        color: white;
        font-size: 1.5rem;
    }}
    
    .mini-stat-content {{
        flex: 1;
    }}
    
    .mini-stat-value {{
        font-size: 1.25rem;
        font-weight: 700;
        color: {colors['text']};
        margin: 0;
    }}
    
    .mini-stat-label {{
        font-size: 0.85rem;
        color: {colors['subtext']};
        margin: 0;
    }}
    
    /* Loading spinner with better animation */
    .loading {{
        display: flex;
        justify-content: center;
        align-items: center;
        height: 200px;
    }}
    
    .loading-spinner {{
        width: 40px;
        height: 40px;
        border: 4px solid rgba(0, 0, 0, 0.1);
        border-left-color: {colors['primary']};
        border-radius: 50%;
        animation: spin 1s linear infinite;
    }}
    
    @keyframes spin {{
        to {{ transform: rotate(360deg); }}
    }}
    
    /* KPI Cards */
    .kpi-card {{
        background-color: {colors['card']};
        border-radius: 16px;
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.08);
        padding: 1.5rem;
        transition: all 0.3s ease;
        height: 100%;
        position: relative;
        display: flex;
        flex-direction: column;
    }}
    
    .kpi-card:hover {{
        transform: translateY(-5px);
        box-shadow: 0 10px 25px rgba(0, 0, 0, 0.12);
    }}
    
    .kpi-card-header {{
        display: flex;
        justify-content: space-between;
        align-items: center;
        margin-bottom: 1rem;
    }}
    
    .kpi-card-title {{
        font-size: 1rem;
        font-weight: 600;
        color: {colors['subtext']};
        margin: 0;
    }}
    
    .kpi-card-icon {{
        width: 40px;
        height: 40px;
        border-radius: 10px;
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 1.25rem;
    }}
    
    .kpi-card-icon.success {{ background-color: rgba({int(colors['success'][1:3], 16)}, {int(colors['success'][3:5], 16)}, {int(colors['success'][5:7], 16)}, 0.1); color: {colors['success']}; }}
    .kpi-card-icon.warning {{ background-color: rgba({int(colors['warning'][1:3], 16)}, {int(colors['warning'][3:5], 16)}, {int(colors['warning'][5:7], 16)}, 0.1); color: {colors['warning']}; }}
    .kpi-card-icon.danger {{ background-color: rgba({int(colors['danger'][1:3], 16)}, {int(colors['danger'][3:5], 16)}, {int(colors['danger'][5:7], 16)}, 0.1); color: {colors['danger']}; }}
    .kpi-card-icon.info {{ background-color: rgba({int(colors['info'][1:3], 16)}, {int(colors['info'][3:5], 16)}, {int(colors['info'][5:7], 16)}, 0.1); color: {colors['info']}; }}
    .kpi-card-icon.primary {{ background-color: rgba({int(colors['primary'][1:3], 16)}, {int(colors['primary'][3:5], 16)}, {int(colors['primary'][5:7], 16)}, 0.1); color: {colors['primary']}; }}
    
    .kpi-card-value {{
        font-size: 2.25rem;
        font-weight: 700;
        color: {colors['text']};
        margin: 0.5rem 0;
    }}
    
    .kpi-card-delta {{
        display: flex;
        align-items: center;
        font-size: 0.85rem;
        font-weight: 500;
        margin-bottom: 0.5rem;
    }}
    
    .kpi-card-delta.positive {{ color: {colors['success']}; }}
    .kpi-card-delta.negative {{ color: {colors['danger']}; }}
    
    .kpi-card-footer {{
        margin-top: auto;
        font-size: 0.85rem;
        color: {colors['subtext']};
        padding-top: 0.75rem;
        border-top: 1px solid {colors['border']};
    }}
    
    /* New style for sidebar */
    .css-1d391kg, .css-163ttbj, .css-1wrcr25 {{ /* Sidebar background classes */
        background-image: linear-gradient(180deg, {colors['gradient_start']}, {colors['gradient_end']});
    }}
    
    /* Responsividad avanzada */
    @media (max-width: 768px) {{
        .metric-value {{
            font-size: 2rem;
        }}
        
        .title {{
            font-size: 2rem;
        }}
        
        .subtitle {{
            font-size: 1.5rem;
        }}
        
        .banner {{
            padding: 2rem 1.5rem;
        }}
        
        .banner-title {{
            font-size: 2rem;
        }}
        
        .dashboard-card {{
            margin-bottom: 1rem;
        }}
    }}
    
    @media (max-width: 576px) {{
        .metric-value {{
            font-size: 1.75rem;
        }}
        
        .title {{
            font-size: 1.75rem;
        }}
        
        .subtitle {{
            font-size: 1.25rem;
        }}
        
        .banner {{
            padding: 1.5rem 1rem;
        }}
        
        .banner-title {{
            font-size: 1.5rem;
        }}
    }}
</style>
"""

# Aplicar estilos CSS
st.markdown(load_css(COLORS), unsafe_allow_html=True)

# --- FUNCIONES UTILITARIAS ---

# Función para mostrar mensaje de carga
def show_loading(message="Procesando datos..."):
    with st.spinner(message):
        # Placeholder para una barra de progreso o animación personalizada
        progress_placeholder = st.empty()
        progress_placeholder.markdown(f"""
        <div class="loading">
            <div class="loading-spinner"></div>
        </div>
        <p style="text-align: center; color: {COLORS['subtext']};">{message}</p>
        """, unsafe_allow_html=True)
        return progress_placeholder

# Función para mostrar un banner de bienvenida
def show_welcome_banner():
    st.markdown(f"""
    <div class="banner fade-in">
        <div class="banner-content">
            <h1 class="banner-title">Smart Fuel Analytics 3.0</h1>
            <p class="banner-subtitle">Plataforma avanzada para análisis y optimización de consumo de combustible.</p>
        </div>
    </div>
    """, unsafe_allow_html=True)

# Función para crear KPI cards
def create_kpi_card(title, value, delta=None, delta_suffix="%", icon="📊", color_class="primary", footnote=None):
    delta_html = ""
    if delta is not None:
        delta_class = "positive" if delta >= 0 else "negative"
        delta_icon = "↑" if delta >= 0 else "↓"
        delta_html = f"""
        <div class="kpi-card-delta {delta_class}">
            {delta_icon} {abs(delta):.2f}{delta_suffix}
        </div>
        """
    
    footnote_html = f"""
    <div class="kpi-card-footer">
        {footnote}
    </div>
    """ if footnote else ""
    
    return f"""
    <div class="kpi-card">
        <div class="kpi-card-header">
            <div class="kpi-card-title">{title}</div>
            <div class="kpi-card-icon {color_class}">{icon}</div>
        </div>
        <div class="kpi-card-value">{value}</div>
        {delta_html}
        {footnote_html}
    </div>
    """

# Función para mostrar alertas mejoradas
def show_alert(message, type="info", title=None, icon=None):
    if not icon:
        icons = {
            "info": "ℹ️",
            "success": "✅",
            "warning": "⚠️",
            "danger": "🚨"
        }
        icon = icons.get(type, "ℹ️")
    
    title_html = f"<h4>{title}</h4>" if title else ""
    
    return f"""
    <div class="alert alert-{type}">
        <div class="alert-icon">{icon}</div>
        <div class="alert-content">
            {title_html}
            <p>{message}</p>
        </div>
    </div>
    """

# Función para crear mini estadísticas
def create_mini_stat(icon, value, label):
    return f"""
    <div class="mini-stat">
        <div class="mini-stat-icon">{icon}</div>
        <div class="mini-stat-content">
            <p class="mini-stat-value">{value}</p>
            <p class="mini-stat-label">{label}</p>
        </div>
    </div>
    """

# Función para cargar y preprocesar datos con mejoras
@st.cache_data(ttl=3600)
def load_data(uploaded_file):
    try:
        # Iniciar temporizador para medir rendimiento
        start_time = time.time()
        
        import os
        import pandas as pd
        import numpy as np
        import datetime
        from sklearn.ensemble import IsolationForest
        from sklearn.linear_model import LinearRegression
        import streamlit as st
        
        # Determina la extensión del archivo
        file_extension = os.path.splitext(uploaded_file.name)[1].lower()
        
        # Mensaje de estado
        st.info(f"📂 Detectando formato de archivo: {file_extension}")
        
        # Establecer parámetros comunes para leer el archivo
        read_params = {'header': 2}
        
        # Manejo específico según el tipo de archivo
        if file_extension == '.xlsx':
            try:
                # Importar explícitamente openpyxl para verificar que esté disponible
                import openpyxl
                # Para archivos .xlsx usa openpyxl
                st.info("🔍 Leyendo archivo Excel con openpyxl...")
                df = pd.read_excel(uploaded_file, engine='openpyxl', **read_params)
            except ImportError:
                st.error("❌ Falta la dependencia 'openpyxl'. Instálala con: pip install openpyxl")
                return None
            except Exception as e:
                st.error(f"❌ Error al leer archivo .xlsx: {e}")
                return None
                
        elif file_extension == '.xls':
            try:
                # Importar explícitamente xlrd para verificar que esté disponible
                import xlrd
                # Verificar la versión de xlrd
                if xlrd.__VERSION__ < '2.0.1':
                    st.warning(f"⚠️ Versión de xlrd detectada: {xlrd.__VERSION__}. Se recomienda ≥2.0.1 para archivos .xls")
                
                # Para archivos .xls usa xlrd
                st.info("🔍 Leyendo archivo Excel con xlrd...")
                df = pd.read_excel(uploaded_file, engine='xlrd', **read_params)
            except ImportError:
                st.error("❌ Falta la dependencia 'xlrd'. Instálala con: pip install \"xlrd>=2.0.1\"")
                return None
            except Exception as e:
                st.error(f"❌ Error al leer archivo .xls: {e}")
                st.error("Si estás usando un archivo .xlsx con extensión .xls, renómbralo a .xlsx e intenta de nuevo.")
                return None
                
        elif file_extension in ['.csv', '.txt']:
            try:
                # Para archivos CSV, intentar con diferentes codificaciones
                encodings = ['utf-8', 'latin1', 'ISO-8859-1']
                separators = [',', ';', '\t']
                
                # Probar diferentes combinaciones de codificación y separador
                for encoding in encodings:
                    for sep in separators:
                        try:
                            st.info(f"🔍 Intentando leer CSV con codificación {encoding} y separador '{sep}'...")
                            df = pd.read_csv(uploaded_file, encoding=encoding, sep=sep, **read_params)
                            if len(df.columns) > 1:  # Si tiene más de una columna, consideramos que se leyó correctamente
                                st.success(f"✅ Archivo CSV leído correctamente con codificación {encoding} y separador '{sep}'")
                                break
                        except Exception:
                            continue
                    else:
                        continue
                    break
                else:
                    st.error("❌ No se pudo leer el archivo CSV con ninguna combinación de codificación y separador.")
                    return None
            except Exception as e:
                st.error(f"❌ Error al leer archivo CSV: {e}")
                return None
        else:
            # Si la extensión no es conocida, intentar autodetectar
            try:
                st.info("🔍 Intentando autodetectar formato de archivo...")
                # Intenta primero con openpyxl
                try:
                    import openpyxl
                    df = pd.read_excel(uploaded_file, engine='openpyxl', **read_params)
                except:
                    # Si falla, intenta con xlrd
                    import xlrd
                    df = pd.read_excel(uploaded_file, engine='xlrd', **read_params)
            except Exception as e:
                st.error(f"❌ Formato de archivo no reconocido: {file_extension}")
                st.error(f"Error al intentar leer el archivo: {e}")
                st.error("Por favor, utiliza archivos Excel (.xlsx, .xls) o CSV (.csv)")
                return None
        
        # Verificar si se cargaron datos
        if df.empty:
            st.error("❌ El archivo está vacío o no contiene datos válidos.")
            return None
        
        # Mostrar estadísticas de carga preliminares
        st.info(f"📊 Datos cargados: {len(df)} filas, {len(df.columns)} columnas")
        
        # Registrar las columnas detectadas
        st.session_state['detected_columns'] = list(df.columns)
        
        # Limpiar nombres de columnas
        df.columns = [col.strip() if isinstance(col, str) else col for col in df.columns]
        
        # Verificar la presencia de columnas clave
        columnas_esperadas = ['Fecha', 'Hora', 'Cantidad litros', 'Terminal', 'Número interno']
        columnas_faltantes = [col for col in columnas_esperadas if col not in df.columns]
        
        if columnas_faltantes:
            st.warning(f"⚠️ Algunas columnas esperadas no fueron encontradas: {', '.join(columnas_faltantes)}")
            # Intentar detectar columnas similares (por ejemplo, "Litros" en lugar de "Cantidad litros")
            posibles_coincidencias = {}
            for col_faltante in columnas_faltantes:
                for col_existente in df.columns:
                    if isinstance(col_existente, str) and col_faltante.lower() in col_existente.lower():
                        posibles_coincidencias[col_faltante] = col_existente
            
            if posibles_coincidencias:
                st.info("🔄 Se encontraron posibles coincidencias para algunas columnas faltantes:")
                for original, coincidencia in posibles_coincidencias.items():
                    st.info(f"   • '{original}' podría ser '{coincidencia}'")
                    # Renombrar automáticamente
                    if coincidencia in df.columns:
                        df = df.rename(columns={coincidencia: original})
                        st.success(f"✅ Se renombró '{coincidencia}' a '{original}'")
        
        # Limpiar datos: eliminar filas completamente vacías
        df = df.dropna(how='all')
        
        # Preprocesamiento avanzado para columnas de fecha y hora
        if 'Fecha' in df.columns:
            # Detectar formato de fecha
            st.info("🔄 Procesando columna de Fecha...")
            
            # Convertir fechas utilizando múltiples formatos
            fecha_original = df['Fecha'].copy()
            formatos_fecha = [
                '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%m-%d-%Y',
                '%d.%m.%Y', '%Y.%m.%d', '%d %b %Y', '%d %B %Y'
            ]
            
            # Intentar convertir fechas con diferentes formatos
            for formato in formatos_fecha:
                try:
                    df['Fecha'] = pd.to_datetime(df['Fecha'], format=formato, errors='coerce')
                    if df['Fecha'].notna().sum() > len(df) * 0.5:  # Si más del 50% se convierte correctamente
                        break
                except Exception:
                    continue
            
            # Si falló, intentar conversión automática
            if df['Fecha'].isna().all():
                df['Fecha'] = pd.to_datetime(fecha_original, errors='coerce')
            
            # Verificar si se convirtieron correctamente
            pct_fechas_validas = df['Fecha'].notna().mean() * 100
            st.info(f"📅 Fechas convertidas: {pct_fechas_validas:.1f}% válidas")
            
            # Agregar columnas derivadas de fecha
            df['Año'] = df['Fecha'].dt.year
            df['Mes'] = df['Fecha'].dt.month
            df['Mes Nombre'] = df['Fecha'].dt.month_name()
            df['Día'] = df['Fecha'].dt.day
            df['Día Semana'] = df['Fecha'].dt.day_name()
            df['Semana del Año'] = df['Fecha'].dt.isocalendar().week
            df['Trimestre'] = df['Fecha'].dt.quarter
            
            # Columna para día laboral vs fin de semana
            df['Es Fin de Semana'] = df['Fecha'].dt.dayofweek >= 5
        
        if 'Hora' in df.columns:
            # Convertir hora a datetime con manejo avanzado de formatos
            st.info("🔄 Procesando columna de Hora...")
            
            hora_original = df['Hora'].copy()
            formatos_hora = ['%H:%M:%S', '%H:%M', '%I:%M:%S %p', '%I:%M %p']
            
            for formato in formatos_hora:
                try:
                    df['Hora'] = pd.to_datetime(df['Hora'], format=formato, errors='coerce').dt.time
                    if df['Hora'].notna().sum() > len(df) * 0.5:  # Si más del 50% se convierte correctamente
                        break
                except Exception:
                    continue
            
            # Si los formatos específicos fallan, intentar detección automática
            if df['Hora'].isna().all():
                try:
                    df['Hora'] = pd.to_datetime(hora_original, errors='coerce').dt.time
                except Exception:
                    # Como último recurso, intentar extraer horas y minutos
                    try:
                        def extraer_hora(x):
                            if pd.isna(x):
                                return None
                            x = str(x)
                            # Buscar patrones de hora:minuto
                            match = re.search(r'(\d{1,2})[:\.](\d{1,2})', x)
                            if match:
                                hora, minuto = int(match.group(1)), int(match.group(2))
                                if 0 <= hora < 24 and 0 <= minuto < 60:
                                    return datetime.time(hour=hora, minute=minuto)
                            return None
                        
                        df['Hora'] = hora_original.apply(extraer_hora)
                    except Exception as e:
                        st.warning(f"⚠️ No se pudo procesar completamente la columna 'Hora': {e}")
            
            # Verificar si se convirtieron correctamente
            pct_horas_validas = df['Hora'].notna().mean() * 100
            st.info(f"🕒 Horas convertidas: {pct_horas_validas:.1f}% válidas")
            
            # Agregar columna de hora numérica de manera segura
            def convert_to_numeric_hour(x):
                try:
                    if hasattr(x, 'hour') and hasattr(x, 'minute'):
                        return x.hour + x.minute/60
                    return None
                except:
                    return None
                    
            df['Hora Numérica'] = df['Hora'].apply(convert_to_numeric_hour)
            
            # Clasificar por período del día de forma más detallada
            def clasificar_periodo(hora):
                if hora is None:
                    return None
                elif isinstance(hora, datetime.time):
                    hora_num = hora.hour
                else:
                    try:
                        hora_num = int(hora)
                    except:
                        return None
                
                if 0 <= hora_num < 6:
                    return 'Madrugada'
                elif 6 <= hora_num < 12:
                    return 'Mañana'
                elif 12 <= hora_num < 18:
                    return 'Tarde'
                else:
                    return 'Noche'
            
            df['Período'] = df['Hora'].apply(clasificar_periodo)
            
            # Clasificar en horas pico y no pico (personalizable según necesidades)
            def es_hora_pico(hora):
                if hora is None:
                    return False
                elif isinstance(hora, datetime.time):
                    hora_num = hora.hour
                else:
                    try:
                        hora_num = int(hora)
                    except:
                        return False
                
                # Definir rangos de hora pico (7-9 AM y 5-7 PM)
                return (7 <= hora_num < 10) or (17 <= hora_num < 20)
            
            df['Hora Pico'] = df['Hora'].apply(es_hora_pico)
        
        # Limpieza y conversión avanzada de columnas numéricas
        numeric_columns = [
            'Cantidad litros', 'Odómetro', 'Estanque', 
            'Valor', 'Precio unitario', 'Kilómetros'
        ]
        
        for col in numeric_columns:
            if col in df.columns:
                st.info(f"🔄 Procesando columna numérica: {col}")
                # Guardar valores originales
                original_values = df[col].copy()
                
                # Intento inicial: conversión directa
                df[col] = pd.to_numeric(df[col], errors='coerce')
                
                # Si hay demasiados NaN, intentar limpiar y reformatear
                if df[col].isna().mean() > 0.3:  # Si más del 30% son NaN
                    # Limpiar formatos de número con comas, puntos, etc.
                    def clean_numeric(x):
                        if pd.isna(x):
                            return np.nan
                        if isinstance(x, (int, float)):
                            return x
                        try:
                            # Convertir a string y limpiar
                            x = str(x).replace(',', '.').strip()
                            # Manejar formatos con separador de miles (1.234,56 → 1234.56)
                            if '.' in x and ',' in x:
                                if x.find('.') < x.find(','):  # Formato 1.234,56
                                    x = x.replace('.', '').replace(',', '.')
                            return float(x)
                        except:
                            return np.nan
                    
                    df[col] = original_values.apply(clean_numeric)
                
                # Estadísticas de conversión
                pct_numeros_validos = df[col].notna().mean() * 100
                st.info(f"🔢 {col}: {pct_numeros_validos:.1f}% valores numéricos válidos")
        
        # Detección mejorada de malas cargas
        st.info("🔄 Identificando malas cargas...")
        df['Mala Carga'] = False  # Inicializar columna
        
        # Método 1: Detectar por texto en columna "Tipo"
        if 'Tipo' in df.columns:
            patron_mala_carga = r'(masiva|masa|mala|carga\s*masiva|carga\s*mala)'
            df['Mala Carga'] = df['Mala Carga'] | (df['Tipo'].str.contains(patron_mala_carga, case=False, na=False, regex=True))
        
        # Método 2: Buscar en la última columna (como solicitado)
        ultima_columna = df.columns[-1]
        if not ultima_columna in ['Mala Carga', 'Sobreconsumo', 'Outlier Extremo']:
            patron_ultima_columna = r'(masiva|masa|mala|carga\s*masiva|carga\s*mala)'
            df['Mala Carga'] = df['Mala Carga'] | (df[ultima_columna].astype(str).str.contains(patron_ultima_columna, case=False, na=False, regex=True))
        
        # Método 3: Buscar en todas las columnas de texto
        for col in df.columns:
            if df[col].dtype == 'object' and col not in ['Fecha', 'Hora', 'Mala Carga']:
                # Comprobar si la columna contiene "Carga Masiva" explícitamente
                if df[col].astype(str).str.contains('Carga Masiva', case=False, na=False).any():
                    patron = r'(carga\s*masiva|carga\s*mala)'
                    df['Mala Carga'] = df['Mala Carga'] | (df[col].astype(str).str.contains(patron, case=False, na=False, regex=True))
        
        # Método 4: Detección por anomalía en cantidades
        if 'Cantidad litros' in df.columns:
            # Detectar valores extremadamente altos (outliers)
            q1 = df['Cantidad litros'].quantile(0.25)
            q3 = df['Cantidad litros'].quantile(0.75)
            iqr = q3 - q1
            upper_bound = q3 + 3 * iqr  # Umbral más conservador para evitar falsos positivos
            
            # Marcar como posibles malas cargas los valores extremos
            df['Posible Mala Carga'] = df['Cantidad litros'] > upper_bound
            
            # Contar cuántas posibles malas cargas se detectaron
            n_nuevas_malas = df['Posible Mala Carga'].sum()
            if n_nuevas_malas > 0:
                st.info(f"ℹ️ Se detectaron {n_nuevas_malas} posibles malas cargas adicionales por valores extremos.")
                
                # Solo asignar a Mala Carga si no están ya marcadas
                df.loc[df['Posible Mala Carga'] & ~df['Mala Carga'], 'Mala Carga'] = True
        
        # Número total de malas cargas detectadas
        n_malas_cargas = df['Mala Carga'].sum()
        st.info(f"🛑 Total de malas cargas detectadas: {n_malas_cargas} ({n_malas_cargas/len(df)*100:.2f}%)")
        
        # Detección avanzada de anomalías en consumo por modelo
        if 'Modelo chasis' in df.columns and 'Cantidad litros' in df.columns:
            st.info("🔄 Analizando patrones de consumo por modelo...")
            
            # Limpiar y normalizar nombres de modelos
            df['Modelo chasis'] = df['Modelo chasis'].astype(str).str.strip().str.upper()
            
            # Agrupar modelos similares
            def normalizar_modelo(modelo):
                if pd.isna(modelo) or modelo == 'nan':
                    return 'SIN MODELO'
                modelo = str(modelo).upper().strip()
                # Eliminar espacios excesivos
                modelo = re.sub(r'\s+', ' ', modelo)
                # Buscar patrones comunes
                for patron in ['MERCEDES', 'BENZ', 'MB']:
                    if patron in modelo:
                        return 'MERCEDES BENZ'
                for patron in ['VOLVO', 'VOL']:
                    if patron in modelo:
                        return 'VOLVO'
                for patron in ['SCANIA', 'SCA']:
                    if patron in modelo:
                        return 'SCANIA'
                # Si no se ha normalizado, devolver el original
                return modelo
            
            df['Modelo Normalizado'] = df['Modelo chasis'].apply(normalizar_modelo)
            
            # Estadísticas por modelo normalizado
            modelo_stats = df.groupby('Modelo Normalizado')['Cantidad litros'].agg(['mean', 'std', 'count']).reset_index()
            modelo_stats.columns = ['Modelo Normalizado', 'Promedio Litros', 'Desviación Estándar', 'Cantidad']
            
            # Filtrar modelos con suficientes datos
            modelo_stats = modelo_stats[modelo_stats['Cantidad'] >= 5]
            
            # Crear diccionarios para mapear estadísticas a cada registro
            promedio_por_modelo = dict(zip(modelo_stats['Modelo Normalizado'], modelo_stats['Promedio Litros']))
            std_por_modelo = dict(zip(modelo_stats['Modelo Normalizado'], modelo_stats['Desviación Estándar']))
            
            # Asignar estadísticas a cada registro
            df['Promedio Modelo'] = df['Modelo Normalizado'].map(promedio_por_modelo)
            df['Desviación Modelo'] = df['Modelo Normalizado'].map(std_por_modelo)
            
            # Evitar divisiones por cero
            df['Z-Score'] = np.where(
                (df['Desviación Modelo'].notna()) & (df['Desviación Modelo'] > 0),
                (df['Cantidad litros'] - df['Promedio Modelo']) / df['Desviación Modelo'],
                np.nan
            )
            
            # Mejorar detección de sobreconsumo con umbrales adaptativos
            df['Umbral Sobreconsumo'] = df['Promedio Modelo'] + 2 * df['Desviación Modelo']
            df['Sobreconsumo'] = (df['Cantidad litros'] > df['Umbral Sobreconsumo']) & (~df['Mala Carga'])
            
            # Detección de outliers extremos por modelo
            st.info("🔄 Detectando outliers extremos...")
            df['Outlier Extremo'] = False  # Inicializar columna
            
            # Usar Isolation Forest para cada modelo
            for modelo in df['Modelo Normalizado'].dropna().unique():
                modelo_df = df[df['Modelo Normalizado'] == modelo]
                if len(modelo_df) >= 10:  # Necesitamos suficientes datos
                    try:
                        # Preparar datos para el modelo
                        litros = modelo_df['Cantidad litros'].values.reshape(-1, 1)
                        litros_validos = ~np.isnan(litros).any(axis=1)
                        
                        if np.sum(litros_validos) >= 10:
                            litros_clean = litros[litros_validos]
                            
                            # Ajustar modelo de detección de anomalías
                            iso = IsolationForest(contamination=0.05, random_state=42)
                            outliers = iso.fit_predict(litros_clean)
                            
                            # Identificar índices de outliers
                            idx_in_subset = np.where(outliers == -1)[0]
                            idx_in_original = modelo_df.index[litros_validos][idx_in_subset]
                            
                            # Marcar outliers
                            df.loc[idx_in_original, 'Outlier Extremo'] = True
                    except Exception as e:
                        st.warning(f"⚠️ No se pudo calcular outliers para el modelo {modelo}: {e}")
        else:
            # Detección simple si no hay datos de modelo
            if 'Cantidad litros' in df.columns:
                try:
                    # Calcular estadísticas generales
                    mean_litros = df['Cantidad litros'].mean()
                    std_litros = df['Cantidad litros'].std()
                    
                    # Evitar operaciones con NaN
                    if not pd.isna(mean_litros) and not pd.isna(std_litros) and std_litros > 0:
                        # Definir umbral de sobreconsumo
                        df['Umbral Sobreconsumo'] = mean_litros + 2 * std_litros
                        df['Sobreconsumo'] = (df['Cantidad litros'] > df['Umbral Sobreconsumo']) & (~df['Mala Carga'])
                        
                        # Detección de outliers extremos global
                        litros = df['Cantidad litros'].dropna().values.reshape(-1, 1)
                        if len(litros) >= 10:
                            iso = IsolationForest(contamination=0.05, random_state=42)
                            outliers = iso.fit_predict(litros)
                            idx_in_subset = np.where(outliers == -1)[0]
                            outlier_values = df['Cantidad litros'].dropna().iloc[idx_in_subset].index
                            df.loc[outlier_values, 'Outlier Extremo'] = True
                    else:
                        df['Sobreconsumo'] = False
                except Exception as e:
                    st.warning(f"⚠️ Error en detección de anomalías: {e}")
                    df['Sobreconsumo'] = False
            else:
                df['Sobreconsumo'] = False
                
            df['Outlier Extremo'] = False
        
        # Conteo de anomalías detectadas
        n_sobreconsumo = df['Sobreconsumo'].sum()
        n_outliers = df['Outlier Extremo'].sum()
        st.info(f"⚠️ Anomalías detectadas: {n_sobreconsumo} sobreconsumos, {n_outliers} outliers extremos")
        
        # Cálculo avanzado de rendimiento (si hay información de odómetro)
        if 'Número interno' in df.columns and 'Odómetro' in df.columns and 'Cantidad litros' in df.columns:
            st.info("🔄 Calculando métricas de rendimiento...")
            
            # Limpiar y normalizar identificadores de buses
            df['Número interno'] = df['Número interno'].astype(str).str.strip().str.upper()
            
            # Ordenar primero para calcular correctamente los cambios en odómetro
            df = df.sort_values(['Número interno', 'Fecha', 'Hora'])
            
            # Calcular diferencias de odómetro entre cargas del mismo bus
            df['Odómetro Anterior'] = df.groupby('Número interno')['Odómetro'].shift(1)
            df['Kilómetros Recorridos'] = df['Odómetro'] - df['Odómetro Anterior']
            
            # Filtrar valores no válidos con criterios más detallados
            # Mínimo de kilómetros (para evitar errores de registro)
            min_km = 5
            # Máximo razonable de kilómetros entre cargas
            max_km = 1200
            
            # Identificar valores sospechosos
            df['Km Sospechosos'] = (df['Kilómetros Recorridos'] < min_km) | (df['Kilómetros Recorridos'] > max_km)
            
            # Aplicar filtros
            df.loc[df['Km Sospechosos'], 'Kilómetros Recorridos'] = np.nan
            
            # Cálculo avanzado de rendimiento (km/l)
            df['Rendimiento'] = np.where(
                (df['Cantidad litros'] > 0) & (df['Kilómetros Recorridos'].notna()),
                df['Kilómetros Recorridos'] / df['Cantidad litros'],
                np.nan
            )
            
            # Filtrar rendimientos no razonables (basados en distribución estadística)
            q1_rend = df['Rendimiento'].quantile(0.10)
            q3_rend = df['Rendimiento'].quantile(0.90)
            iqr_rend = q3_rend - q1_rend
            lower_bound_rend = max(0.5, q1_rend - 1.5 * iqr_rend)  # Mínimo de 0.5 km/l
            upper_bound_rend = min(20, q3_rend + 1.5 * iqr_rend)  # Máximo de 20 km/l
            
            # Marcar valores anómalos
            df['Rendimiento Anómalo'] = (df['Rendimiento'] < lower_bound_rend) | (df['Rendimiento'] > upper_bound_rend)
            
            # Filtrar
            df.loc[df['Rendimiento Anómalo'], 'Rendimiento'] = np.nan
            
            # Estadísticas de cálculo
            pct_rendimiento_valido = df['Rendimiento'].notna().mean() * 100
            st.info(f"🚌 Rendimiento: {pct_rendimiento_valido:.1f}% de registros con rendimiento válido")
            
            if 'Modelo Normalizado' in df.columns:
                # Calcular rendimiento promedio por modelo normalizado
                rendimiento_por_modelo = df.groupby('Modelo Normalizado')['Rendimiento'].mean().to_dict()
                df['Rendimiento Promedio Modelo'] = df['Modelo Normalizado'].map(rendimiento_por_modelo)
                
                # Calcular desviación de rendimiento con manejo de NaN
                def calcular_desviacion(row):
                    if pd.isna(row['Rendimiento']) or pd.isna(row['Rendimiento Promedio Modelo']) or row['Rendimiento Promedio Modelo'] == 0:
                        return np.nan
                    return ((row['Rendimiento'] - row['Rendimiento Promedio Modelo']) / row['Rendimiento Promedio Modelo']) * 100
                
                df['Desviación Rendimiento'] = df.apply(calcular_desviacion, axis=1)
                
                # Categorizar eficiencia basada en la desviación
                def categorizar_eficiencia(desviacion):
                    if pd.isna(desviacion):
                        return None
                    elif desviacion < -15:
                        return 'Baja'
                    elif desviacion < -5:
                        return 'Regular'
                    elif desviacion < 5:
                        return 'Normal'
                    elif desviacion < 15:
                        return 'Buena'
                    else:
                        return 'Excelente'
                
                df['Categoría Eficiencia'] = df['Desviación Rendimiento'].apply(categorizar_eficiencia)
        
        # Añadir información enriquecida de terminales
        if 'Terminal' in df.columns:
            st.info("🔄 Procesando información de terminales...")
            
            # Limpiar y normalizar nombres de terminal
            df['Terminal'] = df['Terminal'].astype(str).str.strip().str.upper()
            
            # Normalizamos nombres similares
            def normalizar_terminal(terminal):
                if pd.isna(terminal) or terminal == 'nan':
                    return 'SIN TERMINAL'
                terminal = str(terminal).upper().strip()
                # Eliminar espacios excesivos y caracteres especiales
                terminal = re.sub(r'\s+', ' ', terminal)
                return terminal
            
            df['Terminal Normalizada'] = df['Terminal'].apply(normalizar_terminal)
            
            # Conteo y estadísticas por terminal
            terminal_count = df.groupby('Terminal Normalizada').size().to_dict()
            terminal_avg_consumo = df.groupby('Terminal Normalizada')['Cantidad litros'].mean().to_dict() if 'Cantidad litros' in df.columns else {}
            
            df['Cargas Terminal'] = df['Terminal Normalizada'].map(terminal_count)
            df['Consumo Promedio Terminal'] = df['Terminal Normalizada'].map(terminal_avg_consumo)
            
            # Añadir ranking de terminales por volumen
            terminal_ranking = df.groupby('Terminal Normalizada').size().sort_values(ascending=False).reset_index()
            terminal_ranking['Ranking'] = range(1, len(terminal_ranking) + 1)
            terminal_ranking_dict = dict(zip(terminal_ranking['Terminal Normalizada'], terminal_ranking['Ranking']))
            df['Ranking Terminal'] = df['Terminal Normalizada'].map(terminal_ranking_dict)
        
        # Análisis del llenado de estanque con detección mejorada
        if 'Cantidad litros' in df.columns:
            st.info("🔄 Analizando patrones de llenado...")
            
            # Método 1: Si existe columna "Estanque" o similar
            if 'Estanque' in df.columns and df['Estanque'].notna().any():
                # Evitar divisiones por cero
                df['Porcentaje Llenado'] = np.where(
                    df['Estanque'] > 0,
                    (df['Cantidad litros'] / df['Estanque']) * 100,
                    np.nan
                )
            # Método 2: Estimación por modelo
            elif 'Modelo Normalizado' in df.columns:
                # Estimar capacidad de estanque por modelo
                estanque_estimado = {}
                for modelo in df['Modelo Normalizado'].dropna().unique():
                    # Tomar el percentil 95 de cargas como estimación del estanque
                    subset = df[df['Modelo Normalizado'] == modelo]['Cantidad litros']
                    if len(subset) >= 10:
                        estanque_estimado[modelo] = subset.quantile(0.95)
                
                df['Estanque Estimado'] = df['Modelo Normalizado'].map(estanque_estimado)
                
                # Calcular porcentaje relativo al estanque estimado
                df['Porcentaje Llenado'] = np.where(
                    df['Estanque Estimado'].notna() & (df['Estanque Estimado'] > 0),
                    (df['Cantidad litros'] / df['Estanque Estimado']) * 100,
                    np.nan
                )
            # Método 3: Estimación por bus
            elif 'Número interno' in df.columns:
                # Estimar capacidad de estanque por bus
                estanque_por_bus = {}
                for bus in df['Número interno'].dropna().unique():
                    # Tomar el percentil 95 de cargas como estimación del estanque
                    subset = df[df['Número interno'] == bus]['Cantidad litros']
                    if len(subset) >= 5:
                        estanque_por_bus[bus] = subset.quantile(0.95)
                
                df['Estanque Estimado'] = df['Número interno'].map(estanque_por_bus)
                
                # Calcular porcentaje relativo al estanque estimado
                df['Porcentaje Llenado'] = np.where(
                    df['Estanque Estimado'].notna() & (df['Estanque Estimado'] > 0),
                    (df['Cantidad litros'] / df['Estanque Estimado']) * 100,
                    np.nan
                )
            
            # Clasificar nivel de llenado si se pudo calcular
            if 'Porcentaje Llenado' in df.columns:
                def clasificar_llenado(porcentaje):
                    if pd.isna(porcentaje):
                        return None
                    elif porcentaje < 25:
                        return 'Muy Bajo'
                    elif porcentaje < 50:
                        return 'Bajo'
                    elif porcentaje < 75:
                        return 'Medio'
                    elif porcentaje < 90:
                        return 'Alto'
                    else:
                        return 'Completo'
                
                df['Nivel Llenado'] = df['Porcentaje Llenado'].apply(clasificar_llenado)
                
                # Patrón de llenado: completo vs parcial
                df['Llenado Completo'] = df['Porcentaje Llenado'] >= 85 if 'Porcentaje Llenado' in df.columns else False
        
        # Análisis de personal mejorado
        if 'Nombre conductor' in df.columns:
            st.info("🔄 Analizando patrones de conductores...")
            
            # Limpiar y normalizar nombres
            df['Nombre conductor'] = df['Nombre conductor'].astype(str).str.strip().str.upper()
            
            # Normalizar
            def normalizar_nombre(nombre):
                if pd.isna(nombre) or nombre.lower() == 'nan':
                    return 'SIN NOMBRE'
                nombre = str(nombre).upper().strip()
                # Eliminar espacios excesivos y normalizar
                nombre = re.sub(r'\s+', ' ', nombre)
                return nombre
            
            df['Conductor Normalizado'] = df['Nombre conductor'].apply(normalizar_nombre)
            
            # Estadísticas por conductor
            conductor_stats = {}
            
            for conductor in df['Conductor Normalizado'].dropna().unique():
                subset = df[df['Conductor Normalizado'] == conductor]
                
                if 'Sobreconsumo' in df.columns:
                    tasa_sobreconsumo = subset['Sobreconsumo'].mean() * 100
                    conductor_stats[conductor] = tasa_sobreconsumo
            
            df['Tasa Sobreconsumo Conductor'] = df['Conductor Normalizado'].map(conductor_stats)
            
            # Categorizando conductores
            def categorizar_conductor(tasa):
                if pd.isna(tasa):
                    return None
                elif tasa < 2:
                    return 'Excelente'
                elif tasa < 5:
                    return 'Bueno'
                elif tasa < 10:
                    return 'Regular'
                else:
                    return 'Atención Requerida'
            
            if 'Tasa Sobreconsumo Conductor' in df.columns:
                df['Categoría Conductor'] = df['Tasa Sobreconsumo Conductor'].apply(categorizar_conductor)
        
        # Análisis por planillero
        if 'Nombre Planillero' in df.columns:
            st.info("🔄 Analizando patrones de planilleros...")
            
            # Limpiar y normalizar nombres
            df['Nombre Planillero'] = df['Nombre Planillero'].astype(str).str.strip().str.upper()
            
            # Normalizar
            df['Planillero Normalizado'] = df['Nombre Planillero'].apply(normalizar_nombre)
            
            # Estadísticas por planillero
            planillero_stats = {}
            
            for planillero in df['Planillero Normalizado'].dropna().unique():
                subset = df[df['Planillero Normalizado'] == planillero]
                
                if 'Mala Carga' in df.columns:
                    tasa_malas_cargas = subset['Mala Carga'].mean() * 100
                    planillero_stats[planillero] = tasa_malas_cargas
            
            df['Tasa Malas Cargas Planillero'] = df['Planillero Normalizado'].map(planillero_stats)
            
            # Categorizando planilleros
            def categorizar_planillero(tasa):
                if pd.isna(tasa):
                    return None
                elif tasa < 1:
                    return 'Excelente'
                elif tasa < 3:
                    return 'Bueno'
                elif tasa < 7:
                    return 'Regular'
                else:
                    return 'Atención Requerida'
            
            if 'Tasa Malas Cargas Planillero' in df.columns:
                df['Categoría Planillero'] = df['Tasa Malas Cargas Planillero'].apply(categorizar_planillero)
        
        # Calcular KPIs adicionales y métricas avanzadas
        if 'Cantidad litros' in df.columns and 'Fecha' in df.columns and df['Fecha'].notna().any():
            try:
                # Agrupar por día para análisis temporal
                litros_por_dia = df.groupby(df['Fecha'].dt.date)['Cantidad litros'].sum().reset_index()
                litros_por_dia.columns = ['Fecha', 'Total Litros']
                
                # Agrupar por semana para tendencias más estables
                df['Semana'] = df['Fecha'].dt.strftime('%Y-%U')
                litros_por_semana = df.groupby('Semana')['Cantidad litros'].sum().reset_index()
                
                # Agrupar por mes para tendencias a largo plazo
                df['Mes'] = df['Fecha'].dt.strftime('%Y-%m')
                litros_por_mes = df.groupby('Mes')['Cantidad litros'].sum().reset_index()
                
                # Guardar en session_state para usar en análisis posterior
                st.session_state['litros_por_dia'] = litros_por_dia
                st.session_state['litros_por_semana'] = litros_por_semana
                st.session_state['litros_por_mes'] = litros_por_mes
                
                # Análisis de tendencia con regresión robusta
                if len(litros_por_dia) > 7:
                    X = np.array(range(len(litros_por_dia))).reshape(-1, 1)
                    y = litros_por_dia['Total Litros'].values
                    
                    # Regresión robusta (menos sensible a outliers)
                    try:
                        modelo_robusto = HuberRegressor()
                        modelo_robusto.fit(X, y)
                        tendencia_robusta = modelo_robusto.coef_[0]
                        st.session_state['tendencia_consumo_robusta'] = tendencia_robusta
                    except:
                        # En caso de error, usar regresión lineal estándar
                        modelo = LinearRegression()
                        modelo.fit(X, y)
                        tendencia = modelo.coef_[0]
                        st.session_state['tendencia_consumo'] = tendencia
                    
                    st.session_state['promedio_diario'] = litros_por_dia['Total Litros'].mean()
                    
                    # Guardar información adicional para análisis
                    st.session_state['dias_analizados'] = len(litros_por_dia)
                    st.session_state['consumo_maximo'] = litros_por_dia['Total Litros'].max()
                    st.session_state['consumo_minimo'] = litros_por_dia['Total Litros'].min()
                    st.session_state['consumo_mediana'] = litros_por_dia['Total Litros'].median()
                    st.session_state['consumo_std'] = litros_por_dia['Total Litros'].std()
                    
                    # Calcular estacionalidad por día de semana
                    if 'Día Semana' in df.columns:
                        consumo_por_dia_semana = df.groupby('Día Semana')['Cantidad litros'].mean().to_dict()
                        st.session_state['consumo_por_dia_semana'] = consumo_por_dia_semana
            except Exception as e:
                st.warning(f"⚠️ No se pudieron calcular algunos KPIs de consumo diario: {e}")
        
        # Análisis de distancia recorrida y eficiencia
        if 'Kilómetros Recorridos' in df.columns and 'Número interno' in df.columns:
            try:
                # Kilometraje por bus
                km_por_bus = df.groupby('Número interno')['Kilómetros Recorridos'].sum().reset_index()
                km_por_bus.columns = ['Número interno', 'Km Totales']
                
                # Top buses por kilometraje
                top_buses_km = km_por_bus.sort_values('Km Totales', ascending=False).head(20)
                st.session_state['top_buses_km'] = top_buses_km
                
                # Promedio de kilómetros diarios por bus
                if 'Fecha' in df.columns:
                    # Calcular días entre primera y última fecha por bus
                    fecha_min_max = df.groupby('Número interno')['Fecha'].agg(['min', 'max']).reset_index()
                    fecha_min_max['Dias'] = (fecha_min_max['max'] - fecha_min_max['min']).dt.days + 1
                    fecha_min_max['Dias'] = fecha_min_max['Dias'].replace(0, 1)  # Evitar división por cero
                    
                    # Unir con km totales
                    km_diarios = pd.merge(km_por_bus, fecha_min_max[['Número interno', 'Dias']], on='Número interno')
                    km_diarios['Km Diarios'] = km_diarios['Km Totales'] / km_diarios['Dias']
                    
                    st.session_state['km_diarios_por_bus'] = km_diarios
            except Exception as e:
                st.warning(f"⚠️ Error en análisis de distancia: {e}")
        
        # Detectar patrones y correlaciones adicionales
        if 'Cantidad litros' in df.columns and 'Hora Numérica' in df.columns:
            try:
                # Correlación entre hora y cantidad de litros
                correlacion_hora_litros = df['Cantidad litros'].corr(df['Hora Numérica'])
                st.session_state['correlacion_hora_litros'] = correlacion_hora_litros
                
                # Comparar cargas en días laborales vs fines de semana
                if 'Es Fin de Semana' in df.columns:
                    promedio_laboral = df[~df['Es Fin de Semana']]['Cantidad litros'].mean()
                    promedio_finde = df[df['Es Fin de Semana']]['Cantidad litros'].mean()
                    st.session_state['promedio_laboral'] = promedio_laboral
                    st.session_state['promedio_finde'] = promedio_finde
                    st.session_state['diff_finde_laboral'] = (promedio_finde / promedio_laboral - 1) * 100
            except Exception as e:
                st.warning(f"⚠️ Error en análisis de patrones: {e}")
        
        # Tiempo de ejecución
        end_time = time.time()
        execution_time = end_time - start_time
        st.info(f"⏱️ Tiempo de procesamiento: {execution_time:.2f} segundos")
        
        # Mensaje de éxito
        total_outliers = df['Outlier Extremo'].sum() + df['Sobreconsumo'].sum()
        
        st.success(f"""
        ✅ Archivo cargado y analizado exitosamente:
        - {len(df)} registros procesados
        - {df['Mala Carga'].sum()} malas cargas detectadas
        - {total_outliers} anomalías identificadas
        """)
        
        return df
        
    except Exception as e:
        st.error(f"❌ Error al procesar el archivo: {e}")
        st.error("Revise que el archivo tenga el formato correcto y todas las columnas necesarias.")
        import traceback
        st.error(f"Detalles adicionales: {traceback.format_exc()}")
        return None

# Función para generar gráfico de evolución temporal mejorado
def plot_time_series(df, y_column, title, color=COLORS['primary'], show_trend=True, show_annotations=True, range_selector=True):
    if 'Fecha' not in df.columns or y_column not in df.columns:
        return None
    
    # Agrupar por fecha
    df_agg = df.groupby(df['Fecha'].dt.date)[y_column].sum().reset_index()
    
    # Crear figura base
    fig = px.line(
        df_agg, 
        x='Fecha', 
        y=y_column,
        title=title,
        markers=True,
    )
    
    # Personalizar diseño
    fig.update_traces(
        line=dict(color=color, width=3),
        marker=dict(color=color, size=8)
    )
    
    # Añadir línea de tendencia si hay suficientes datos
    if show_trend and len(df_agg) > 2:
        X = np.array(range(len(df_agg))).reshape(-1, 1)
        y = df_agg[y_column].values
        
        try:
            # Intentar primero con regresión robusta
            modelo = HuberRegressor()
            modelo.fit(X, y)
            tendencia = modelo.coef_[0]
            tendencia_y = modelo.predict(X)
        except:
            # Fallback a regresión lineal estándar
            modelo = LinearRegression()
            modelo.fit(X, y)
            tendencia = modelo.coef_[0]
            tendencia_y = modelo.predict(X)
        
        # Añadir línea de tendencia
        fig.add_trace(
            go.Scatter(
                x=df_agg['Fecha'],
                y=tendencia_y,
                mode='lines',
                name='Tendencia',
                line=dict(color='red', width=2, dash='dash')
            )
        )
        
        # Añadir anotación con la tendencia
        if show_annotations:
            promedio = np.mean(y)
            porcentaje_cambio = (tendencia / promedio) * 100
            
            tendencia_texto = f"Tendencia: {'+' if tendencia > 0 else ''}{tendencia:.2f} ({'+' if tendencia > 0 else ''}{porcentaje_cambio:.2f}% diario)"
            
            fig.add_annotation(
                x=df_agg['Fecha'].iloc[-1],
                y=tendencia_y[-1],
                text=tendencia_texto,
                showarrow=True,
                arrowhead=2,
                arrowsize=1,
                arrowwidth=2,
                arrowcolor='red',
                ax=50,
                ay=-30
            )
    
    # Añadir rangos de selección
    if range_selector:
        fig.update_layout(
            xaxis=dict(
                rangeselector=dict(
                    buttons=list([
                        dict(count=7, label="7d", step="day", stepmode="backward"),
                        dict(count=1, label="1m", step="month", stepmode="backward"),
                        dict(count=3, label="3m", step="month", stepmode="backward"),
                        dict(step="all")
                    ])
                ),
                rangeslider=dict(visible=True),
                type="date"
            )
        )
    
    fig.update_layout(
        xaxis_title="Fecha",
        yaxis_title=y_column,
        plot_bgcolor='white',
        hovermode='x unified',
        height=450,
        margin=dict(l=40, r=40, t=60, b=40),
    )
    
    return fig

# Función mejorada para detectar anomalías
def detect_anomalies(df, column, group_by=None, threshold=3, method='zscore'):
    """
    Detecta anomalías usando diversos métodos.
    
    Args:
        df: DataFrame con los datos
        column: Columna a analizar
        group_by: Columna para agrupar (opcional)
        threshold: Umbral para detectar anomalías
        method: Método de detección ('zscore', 'iqr', 'isolation_forest')
    
    Returns:
        DataFrame con columna 'Es_Anomalía' añadida
    """
    result = df.copy()
    
    # Método por Z-Score (desviaciones estándar)
    if method == 'zscore':
        if group_by:
            for name, group in df.groupby(group_by):
                z_scores = np.abs(stats.zscore(group[column], nan_policy='omit'))
                result.loc[group.index, 'Es_Anomalía'] = z_scores > threshold
        else:
            z_scores = np.abs(stats.zscore(df[column], nan_policy='omit'))
            result['Es_Anomalía'] = z_scores > threshold
    
    # Método por rango intercuartil (IQR)
    elif method == 'iqr':
        if group_by:
            for name, group in df.groupby(group_by):
                q1 = group[column].quantile(0.25)
                q3 = group[column].quantile(0.75)
                iqr = q3 - q1
                lower_bound = q1 - threshold * iqr
                upper_bound = q3 + threshold * iqr
                result.loc[group.index, 'Es_Anomalía'] = (group[column] < lower_bound) | (group[column] > upper_bound)
        else:
            q1 = df[column].quantile(0.25)
            q3 = df[column].quantile(0.75)
            iqr = q3 - q1
            lower_bound = q1 - threshold * iqr
            upper_bound = q3 + threshold * iqr
            result['Es_Anomalía'] = (df[column] < lower_bound) | (df[column] > upper_bound)
    
    # Método por Isolation Forest (machine learning)
    elif method == 'isolation_forest':
        from sklearn.ensemble import IsolationForest
        
        if group_by:
            for name, group in df.groupby(group_by):
                if len(group) > 10:  # Necesitamos suficientes datos
                    try:
                        iso = IsolationForest(contamination=float(threshold/100), random_state=42)
                        result.loc[group.index, 'Es_Anomalía'] = iso.fit_predict(group[[column]]) == -1
                    except:
                        # Si falla, usar z-score como fallback
                        z_scores = np.abs(stats.zscore(group[column], nan_policy='omit'))
                        result.loc[group.index, 'Es_Anomalía'] = z_scores > 3
        else:
            if len(df) > 10:
                try:
                    iso = IsolationForest(contamination=float(threshold/100), random_state=42)
                    result['Es_Anomalía'] = iso.fit_predict(df[[column]].fillna(df[column].mean()).values) == -1
                except:
                    # Si falla, usar z-score como fallback
                    z_scores = np.abs(stats.zscore(df[column], nan_policy='omit'))
                    result['Es_Anomalía'] = z_scores > 3
    
    return result

# Función mejorada para crear mapa de calor por horario
def create_heatmap(df, value_column='Cantidad litros', normalized=False, title=None):
    if 'Día Semana' not in df.columns or 'Hora Numérica' not in df.columns:
        return None
    
    # Mapear días de la semana a números
    dias_orden = {
        'Monday': 0, 'Tuesday': 1, 'Wednesday': 2, 'Thursday': 3, 
        'Friday': 4, 'Saturday': 5, 'Sunday': 6
    }
    
    # Versión en español
    dias_orden_es = {
        'Lunes': 0, 'Martes': 1, 'Miércoles': 2, 'Jueves': 3, 
        'Viernes': 4, 'Sábado': 5, 'Domingo': 6
    }
    
    # Determinar qué mapeo usar
    if df['Día Semana'].iloc[0] in dias_orden:
        df['Día Orden'] = df['Día Semana'].map(dias_orden)
        dias_map = dias_orden
    else:
        df['Día Orden'] = df['Día Semana'].map(dias_orden_es)
        dias_map = dias_orden_es
    
    # Redondear horas a la más cercana
    df['Hora Redondeada'] = np.round(df['Hora Numérica']).astype(int)
    
    # Agrupar por día y hora
    pivot = df.pivot_table(
        index='Día Semana', 
        columns='Hora Redondeada',
        values=value_column,
        aggfunc='mean'
    )
    
    # Normalizar valores (opcional)
    if normalized:
        # Normalizar para cada día de la semana
        pivot = pivot.div(pivot.max(axis=1), axis=0).fillna(0)
    
    # Ordenar días
    if df['Día Semana'].iloc[0] in dias_orden:
        pivot = pivot.reindex(sorted(pivot.index, key=lambda x: dias_orden.get(x, 7)))
    else:
        pivot = pivot.reindex(sorted(pivot.index, key=lambda x: dias_orden_es.get(x, 7)))
    
    # Crear título si no se proporciona
    if title is None:
        if normalized:
            title = f"Mapa de calor normalizado: {value_column} por día y hora"
        else:
            title = f"Mapa de calor: {value_column} por día y hora"
    
    # Crear heatmap
    fig = px.imshow(
        pivot,
        labels=dict(x="Hora del día", y="Día de la semana", color=value_column),
        x=pivot.columns,
        y=pivot.index,
        color_continuous_scale="Viridis",
        aspect="auto"
    )
    
    fig.update_layout(
        title=title,
        height=470,
        margin=dict(l=40, r=40, b=40, t=50),
        plot_bgcolor='white',
        xaxis=dict(
            tickmode='linear',
            tick0=0,
            dtick=2,  # Mostrar etiquetas cada 2 horas
            tickangle=0
        )
    )
    
    # Añadir anotaciones para valores clave
    max_value = pivot.max().max()
    min_value = pivot[pivot > 0].min().min() if pivot[pivot > 0].min().min() is not np.nan else 0
    
    max_idx = np.where(pivot.values == max_value)
    if len(max_idx[0]) > 0 and len(max_idx[1]) > 0:
        row_idx, col_idx = max_idx[0][0], max_idx[1][0]
        day = pivot.index[row_idx]
        hour = pivot.columns[col_idx]
        
        fig.add_annotation(
            x=hour,
            y=day,
            text=f"Máx: {max_value:.2f}",
            showarrow=True,
            arrowhead=2,
            arrowsize=1,
            arrowwidth=2,
            arrowcolor="white",
            font=dict(color="white", size=10),
            bordercolor="white",
            bgcolor="rgba(0,0,0,0.5)",
            borderwidth=1,
            borderpad=4,
            ax=25,
            ay=-25
        )
    
    return fig

# Función mejorada para crear gráfico de barras por categoría
def plot_bar_chart(df, x_column, y_column, title, color=None, horizontal=False, top_n=None, 
                   show_values=True, color_by_value=True, add_mean_line=True):
    if x_column not in df.columns or y_column not in df.columns:
        return None
    
    # Agrupar datos
    df_agg = df.groupby(x_column)[y_column].sum().reset_index()
    
    # Ordenar por valor y limitar a top_n si se especifica
    df_agg = df_agg.sort_values(y_column, ascending=False)
    
    if top_n and len(df_agg) > top_n:
        df_agg = df_agg.head(top_n)
    
    # Crear gráfico
    if color_by_value:
        # Usar un gradiente de color basado en los valores
        if horizontal:
            fig = px.bar(
                df_agg,
                y=x_column,
                x=y_column,
                title=title,
                color=y_column,
                color_continuous_scale='Viridis',
                orientation='h',
                text=y_column if show_values else None
            )
        else:
            fig = px.bar(
                df_agg,
                x=x_column,
                y=y_column,
                title=title,
                color=y_column,
                color_continuous_scale='Viridis',
                text=y_column if show_values else None
            )
    else:
        # Usar un color uniforme
        bar_color = color if color else COLORS['primary']
        
        if horizontal:
            fig = px.bar(
                df_agg,
                y=x_column,
                x=y_column,
                title=title,
                color_discrete_sequence=[bar_color],
                orientation='h',
                text=y_column if show_values else None
            )
        else:
            fig = px.bar(
                df_agg,
                x=x_column,
                y=y_column,
                title=title,
                color_discrete_sequence=[bar_color],
                text=y_column if show_values else None
            )
    
    # Formato para los valores en las barras
    if show_values:
        fig.update_traces(
            texttemplate='%{text:.1f}',
            textposition='outside'
        )
    
    # Añadir línea para el promedio
    if add_mean_line:
        promedio = df_agg[y_column].mean()
        
        if horizontal:
            fig.add_vline(
                x=promedio,
                line_dash="dash",
                line_color="red",
                annotation_text=f"Media: {promedio:.2f}",
                annotation_position="top right"
            )
        else:
            fig.add_hline(
                y=promedio,
                line_dash="dash",
                line_color="red",
                annotation_text=f"Media: {promedio:.2f}",
                annotation_position="top right"
            )
    
    # Personalizar diseño
    fig.update_layout(
        plot_bgcolor='white',
        height=450,
        margin=dict(l=40, r=40, t=60, b=80),
        xaxis_tickangle=-45 if not horizontal and len(df_agg) > 5 else 0
    )
    
    return fig

# Función para crear gráficos circulares mejorados
def plot_pie_chart(df, category_column, value_column=None, title=None, hole=0.4, show_percents=True):
    if category_column not in df.columns:
        return None
    
    # Si no se proporciona columna de valor, se cuenta la frecuencia
    if value_column is None or value_column not in df.columns:
        df_agg = df.groupby(category_column).size().reset_index()
        df_agg.columns = [category_column, 'Count']
        values = 'Count'
        if title is None:
            title = f"Distribución por {category_column}"
    else:
        df_agg = df.groupby(category_column)[value_column].sum().reset_index()
        values = value_column
        if title is None:
            title = f"{value_column} por {category_column}"
    
    # Ordenar por valor (descendente)
    df_agg = df_agg.sort_values(values, ascending=False)
    
    # Crear gráfico
    fig = px.pie(
        df_agg,
        names=category_column,
        values=values,
        title=title,
        hole=hole,
        color_discrete_sequence=px.colors.sequential.Viridis
    )
    
    # Configurar información textual
    if show_percents:
        fig.update_traces(
            textposition='inside',
            textinfo='percent+label'
        )
    else:
        fig.update_traces(
            textposition='inside',
            textinfo='label+value'
        )
    
    # Estilo
    fig.update_layout(
        plot_bgcolor='white',
        height=450,
        margin=dict(l=20, r=20, t=50, b=20),
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=-0.2,
            xanchor="center",
            x=0.5
        )
    )
    
    return fig

# Función para generar gráficos de correlación
def plot_correlation(df, x_column, y_column, color_column=None, size_column=None, title=None, 
                     add_trendline=True, add_annotations=True):
    """
    Crea un gráfico de dispersión para analizar la correlación entre dos variables.
    
    Args:
        df: DataFrame con los datos
        x_column: Columna para el eje X
        y_column: Columna para el eje Y
        color_column: Columna para colorear puntos (opcional)
        size_column: Columna para el tamaño de los puntos (opcional)
        title: Título del gráfico
        add_trendline: Si se debe añadir línea de tendencia
        add_annotations: Si se deben añadir anotaciones con estadísticas
    """
    if x_column not in df.columns or y_column not in df.columns:
        return None
    
    # Filtrar valores válidos
    df_clean = df.dropna(subset=[x_column, y_column])
    
    if len(df_clean) < 2:
        return None
    
    # Título predeterminado
    if title is None:
        title = f"Correlación entre {x_column} y {y_column}"
    
    # Crear gráfico base
    if color_column is not None and color_column in df.columns:
        if size_column is not None and size_column in df.columns:
            fig = px.scatter(
                df_clean,
                x=x_column,
                y=y_column,
                color=color_column,
                size=size_column,
                title=title,
                opacity=0.7
            )
        else:
            fig = px.scatter(
                df_clean,
                x=x_column,
                y=y_column,
                color=color_column,
                title=title,
                opacity=0.7
            )
    else:
        fig = px.scatter(
            df_clean,
            x=x_column,
            y=y_column,
            title=title,
            opacity=0.7,
            color_discrete_sequence=[COLORS['primary']]
        )
    
    # Añadir línea de tendencia
    if add_trendline and len(df_clean) >= 3:
        try:
            # Calcular coeficiente de correlación
            corr = df_clean[x_column].corr(df_clean[y_column])
            
            # Añadir línea de tendencia
            fig.update_layout(
                shapes=[
                    dict(
                        type='line',
                        xref='x',
                        yref='y',
                        x0=df_clean[x_column].min(),
                        y0=np.polyval(np.polyfit(df_clean[x_column], df_clean[y_column], 1), df_clean[x_column].min()),
                        x1=df_clean[x_column].max(),
                        y1=np.polyval(np.polyfit(df_clean[x_column], df_clean[y_column], 1), df_clean[x_column].max()),
                        line=dict(color='red', width=2, dash='dash'),
                    )
                ]
            )
            
            # Añadir anotación con correlación
            if add_annotations:
                corr_text = f"Correlación: {corr:.3f}"
                fig.add_annotation(
                    x=0.95,
                    y=0.05,
                    xref="paper",
                    yref="paper",
                    text=corr_text,
                    showarrow=False,
                    font=dict(size=12, color="black"),
                    bgcolor="white",
                    bordercolor="black",
                    borderwidth=1,
                    borderpad=4,
                    align="right"
                )
        except Exception as e:
            print(f"Error al añadir línea de tendencia: {e}")
    
    # Personalizar diseño
    fig.update_layout(
        plot_bgcolor='white',
        height=450,
        margin=dict(l=40, r=40, t=60, b=40),
        xaxis=dict(
            title=x_column,
            zeroline=True,
            zerolinewidth=1,
            zerolinecolor='lightgray',
            showgrid=True,
            gridwidth=1,
            gridcolor='lightgray'
        ),
        yaxis=dict(
            title=y_column,
            zeroline=True,
            zerolinewidth=1,
            zerolinecolor='lightgray',
            showgrid=True,
            gridwidth=1,
            gridcolor='lightgray'
        )
    )
    
    return fig

# Función para exportar a Excel con opciones avanzadas
def export_to_excel(df, filename=None, include_plots=False):
    """
    Exporta los análisis a un archivo Excel con formato mejorado.
    
    Args:
        df: DataFrame con los datos
        filename: Nombre del archivo (opcional)
        include_plots: Si se deben incluir gráficos (opcional)
    
    Returns:
        BytesIO con el archivo Excel
    """
    output = BytesIO()
    
    try:
        # Para Excel con estilos
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import BarChart, Reference, LineChart, PieChart, ScatterChart
        from openpyxl.chart.series import DataPoint
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Crear libro de trabajo
        workbook = openpyxl.Workbook()
        
        # Eliminar hoja predeterminada
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        
        # Definir estilos
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4361EE', end_color='4361EE', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        normal_font = Font(name='Arial', size=11)
        normal_alignment = Alignment(horizontal='left', vertical='center')
        
        number_alignment = Alignment(horizontal='right', vertical='center')
        
        border = Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0')
        )
        
        # ----- RESUMEN GENERAL -----
        ws_resumen = workbook.create_sheet(title='Resumen General')
        
        # Título
        ws_resumen.merge_cells('A1:C1')
        ws_resumen['A1'] = 'INFORME DE ANÁLISIS DE COMBUSTIBLE'
        ws_resumen['A1'].font = Font(name='Arial', size=16, bold=True, color='4361EE')
        ws_resumen['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Fecha de generación
        ws_resumen['A2'] = f'Generado: {datetime.datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws_resumen['A2'].font = Font(name='Arial', size=10, italic=True)
        
        # Encabezados
        ws_resumen['A4'] = 'Métrica'
        ws_resumen['B4'] = 'Valor'
        ws_resumen['C4'] = 'Observación'
        
        for cell in ['A4', 'B4', 'C4']:
            ws_resumen[cell].font = header_font
            ws_resumen[cell].fill = header_fill
            ws_resumen[cell].alignment = header_alignment
            ws_resumen[cell].border = border
        
        # Datos de resumen
        metricas = [
            ('Total Cargas', len(df), ''),
            ('Total Litros', f"{df['Cantidad litros'].sum():,.2f}" if 'Cantidad litros' in df.columns else 'N/A', ''),
            ('Promedio Litros/Carga', f"{df['Cantidad litros'].mean():,.2f}" if 'Cantidad litros' in df.columns else 'N/A', ''),
            ('Malas Cargas', f"{df['Mala Carga'].sum()}" if 'Mala Carga' in df.columns else 'N/A', 
             f"({df['Mala Carga'].mean()*100:.2f}%)" if 'Mala Carga' in df.columns else ''),
            ('Cargas con Sobreconsumo', f"{df['Sobreconsumo'].sum()}" if 'Sobreconsumo' in df.columns else 'N/A',
             f"({df['Sobreconsumo'].mean()*100:.2f}%)" if 'Sobreconsumo' in df.columns else ''),
            ('Litros en Sobreconsumo', f"{df.loc[df['Sobreconsumo'] == True, 'Cantidad litros'].sum():,.2f}" if 'Sobreconsumo' in df.columns and 'Cantidad litros' in df.columns else 'N/A', ''),
            ('Total Buses', f"{df['Número interno'].nunique()}" if 'Número interno' in df.columns else 'N/A', ''),
            ('Total Terminales', f"{df['Terminal'].nunique()}" if 'Terminal' in df.columns else 'N/A', ''),
            ('Total Conductores', f"{df['Nombre conductor'].nunique()}" if 'Nombre conductor' in df.columns else 'N/A', ''),
            ('Rendimiento Promedio', f"{df['Rendimiento'].mean():,.2f} km/l" if 'Rendimiento' in df.columns else 'N/A', ''),
            ('Período Analizado', f"{df['Fecha'].min().strftime('%d/%m/%Y')} al {df['Fecha'].max().strftime('%d/%m/%Y')}" if 'Fecha' in df.columns else 'N/A', '')
        ]
        
        for i, (metrica, valor, obs) in enumerate(metricas, start=5):
            ws_resumen[f'A{i}'] = metrica
            ws_resumen[f'B{i}'] = valor
            ws_resumen[f'C{i}'] = obs
            
            # Estilo
            for col in ['A', 'B', 'C']:
                ws_resumen[f'{col}{i}'].font = normal_font
                ws_resumen[f'{col}{i}'].border = border
            
            if i % 2 == 0:  # Filas alternas
                for col in ['A', 'B', 'C']:
                    ws_resumen[f'{col}{i}'].fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        
        # Ajustar anchos de columna
        ws_resumen.column_dimensions['A'].width = 30
        ws_resumen.column_dimensions['B'].width = 20
        ws_resumen.column_dimensions['C'].width = 30
        
        # ----- CONSUMO DIARIO -----
        if 'Fecha' in df.columns and 'Cantidad litros' in df.columns:
            ws_diario = workbook.create_sheet(title='Consumo Diario')
            
            # Agrupar datos
            consumo_diario = df.groupby(df['Fecha'].dt.date).agg({
                'Cantidad litros': ['sum', 'mean', 'count'],
                'Mala Carga': 'sum' if 'Mala Carga' in df.columns else 'count',
                'Sobreconsumo': 'sum' if 'Sobreconsumo' in df.columns else 'count'
            }).reset_index()
            
            # Aplanar multiíndice
            consumo_diario.columns = ['Fecha', 'Total Litros', 'Promedio Litros', 'Cantidad Cargas', 'Malas Cargas', 'Sobreconsumos']
            
            # Calcular porcentajes
            consumo_diario['% Malas Cargas'] = (consumo_diario['Malas Cargas'] / consumo_diario['Cantidad Cargas'] * 100).round(2)
            consumo_diario['% Sobreconsumos'] = (consumo_diario['Sobreconsumos'] / consumo_diario['Cantidad Cargas'] * 100).round(2)
            
            # Ordenar por fecha (descendente)
            consumo_diario = consumo_diario.sort_values('Fecha', ascending=False)
            
            # Escribir encabezados
            encabezados = ['Fecha', 'Total Litros', 'Promedio Litros', 'Cantidad Cargas', 'Malas Cargas', '% Malas', 'Sobreconsumos', '% Sobre']
            for i, encabezado in enumerate(encabezados, start=1):
                cell = ws_diario.cell(row=1, column=i, value=encabezado)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Escribir datos
            for i, row in enumerate(dataframe_to_rows(consumo_diario, index=False, header=False), start=2):
                for j, value in enumerate(row, start=1):
                    if j == 1:  # Fecha
                        cell = ws_diario.cell(row=i, column=j, value=value)
                        cell.number_format = 'DD/MM/YYYY'
                    elif j in [2, 3]:  # Valores numéricos con decimales
                        cell = ws_diario.cell(row=i, column=j, value=value)
                        cell.number_format = '#,##0.00'
                    elif j in [7, 8]:  # Porcentajes
                        cell = ws_diario.cell(row=i, column=j, value=value)
                        cell.number_format = '0.00%'
                    else:  # Valores enteros
                        cell = ws_diario.cell(row=i, column=j, value=value)
                        cell.number_format = '#,##0'
                    
                    # Estilo común
                    cell.font = normal_font
                    cell.border = border
                    cell.alignment = number_alignment if j > 1 else normal_alignment
                
                # Filas alternas
                if i % 2 == 0:
                    for j in range(1, len(encabezados) + 1):
                        ws_diario.cell(row=i, column=j).fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
            
            # Ajustar anchos de columna
            for i, ancho in enumerate([15, 15, 15, 15, 15, 10, 15, 10], start=1):
                ws_diario.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
            
            # Añadir gráfico si se solicita
            if include_plots:
                chart = LineChart()
                chart.title = "Evolución de Consumo Diario"
                chart.style = 2
                chart.x_axis.title = "Fecha"
                chart.y_axis.title = "Litros"
                
                # Invertir orden para gráfico
                chart_data = consumo_diario.sort_values('Fecha')
                
                # Agregar datos al gráfico
                data = Reference(ws_diario, min_col=2, min_row=1, max_row=len(chart_data) + 1, max_col=2)
                cats = Reference(ws_diario, min_col=1, min_row=2, max_row=len(chart_data) + 1)
                
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                
                # Estilo de línea
                s = chart.series[0]
                s.graphicalProperties.line.solidFill = "4361EE"
                s.graphicalProperties.line.width = 20000  # ancho en EMUs
                
                # Añadir gráfico a la hoja
                ws_diario.add_chart(chart, "J2")
        
        # ----- ANÁLISIS POR TERMINAL -----
        if 'Terminal' in df.columns:
            ws_terminal = workbook.create_sheet(title='Por Terminal')
            
            # Agrupar datos
            terminal_stats = df.groupby('Terminal').agg({
                'Cantidad litros': ['sum', 'mean', 'count'] if 'Cantidad litros' in df.columns else 'count',
                'Mala Carga': 'sum' if 'Mala Carga' in df.columns else 'count',
                'Sobreconsumo': 'sum' if 'Sobreconsumo' in df.columns else 'count'
            }).reset_index()
            
            # Aplanar multiíndice
            if 'Cantidad litros' in df.columns:
                terminal_stats.columns = ['Terminal', 'Total Litros', 'Promedio Litros', 'Cantidad Cargas', 'Malas Cargas', 'Sobreconsumos']
                
                # Calcular porcentajes e índice de eficiencia
                terminal_stats['% Malas Cargas'] = (terminal_stats['Malas Cargas'] / terminal_stats['Cantidad Cargas'] * 100).round(2)
                terminal_stats['% Sobreconsumos'] = (terminal_stats['Sobreconsumos'] / terminal_stats['Cantidad Cargas'] * 100).round(2)
                terminal_stats['Índice Eficiencia'] = (100 - (terminal_stats['% Sobreconsumos'] * 0.7 + terminal_stats['% Malas Cargas'] * 0.3)).round(2)
                terminal_stats['Índice Eficiencia'] = terminal_stats['Índice Eficiencia'].clip(0, 100)
                
                # Ordenar por índice de eficiencia (descendente)
                terminal_stats = terminal_stats.sort_values('Índice Eficiencia', ascending=False)
                
                # Escribir encabezados
                encabezados = ['Terminal', 'Total Litros', 'Promedio Litros', 'Cantidad Cargas', 
                             'Malas Cargas', '% Malas', 'Sobreconsumos', '% Sobre', 'Índice Eficiencia']
            else:
                terminal_stats.columns = ['Terminal', 'Cantidad Cargas', 'Malas Cargas', 'Sobreconsumos']
                encabezados = ['Terminal', 'Cantidad Cargas', 'Malas Cargas', 'Sobreconsumos']
                terminal_stats = terminal_stats.sort_values('Cantidad Cargas', ascending=False)
            
            # Escribir encabezados
            for i, encabezado in enumerate(encabezados, start=1):
                cell = ws_terminal.cell(row=1, column=i, value=encabezado)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Escribir datos
            for i, row in enumerate(dataframe_to_rows(terminal_stats, index=False, header=False), start=2):
                for j, value in enumerate(row, start=1):
                    if j == 1:  # Terminal
                        cell = ws_terminal.cell(row=i, column=j, value=value)
                    elif j in [2, 3]:  # Valores numéricos con decimales
                        cell = ws_terminal.cell(row=i, column=j, value=value)
                        cell.number_format = '#,##0.00'
                    elif j in [6, 7, 9]:  # Porcentajes
                        cell = ws_terminal.cell(row=i, column=j, value=value)
                        cell.number_format = '0.00%'
                    else:  # Valores enteros
                        cell = ws_terminal.cell(row=i, column=j, value=value)
                        cell.number_format = '#,##0'
                    
                    # Estilo común
                    cell.font = normal_font
                    cell.border = border
                    cell.alignment = number_alignment if j > 1 else normal_alignment
                
                # Colorear según eficiencia
                if 'Índice Eficiencia' in terminal_stats.columns:
                    eficiencia = row[-1]  # Último valor (Índice Eficiencia)
                    if eficiencia >= 90:
                        color = '43AA8B'  # Verde
                    elif eficiencia >= 75:
                        color = '90BE6D'  # Verde claro
                    elif eficiencia >= 60:
                        color = 'F9C74F'  # Amarillo
                    elif eficiencia >= 40:
                        color = 'F8961E'  # Naranja
                    else:
                        color = 'F94144'  # Rojo
                    
                    ws_terminal.cell(row=i, column=len(encabezados)).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    # Color de texto adaptativo para mejor contraste
                    if eficiencia >= 60:
                        ws_terminal.cell(row=i, column=len(encabezados)).font = Font(name='Arial', size=11, color='000000')
                    else:
                        ws_terminal.cell(row=i, column=len(encabezados)).font = Font(name='Arial', size=11, color='FFFFFF')
                
                # Filas alternas
                if i % 2 == 0:
                    for j in range(1, len(encabezados) + 1):
                        if not (j == len(encabezados) and 'Índice Eficiencia' in terminal_stats.columns):  # No cambiar el color del índice
                            ws_terminal.cell(row=i, column=j).fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
            
            # Ajustar anchos de columna
            for i, ancho in enumerate([30, 15, 15, 15, 15, 10, 15, 10, 18], start=1):
                if i <= len(encabezados):
                    ws_terminal.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
            
            # Añadir gráfico si se solicita y hay suficientes datos
            if include_plots and 'Cantidad litros' in df.columns and len(terminal_stats) > 1:
                chart = BarChart()
                chart.type = "col"
                chart.style = 10
                chart.title = "Consumo Total por Terminal"
                chart.y_axis.title = "Litros"
                
                # Limitar a top 10 terminales para el gráfico
                chart_data = terminal_stats.sort_values('Total Litros', ascending=False).head(10)
                
                # Datos para el gráfico
                data = Reference(ws_terminal, min_col=2, min_row=1, max_row=len(chart_data) + 1, max_col=2)
                cats = Reference(ws_terminal, min_col=1, min_row=2, max_row=len(chart_data) + 1)
                
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                
                # Colores personalizados para cada barra
                colors = ['4361EE', '3A0CA3', 'F72585', '4CC9F0', '4895EF', 'F8961E', 'F94144', '90BE6D', '43AA8B', '577590']
                for i, (idx, color) in enumerate(zip(range(len(chart_data)), colors)):
                    pt = DataPoint(idx=idx)
                    pt.graphicalProperties.solidFill = color
                    chart.series[0].dPt.append(pt)
                
                # Añadir gráfico a la hoja
                ws_terminal.add_chart(chart, "K2")
        
        # ----- ANÁLISIS POR MODELO -----
        if 'Modelo chasis' in df.columns or 'Modelo Normalizado' in df.columns:
            ws_modelo = workbook.create_sheet(title='Por Modelo')
            
            # Determinar qué columna usar
            modelo_col = 'Modelo Normalizado' if 'Modelo Normalizado' in df.columns else 'Modelo chasis'
            
            # Agrupar datos
            modelo_stats = df.groupby(modelo_col).agg({
                'Cantidad litros': ['sum', 'mean', 'count'] if 'Cantidad litros' in df.columns else 'count',
                'Sobreconsumo': 'sum' if 'Sobreconsumo' in df.columns else 'count',
                'Rendimiento': ['mean', 'median', 'std'] if 'Rendimiento' in df.columns else 'count'
            }).reset_index()
            
            # Aplanar multiíndice
            if 'Cantidad litros' in df.columns and 'Rendimiento' in df.columns:
                modelo_stats.columns = ['Modelo', 'Total Litros', 'Promedio Litros', 'Cantidad Cargas', 
                                      'Sobreconsumos', 'Rendimiento Promedio', 'Rendimiento Mediana', 
                                      'Rendimiento Desviación']
                
                # Calcular porcentajes
                modelo_stats['% Sobreconsumos'] = (modelo_stats['Sobreconsumos'] / modelo_stats['Cantidad Cargas'] * 100).round(2)
                
                # Ordenar por rendimiento (descendente)
                modelo_stats = modelo_stats.sort_values('Rendimiento Promedio', ascending=False)
                
                # Escribir encabezados
                encabezados = ['Modelo', 'Total Litros', 'Promedio Litros', 'Cantidad Cargas', 
                             'Sobreconsumos', '% Sobre', 'Rendimiento Promedio', 'Rendimiento Mediana', 
                             'Rendimiento Desviación']
            elif 'Cantidad litros' in df.columns:
                modelo_stats.columns = ['Modelo', 'Total Litros', 'Promedio Litros', 'Cantidad Cargas', 
                                      'Sobreconsumos', 'Rendimiento Count']
                
                # Calcular porcentajes
                modelo_stats['% Sobreconsumos'] = (modelo_stats['Sobreconsumos'] / modelo_stats['Cantidad Cargas'] * 100).round(2)
                
                # Ordenar por total litros
                modelo_stats = modelo_stats.sort_values('Total Litros', ascending=False)
                
                # Escribir encabezados
                encabezados = ['Modelo', 'Total Litros', 'Promedio Litros', 'Cantidad Cargas', 
                             'Sobreconsumos', '% Sobre', 'Rendimiento Count']
            else:
                modelo_stats.columns = ['Modelo', 'Cantidad Cargas', 'Sobreconsumos', 'Rendimiento Count']
                
                # Ordenar por cantidad de cargas
                modelo_stats = modelo_stats.sort_values('Cantidad Cargas', ascending=False)
                
                # Escribir encabezados
                encabezados = ['Modelo', 'Cantidad Cargas', 'Sobreconsumos', 'Rendimiento Count']
            
            # Escribir encabezados
            for i, encabezado in enumerate(encabezados, start=1):
                cell = ws_modelo.cell(row=1, column=i, value=encabezado)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Escribir datos
            for i, row in enumerate(dataframe_to_rows(modelo_stats, index=False, header=False), start=2):
                for j, value in enumerate(row, start=1):
                    if j == 1:  # Modelo
                        cell = ws_modelo.cell(row=i, column=j, value=value)
                    elif j in [2, 3, 7, 8, 9]:  # Valores numéricos con decimales
                        cell = ws_modelo.cell(row=i, column=j, value=value)
                        cell.number_format = '#,##0.00'
                    elif j == 6:  # Porcentajes
                        cell = ws_modelo.cell(row=i, column=j, value=value)
                        cell.number_format = '0.00%'
                    else:  # Valores enteros
                        cell = ws_modelo.cell(row=i, column=j, value=value)
                        cell.number_format = '#,##0'
                    
                    # Estilo común
                    cell.font = normal_font
                    cell.border = border
                    cell.alignment = number_alignment if j > 1 else normal_alignment
                
                # Filas alternas
                if i % 2 == 0:
                    for j in range(1, len(encabezados) + 1):
                        ws_modelo.cell(row=i, column=j).fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
            
            # Ajustar anchos de columna
            for i, ancho in enumerate([30, 15, 15, 15, 15, 10, 20, 20, 20], start=1):
                if i <= len(encabezados):
                    ws_modelo.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
            
            # Añadir gráfico si se solicita
            if include_plots and 'Rendimiento' in df.columns and len(modelo_stats) > 1:
                chart = BarChart()
                chart.type = "col"
                chart.style = 10
                chart.title = "Rendimiento Promedio por Modelo"
                chart.y_axis.title = "km/L"
                
                # Limitar a top 10 modelos para el gráfico
                chart_data = modelo_stats.head(10)
                
                # Datos para el gráfico
                data = Reference(ws_modelo, min_col=7, min_row=1, max_row=len(chart_data) + 1, max_col=7)
                cats = Reference(ws_modelo, min_col=1, min_row=2, max_row=len(chart_data) + 1)
                
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                
                # Añadir gráfico a la hoja
                ws_modelo.add_chart(chart, "K2")
        
        # ----- MALAS CARGAS -----
        if 'Mala Carga' in df.columns:
            malas_cargas = df[df['Mala Carga'] == True]
            if not malas_cargas.empty:
                ws_malas = workbook.create_sheet(title='Malas Cargas')
                
                # Columnas relevantes para mostrar
                cols_relevantes = [col for col in [
                    'Fecha', 'Hora', 'Terminal', 'Número interno', 'Patente', 
                    'Cantidad litros', 'Tipo', 'Nombre Planillero', 'Nombre supervisor', 
                    ultima_columna  # Incluir la última columna como solicitado
                ] if col in malas_cargas.columns]
                
                # Escribir encabezados
                for i, encabezado in enumerate(cols_relevantes, start=1):
                    cell = ws_malas.cell(row=1, column=i, value=encabezado)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Ordenar por fecha (descendente)
                malas_cargas_sorted = malas_cargas[cols_relevantes].sort_values('Fecha', ascending=False) if 'Fecha' in cols_relevantes else malas_cargas[cols_relevantes]
                
                # Escribir datos
                for i, row in enumerate(dataframe_to_rows(malas_cargas_sorted, index=False, header=False), start=2):
                    for j, value in enumerate(row, start=1):
                        if j == 1 and 'Fecha' in cols_relevantes:  # Fecha
                            cell = ws_malas.cell(row=i, column=j, value=value)
                            cell.number_format = 'DD/MM/YYYY'
                        elif j == 2 and 'Hora' in cols_relevantes:  # Hora
                            cell = ws_malas.cell(row=i, column=j, value=value)
                            cell.number_format = 'HH:MM'
                        elif cols_relevantes[j-1] == 'Cantidad litros':  # Litros
                            cell = ws_malas.cell(row=i, column=j, value=value)
                            cell.number_format = '#,##0.00'
                        else:  # Otros valores
                            cell = ws_malas.cell(row=i, column=j, value=value)
                        
                        # Estilo común
                        cell.font = normal_font
                        cell.border = border
                        cell.alignment = normal_alignment
                    
                    # Filas alternas
                    if i % 2 == 0:
                        for j in range(1, len(cols_relevantes) + 1):
                            ws_malas.cell(row=i, column=j).fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                
                # Ajustar anchos de columna
                anchos_columna = {
                    'Fecha': 15,
                    'Hora': 10,
                    'Terminal': 20,
                    'Número interno': 15,
                    'Patente': 15,
                    'Cantidad litros': 15,
                    'Tipo': 20,
                    'Nombre Planillero': 25,
                    'Nombre supervisor': 25
                }
                
                for i, col in enumerate(cols_relevantes, start=1):
                    ancho = anchos_columna.get(col, 20)  # Ancho predeterminado si no está en el diccionario
                    ws_malas.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
                
                # Añadir análisis por Terminal y Planillero
                if 'Terminal' in cols_relevantes:
                    # Añadir espacio
                    ws_malas.cell(row=len(malas_cargas) + 4, column=1, value="ANÁLISIS POR TERMINAL")
                    ws_malas.cell(row=len(malas_cargas) + 4, column=1).font = Font(name='Arial', size=14, bold=True)
                    
                    # Encabezados análisis
                    encabezados_analisis = ['Terminal', 'Malas Cargas', 'Total Cargas', '% Malas Cargas']
                    for i, encabezado in enumerate(encabezados_analisis, start=1):
                        cell = ws_malas.cell(row=len(malas_cargas) + 5, column=i, value=encabezado)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = border
                    
                    # Calcular estadísticas por terminal
                    terminal_counts = malas_cargas['Terminal'].value_counts().reset_index()
                    terminal_counts.columns = ['Terminal', 'Malas Cargas']
                    
                    total_por_terminal = df.groupby('Terminal').size().reset_index()
                    total_por_terminal.columns = ['Terminal', 'Total Cargas']
                    
                    terminal_analysis = pd.merge(terminal_counts, total_por_terminal, on='Terminal')
                    terminal_analysis['Porcentaje'] = (terminal_analysis['Malas Cargas'] / terminal_analysis['Total Cargas'] * 100).round(2)
                    terminal_analysis = terminal_analysis.sort_values('Porcentaje', ascending=False)
                    
                    # Escribir datos de análisis
                    for i, row in enumerate(dataframe_to_rows(terminal_analysis, index=False, header=False), start=len(malas_cargas) + 6):
                        for j, value in enumerate(row, start=1):
                            if j == 1:  # Terminal
                                cell = ws_malas.cell(row=i, column=j, value=value)
                            elif j == 4:  # Porcentaje
                                cell = ws_malas.cell(row=i, column=j, value=value/100)  # Convertir a decimal para formato de porcentaje
                                cell.number_format = '0.00%'
                            else:  # Valores enteros
                                cell = ws_malas.cell(row=i, column=j, value=value)
                                cell.number_format = '#,##0'
                            
                            # Estilo común
                            cell.font = normal_font
                            cell.border = border
                            cell.alignment = number_alignment if j > 1 else normal_alignment
                        
                        # Colorear según porcentaje
                        porcentaje = row[3]  # Porcentaje
                        if porcentaje >= 10:
                            color = 'F94144'  # Rojo (alto)
                        elif porcentaje >= 5:
                            color = 'F8961E'  # Naranja (medio)
                        elif porcentaje >= 2:
                            color = 'F9C74F'  # Amarillo (bajo)
                        else:
                            color = '90BE6D'  # Verde (muy bajo)
                        
                        ws_malas.cell(row=i, column=4).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        
        # ----- SOBRECONSUMO -----
        if 'Sobreconsumo' in df.columns:
            sobreconsumo = df[df['Sobreconsumo'] == True]
            if not sobreconsumo.empty:
                ws_sobre = workbook.create_sheet(title='Sobreconsumo')
                
                # Columnas relevantes para mostrar
                cols_relevantes = [col for col in [
                    'Fecha', 'Hora', 'Terminal', 'Número interno', 'Patente', 'Modelo chasis',
                    'Cantidad litros', 'Z-Score', 'Promedio Modelo', 'Nombre conductor', 'Nombre supervisor'
                ] if col in sobreconsumo.columns]
                
                # Escribir encabezados
                for i, encabezado in enumerate(cols_relevantes, start=1):
                    cell = ws_sobre.cell(row=1, column=i, value=encabezado)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Ordenar por Z-Score (descendente)
                if 'Z-Score' in cols_relevantes:
                    sobreconsumo_sorted = sobreconsumo[cols_relevantes].sort_values('Z-Score', ascending=False)
                else:
                    sobreconsumo_sorted = sobreconsumo[cols_relevantes]
                
                # Escribir datos
                for i, row in enumerate(dataframe_to_rows(sobreconsumo_sorted, index=False, header=False), start=2):
                    for j, value in enumerate(row, start=1):
                        if j == 1 and 'Fecha' in cols_relevantes:  # Fecha
                            cell = ws_sobre.cell(row=i, column=j, value=value)
                            cell.number_format = 'DD/MM/YYYY'
                        elif j == 2 and 'Hora' in cols_relevantes:  # Hora
                            cell = ws_sobre.cell(row=i, column=j, value=value)
                            cell.number_format = 'HH:MM'
                        elif cols_relevantes[j-1] in ['Cantidad litros', 'Z-Score', 'Promedio Modelo']:  # Valores numéricos
                            cell = ws_sobre.cell(row=i, column=j, value=value)
                            cell.number_format = '#,##0.00'
                        else:  # Otros valores
                            cell = ws_sobre.cell(row=i, column=j, value=value)
                        
                        # Estilo común
                        cell.font = normal_font
                        cell.border = border
                        cell.alignment = normal_alignment
                        
                        # Colorear según Z-Score
                        if cols_relevantes[j-1] == 'Z-Score' and value is not None:
                            if value >= 4:
                                color = 'F94144'  # Rojo (extremo)
                            elif value >= 3:
                                color = 'F8961E'  # Naranja (alto)
                            else:
                                color = 'F9C74F'  # Amarillo (moderado)
                            
                            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    
                    # Filas alternas
                    if i % 2 == 0:
                        for j in range(1, len(cols_relevantes) + 1):
                            if cols_relevantes[j-1] != 'Z-Score':  # No cambiar el color del Z-Score
                                ws_sobre.cell(row=i, column=j).fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                
                # Ajustar anchos de columna
                anchos_columna = {
                    'Fecha': 15,
                    'Hora': 10,
                    'Terminal': 20,
                    'Número interno': 15,
                    'Patente': 15,
                    'Modelo chasis': 20,
                    'Cantidad litros': 15,
                    'Z-Score': 10,
                    'Promedio Modelo': 15,
                    'Nombre conductor': 25,
                    'Nombre supervisor': 25
                }
                
                for i, col in enumerate(cols_relevantes, start=1):
                    ancho = anchos_columna.get(col, 20)  # Ancho predeterminado si no está en el diccionario
                    ws_sobre.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
        
        # ----- RENDIMIENTO -----
        if 'Rendimiento' in df.columns:
            rendimiento_df = df.dropna(subset=['Rendimiento'])
            if not rendimiento_df.empty:
                ws_rendimiento = workbook.create_sheet(title='Rendimiento')
                
                # Estadísticas básicas
                stats = {
                    'Rendimiento Promedio': rendimiento_df['Rendimiento'].mean(),
                    'Rendimiento Mediana': rendimiento_df['Rendimiento'].median(),
                    'Rendimiento Mínimo': rendimiento_df['Rendimiento'].min(),
                    'Rendimiento Máximo': rendimiento_df['Rendimiento'].max(),
                    'Desviación Estándar': rendimiento_df['Rendimiento'].std()
                }
                
                # Escribir estadísticas básicas
                ws_rendimiento.cell(row=1, column=1, value="ESTADÍSTICAS DE RENDIMIENTO")
                ws_rendimiento.cell(row=1, column=1).font = Font(name='Arial', size=14, bold=True)
                
                for i, (stat, value) in enumerate(stats.items(), start=2):
                    ws_rendimiento.cell(row=i, column=1, value=stat)
                    ws_rendimiento.cell(row=i, column=2, value=value)
                    ws_rendimiento.cell(row=i, column=2).number_format = '#,##0.00'
                    
                    # Estilo
                    ws_rendimiento.cell(row=i, column=1).font = normal_font
                    ws_rendimiento.cell(row=i, column=2).font = normal_font
                    ws_rendimiento.cell(row=i, column=1).border = border
                    ws_rendimiento.cell(row=i, column=2).border = border
                    ws_rendimiento.cell(row=i, column=2).alignment = number_alignment
                
                # Añadir espacio
                ws_rendimiento.cell(row=8, column=1, value="RENDIMIENTO POR MODELO")
                ws_rendimiento.cell(row=8, column=1).font = Font(name='Arial', size=14, bold=True)
                
                # Análisis por modelo
                if 'Modelo chasis' in df.columns:
                    modelo_col = 'Modelo Normalizado' if 'Modelo Normalizado' in df.columns else 'Modelo chasis'
                    
                    # Agrupar por modelo
                    rendimiento_modelo = rendimiento_df.groupby(modelo_col)['Rendimiento'].agg(['mean', 'median', 'std', 'count']).reset_index()
                    rendimiento_modelo.columns = ['Modelo', 'Promedio', 'Mediana', 'Desviación', 'Cantidad']
                    
                    # Ordenar por promedio (descendente)
                    rendimiento_modelo = rendimiento_modelo.sort_values('Promedio', ascending=False)
                    
                    # Encabezados
                    encabezados = ['Modelo', 'Promedio (km/L)', 'Mediana (km/L)', 'Desviación', 'Cantidad']
                    for i, encabezado in enumerate(encabezados, start=1):
                        cell = ws_rendimiento.cell(row=9, column=i, value=encabezado)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = border
                    
                    # Escribir datos
                    for i, row in enumerate(dataframe_to_rows(rendimiento_modelo, index=False, header=False), start=10):
                        for j, value in enumerate(row, start=1):
                            if j == 1:  # Modelo
                                cell = ws_rendimiento.cell(row=i, column=j, value=value)
                            elif j in [2, 3, 4]:  # Valores numéricos con decimales
                                cell = ws_rendimiento.cell(row=i, column=j, value=value)
                                cell.number_format = '#,##0.00'
                            else:  # Valores enteros
                                cell = ws_rendimiento.cell(row=i, column=j, value=value)
                                cell.number_format = '#,##0'
                            
                            # Estilo común
                            cell.font = normal_font
                            cell.border = border
                            cell.alignment = number_alignment if j > 1 else normal_alignment
                        
                        # Filas alternas
                        if i % 2 == 0:
                            for j in range(1, len(encabezados) + 1):
                                ws_rendimiento.cell(row=i, column=j).fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                    
                    # Análisis por bus
                    if 'Número interno' in df.columns:
                        # Añadir espacio
                        last_row = 10 + len(rendimiento_modelo)
                        ws_rendimiento.cell(row=last_row + 2, column=1, value="TOP 20 BUSES POR RENDIMIENTO")
                        ws_rendimiento.cell(row=last_row + 2, column=1).font = Font(name='Arial', size=14, bold=True)
                        
                        # Agrupar por bus
                        rendimiento_bus = rendimiento_df.groupby('Número interno')['Rendimiento'].agg(['mean', 'count']).reset_index()
                        rendimiento_bus.columns = ['Bus', 'Rendimiento', 'Cargas']
                        
                        # Filtrar buses con suficientes cargas
                        rendimiento_bus = rendimiento_bus[rendimiento_bus['Cargas'] >= 5]
                        
                        # Ordenar por rendimiento (descendente)
                        rendimiento_bus = rendimiento_bus.sort_values('Rendimiento', ascending=False)
                        
                        # Tomar top 20
                        top_buses = rendimiento_bus.head(20)
                        
                        # Encabezados
                        encabezados = ['Bus', 'Rendimiento (km/L)', 'Cargas']
                        for i, encabezado in enumerate(encabezados, start=1):
                            cell = ws_rendimiento.cell(row=last_row + 3, column=i, value=encabezado)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                            cell.border = border
                        
                        # Escribir datos
                        for i, row in enumerate(dataframe_to_rows(top_buses, index=False, header=False), start=last_row + 4):
                            for j, value in enumerate(row, start=1):
                                if j == 1:  # Bus
                                    cell = ws_rendimiento.cell(row=i, column=j, value=value)
                                elif j == 2:  # Rendimiento
                                    cell = ws_rendimiento.cell(row=i, column=j, value=value)
                                    cell.number_format = '#,##0.00'
                                else:  # Cargas
                                    cell = ws_rendimiento.cell(row=i, column=j, value=value)
                                    cell.number_format = '#,##0'
                                
                                # Estilo común
                                cell.font = normal_font
                                cell.border = border
                                cell.alignment = number_alignment if j > 1 else normal_alignment
                            
                            # Filas alternas
                            if i % 2 == 0:
                                for j in range(1, len(encabezados) + 1):
                                    ws_rendimiento.cell(row=i, column=j).fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                
                # Ajustar anchos de columna
                ws_rendimiento.column_dimensions['A'].width = 30
                ws_rendimiento.column_dimensions['B'].width = 20
                ws_rendimiento.column_dimensions['C'].width = 20
                ws_rendimiento.column_dimensions['D'].width = 20
                ws_rendimiento.column_dimensions['E'].width = 15
        
        # ----- POR CONDUCTOR -----
        if 'Nombre conductor' in df.columns:
            ws_conductor = workbook.create_sheet(title='Por Conductor')
            
            # Agrupar por conductor
            conductor_stats = df.groupby('Nombre conductor').agg({
                'Cantidad litros': ['sum', 'mean', 'count'] if 'Cantidad litros' in df.columns else 'count',
                'Sobreconsumo': 'sum' if 'Sobreconsumo' in df.columns else 'count'
            }).reset_index()
            
            # Aplanar multiíndice
            if 'Cantidad litros' in df.columns:
                conductor_stats.columns = ['Conductor', 'Total Litros', 'Promedio Litros', 'Cantidad Cargas', 'Sobreconsumos']
                
                # Calcular porcentaje de sobreconsumo
                conductor_stats['% Sobreconsumos'] = (conductor_stats['Sobreconsumos'] / conductor_stats['Cantidad Cargas'] * 100).round(2)
                
                # Ordenar por porcentaje de sobreconsumo (descendente)
                conductor_stats = conductor_stats.sort_values('% Sobreconsumos', ascending=False)
                
                # Escribir encabezados
                encabezados = ['Conductor', 'Total Litros', 'Promedio Litros', 'Cantidad Cargas', 'Sobreconsumos', '% Sobre']
            else:
                conductor_stats.columns = ['Conductor', 'Cantidad Cargas', 'Sobreconsumos']
                
                # Calcular porcentaje de sobreconsumo
                conductor_stats['% Sobreconsumos'] = (conductor_stats['Sobreconsumos'] / conductor_stats['Cantidad Cargas'] * 100).round(2)
                
                # Ordenar por porcentaje de sobreconsumo (descendente)
                conductor_stats = conductor_stats.sort_values('% Sobreconsumos', ascending=False)
                
                # Escribir encabezados
                encabezados = ['Conductor', 'Cantidad Cargas', 'Sobreconsumos', '% Sobre']
            
            # Escribir encabezados
            for i, encabezado in enumerate(encabezados, start=1):
                cell = ws_conductor.cell(row=1, column=i, value=encabezado)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Escribir datos
            for i, row in enumerate(dataframe_to_rows(conductor_stats, index=False, header=False), start=2):
                for j, value in enumerate(row, start=1):
                    if j == 1:  # Conductor
                        cell = ws_conductor.cell(row=i, column=j, value=value)
                    elif j in [2, 3]:  # Valores numéricos con decimales
                        cell = ws_conductor.cell(row=i, column=j, value=value)
                        cell.number_format = '#,##0.00'
                    elif j == len(encabezados):  # Porcentaje
                        cell = ws_conductor.cell(row=i, column=j, value=value/100)  # Convertir a decimal para formato de porcentaje
                        cell.number_format = '0.00%'
                    else:  # Valores enteros
                        cell = ws_conductor.cell(row=i, column=j, value=value)
                        cell.number_format = '#,##0'
                    
                    # Estilo común
                    cell.font = normal_font
                    cell.border = border
                    cell.alignment = number_alignment if j > 1 else normal_alignment
                
                # Colorear según porcentaje de sobreconsumo
                if j == len(encabezados):
                    porcentaje = row[-1]  # Porcentaje de sobreconsumo
                    if porcentaje >= 15:
                        color = 'F94144'  # Rojo (alto)
                    elif porcentaje >= 10:
                        color = 'F8961E'  # Naranja (medio)
                    elif porcentaje >= 5:
                        color = 'F9C74F'  # Amarillo (bajo)
                    else:
                        color = '90BE6D'  # Verde (muy bajo)
                    
                    ws_conductor.cell(row=i, column=len(encabezados)).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                
                # Filas alternas
                if i % 2 == 0:
                    for j in range(1, len(encabezados)):  # No incluir la última columna (porcentaje)
                        ws_conductor.cell(row=i, column=j).fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
            
            # Ajustar anchos de columna
            for i, ancho in enumerate([30, 15, 15, 15, 15, 10], start=1):
                if i <= len(encabezados):
                    ws_conductor.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
        
        # ----- POR BUS -----
        if 'Número interno' in df.columns:
            ws_bus = workbook.create_sheet(title='Por Bus')
            
            # Agrupar por bus
            bus_stats = df.groupby('Número interno').agg({
                'Cantidad litros': ['sum', 'mean', 'count'] if 'Cantidad litros' in df.columns else 'count',
                'Sobreconsumo': 'sum' if 'Sobreconsumo' in df.columns else 'count',
                'Rendimiento': ['mean', 'median'] if 'Rendimiento' in df.columns else 'count'
            }).reset_index()
            
            # Aplanar multiíndice
            if 'Cantidad litros' in df.columns and 'Rendimiento' in df.columns:
                bus_stats.columns = ['Bus', 'Total Litros', 'Promedio Litros', 'Cantidad Cargas', 
                                   'Sobreconsumos', 'Rendimiento Promedio', 'Rendimiento Mediana']
                
                # Calcular porcentaje de sobreconsumo
                bus_stats['% Sobreconsumos'] = (bus_stats['Sobreconsumos'] / bus_stats['Cantidad Cargas'] * 100).round(2)
                
                # Ordenar por rendimiento (descendente)
                bus_stats = bus_stats.sort_values('Rendimiento Promedio', ascending=False)
                
                # Escribir encabezados
                encabezados = ['Bus', 'Total Litros', 'Promedio Litros', 'Cantidad Cargas', 
                             'Sobreconsumos', '% Sobre', 'Rendimiento Promedio', 'Rendimiento Mediana']
            elif 'Cantidad litros' in df.columns:
                bus_stats.columns = ['Bus', 'Total Litros', 'Promedio Litros', 'Cantidad Cargas', 
                                   'Sobreconsumos', 'Rendimiento Count']
                
                # Calcular porcentaje de sobreconsumo
                bus_stats['% Sobreconsumos'] = (bus_stats['Sobreconsumos'] / bus_stats['Cantidad Cargas'] * 100).round(2)
                
                # Ordenar por total litros (descendente)
                bus_stats = bus_stats.sort_values('Total Litros', ascending=False)
                
                # Escribir encabezados
                encabezados = ['Bus', 'Total Litros', 'Promedio Litros', 'Cantidad Cargas', 
                             'Sobreconsumos', '% Sobre', 'Rendimiento Count']
            else:
                bus_stats.columns = ['Bus', 'Cantidad Cargas', 'Sobreconsumos', 'Rendimiento Count']
                
                # Ordenar por cantidad de cargas (descendente)
                bus_stats = bus_stats.sort_values('Cantidad Cargas', ascending=False)
                
                # Escribir encabezados
                encabezados = ['Bus', 'Cantidad Cargas', 'Sobreconsumos', 'Rendimiento Count']
            
            # Escribir encabezados
            for i, encabezado in enumerate(encabezados, start=1):
                cell = ws_bus.cell(row=1, column=i, value=encabezado)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Escribir datos
            for i, row in enumerate(dataframe_to_rows(bus_stats, index=False, header=False), start=2):
                for j, value in enumerate(row, start=1):
                    if j == 1:  # Bus
                        cell = ws_bus.cell(row=i, column=j, value=value)
                    elif j in [2, 3, 7, 8]:  # Valores numéricos con decimales
                        cell = ws_bus.cell(row=i, column=j, value=value)
                        cell.number_format = '#,##0.00'
                    elif j == 6:  # Porcentaje
                        cell = ws_bus.cell(row=i, column=j, value=value/100)  # Convertir a decimal para formato de porcentaje
                        cell.number_format = '0.00%'
                    else:  # Valores enteros
                        cell = ws_bus.cell(row=i, column=j, value=value)
                        cell.number_format = '#,##0'
                    
                    # Estilo común
                    cell.font = normal_font
                    cell.border = border
                    cell.alignment = number_alignment if j > 1 else normal_alignment
                
                # Filas alternas
                if i % 2 == 0:
                    for j in range(1, len(encabezados) + 1):
                        ws_bus.cell(row=i, column=j).fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
            
            # Ajustar anchos de columna
            for i, ancho in enumerate([15, 15, 15, 15, 15, 10, 20, 20], start=1):
                if i <= len(encabezados):
                    ws_bus.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
        
        # ----- DATOS COMPLETOS -----
        ws_datos = workbook.create_sheet(title='Datos Completos')
        
        # Determinar columnas a incluir (evitar columnas derivadas)
        cols_a_excluir = [
            'Día Orden', 'Semana', 'Posible Mala Carga', 'Rendimiento Anómalo',
            'Km Sospechosos', 'Cargas Terminal', 'Índice Eficiencia'
        ]
        
        cols_a_incluir = [col for col in df.columns if col not in cols_a_excluir]
        
        # Escribir encabezados
        for i, encabezado in enumerate(cols_a_incluir, start=1):
            cell = ws_datos.cell(row=1, column=i, value=encabezado)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Ordenar datos por fecha si está disponible
        if 'Fecha' in df.columns:
            datos_a_escribir = df[cols_a_incluir].sort_values('Fecha', ascending=False)
        else:
            datos_a_escribir = df[cols_a_incluir]
        
        # Escribir datos
        for i, row in enumerate(dataframe_to_rows(datos_a_escribir.head(1000), index=False, header=False), start=2):  # Limitar a 1000 filas para evitar archivos demasiado grandes
            for j, value in enumerate(row, start=1):
                # Manejar tipos específicos
                if cols_a_incluir[j-1] == 'Fecha' and value is not None:
                    cell = ws_datos.cell(row=i, column=j, value=value)
                    cell.number_format = 'DD/MM/YYYY'
                elif cols_a_incluir[j-1] == 'Hora' and value is not None:
                    cell = ws_datos.cell(row=i, column=j, value=value)
                    cell.number_format = 'HH:MM'
                elif cols_a_incluir[j-1] in ['Cantidad litros', 'Rendimiento', 'Z-Score', 'Promedio Modelo', 'Desviación Modelo'] and value is not None:
                    cell = ws_datos.cell(row=i, column=j, value=value)
                    cell.number_format = '#,##0.00'
                elif cols_a_incluir[j-1] in ['Mala Carga', 'Sobreconsumo', 'Outlier Extremo', 'Es Fin de Semana', 'Hora Pico', 'Llenado Completo'] and value is not None:
                    cell = ws_datos.cell(row=i, column=j, value=value)
                    # Resaltar valores True
                    if value == True:
                        if cols_a_incluir[j-1] == 'Mala Carga':
                            cell.fill = PatternFill(start_color='F94144', end_color='F94144', fill_type='solid')
                        elif cols_a_incluir[j-1] == 'Sobreconsumo':
                            cell.fill = PatternFill(start_color='F8961E', end_color='F8961E', fill_type='solid')
                        elif cols_a_incluir[j-1] == 'Outlier Extremo':
                            cell.fill = PatternFill(start_color='F9C74F', end_color='F9C74F', fill_type='solid')
                else:
                    cell = ws_datos.cell(row=i, column=j, value=value)
                
                # Estilo común
                cell.font = normal_font
                cell.border = border
        
        # Ajustar anchos de columna automáticamente
        for i, col in enumerate(cols_a_incluir, start=1):
            # Ancho predeterminado basado en el tipo de columna
            if col in ['Fecha', 'Hora', 'Día', 'Mes']:
                ancho = 12
            elif col in ['Terminal', 'Modelo chasis', 'Nombre conductor', 'Nombre Planillero', 'Nombre supervisor']:
                ancho = 25
            elif col in ['Número interno', 'Patente']:
                ancho = 15
            elif col in ['Cantidad litros', 'Rendimiento', 'Z-Score', 'Promedio Modelo']:
                ancho = 15
            else:
                ancho = 15
            
            ws_datos.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
        
        # Guardar el libro
        workbook.save(output)
        
    except Exception as e:
        # Si falla la exportación avanzada, intentar con pandas básico
        print(f"Error en la exportación avanzada: {e}")
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Resumen general
            resumen = pd.DataFrame({
                'Métrica': [
                    'Total Cargas', 
                    'Total Litros', 
                    'Promedio Litros/Carga', 
                    'Malas Cargas', 
                    'Cargas con Sobreconsumo',
                    'Litros en Sobreconsumo',
                    'Total Buses',
                    'Total Terminales',
                    'Total Conductores'
                ],
                'Valor': [
                    len(df),
                    f"{df['Cantidad litros'].sum():,.2f}" if 'Cantidad litros' in df.columns else 'N/A',
                    f"{df['Cantidad litros'].mean():,.2f}" if 'Cantidad litros' in df.columns else 'N/A',
                    f"{df['Mala Carga'].sum()}" if 'Mala Carga' in df.columns else 'N/A',
                    f"{df['Sobreconsumo'].sum()}" if 'Sobreconsumo' in df.columns else 'N/A',
                    f"{df.loc[df['Sobreconsumo'] == True, 'Cantidad litros'].sum():,.2f}" if 'Sobreconsumo' in df.columns and 'Cantidad litros' in df.columns else 'N/A',
                    f"{df['Número interno'].nunique()}" if 'Número interno' in df.columns else 'N/A',
                    f"{df['Terminal'].nunique()}" if 'Terminal' in df.columns else 'N/A',
                    f"{df['Nombre conductor'].nunique()}" if 'Nombre conductor' in df.columns else 'N/A'
                ]
            })
            resumen.to_excel(writer, sheet_name='Resumen General', index=False)
            
            # Consumo diario
            if 'Fecha' in df.columns and 'Cantidad litros' in df.columns:
                consumo_diario = df.groupby(df['Fecha'].dt.date)['Cantidad litros'].agg(['sum', 'mean', 'count']).reset_index()
                consumo_diario.columns = ['Fecha', 'Total Litros', 'Promedio Litros', 'Cantidad Cargas']
                consumo_diario.to_excel(writer, sheet_name='Consumo Diario', index=False)
            
            # Análisis por terminal
            if 'Terminal' in df.columns:
                terminal_stats = df.groupby('Terminal').agg({
                    'Cantidad litros': ['sum', 'mean', 'count'] if 'Cantidad litros' in df.columns else 'count',
                    'Mala Carga': 'sum' if 'Mala Carga' in df.columns else 'count',
                    'Sobreconsumo': 'sum' if 'Sobreconsumo' in df.columns else 'count'
                }).reset_index()
                
                # Aplanar los nombres de columnas multiíndice
                terminal_stats.columns = ['Terminal', 'Total Litros', 'Promedio Litros', 'Cantidad Cargas', 
                                        'Malas Cargas', 'Sobreconsumos']
                
                terminal_stats.to_excel(writer, sheet_name='Por Terminal', index=False)
            
            # Análisis por modelo
            if 'Modelo chasis' in df.columns:
                modelo_stats = df.groupby('Modelo chasis').agg({
                    'Cantidad litros': ['sum', 'mean', 'count'] if 'Cantidad litros' in df.columns else 'count',
                    'Sobreconsumo': 'sum' if 'Sobreconsumo' in df.columns else 'count',
                    'Rendimiento': ['mean', 'median', 'std'] if 'Rendimiento' in df.columns else 'count'
                }).reset_index()
                
                # Aplanar los nombres de columnas multiíndice
                if 'Rendimiento' in df.columns:
                    modelo_stats.columns = ['Modelo', 'Total Litros', 'Promedio Litros', 'Cantidad Cargas', 
                                            'Sobreconsumos', 'Rendimiento Promedio', 'Rendimiento Mediana', 
                                            'Rendimiento Desviación']
                else:
                    modelo_stats.columns = ['Modelo', 'Total Litros', 'Promedio Litros', 'Cantidad Cargas', 
                                            'Sobreconsumos', 'Rendimiento Count']
                
                modelo_stats.to_excel(writer, sheet_name='Por Modelo', index=False)
            
            # Malas cargas
            if 'Mala Carga' in df.columns:
                malas_cargas = df[df['Mala Carga'] == True]
                if not malas_cargas.empty:
                    cols_relevantes = [col for col in [
                        'Fecha', 'Hora', 'Terminal', 'Número interno', 'Patente', 
                        'Cantidad litros', 'Tipo', 'Nombre Planillero', 'Nombre supervisor'
                    ] if col in malas_cargas.columns]
                    malas_cargas[cols_relevantes].to_excel(writer, sheet_name='Malas Cargas', index=False)
            
            # Sobreconsumo
            if 'Sobreconsumo' in df.columns:
                sobreconsumo = df[df['Sobreconsumo'] == True]
                if not sobreconsumo.empty:
                    cols_relevantes = [col for col in [
                        'Fecha', 'Hora', 'Terminal', 'Número interno', 'Patente', 'Modelo chasis',
                        'Cantidad litros', 'Z-Score', 'Promedio Modelo', 'Nombre conductor', 'Nombre supervisor'
                    ] if col in sobreconsumo.columns]
                    sobreconsumo[cols_relevantes].to_excel(writer, sheet_name='Sobreconsumo', index=False)
            
            # Por conductor
            if 'Nombre conductor' in df.columns:
                conductor_stats = df.groupby('Nombre conductor').agg({
                    'Cantidad litros': ['sum', 'mean', 'count'] if 'Cantidad litros' in df.columns else 'count',
                    'Sobreconsumo': 'sum' if 'Sobreconsumo' in df.columns else 'count'
                }).reset_index()
                
                # Aplanar los nombres de columnas multiíndice
                conductor_stats.columns = ['Conductor', 'Total Litros', 'Promedio Litros', 'Cantidad Cargas', 'Sobreconsumos']
                
                conductor_stats.to_excel(writer, sheet_name='Por Conductor', index=False)
            
            # Por bus
            if 'Número interno' in df.columns:
                bus_stats = df.groupby('Número interno').agg({
                    'Cantidad litros': ['sum', 'mean', 'count'] if 'Cantidad litros' in df.columns else 'count',
                    'Sobreconsumo': 'sum' if 'Sobreconsumo' in df.columns else 'count',
                    'Rendimiento': ['mean', 'median'] if 'Rendimiento' in df.columns else 'count'
                }).reset_index()
                
                # Aplanar los nombres de columnas multiíndice
                if 'Rendimiento' in df.columns:
                    bus_stats.columns = ['Bus', 'Total Litros', 'Promedio Litros', 'Cantidad Cargas', 
                                        'Sobreconsumos', 'Rendimiento Promedio', 'Rendimiento Mediana']
                else:
                    bus_stats.columns = ['Bus', 'Total Litros', 'Promedio Litros', 'Cantidad Cargas', 
                                        'Sobreconsumos', 'Rendimiento Count']
                
                bus_stats.to_excel(writer, sheet_name='Por Bus', index=False)
            
            # Horas pico
            if 'Hora Numérica' in df.columns and 'Cantidad litros' in df.columns:
                hora_stats = df.groupby(np.round(df['Hora Numérica']).astype(int)).agg({
                    'Cantidad litros': ['sum', 'count'],
                    'Mala Carga': 'sum' if 'Mala Carga' in df.columns else 'count'
                }).reset_index()
                
                # Aplanar los nombres de columnas multiíndice
                hora_stats.columns = ['Hora', 'Total Litros', 'Cantidad Cargas', 'Malas Cargas']
                
                hora_stats.to_excel(writer, sheet_name='Por Hora', index=False)
            
            # Datos completos
            df.to_excel(writer, sheet_name='Datos Completos', index=False)
    
    output.seek(0)
    return output

# Función para mostrar un dashboard dinámico
def show_dashboard(df, metrics=None, charts=None, filters=None):
    """
    Muestra un dashboard interactivo con métricas y gráficos personalizables.
    
    Args:
        df: DataFrame con los datos
        metrics: Lista de métricas a mostrar
        charts: Lista de gráficos a mostrar
        filters: Lista de filtros a aplicar
    """
    # Contenedor para todo el dashboard
    dashboard = st.container()
    
    with dashboard:
        # Aplicar filtros si se proporcionan
        filtered_df = df.copy()
        if filters:
            # Crear contenedor para filtros
            filter_container = st.container()
            with filter_container:
                st.markdown(f"<h3 class='section-title'>Filtros</h3>", unsafe_allow_html=True)
                col1, col2, col3 = st.columns(3)
                
                # Filtro de fecha
                if 'Fecha' in df.columns and 'fecha' in filters:
                    with col1:
                        min_date = df['Fecha'].min().date()
                        max_date = df['Fecha'].max().date()
                        date_range = st.date_input(
                            "Rango de fechas",
                            value=(min_date, max_date),
                            min_value=min_date,
                            max_value=max_date
                        )
                        
                        if len(date_range) == 2:
                            start_date, end_date = date_range
                            filtered_df = filtered_df[(filtered_df['Fecha'].dt.date >= start_date) & 
                                                     (filtered_df['Fecha'].dt.date <= end_date)]
                
                # Filtro de terminal
                if 'Terminal' in df.columns and 'terminal' in filters:
                    with col2:
                        terminals = ['Todas'] + sorted(df['Terminal'].unique().tolist())
                        selected_terminal = st.selectbox("Terminal", terminals)
                        
                        if selected_terminal != 'Todas':
                            filtered_df = filtered_df[filtered_df['Terminal'] == selected_terminal]
                
                # Filtro de modelo
                if 'Modelo chasis' in df.columns and 'modelo' in filters:
                    with col3:
                        modelos = ['Todos'] + sorted(df['Modelo chasis'].unique().tolist())
                        selected_modelo = st.selectbox("Modelo", modelos)
                        
                        if selected_modelo != 'Todos':
                           filtered_df = filtered_df[filtered_df['Modelo chasis'] == selected_modelo]
                
                # Filtro de bus
                if 'Número interno' in df.columns and 'bus' in filters:
                    with col1:
                        buses = ['Todos'] + sorted(df['Número interno'].unique().tolist())
                        selected_bus = st.selectbox("Bus", buses)
                        
                        if selected_bus != 'Todos':
                            filtered_df = filtered_df[filtered_df['Número interno'] == selected_bus]
                
                # Filtro de conductor
                if 'Nombre conductor' in df.columns and 'conductor' in filters:
                    with col2:
                        conductores = ['Todos'] + sorted(df['Nombre conductor'].unique().tolist())
                        selected_conductor = st.selectbox("Conductor", conductores)
                        
                        if selected_conductor != 'Todos':
                            filtered_df = filtered_df[filtered_df['Nombre conductor'] == selected_conductor]
                
                # Mostrar cantidad de registros filtrados
                st.markdown(f"<div style='margin-top: 10px; text-align: center; padding: 10px; background-color: {COLORS['light']}; border-radius: 10px;'>Mostrando {len(filtered_df)} de {len(df)} registros ({len(filtered_df)/len(df)*100:.1f}%)</div>", unsafe_allow_html=True)
        
        # Mostrar métricas
        if metrics:
            metrics_container = st.container()
            with metrics_container:
                st.markdown(f"<h3 class='section-title'>Métricas Clave</h3>", unsafe_allow_html=True)
                cols = st.columns(len(metrics))
                
                for i, metric in enumerate(metrics):
                    with cols[i]:
                        if metric == 'total_cargas':
                            value = len(filtered_df)
                            delta = None
                            if 'Fecha' in filtered_df.columns:
                                # Calcular tendencia
                                cargas_por_dia = filtered_df.groupby(filtered_df['Fecha'].dt.date).size()
                                if len(cargas_por_dia) > 1:
                                    first_half = cargas_por_dia.iloc[:len(cargas_por_dia)//2].mean()
                                    second_half = cargas_por_dia.iloc[len(cargas_por_dia)//2:].mean()
                                    delta = ((second_half / first_half) - 1) * 100 if first_half > 0 else 0
                            
                            st.markdown(create_kpi_card(
                                "Total de Cargas", 
                                f"{value:,}",
                                delta,
                                icon="🚌",
                                color_class="primary",
                                footnote=f"Durante el período analizado"
                            ), unsafe_allow_html=True)
                        
                        elif metric == 'total_litros' and 'Cantidad litros' in filtered_df.columns:
                            value = filtered_df['Cantidad litros'].sum()
                            delta = None
                            if 'Fecha' in filtered_df.columns:
                                # Calcular tendencia
                                litros_por_dia = filtered_df.groupby(filtered_df['Fecha'].dt.date)['Cantidad litros'].sum()
                                if len(litros_por_dia) > 1:
                                    first_half = litros_por_dia.iloc[:len(litros_por_dia)//2].mean()
                                    second_half = litros_por_dia.iloc[len(litros_por_dia)//2:].mean()
                                    delta = ((second_half / first_half) - 1) * 100 if first_half > 0 else 0
                            
                            st.markdown(create_kpi_card(
                                "Total de Litros", 
                                f"{value:,.0f}",
                                delta,
                                icon="⛽",
                                color_class="info",
                                footnote=f"Consumo total de combustible"
                            ), unsafe_allow_html=True)
                        
                        elif metric == 'promedio_carga' and 'Cantidad litros' in filtered_df.columns:
                            value = filtered_df['Cantidad litros'].mean()
                            global_avg = df['Cantidad litros'].mean()
                            delta = ((value / global_avg) - 1) * 100 if global_avg > 0 else 0
                            
                            st.markdown(create_kpi_card(
                                "Promedio por Carga", 
                                f"{value:.2f} L",
                                delta,
                                icon="📊",
                                color_class="success",
                                footnote=f"Media: {df['Cantidad litros'].median():.2f} L"
                            ), unsafe_allow_html=True)
                        
                        elif metric == 'malas_cargas' and 'Mala Carga' in filtered_df.columns:
                            value = filtered_df['Mala Carga'].sum()
                            pct = (value / len(filtered_df)) * 100 if len(filtered_df) > 0 else 0
                            
                            # Determinar color según porcentaje
                            color_class = "success" if pct < 2 else "warning" if pct < 5 else "danger"
                            
                            st.markdown(create_kpi_card(
                                "Malas Cargas", 
                                f"{value:,}",
                                -pct,  # Delta negativo (menos es mejor)
                                icon="❌",
                                color_class=color_class,
                                footnote=f"{pct:.2f}% del total de cargas"
                            ), unsafe_allow_html=True)
                        
                        elif metric == 'sobreconsumo' and 'Sobreconsumo' in filtered_df.columns:
                            value = filtered_df['Sobreconsumo'].sum()
                            pct = (value / len(filtered_df)) * 100 if len(filtered_df) > 0 else 0
                            
                            # Determinar color según porcentaje
                            color_class = "success" if pct < 5 else "warning" if pct < 10 else "danger"
                            
                            st.markdown(create_kpi_card(
                                "Sobreconsumos", 
                                f"{value:,}",
                                -pct,  # Delta negativo (menos es mejor)
                                icon="⚠️",
                                color_class=color_class,
                                footnote=f"{pct:.2f}% del total de cargas"
                            ), unsafe_allow_html=True)
                        
                        elif metric == 'rendimiento' and 'Rendimiento' in filtered_df.columns:
                            value = filtered_df['Rendimiento'].mean()
                            global_avg = df['Rendimiento'].mean()
                            delta = ((value / global_avg) - 1) * 100 if global_avg > 0 else 0
                            
                            st.markdown(create_kpi_card(
                                "Rendimiento Promedio", 
                                f"{value:.2f} km/L",
                                delta,
                                icon="🔄",
                                color_class="primary",
                                footnote=f"Media: {filtered_df['Rendimiento'].median():.2f} km/L"
                            ), unsafe_allow_html=True)
        
        # Mostrar gráficos
        if charts:
            charts_container = st.container()
            with charts_container:
                st.markdown(f"<h3 class='section-title'>Visualizaciones</h3>", unsafe_allow_html=True)
                
                for chart in charts:
                    if chart == 'evolucion_consumo' and 'Fecha' in filtered_df.columns and 'Cantidad litros' in filtered_df.columns:
                        st.subheader("Evolución del Consumo")
                        fig = plot_time_series(filtered_df, 'Cantidad litros', 'Evolución del Consumo Diario')
                        if fig:
                            st.plotly_chart(fig, use_container_width=True)
                    
                    elif chart == 'consumo_terminal' and 'Terminal' in filtered_df.columns and 'Cantidad litros' in filtered_df.columns:
                        st.subheader("Consumo por Terminal")
                        fig = plot_bar_chart(filtered_df, 'Terminal', 'Cantidad litros', 'Total Litros por Terminal', top_n=10)
                        if fig:
                            st.plotly_chart(fig, use_container_width=True)
                    
                    elif chart == 'consumo_modelo' and 'Modelo chasis' in filtered_df.columns and 'Cantidad litros' in filtered_df.columns:
                        st.subheader("Consumo por Modelo")
                        fig = plot_bar_chart(filtered_df, 'Modelo chasis', 'Cantidad litros', 'Total Litros por Modelo de Chasis', top_n=10)
                        if fig:
                            st.plotly_chart(fig, use_container_width=True)
                    
                    elif chart == 'heatmap_horario' and 'Hora Numérica' in filtered_df.columns and 'Día Semana' in filtered_df.columns:
                        st.subheader("Distribución por Horario")
                        # Selector para variable a mostrar
                        variable = 'Cantidad de Cargas'
                        if 'Cantidad litros' in filtered_df.columns:
                            variable = st.radio("Variable a visualizar:", ["Cantidad de Cargas", "Cantidad litros"], horizontal=True)
                        
                        if variable == "Cantidad de Cargas":
                            fig = create_heatmap(filtered_df, title="Mapa de calor: Cantidad de Cargas por día y hora")
                        else:
                            fig = create_heatmap(filtered_df, 'Cantidad litros', title="Mapa de calor: Consumo de Combustible por día y hora")
                        
                        if fig:
                            st.plotly_chart(fig, use_container_width=True)
                    
                    elif chart == 'distribucion_litros' and 'Cantidad litros' in filtered_df.columns:
                        st.subheader("Distribución de Cargas")
                        fig = px.histogram(
                            filtered_df, 
                            x='Cantidad litros',
                            nbins=30,
                            marginal='box',
                            title='Distribución de la cantidad de litros por carga',
                            color_discrete_sequence=[COLORS['primary']]
                        )
                        
                        fig.update_layout(
                            xaxis_title="Litros",
                            yaxis_title="Frecuencia",
                            plot_bgcolor='white'
                        )
                        
                        # Marcar anomalías
                        if 'Sobreconsumo' in filtered_df.columns:
                            # Umbral para sobreconsumo
                            umbral = filtered_df[~filtered_df['Sobreconsumo']]['Cantidad litros'].mean() * 1.2
                            
                            fig.add_vline(
                                x=umbral, 
                                line_dash="dash", 
                                line_color="red",
                                annotation_text="Umbral Sobreconsumo"
                            )
                        
                        st.plotly_chart(fig, use_container_width=True)
                    
                    elif chart == 'top_buses' and 'Número interno' in filtered_df.columns and 'Cantidad litros' in filtered_df.columns:
                        st.subheader("Top Buses por Consumo")
                        # Agrupar por bus
                        top_buses = filtered_df.groupby('Número interno')['Cantidad litros'].sum().reset_index()
                        top_buses = top_buses.sort_values('Cantidad litros', ascending=False).head(10)
                        
                        fig = px.bar(
                            top_buses,
                            x='Número interno',
                            y='Cantidad litros',
                            title='Top 10 Buses con Mayor Consumo Total',
                            color='Cantidad litros',
                            color_continuous_scale='Viridis',
                            text='Cantidad litros'
                        )
                        
                        fig.update_traces(
                            texttemplate='%{text:.1f}',
                            textposition='outside'
                        )
                        
                        fig.update_layout(
                            xaxis_title="Número de Bus",
                            yaxis_title="Total Litros",
                            plot_bgcolor='white'
                        )
                        
                        st.plotly_chart(fig, use_container_width=True)
                    
                    elif chart == 'rendimiento_modelo' and 'Modelo chasis' in filtered_df.columns and 'Rendimiento' in filtered_df.columns:
                        st.subheader("Rendimiento por Modelo")
                        # Agrupar por modelo
                        rendimiento_modelo = filtered_df.groupby('Modelo chasis')['Rendimiento'].mean().reset_index()
                        rendimiento_modelo = rendimiento_modelo.sort_values('Rendimiento', ascending=False)
                        
                        fig = px.bar(
                            rendimiento_modelo,
                            x='Modelo chasis',
                            y='Rendimiento',
                            title='Rendimiento Promedio por Modelo de Chasis',
                            color='Rendimiento',
                            color_continuous_scale='Viridis',
                            text='Rendimiento'
                        )
                        
                        fig.update_traces(
                            texttemplate='%{text:.2f}',
                            textposition='outside'
                        )
                        
                        fig.update_layout(
                            xaxis_title="Modelo de Chasis",
                            yaxis_title="Rendimiento (km/L)",
                            plot_bgcolor='white'
                        )
                        
                        st.plotly_chart(fig, use_container_width=True)

# Función para generar informes automáticos basados en patrones detectados
def generate_insights(df):
    """
    Genera insights automáticos a partir de los datos analizados.
    
    Args:
        df: DataFrame con los datos procesados
    
    Returns:
        Lista de insights con formato (título, descripción, tipo, importancia)
    """
    insights = []
    
    # Verificar columnas disponibles
    has_litros = 'Cantidad litros' in df.columns
    has_fecha = 'Fecha' in df.columns
    has_terminal = 'Terminal' in df.columns
    has_modelo = 'Modelo chasis' in df.columns
    has_rendimiento = 'Rendimiento' in df.columns
    has_mala_carga = 'Mala Carga' in df.columns
    has_sobreconsumo = 'Sobreconsumo' in df.columns
    
    # 1. Insight: Tendencia general de consumo
    if has_litros and has_fecha:
        try:
            # Agrupar por fecha
            consumo_diario = df.groupby(df['Fecha'].dt.date)['Cantidad litros'].sum().reset_index()
            
            if len(consumo_diario) >= 7:  # Al menos una semana de datos
                # Calcular tendencia
                X = np.array(range(len(consumo_diario))).reshape(-1, 1)
                y = consumo_diario['Cantidad litros'].values
                
                modelo = LinearRegression()
                modelo.fit(X, y)
                tendencia = modelo.coef_[0]
                
                # Porcentaje de cambio diario
                promedio = consumo_diario['Cantidad litros'].mean()
                porcentaje = (tendencia / promedio) * 100
                
                # Generar insight
                if abs(porcentaje) > 1:  # Solo si es significativo
                    if tendencia > 0:
                        titulo = "🔼 Tendencia de consumo al alza"
                        descripcion = f"El consumo de combustible está aumentando un {porcentaje:.2f}% diario. Evalúe si esto responde a un aumento en la operación o podría indicar una reducción en la eficiencia."
                        tipo = "warning"
                        importancia = 8 if porcentaje > 3 else 6
                    else:
                        titulo = "🔽 Tendencia de consumo a la baja"
                        descripcion = f"El consumo de combustible está disminuyendo un {abs(porcentaje):.2f}% diario. Esto podría indicar una mejora en la eficiencia o una reducción en la operación."
                        tipo = "success"
                        importancia = 7 if abs(porcentaje) > 3 else 5
                    
                    insights.append((titulo, descripcion, tipo, importancia))
        except Exception as e:
            print(f"Error al generar insight de tendencia: {e}")
    
    # 2. Insight: Malas cargas
    if has_mala_carga:
        try:
            total_malas = df['Mala Carga'].sum()
            porcentaje_malas = (total_malas / len(df)) * 100
            
            if porcentaje_malas > 2:
                # Título y descripción
                titulo = "❌ Nivel elevado de malas cargas"
                descripcion = f"Se detectaron {total_malas} malas cargas ({porcentaje_malas:.2f}% del total). "
                
                # Análisis por terminal si está disponible
                if has_terminal:
                    malas_por_terminal = df[df['Mala Carga']].groupby('Terminal').size()
                    total_por_terminal = df.groupby('Terminal').size()
                    pct_por_terminal = (malas_por_terminal / total_por_terminal * 100).sort_values(ascending=False)
                    
                    if not pct_por_terminal.empty:
                        peor_terminal = pct_por_terminal.index[0]
                        pct_peor = pct_por_terminal.iloc[0]
                        
                        if pct_peor > 5:
                            descripcion += f"La terminal {peor_terminal} muestra el mayor porcentaje ({pct_peor:.2f}%) de malas cargas."
                
                # Tipo e importancia
                if porcentaje_malas > 5:
                    tipo = "danger"
                    importancia = 9
                else:
                    tipo = "warning"
                    importancia = 7
                
                insights.append((titulo, descripcion, tipo, importancia))
        except Exception as e:
            print(f"Error al generar insight de malas cargas: {e}")
    
    # 3. Insight: Sobreconsumos
    if has_sobreconsumo and has_litros:
        try:
            total_sobre = df['Sobreconsumo'].sum()
            porcentaje_sobre = (total_sobre / len(df)) * 100
            
            if porcentaje_sobre > 5:
                # Obtener litros en sobreconsumo
                litros_sobre = df.loc[df['Sobreconsumo'], 'Cantidad litros'].sum()
                pct_litros = (litros_sobre / df['Cantidad litros'].sum()) * 100
                
                # Título y descripción
                titulo = "⚠️ Nivel elevado de sobreconsumo"
                descripcion = f"Se detectaron {total_sobre} cargas con sobreconsumo ({porcentaje_sobre:.2f}% del total), representando {litros_sobre:.2f} litros ({pct_litros:.2f}% del combustible total). "
                
                # Análisis por modelo si está disponible
                if has_modelo:
                    sobre_por_modelo = df[df['Sobreconsumo']].groupby('Modelo chasis').size()
                    total_por_modelo = df.groupby('Modelo chasis').size()
                    pct_por_modelo = (sobre_por_modelo / total_por_modelo * 100).sort_values(ascending=False)
                    
                    if not pct_por_modelo.empty:
                        peor_modelo = pct_por_modelo.index[0]
                        pct_peor = pct_por_modelo.iloc[0]
                        
                        if pct_peor > 10:
                            descripcion += f"El modelo {peor_modelo} muestra el mayor porcentaje ({pct_peor:.2f}%) de sobreconsumos."
                
                # Tipo e importancia
                if porcentaje_sobre > 10:
                    tipo = "danger"
                    importancia = 9
                else:
                    tipo = "warning"
                    importancia = 7
                
                insights.append((titulo, descripcion, tipo, importancia))
        except Exception as e:
            print(f"Error al generar insight de sobreconsumos: {e}")
    
    # 4. Insight: Rendimiento
    if has_rendimiento and has_modelo:
        try:
            rendimiento_global = df['Rendimiento'].mean()
            
            # Análisis por modelo
            rendimiento_por_modelo = df.groupby('Modelo chasis')['Rendimiento'].agg(['mean', 'count']).reset_index()
            rendimiento_por_modelo = rendimiento_por_modelo[rendimiento_por_modelo['count'] >= 5]  # Al menos 5 mediciones
            
            if not rendimiento_por_modelo.empty:
                # Encontrar modelos con rendimiento destacado (positivo o negativo)
                rendimiento_por_modelo['diff_pct'] = (rendimiento_por_modelo['mean'] / rendimiento_global - 1) * 100
                
                # Mejor modelo
                mejor_modelo = rendimiento_por_modelo.sort_values('mean', ascending=False).iloc[0]
                if mejor_modelo['diff_pct'] > 10:  # Al menos 10% mejor que el promedio
                    titulo = "🌟 Modelo con rendimiento destacado"
                    descripcion = f"El modelo {mejor_modelo['Modelo chasis']} muestra un rendimiento promedio de {mejor_modelo['mean']:.2f} km/L, un {mejor_modelo['diff_pct']:.2f}% superior al promedio global ({rendimiento_global:.2f} km/L)."
                    tipo = "success"
                    importancia = 7
                    
                    insights.append((titulo, descripcion, tipo, importancia))
                
                # Peor modelo
                peor_modelo = rendimiento_por_modelo.sort_values('mean', ascending=True).iloc[0]
                if peor_modelo['diff_pct'] < -10:  # Al menos 10% peor que el promedio
                    titulo = "⚠️ Modelo con bajo rendimiento"
                    descripcion = f"El modelo {peor_modelo['Modelo chasis']} muestra un rendimiento promedio de {peor_modelo['mean']:.2f} km/L, un {abs(peor_modelo['diff_pct']):.2f}% inferior al promedio global ({rendimiento_global:.2f} km/L)."
                    tipo = "warning"
                    importancia = 6
                    
                    insights.append((titulo, descripcion, tipo, importancia))
        except Exception as e:
            print(f"Error al generar insight de rendimiento: {e}")
    
    # 5. Insight: Patrones de horario
    if has_fecha and has_litros and 'Hora Numérica' in df.columns:
        try:
            # Consumo por hora
            consumo_por_hora = df.groupby(np.round(df['Hora Numérica']).astype(int))['Cantidad litros'].mean()
            
            if not consumo_por_hora.empty:
                # Pico de consumo
                hora_pico = consumo_por_hora.idxmax()
                valor_pico = consumo_por_hora.max()
                promedio = consumo_por_hora.mean()
                
                if valor_pico > promedio * 1.2:  # Al menos 20% mayor que el promedio
                    titulo = "⏰ Patrón horario de consumo"
                    descripcion = f"Se detectó un pico de consumo promedio a las {hora_pico}:00 horas, con {valor_pico:.2f} litros por carga, un {(valor_pico/promedio-1)*100:.2f}% superior al promedio horario."
                    tipo = "info"
                    importancia = 5
                    
                    insights.append((titulo, descripcion, tipo, importancia))
        except Exception as e:
            print(f"Error al generar insight de patrones horarios: {e}")
    
    # 6. Insight: Eficiencia por terminal
    if has_terminal and has_litros:
        try:
            # Consumo promedio por terminal
            consumo_por_terminal = df.groupby('Terminal')['Cantidad litros'].mean().sort_values()
            
            if len(consumo_por_terminal) >= 3:  # Al menos 3 terminales para comparar
                mejor_terminal = consumo_por_terminal.index[0]
                mejor_valor = consumo_por_terminal.iloc[0]
                
                peor_terminal = consumo_por_terminal.index[-1]
                peor_valor = consumo_por_terminal.iloc[-1]
                
                diff_pct = ((peor_valor / mejor_valor) - 1) * 100
                
                if diff_pct > 15:  # Diferencia significativa
                    titulo = "🏆 Diferencias de eficiencia entre terminales"
                    descripcion = f"La terminal {mejor_terminal} muestra el menor consumo promedio ({mejor_valor:.2f} L/carga), mientras que {peor_terminal} tiene el mayor ({peor_valor:.2f} L/carga), un {diff_pct:.2f}% más alto."
                    tipo = "info"
                    importancia = 6
                    
                    insights.append((titulo, descripcion, tipo, importancia))
        except Exception as e:
            print(f"Error al generar insight de eficiencia por terminal: {e}")
    
    # Ordenar insights por importancia (descendente)
    insights.sort(key=lambda x: x[3], reverse=True)
    
    return insights

# Función principal
def main():
    # Configuración global de la página
    st.markdown(load_css(COLORS), unsafe_allow_html=True)
    
    # Variables de sesión para temas
    if 'theme' not in st.session_state:
        st.session_state['theme'] = DEFAULT_THEME
    
    # Logo y título principal
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("<h1 class='title'>Smart Fuel Analytics 3.0</h1>", unsafe_allow_html=True)
        st.markdown("<p style='text-align: center; margin-top: -1rem; color: #757575;'>Plataforma Avanzada de Análisis de Combustible</p>", unsafe_allow_html=True)
    
    # Configuración del sidebar
    with st.sidebar:
        st.image("https://www.svgrepo.com/show/530443/fuel.svg", width=80)
        st.markdown("<h2 style='text-align: center; margin-top: 0;'>Control de Flotas</h2>", unsafe_allow_html=True)
        
        # Selector de tema
        st.markdown("### 🎨 Apariencia")
        selected_theme = st.selectbox(
            "Seleccionar tema",
            options=list(THEMES.keys()),
            format_func=lambda x: x.capitalize(),
            index=list(THEMES.keys()).index(st.session_state['theme'])
        )
        
        # Actualizar tema si cambia
        if selected_theme != st.session_state['theme']:
            st.session_state['theme'] = selected_theme
            COLORS = THEMES[selected_theme]
            st.markdown(load_css(COLORS), unsafe_allow_html=True)
            st.rerun()
        
        # Sección de carga de archivo
        st.markdown("### 📤 Cargar Datos")
        uploaded_file = st.file_uploader("Selecciona el archivo Excel", type=["xlsx", "xls", "csv"], 
                                        help="El archivo debe tener los datos a partir de la fila A3")
        
        if uploaded_file is not None:
            st.success("✅ Archivo cargado correctamente")
            
            # Botón de análisis
            st.markdown("### 🔍 Analizar Datos")
            analyze_btn = st.button("Iniciar Análisis Avanzado", type="primary", use_container_width=True)
            
            # Navegación
            st.markdown("### 📊 Secciones de Análisis")
            
            # Definir las secciones disponibles
            sections = [
                "Dashboard Principal",
                "Análisis por Terminal",
                "Análisis por Buses",
                "Análisis por Personal",
                "Malas Cargas",
                "Detección de Sobreconsumo",
                "Análisis Temporal",
                "Rendimiento",
                "Exportar Resultados"
            ]
            
            # Radio buttons para selección de sección
            selected_section = st.radio("Ir a:", sections)
            
            # Guardar selección en session_state
            st.session_state['current_section'] = selected_section
        
        # Separador
        st.markdown("---")
        
        # Sección de créditos
        st.markdown("<p style='text-align: center; color: #757575; font-size: 0.8rem;'>Smart Fuel Analytics v3.0<br>© 2025 - Todos los derechos reservados</p>", unsafe_allow_html=True)
    
    # Pantalla principal cuando no hay archivo cargado
    if uploaded_file is None:
        show_welcome_banner()
        
        # Características principales
        st.markdown("<h2 class='subtitle'>Características principales</h2>", unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.markdown(f"""
            <div class="card">
                <h3 style="color: {COLORS['primary']};">📊 Análisis Preciso</h3>
                <p>Detección automática de patrones de consumo, identificación de anomalías y visualizaciones detalladas con tecnología de machine learning.</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
            <div class="card">
                <h3 style="color: {COLORS['primary']};">⚠️ Alertas Inteligentes</h3>
                <p>Sistema avanzado de detección de sobreconsumo, malas cargas y anomalías operativas con tableros de control en tiempo real.</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            st.markdown(f"""
            <div class="card">
                <h3 style="color: {COLORS['primary']};">📱 Informes Completos</h3>
                <p>Informes personalizables por terminal, buses, personal y períodos, exportables a Excel con formatos profesionales.</p>
            </div>
            """, unsafe_allow_html=True)
        
        # Cómo funciona
        st.markdown("<h2 class='subtitle'>¿Cómo funciona?</h2>", unsafe_allow_html=True)
        
        col1, col2 = st.columns([3, 2])
        
        with col1:
            st.markdown(f"""
            <div class="card">
                <h3 style="color: {COLORS['primary']};">Proceso de Análisis</h3>
                <p>Smart Fuel Analytics 3.0 utiliza algoritmos avanzados para procesar y analizar sus datos de combustible:</p>
                <ol>
                    <li><strong>Carga de datos:</strong> Suba sus archivos Excel o CSV con registros de cargas de combustible.</li>
                    <li><strong>Preprocesamiento automático:</strong> El sistema limpia, normaliza y enriquece los datos, detectando automáticamente formatos y columnas.</li>
                    <li><strong>Detección de anomalías:</strong> Utilizamos técnicas de IA como Isolation Forest y análisis estadístico para identificar patrones anómalos.</li>
                    <li><strong>Análisis multidimensional:</strong> Examinamos el consumo por terminal, bus, modelo, conductor y períodos de tiempo.</li>
                    <li><strong>Generación de informes:</strong> Visualizaciones interactivas, métricas clave y reportes exportables para una toma de decisiones informada.</li>
                </ol>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
            <div class="card">
                <h3 style="color: {COLORS['primary']};">Beneficios Clave</h3>
                <ul>
                    <li><strong>Ahorro de combustible</strong> mediante la identificación de ineficiencias</li>
                    <li><strong>Reducción de errores operativos</strong> a través de la detección de malas cargas</li>
                    <li><strong>Optimización de flota</strong> con análisis de rendimiento por modelo y bus</li>
                    <li><strong>Mejora de procesos</strong> en terminales y personal</li>
                    <li><strong>Toma de decisiones basada en datos</strong> con métricas claras y precisas</li>
                </ul>
                <div style="text-align: center; margin-top: 1.5rem;">
                    <p>👈 Comience cargando un archivo desde el panel lateral</p>
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        # Instrucciones
        st.markdown("<h2 class='subtitle'>Instrucciones de uso</h2>", unsafe_allow_html=True)
        
        st.markdown(f"""
        <div class="card">
            <ol>
                <li><strong>Sube tu archivo Excel o CSV</strong> con datos de carga de combustible desde el panel lateral</li>
                <li>Haz clic en <strong>Iniciar Análisis Avanzado</strong> para procesar los datos con nuestros algoritmos</li>
                <li>Navega entre las diferentes secciones para ver análisis específicos y visualizaciones detalladas</li>
                <li>Utiliza los filtros interactivos para profundizar en aspectos particulares de tus datos</li>
                <li>Exporta los resultados a Excel para compartir o guardar cuando sea necesario</li>
            </ol>
            <p style="margin-top: 1rem;"><strong>Formato esperado del archivo:</strong></p>
            <ul>
                <li>Título en fila A3</li>
                <li>Datos desde la fila A4</li>
                <li>Columnas principales: Turno, Fecha, Hora, Terminal, Número interno, Patente, Cantidad litros, etc.</li>
                <li>El sistema detectará automáticamente "Carga Masiva" en cualquier columna, incluida la última</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
    
    elif analyze_btn or 'data' in st.session_state:
        # Cargar y procesar datos si es necesario
        if analyze_btn or 'data' not in st.session_state:
            # Mostrar indicador de carga
            progress_placeholder = show_loading("Analizando datos con algoritmos avanzados... Por favor espere")
            
            df = load_data(uploaded_file)
            if df is not None:
                st.session_state['data'] = df
                
                # Generar insights automáticos
                st.session_state['insights'] = generate_insights(df)
            
            # Quitar indicador de carga
            progress_placeholder.empty()
        
        # Usar datos ya cargados
        if 'data' in st.session_state:
            df = st.session_state['data']
            
            # Mostrar la sección seleccionada
            if 'current_section' in st.session_state:
                section = st.session_state['current_section']
                
                # === DASHBOARD PRINCIPAL ===
                if section == "Dashboard Principal":
                    st.markdown("<h2 class='subtitle'>Dashboard General</h2>", unsafe_allow_html=True)
                    
                    # Mostrar insights automáticos
                    if 'insights' in st.session_state and st.session_state['insights']:
                        with st.expander("📋 Hallazgos Automáticos Detectados", expanded=True):
                            for titulo, descripcion, tipo, _ in st.session_state['insights'][:3]:  # Mostrar solo los 3 más importantes
                                st.markdown(show_alert(descripcion, tipo, titulo), unsafe_allow_html=True)
                            
                            if len(st.session_state['insights']) > 3:
                                with st.expander("Ver todos los hallazgos"):
                                    for titulo, descripcion, tipo, _ in st.session_state['insights'][3:]:
                                        st.markdown(show_alert(descripcion, tipo, titulo), unsafe_allow_html=True)
                    
                    # Resumen del período analizado
                    if 'Fecha' in df.columns:
                        fecha_min = df['Fecha'].min().strftime('%d/%m/%Y')
                        fecha_max = df['Fecha'].max().strftime('%d/%m/%Y')
                        dias_totales = (df['Fecha'].max() - df['Fecha'].min()).days + 1
                        
                        st.markdown(f"""
                        <div class="card" style="margin-bottom: 1.5rem;">
                            <h3 style="color: {COLORS['primary']};">📅 Período Analizado</h3>
                            <p>Desde: <strong>{fecha_min}</strong> hasta: <strong>{fecha_max}</strong> ({dias_totales} días)</p>
                            <p>Total de registros: <strong>{len(df):,}</strong></p>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    # Métricas principales en tarjetas
                    st.markdown("<h3 class='section-title'>Métricas Principales</h3>", unsafe_allow_html=True)
                    
                    col1, col2, col3, col4 = st.columns(4)
                    
                    with col1:
                        total_cargas = len(df)
                        st.markdown(create_kpi_card(
                            "Total de Cargas", 
                            f"{total_cargas:,}",
                            None,
                            icon="🚌",
                            color_class="primary",
                            footnote=f"En {dias_totales} días" if 'Fecha' in df.columns else None
                        ), unsafe_allow_html=True)
                    
                    with col2:
                        if 'Cantidad litros' in df.columns:
                            total_litros = df['Cantidad litros'].sum()
                            st.markdown(create_kpi_card(
                                "Total Litros", 
                                f"{total_litros:,.0f}",
                                None,
                                icon="⛽",
                                color_class="info",
                                footnote=f"Promedio: {df['Cantidad litros'].mean():.2f} L/carga"
                            ), unsafe_allow_html=True)
                    
                    with col3:
                        if 'Mala Carga' in df.columns:
                            malas_cargas = df['Mala Carga'].sum()
                            porcentaje_malas = (malas_cargas / len(df)) * 100 if len(df) > 0 else 0
                            
                            # Color dinámico basado en el porcentaje
                            color_class = "success" if porcentaje_malas < 2 else "warning" if porcentaje_malas < 5 else "danger"
                            
                            st.markdown(create_kpi_card(
                                "Malas Cargas", 
                                f"{malas_cargas}",
                                -porcentaje_malas,  # Delta negativo (menos es mejor)
                                icon="❌",
                                color_class=color_class,
                                footnote=f"{porcentaje_malas:.2f}% del total"
                            ), unsafe_allow_html=True)
                    
                    with col4:
                        if 'Sobreconsumo' in df.columns:
                            sobreconsumos = df['Sobreconsumo'].sum()
                            porcentaje_sobre = (sobreconsumos / len(df)) * 100 if len(df) > 0 else 0
                            
                            # Color dinámico basado en el porcentaje
                            color_class = "success" if porcentaje_sobre < 5 else "warning" if porcentaje_sobre < 10 else "danger"
                            
                            st.markdown(create_kpi_card(
                                "Alertas Sobreconsumo", 
                                f"{sobreconsumos}",
                                -porcentaje_sobre,  # Delta negativo (menos es mejor)
                                icon="⚠️",
                                color_class=color_class,
                                footnote=f"{porcentaje_sobre:.2f}% del total"
                            ), unsafe_allow_html=True)
                    
                    # Segunda fila de métricas
                    col1, col2, col3, col4 = st.columns(4)
                    
                    with col1:
                        if 'Número interno' in df.columns:
                            total_buses = df['Número interno'].nunique()
                            st.markdown(create_kpi_card(
                                "Total Buses", 
                                f"{total_buses}",
                                None,
                                icon="🚍",
                                color_class="success",
                                footnote=f"Promedio: {len(df)/total_buses:.1f} cargas/bus" if total_buses > 0 else None
                            ), unsafe_allow_html=True)
                    
                    with col2:
                        if 'Terminal' in df.columns:
                            total_terminales = df['Terminal'].nunique()
                            st.markdown(create_kpi_card(
                                "Total Terminales", 
                                f"{total_terminales}",
                                None,
                                icon="🏢",
                                color_class="info",
                                footnote=f"Promedio: {len(df)/total_terminales:.1f} cargas/terminal" if total_terminales > 0 else None
                            ), unsafe_allow_html=True)
                    
                    with col3:
                        if 'Rendimiento' in df.columns:
                            rendimiento_medio = df['Rendimiento'].mean()
                            rendimiento_mediana = df['Rendimiento'].median()
                            st.markdown(create_kpi_card(
                                "Rendimiento Promedio", 
                                f"{rendimiento_medio:.2f} km/L",
                                None,
                                icon="🔄",
                                color_class="primary",
                                footnote=f"Mediana: {rendimiento_mediana:.2f} km/L"
                            ), unsafe_allow_html=True)
                    
                    with col4:
                        if 'Cantidad litros' in df.columns and 'Sobreconsumo' in df.columns:
                            litros_sobreconsumo = df.loc[df['Sobreconsumo'], 'Cantidad litros'].sum()
                            porcentaje_litros = (litros_sobreconsumo / df['Cantidad litros'].sum()) * 100 if df['Cantidad litros'].sum() > 0 else 0
                            
                            color_class = "success" if porcentaje_litros < 5 else "warning" if porcentaje_litros < 10 else "danger"
                            
                            st.markdown(create_kpi_card(
                                "Litros en Sobreconsumo", 
                                f"{litros_sobreconsumo:,.0f}",
                                -porcentaje_litros,  # Delta negativo (menos es mejor)
                                icon="💧",
                                color_class=color_class,
                                footnote=f"{porcentaje_litros:.2f}% del total"
                            ), unsafe_allow_html=True)
                    
                    # Gráfico de evolución temporal de consumo
                    st.markdown("<h3 class='section-title'>Evolución del Consumo</h3>", unsafe_allow_html=True)
                    
                    tab1, tab2 = st.tabs(["📊 Gráfico", "📋 Datos"])
                    
                    with tab1:
                        if 'Fecha' in df.columns and 'Cantidad litros' in df.columns:
                            # Crear gráfico de tendencia
                            fig = plot_time_series(df, 'Cantidad litros', 'Evolución del Consumo Diario')
                            
                            if fig:
                                st.plotly_chart(fig, use_container_width=True)
                            
                            # Mostrar tendencia si está disponible
                            if 'tendencia_consumo_robusta' in st.session_state:
                                tendencia = st.session_state['tendencia_consumo_robusta']
                                promedio = st.session_state['promedio_diario']
                                
                                # Calcular porcentaje de cambio diario
                                porcentaje = (tendencia / promedio) * 100
                                
                                # Crear alerta según la tendencia
                                if tendencia > 0 and abs(porcentaje) > 2:
                                    st.markdown(show_alert(
                                        f"El consumo de combustible muestra una tendencia creciente. Recomendamos revisar la eficiencia de la flota.",
                                        "warning",
                                        f"Tendencia al Alza: +{abs(porcentaje):.2f}% diario",
                                        "📈"
                                    ), unsafe_allow_html=True)
                                elif tendencia < 0 and abs(porcentaje) > 2:
                                    st.markdown(show_alert(
                                        f"El consumo de combustible muestra una tendencia decreciente. Las medidas de eficiencia parecen estar funcionando.",
                                        "success",
                                        f"Tendencia a la Baja: -{abs(porcentaje):.2f}% diario",
                                        "📉"
                                    ), unsafe_allow_html=True)
                    
                    with tab2:
                        if 'Fecha' in df.columns and 'Cantidad litros' in df.columns:
                            # Agrupar por fecha
                            consumo_diario = df.groupby(df['Fecha'].dt.date).agg({
                                'Cantidad litros': ['sum', 'mean', 'count'],
                                'Mala Carga': 'sum' if 'Mala Carga' in df.columns else 'count',
                                'Sobreconsumo': 'sum' if 'Sobreconsumo' in df.columns else 'count'
                            }).reset_index()
                            
                            # Aplanar los nombres de columnas multiíndice
                            consumo_diario.columns = ['Fecha', 'Total Litros', 'Promedio Litros', 'Cantidad Cargas', 'Malas Cargas', 'Sobreconsumos']
                            
                            # Calcular porcentajes
                            consumo_diario['% Malas Cargas'] = (consumo_diario['Malas Cargas'] / consumo_diario['Cantidad Cargas'] * 100).round(2)
                            consumo_diario['% Sobreconsumos'] = (consumo_diario['Sobreconsumos'] / consumo_diario['Cantidad Cargas'] * 100).round(2)
                            
                            # Mostrar tabla
                            st.dataframe(
                                consumo_diario.sort_values('Fecha', ascending=False),
                                column_config={
                                    "Fecha": st.column_config.DateColumn("Fecha"),
                                    "Total Litros": st.column_config.NumberColumn("Total Litros", format="%.2f"),
                                    "Promedio Litros": st.column_config.NumberColumn("Prom. L/Carga", format="%.2f"),
                                    "Cantidad Cargas": st.column_config.NumberColumn("# Cargas", format="%d"),
                                    "Malas Cargas": st.column_config.NumberColumn("Malas", format="%d"),
                                    "% Malas Cargas": st.column_config.NumberColumn("% Malas", format="%.2f%%"),
                                    "Sobreconsumos": st.column_config.NumberColumn("Sobre", format="%d"),
                                    "% Sobreconsumos": st.column_config.NumberColumn("% Sobre", format="%.2f%%")
                                },
                                use_container_width=True
                            )
                        else:
                            st.warning("No se encontraron datos suficientes para mostrar la evolución temporal.")
                    
                    # Distribución por Terminal y Modelo
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        st.markdown("<h3 class='section-title'>Consumo por Terminal</h3>", unsafe_allow_html=True)
                        
                        if 'Terminal' in df.columns and 'Cantidad litros' in df.columns:
                            fig = plot_bar_chart(df, 'Terminal', 'Cantidad litros', 'Total Litros por Terminal', top_n=10)
                            if fig:
                                st.plotly_chart(fig, use_container_width=True)
                    
                    with col2:
                        st.markdown("<h3 class='section-title'>Consumo por Modelo</h3>", unsafe_allow_html=True)
                        
                        if 'Modelo chasis' in df.columns and 'Cantidad litros' in df.columns:
                            fig = plot_bar_chart(df, 'Modelo chasis', 'Cantidad litros', 'Total Litros por Modelo de Chasis', top_n=10)
                            if fig:
                                st.plotly_chart(fig, use_container_width=True)
                    
                    # Mapa de calor por horario
                    st.markdown("<h3 class='section-title'>Distribución de Cargas por Horario</h3>", unsafe_allow_html=True)
                    
                    if 'Hora Numérica' in df.columns and 'Día Semana' in df.columns and 'Cantidad litros' in df.columns:
                        # Selector para variable a visualizar
                        tipo_variable = st.radio("Variable a visualizar:", ["Cantidad de Cargas", "Cantidad de Litros", "Litros Promedio/Carga"], horizontal=True)
                        
                        if tipo_variable == "Cantidad de Cargas":
                            fig = create_heatmap(df, title="Mapa de calor: Cantidad de Cargas por día y hora")
                        elif tipo_variable == "Cantidad de Litros":
                            fig = create_heatmap(df, 'Cantidad litros', title="Mapa de calor: Total de Litros por día y hora")
                        else:  # Litros Promedio/Carga
                            # Agrupar por día y hora, calculando promedios
                            heatmap_df = df.copy()
                            heatmap_df['Promedio Litros'] = heatmap_df['Cantidad litros']
                            fig = create_heatmap(heatmap_df, 'Promedio Litros', title="Mapa de calor: Litros Promedio por día y hora")
                        
                        if fig:
                            st.plotly_chart(fig, use_container_width=True)
                            
                            # Añadir interpretación del mapa de calor
                            st.markdown(f"""
                            <div class="card">
                                <h4>Interpretación del Mapa de Calor</h4>
                                <p>Este mapa de calor muestra la distribución de {tipo_variable.lower()} por día de la semana y hora del día. 
                                Los colores más intensos indican mayor actividad en ese horario específico.</p>
                                <ul>
                                    <li><strong>Utilidad:</strong> Identifique patrones de carga, optimice la distribución de personal y planifique mantenimientos en horas de menor actividad.</li>
                                    <li><strong>Recomendación:</strong> Compare este mapa con los horarios de operación para detectar ineficiencias o concentraciones excesivas de cargas.</li>
                                </ul>
                            </div>
                            """, unsafe_allow_html=True)
                    
                    # Distribución de cargas
                    st.markdown("<h3 class='section-title'>Distribución de Cantidad de Litros</h3>", unsafe_allow_html=True)
                    
                    if 'Cantidad litros' in df.columns:
                        tab1, tab2 = st.tabs(["📊 Gráfico", "📋 Estadísticas"])
                        
                        with tab1:
                            fig = px.histogram(
                                df, 
                                x='Cantidad litros',
                                nbins=30,
                                marginal='box',
                                title='Distribución de la cantidad de litros por carga',
                                color_discrete_sequence=[COLORS['primary']]
                            )
                            
                            fig.update_layout(
                                xaxis_title="Litros",
                                yaxis_title="Frecuencia",
                                plot_bgcolor='white'
                            )
                            
                            # Marcar anomalías
                            if 'Sobreconsumo' in df.columns:
                                # Umbral para sobreconsumo
                                umbral = df[~df['Sobreconsumo']]['Cantidad litros'].mean() * 1.2
                                
                                fig.add_vline(
                                    x=umbral, 
                                    line_dash="dash", 
                                    line_color="red",
                                    annotation_text="Umbral Sobreconsumo"
                                )
                            
                            st.plotly_chart(fig, use_container_width=True)
                        
                        with tab2:
                            # Calcular estadísticas
                            stats = {
                                "Promedio": df['Cantidad litros'].mean(),
                                "Mediana": df['Cantidad litros'].median(),
                                "Desv. Estándar": df['Cantidad litros'].std(),
                                "Mínimo": df['Cantidad litros'].min(),
                                "Máximo": df['Cantidad litros'].max(),
                                "Percentil 25%": df['Cantidad litros'].quantile(0.25),
                                "Percentil 75%": df['Cantidad litros'].quantile(0.75),
                                "Moda": df['Cantidad litros'].mode().iloc[0] if not df['Cantidad litros'].mode().empty else None
                            }
                            
                            # Mostrar estadísticas
                            st.markdown(f"""
                            <div class="card">
                                <h4>Estadísticas Descriptivas</h4>
                                <div style="display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 20px;">
                                    {create_mini_stat("📊", f"{stats['Promedio']:.2f} L", "Promedio")}
                                    {create_mini_stat("📏", f"{stats['Mediana']:.2f} L", "Mediana")}
                                    {create_mini_stat("📈", f"{stats['Desv. Estándar']:.2f} L", "Desviación Estándar")}
                                    {create_mini_stat("📉", f"{stats['Mínimo']:.2f} L", "Mínimo")}
                                    {create_mini_stat("📈", f"{stats['Máximo']:.2f} L", "Máximo")}
                                    {create_mini_stat("🔢", f"{stats['Moda']:.2f} L", "Moda")}
                                </div>
                            </div>
                            """, unsafe_allow_html=True)
                            
                            # Mostrar distribución por rangos
                            rangos = [0, 50, 100, 150, 200, 250, 300, float('inf')]
                            etiquetas = ['0-50', '50-100', '100-150', '150-200', '200-250', '250-300', '>300']
                            
                            df['Rango Litros'] = pd.cut(df['Cantidad litros'], bins=rangos, labels=etiquetas, right=False)
                            rango_counts = df['Rango Litros'].value_counts().sort_index()
                            
                            # Crear gráfico de barras para rangos
                            fig = px.bar(
                                x=rango_counts.index,
                                y=rango_counts.values,
                                title="Distribución por Rangos de Litros",
                                labels={'x': 'Rango de Litros', 'y': 'Cantidad de Cargas'},
                                color=rango_counts.values,
                                color_continuous_scale='Viridis'
                            )
                            
                            st.plotly_chart(fig, use_container_width=True)
                    
                    # Resumen de alertas
                    st.markdown("<h3 class='section-title'>Resumen de Alertas</h3>", unsafe_allow_html=True)
                    
                    # Crear un dataframe para el resumen
                    alertas = []
                    
                    if 'Mala Carga' in df.columns:
                        malas_cargas = df['Mala Carga'].sum()
                        if malas_cargas > 0:
                            porcentaje = (malas_cargas / len(df)) * 100
                            nivel = "Alto" if porcentaje > 5 else "Medio" if porcentaje > 2 else "Bajo"
                            alertas.append({
                                "Tipo": "Malas Cargas",
                                "Cantidad": malas_cargas,
                                "Porcentaje": f"{porcentaje:.2f}%",
                                "Nivel": nivel,
                                "Impacto": "Alto" if nivel == "Alto" else "Medio" if nivel == "Medio" else "Bajo"
                            })
                    
                    if 'Sobreconsumo' in df.columns:
                        sobreconsumos = df['Sobreconsumo'].sum()
                        if sobreconsumos > 0:
                            porcentaje = (sobreconsumos / len(df)) * 100
                            nivel = "Alto" if porcentaje > 10 else "Medio" if porcentaje > 5 else "Bajo"
                            alertas.append({
                                "Tipo": "Sobreconsumos",
                                "Cantidad": sobreconsumos,
                                "Porcentaje": f"{porcentaje:.2f}%",
                                "Nivel": nivel,
                                "Impacto": "Alto" if nivel == "Alto" else "Medio" if nivel == "Medio" else "Bajo"
                            })
                    
                    if 'Outlier Extremo' in df.columns:
                        outliers = df['Outlier Extremo'].sum()
                        if outliers > 0:
                            porcentaje = (outliers / len(df)) * 100
                            nivel = "Alto" if porcentaje > 3 else "Medio" if porcentaje > 1 else "Bajo"
                            alertas.append({
                                "Tipo": "Outliers Extremos",
                                "Cantidad": outliers,
                                "Porcentaje": f"{porcentaje:.2f}%",
                                "Nivel": nivel,
                                "Impacto": "Alto" if nivel == "Alto" else "Medio" if nivel == "Medio" else "Bajo"
                            })
                    
                    if alertas:
                        alertas_df = pd.DataFrame(alertas)
                        
                        # Mostrar tabla
                        st.dataframe(
                            alertas_df,
                            column_config={
                                "Tipo": st.column_config.TextColumn("Tipo de Alerta"),
                                "Cantidad": st.column_config.NumberColumn("Cantidad", format="%d"),
                                "Porcentaje": st.column_config.TextColumn("% del Total"),
                                "Nivel": st.column_config.TextColumn("Nivel de Alerta"),
                                "Impacto": st.column_config.TextColumn("Impacto Operativo")
                            },
                            use_container_width=True
                        )
                        
                        # Generar recomendaciones basadas en alertas
                        st.markdown(f"""
                        <div class="card">
                            <h4>Recomendaciones Basadas en Alertas</h4>
                            <ul>
                        """, unsafe_allow_html=True)
                        
                        for alerta in alertas:
                            if alerta["Tipo"] == "Malas Cargas" and alerta["Nivel"] in ["Medio", "Alto"]:
                                st.markdown(f"""
                                <li><strong>Malas Cargas:</strong> Revisar los procedimientos de registro y capacitar al personal en la correcta documentación de cargas, especialmente en las terminales con mayor incidencia.</li>
                                """, unsafe_allow_html=True)
                            
                            if alerta["Tipo"] == "Sobreconsumos" and alerta["Nivel"] in ["Medio", "Alto"]:
                                st.markdown(f"""
                                <li><strong>Sobreconsumos:</strong> Programar revisiones mecánicas para los buses con mayor incidencia de sobreconsumo y evaluar patrones de conducción de los operadores con más alertas.</li>
                                """, unsafe_allow_html=True)
                            
                            if alerta["Tipo"] == "Outliers Extremos" and alerta["Nivel"] in ["Medio", "Alto"]:
                                st.markdown(f"""
                                <li><strong>Outliers Extremos:</strong> Investigar las causas de estas desviaciones significativas, que podrían indicar fugas de combustible, errores de registro o problemas mecánicos graves.</li>
                                """, unsafe_allow_html=True)
                        
                        st.markdown(f"""
                            </ul>
                        </div>
                        """, unsafe_allow_html=True)
                    else:
                        st.success("✅ No se detectaron alertas significativas en los datos analizados.")
                    
                    # Top 10 buses con mayor consumo
                    st.markdown("<h3 class='section-title'>Top 10 Buses con Mayor Consumo</h3>", unsafe_allow_html=True)
                    
                    if 'Número interno' in df.columns and 'Cantidad litros' in df.columns:
                        # Agrupar por bus y calcular total de litros
                        top_buses = df.groupby('Número interno')['Cantidad litros'].sum().reset_index()
                        top_buses = top_buses.sort_values('Cantidad litros', ascending=False).head(10)
                        
                        fig = px.bar(
                            top_buses,
                            x='Número interno',
                            y='Cantidad litros',
                            title='Top 10 Buses con Mayor Consumo Total',
                            color='Cantidad litros',
                            color_continuous_scale='Viridis',
                            text='Cantidad litros'
                        )
                        
                        fig.update_traces(
                            texttemplate='%{text:.1f}',
                            textposition='outside'
                        )
                        
                        fig.update_layout(
                            xaxis_title="Número de Bus",
                            yaxis_title="Total Litros",
                            plot_bgcolor='white'
                        )
                        
                        st.plotly_chart(fig, use_container_width=True)
                
                # === ANÁLISIS POR TERMINAL ===
                elif section == "Análisis por Terminal":
                    st.markdown("<h2 class='subtitle'>Análisis por Terminal</h2>", unsafe_allow_html=True)
                    
                    if 'Terminal' in df.columns:
                        # Selector de terminal
                        terminales = sorted(df['Terminal'].unique())
                        col1, col2, col3 = st.columns([3, 1, 1])
                        
                        with col1:
                            terminal_seleccionada = st.selectbox("Seleccionar Terminal", ['Todas'] + list(terminales))
                        
                        with col2:
                            if 'Fecha' in df.columns:
                                fecha_min = df['Fecha'].min().date()
                                fecha_max = df['Fecha'].max().date()
                            else:
                                fecha_min = datetime.date.today() - datetime.timedelta(days=30)
                                fecha_max = datetime.date.today()
                            
                            fecha_inicio = st.date_input("Fecha inicio", fecha_min)
                        
                        with col3:
                            fecha_fin = st.date_input("Fecha fin", fecha_max)
                        
                        # Filtrar datos por terminal seleccionada y fechas
                        if terminal_seleccionada != 'Todas':
                            df_terminal = df[df['Terminal'] == terminal_seleccionada]
                        else:
                            df_terminal = df.copy()
                        
                        if 'Fecha' in df.columns:
                            df_terminal = df_terminal[
                                (df_terminal['Fecha'].dt.date >= fecha_inicio) & 
                                (df_terminal['Fecha'].dt.date <= fecha_fin)
                            ]
                        
                        # Métricas principales
                        col1, col2, col3, col4 = st.columns(4)
                        
                        with col1:
                            cargas_terminal = len(df_terminal)
                            total_cargas = len(df)
                            porcentaje_cargas = (cargas_terminal / total_cargas) * 100 if total_cargas > 0 else 0
                            
                            st.markdown(create_kpi_card(
                                "Total Cargas", 
                                f"{cargas_terminal:,}",
                                porcentaje_cargas,
                                icon="🚌",
                                color_class="primary",
                                footnote=f"{porcentaje_cargas:.2f}% del total"
                            ), unsafe_allow_html=True)
                        
                        with col2:
                            if 'Cantidad litros' in df_terminal.columns:
                                litros_terminal = df_terminal['Cantidad litros'].sum()
                                total_litros = df['Cantidad litros'].sum()
                                porcentaje_litros = (litros_terminal / total_litros) * 100 if total_litros > 0 else 0
                                
                                st.markdown(create_kpi_card(
                                    "Total Litros", 
                                    f"{litros_terminal:,.0f}",
                                    porcentaje_litros,
                                    icon="⛽",
                                    color_class="info",
                                    footnote=f"{porcentaje_litros:.2f}% del total"
                                ), unsafe_allow_html=True)
                        
                        with col3:
                            if 'Cantidad litros' in df_terminal.columns:
                                promedio_terminal = df_terminal['Cantidad litros'].mean()
                                promedio_global = df['Cantidad litros'].mean()
                                diferencia_pct = ((promedio_terminal / promedio_global) - 1) * 100 if promedio_global > 0 else 0
                                
                                color_class = "success" if diferencia_pct < 0 else "warning" if diferencia_pct < 5 else "danger"
                                
                                st.markdown(create_kpi_card(
                                    "Promedio L/Carga", 
                                    f"{promedio_terminal:.2f}",
                                    diferencia_pct,
                                    icon="📊",
                                    color_class=color_class,
                                    footnote=f"vs {promedio_global:.2f} global"
                                ), unsafe_allow_html=True)
                        
                        with col4:
                            if 'Número interno' in df_terminal.columns:
                                buses_terminal = df_terminal['Número interno'].nunique()
                                cargas_por_bus = cargas_terminal / buses_terminal if buses_terminal > 0 else 0
                                
                                st.markdown(create_kpi_card(
                                    "Buses Atendidos", 
                                    f"{buses_terminal:,}",
                                    None,
                                    icon="🚍",
                                    color_class="success",
                                    footnote=f"{cargas_por_bus:.2f} cargas/bus"
                                ), unsafe_allow_html=True)
                        
                        # Métricas de alertas
                        col1, col2, col3 = st.columns(3)
                        
                        with col1:
                            if 'Mala Carga' in df_terminal.columns:
                                malas_terminal = df_terminal['Mala Carga'].sum()
                                porcentaje_malas = (malas_terminal / len(df_terminal)) * 100 if len(df_terminal) > 0 else 0
                                
                                # Comparar con el promedio global
                                porcentaje_global = (df['Mala Carga'].sum() / len(df)) * 100 if len(df) > 0 else 0
                                diferencia = porcentaje_malas - porcentaje_global
                                
                                color = "success" if porcentaje_malas < 2 else "warning" if porcentaje_malas < 5 else "danger"
                                
                                st.markdown(create_kpi_card(
                                    "Malas Cargas", 
                                    f"{malas_terminal} ({porcentaje_malas:.2f}%)",
                                    -diferencia,  # Negativo porque menos es mejor
                                    icon="❌",
                                    color_class=color,
                                    footnote=f"vs {porcentaje_global:.2f}% global"
                                ), unsafe_allow_html=True)
                        
                        with col2:
                            if 'Sobreconsumo' in df_terminal.columns:
                                sobre_terminal = df_terminal['Sobreconsumo'].sum()
                                porcentaje_sobre = (sobre_terminal / len(df_terminal)) * 100 if len(df_terminal) > 0 else 0
                                
                                # Comparar con el promedio global
                                porcentaje_global = (df['Sobreconsumo'].sum() / len(df)) * 100 if len(df) > 0 else 0
                                diferencia = porcentaje_sobre - porcentaje_global
                                
                                color = "success" if porcentaje_sobre < 5 else "warning" if porcentaje_sobre < 10 else "danger"
                                
                                st.markdown(create_kpi_card(
                                    "Sobreconsumos", 
                                    f"{sobre_terminal} ({porcentaje_sobre:.2f}%)",
                                    -diferencia,  # Negativo porque menos es mejor
                                    icon="⚠️",
                                    color_class=color,
                                    footnote=f"vs {porcentaje_global:.2f}% global"
                                ), unsafe_allow_html=True)
                        
                        with col3:
                            if 'Rendimiento' in df_terminal.columns:
                                rendimiento_terminal = df_terminal['Rendimiento'].mean()
                                rendimiento_global = df['Rendimiento'].mean()
                                diferencia_pct = ((rendimiento_terminal / rendimiento_global) - 1) * 100 if rendimiento_global > 0 else 0
                                
                                color_class = "success" if diferencia_pct > 0 else "warning" if diferencia_pct > -10 else "danger"
                                
                                st.markdown(create_kpi_card(
                                    "Rendimiento (km/L)", 
                                    f"{rendimiento_terminal:.2f}",
                                    diferencia_pct,
                                    icon="🔄",
                                    color_class=color_class,
                                    footnote=f"vs {rendimiento_global:.2f} global"
                                ), unsafe_allow_html=True)
                        
                        # Si es "Todas", mostrar comparativa entre terminales
                        if terminal_seleccionada == 'Todas':
                            st.markdown("<h3 class='section-title'>Comparativa entre Terminales</h3>", unsafe_allow_html=True)
                            
                            # Agrupar datos por terminal
                            terminal_stats = df.groupby('Terminal').agg({
                                'Cantidad litros': ['sum', 'mean', 'count'] if 'Cantidad litros' in df.columns else 'count',
                                'Mala Carga': 'sum' if 'Mala Carga' in df.columns else 'count',
                                'Sobreconsumo': 'sum' if 'Sobreconsumo' in df.columns else 'count'
                            }).reset_index()
                            
                            # Aplanar los nombres de columnas multiíndice
                            if 'Cantidad litros' in df.columns:
                                terminal_stats.columns = ['Terminal', 'Total Litros', 'Promedio Litros', 'Cantidad Cargas', 
                                                        'Malas Cargas', 'Sobreconsumos']
                                
                                # Calcular porcentajes e índice de eficiencia
                                terminal_stats['% Malas Cargas'] = (terminal_stats['Malas Cargas'] / terminal_stats['Cantidad Cargas'] * 100).round(2)
                                terminal_stats['% Sobreconsumos'] = (terminal_stats['Sobreconsumos'] / terminal_stats['Cantidad Cargas'] * 100).round(2)
                                terminal_stats['Índice Eficiencia'] = 100 - (terminal_stats['% Sobreconsumos'] * 0.7 + terminal_stats['% Malas Cargas'] * 0.3)
                                terminal_stats['Índice Eficiencia'] = terminal_stats['Índice Eficiencia'].clip(0, 100).round(2)
                                
                                # Añadir ranking
                                terminal_stats['Ranking'] = terminal_stats['Índice Eficiencia'].rank(ascending=False).astype(int)
                                
                                # Tabs para diferentes vistas
                                tab1, tab2, tab3 = st.tabs(["📊 Gráfico", "📋 Tabla", "🔍 Detalles"])
                                
                                with tab1:
                                    col1, col2 = st.columns(2)
                                    
                                    with col1:
                                        # Gráfico de barras comparativo (Litros)
                                        fig = px.bar(
                                            terminal_stats,
                                            x='Terminal',
                                            y='Total Litros',
                                            title='Consumo Total por Terminal',
                                            color='Índice Eficiencia',
                                            color_continuous_scale='RdYlGn',
                                            text='Total Litros'
                                        )
                                        
                                        fig.update_traces(
                                            texttemplate='%{text:.0f}',
                                            textposition='outside'
                                        )
                                        
                                        fig.update_layout(
                                            xaxis_title="Terminal",
                                            yaxis_title="Total Litros",
                                            plot_bgcolor='white',
                                            height=450
                                        )
                                        
                                        st.plotly_chart(fig, use_container_width=True)
                                    
                                    with col2:
                                        # Gráfico de eficiencia
                                        fig = px.bar(
                                            terminal_stats,
                                            x='Terminal',
                                            y='Índice Eficiencia',
                                            title='Índice de Eficiencia por Terminal',
                                            color='Índice Eficiencia',
                                            color_continuous_scale='RdYlGn',
                                            text='Índice Eficiencia'
                                        )
                                        
                                        fig.update_traces(
                                            texttemplate='%{text:.1f}',
                                            textposition='outside'
                                        )
                                        
                                        fig.update_layout(
                                            xaxis_title="Terminal",
                                            yaxis_title="Índice de Eficiencia",
                                            plot_bgcolor='white',
                                            height=450
                                        )
                                        
                                        # Añadir línea de referencia para el promedio
                                        promedio_eficiencia = terminal_stats['Índice Eficiencia'].mean()
                                        fig.add_hline(
                                            y=promedio_eficiencia, 
                                            line_dash="dash", 
                                            line_color="gray",
                                            annotation_text=f"Promedio ({promedio_eficiencia:.2f})"
                                        )
                                        
                                        st.plotly_chart(fig, use_container_width=True)
                                
                                with tab2:
                                    # Ordenar por índice de eficiencia (descendente)
                                    terminal_stats_display = terminal_stats.sort_values('Índice Eficiencia', ascending=False)
                                    
                                    # Mostrar tabla
                                    st.dataframe(
                                        terminal_stats_display,
                                        column_config={
                                            "Terminal": st.column_config.TextColumn("Terminal"),
                                            "Total Litros": st.column_config.NumberColumn("Total Litros", format="%,.2f"),
                                            "Promedio Litros": st.column_config.NumberColumn("Prom. L/Carga", format="%.2f"),
                                            "Cantidad Cargas": st.column_config.NumberColumn("# Cargas", format="%d"),
                                            "Malas Cargas": st.column_config.NumberColumn("Malas", format="%d"),
                                            "% Malas Cargas": st.column_config.NumberColumn("% Malas", format="%.2f%%"),
                                            "Sobreconsumos": st.column_config.NumberColumn("Sobre", format="%d"),
                                            "% Sobreconsumos": st.column_config.NumberColumn("% Sobre", format="%.2f%%"),
                                            "Índice Eficiencia": st.column_config.ProgressColumn("Eficiencia", format="%.2f", min_value=0, max_value=100),
                                            "Ranking": st.column_config.NumberColumn("Ranking", format="%d")
                                        },
                                        use_container_width=True
                                    )
                                
                                with tab3:
                                    # Terminal más eficiente
                                    mejor_terminal = terminal_stats.loc[terminal_stats['Índice Eficiencia'].idxmax()]
                                    
                                    st.markdown(f"""
                                    <div class="card">
                                        <h4>🏆 Terminal más Eficiente: {mejor_terminal['Terminal']}</h4>
                                        <p><strong>Índice de Eficiencia:</strong> {mejor_terminal['Índice Eficiencia']:.2f}/100</p>
                                        <p><strong>Características destacadas:</strong></p>
                                        <ul>
                                            <li>Promedio de litros por carga: {mejor_terminal['Promedio Litros']:.2f}</li>
                                            <li>Tasa de malas cargas: {mejor_terminal['% Malas Cargas']:.2f}%</li>
                                            <li>Tasa de sobreconsumos: {mejor_terminal['% Sobreconsumos']:.2f}%</li>
                                        </ul>
                                    </div>
                                    """, unsafe_allow_html=True)
                                    
                                    # Terminal menos eficiente
                                    peor_terminal = terminal_stats.loc[terminal_stats['Índice Eficiencia'].idxmin()]
                                    
                                    st.markdown(f"""
                                    <div class="card">
                                        <h4>⚠️ Terminal con Mayor Oportunidad de Mejora: {peor_terminal['Terminal']}</h4>
                                        <p><strong>Índice de Eficiencia:</strong> {peor_terminal['Índice Eficiencia']:.2f}/100</p>
                                        <p><strong>Áreas de oportunidad:</strong></p>
                                        <ul>
                                            <li>Promedio de litros por carga: {peor_terminal['Promedio Litros']:.2f}</li>
                                            <li>Tasa de malas cargas: {peor_terminal['% Malas Cargas']:.2f}%</li>
                                            <li>Tasa de sobreconsumos: {peor_terminal['% Sobreconsumos']:.2f}%</li>
                                        </ul>
                                        <p><strong>Recomendaciones:</strong></p>
                                        <ul>
                                            <li>Realizar capacitación al personal sobre procedimientos correctos de registro</li>
                                            <li>Revisar calibración de surtidores</li>
                                            <li>Implementar un sistema de supervisión más estricto</li>
                                        </ul>
                                    </div>
                                    """, unsafe_allow_html=True)
                                
                                # Gráfico tipo radar para comparar terminales
                                st.markdown("<h3 class='section-title'>Comparativa Multidimensional</h3>", unsafe_allow_html=True)
                                
                                # Seleccionar terminales a comparar
                                terminales_comparar = st.multiselect(
                                    "Seleccionar terminales a comparar",
                                    options=terminal_stats['Terminal'].tolist(),
                                    default=terminal_stats.nlargest(3, 'Cantidad Cargas')['Terminal'].tolist()
                                )
                                
                                if terminales_comparar:
                                    # Filtrar datos para las terminales seleccionadas
                                    comparacion = terminal_stats[terminal_stats['Terminal'].isin(terminales_comparar)]
                                    
                                    # Normalizar valores para el gráfico radar
                                    for col in ['Promedio Litros', '% Malas Cargas', '% Sobreconsumos']:
                                        if col == 'Promedio Litros':
                                            # Para litros, menores valores son mejores (normalización inversa)
                                            min_val = comparacion[col].min()
                                            max_val = comparacion[col].max()
                                            if max_val > min_val:
                                                comparacion[f'{col} (Norm)'] = 1 - (comparacion[col] - min_val) / (max_val - min_val)
                                            else:
                                                comparacion[f'{col} (Norm)'] = 1
                                        else:
                                            # Para porcentajes de problemas, menores valores son mejores (normalización inversa)
                                            max_val = comparacion[col].max()
                                            if max_val > 0:
                                                comparacion[f'{col} (Norm)'] = 1 - comparacion[col] / max_val
                                            else:
                                                comparacion[f'{col} (Norm)'] = 1
                                    
                                    # Añadir el índice de eficiencia normalizado
                                    comparacion['Eficiencia (Norm)'] = comparacion['Índice Eficiencia'] / 100
                                    
                                    # Crear el gráfico radar
                                    fig = go.Figure()
                                    
                                    categories = ['Litros (Eficiencia)', 'Malas Cargas (Inverso)', 
                                                'Sobreconsumos (Inverso)', 'Índice Eficiencia']
                                    
                                    # Colores para cada terminal
                                    colors = px.colors.qualitative.Plotly[:len(comparacion)]
                                    
                                    for i, row in comparacion.iterrows():
                                        fig.add_trace(go.Scatterpolar(
                                            r=[row['Promedio Litros (Norm)'], row['% Malas Cargas (Norm)'], 
                                              row['% Sobreconsumos (Norm)'], row['Eficiencia (Norm)']],
                                            theta=categories,
                                            fill='toself',
                                            name=row['Terminal'],
                                            line_color=colors[i % len(colors)]
                                        ))
                                    
                                    fig.update_layout(
                                        polar=dict(
                                            radialaxis=dict(
                                                visible=True,
                                                range=[0, 1]
                                            )
                                        ),
                                        title="Comparativa Multidimensional entre Terminales",
                                        height=500
                                    )
                                    
                                    st.plotly_chart(fig, use_container_width=True)
                                    
                                    # Explicación del gráfico
                                    st.markdown(f"""
                                    <div class="card">
                                        <h4>📊 Interpretación del Gráfico Radar</h4>
                                        <p>Este gráfico permite comparar múltiples dimensiones de desempeño entre terminales:</p>
                                        <ul>
                                            <li><strong>Litros (Eficiencia):</strong> Mayor valor indica menor consumo promedio por carga.</li>
                                            <li><strong>Malas Cargas (Inverso):</strong> Mayor valor indica menor porcentaje de malas cargas.</li>
                                            <li><strong>Sobreconsumos (Inverso):</strong> Mayor valor indica menor porcentaje de sobreconsumos.</li>
                                            <li><strong>Índice Eficiencia:</strong> Mayor valor indica mejor eficiencia general.</li>
                                        </ul>
                                        <p><em>Nota: Todos los valores están normalizados para permitir una comparación justa.</em></p>
                                    </div>
                                    """, unsafe_allow_html=True)
                        else:
                            # Análisis específico de la terminal seleccionada
                            st.markdown(f"<h3 class='section-title'>Análisis Detallado: Terminal {terminal_seleccionada}</h3>", unsafe_allow_html=True)
                            
                            # Evolución temporal del consumo
                            if 'Fecha' in df_terminal.columns and 'Cantidad litros' in df_terminal.columns:
                                st.markdown("<h4>Evolución Temporal</h4>", unsafe_allow_html=True)
                                
                                tab1, tab2 = st.tabs(["📊 Consumo Diario", "📈 Tendencia Acumulada"])
                                
                                with tab1:
                                    fig = plot_time_series(
                                        df_terminal, 
                                        'Cantidad litros', 
                                        f'Evolución del Consumo - Terminal {terminal_seleccionada}'
                                    )
                                    
                                    if fig:
                                        st.plotly_chart(fig, use_container_width=True)
                                
                                with tab2:
                                    # Gráfico de consumo acumulado
                                    consumo_diario = df_terminal.groupby(df_terminal['Fecha'].dt.date)['Cantidad litros'].sum().reset_index()
                                    consumo_diario['Consumo Acumulado'] = consumo_diario['Cantidad litros'].cumsum()
                                    
                                    fig = px.line(
                                        consumo_diario, 
                                        x='Fecha', 
                                        y='Consumo Acumulado',
                                        title=f'Consumo Acumulado - Terminal {terminal_seleccionada}',
                                        markers=True
                                    )
                                    
                                    fig.update_traces(
                                        line=dict(color=COLORS['primary'], width=3),
                                        marker=dict(color=COLORS['primary'], size=8)
                                    )
                                    
                                    fig.update_layout(
                                        xaxis_title="Fecha",
                                        yaxis_title="Litros Acumulados",
                                        plot_bgcolor='white',
                                        hovermode='x unified',
                                        height=400
                                    )
                                    
                                    st.plotly_chart(fig, use_container_width=True)
                            
                            # Distribución por hora y período
                            col1, col2 = st.columns(2)
                            
                            with col1:
                                st.markdown("<h4>Distribución por Hora</h4>", unsafe_allow_html=True)
                                
                                if 'Hora Numérica' in df_terminal.columns:
                                    # Agrupar por hora
                                    hora_stats = df_terminal.groupby(np.round(df_terminal['Hora Numérica']).astype(int)).size().reset_index()
                                    hora_stats.columns = ['Hora', 'Cargas']
                                    
                                    fig = px.bar(
                                        hora_stats,
                                        x='Hora',
                                        y='Cargas',
                                        title='Cargas por Hora del Día',
                                        color='Cargas',
                                       color_continuous_scale='Viridis',
                                        text='Cargas'
                                    )
                                    
                                    fig.update_traces(
                                        texttemplate='%{text}',
                                        textposition='outside'
                                    )
                                    
                                    fig.update_layout(
                                        xaxis_title="Hora del Día",
                                        yaxis_title="Cantidad de Cargas",
                                        plot_bgcolor='white'
                                    )
                                    
                                    st.plotly_chart(fig, use_container_width=True)
                            
                            with col2:
                                st.markdown("<h4>Distribución por Período</h4>", unsafe_allow_html=True)
                                
                                if 'Período' in df_terminal.columns:
                                    # Agrupar por período
                                    periodo_stats = df_terminal.groupby('Período').size().reset_index()
                                    periodo_stats.columns = ['Período', 'Cargas']
                                    
                                    # Ordenar períodos
                                    orden_periodo = {'Madrugada': 0, 'Mañana': 1, 'Tarde': 2, 'Noche': 3}
                                    periodo_stats['Orden'] = periodo_stats['Período'].map(orden_periodo)
                                    periodo_stats = periodo_stats.sort_values('Orden')
                                    
                                    fig = px.pie(
                                        periodo_stats,
                                        names='Período',
                                        values='Cargas',
                                        title='Distribución por Período del Día',
                                        color_discrete_sequence=px.colors.sequential.Viridis
                                    )
                                    
                                    fig.update_layout(
                                        plot_bgcolor='white',
                                        legend=dict(
                                            orientation="h",
                                            yanchor="bottom",
                                            y=-0.2,
                                            xanchor="center",
                                            x=0.5
                                        )
                                    )
                                    
                                    st.plotly_chart(fig, use_container_width=True)
                            
                            # Análisis por modelo
                            if 'Modelo chasis' in df_terminal.columns and 'Cantidad litros' in df_terminal.columns:
                                st.markdown("<h4>Consumo por Modelo de Chasis</h4>", unsafe_allow_html=True)
                                
                                # Agrupar por modelo
                                modelo_stats = df_terminal.groupby('Modelo chasis').agg({
                                    'Cantidad litros': ['sum', 'mean', 'count'],
                                    'Sobreconsumo': 'sum' if 'Sobreconsumo' in df_terminal.columns else 'count'
                                }).reset_index()
                                
                                # Aplanar los nombres de columnas multiíndice
                                modelo_stats.columns = ['Modelo', 'Total Litros', 'Promedio Litros', 'Cantidad Cargas', 'Sobreconsumos']
                                
                                # Calcular porcentajes
                                modelo_stats['% Sobreconsumos'] = (modelo_stats['Sobreconsumos'] / modelo_stats['Cantidad Cargas'] * 100).round(2)
                                
                                # Ordenar por total litros
                                modelo_stats = modelo_stats.sort_values('Total Litros', ascending=False)
                                
                                # Mostrar gráfico y tabla en pestañas
                                tab1, tab2 = st.tabs(["📊 Gráfico", "📋 Tabla"])
                                
                                with tab1:
                                    fig = px.bar(
                                        modelo_stats.head(10),  # Top 10 modelos
                                        x='Modelo',
                                        y='Total Litros',
                                        title=f'Consumo por Modelo - Terminal {terminal_seleccionada}',
                                        color='Promedio Litros',
                                        color_continuous_scale='Viridis',
                                        text='Total Litros'
                                    )
                                    
                                    fig.update_traces(
                                        texttemplate='%{text:.0f}',
                                        textposition='outside'
                                    )
                                    
                                    fig.update_layout(
                                        xaxis_title="Modelo de Chasis",
                                        yaxis_title="Total Litros",
                                        plot_bgcolor='white'
                                    )
                                    
                                    st.plotly_chart(fig, use_container_width=True)
                                
                                with tab2:
                                    # Mostrar tabla
                                    st.dataframe(
                                        modelo_stats,
                                        column_config={
                                            "Modelo": st.column_config.TextColumn("Modelo de Chasis"),
                                            "Total Litros": st.column_config.NumberColumn("Total Litros", format="%,.2f"),
                                            "Promedio Litros": st.column_config.NumberColumn("Promedio L/Carga", format="%.2f"),
                                            "Cantidad Cargas": st.column_config.NumberColumn("Cargas", format="%d"),
                                            "Sobreconsumos": st.column_config.NumberColumn("Sobrec.", format="%d"),
                                            "% Sobreconsumos": st.column_config.NumberColumn("% Sobre", format="%.2f%%")
                                        },
                                        use_container_width=True
                                    )
                            
                            # Mapa de calor por horario
                            st.markdown("<h4>Distribución de Cargas por Horario</h4>", unsafe_allow_html=True)
                            
                            if 'Hora Numérica' in df_terminal.columns and 'Día Semana' in df_terminal.columns:
                                fig = create_heatmap(df_terminal, title=f"Mapa de calor - Terminal {terminal_seleccionada}")
                                if fig:
                                    st.plotly_chart(fig, use_container_width=True)
                            
                            # Análisis de malas cargas y sobreconsumos
                            col1, col2 = st.columns(2)
                            
                            with col1:
                                st.markdown("<h4>Malas Cargas</h4>", unsafe_allow_html=True)
                                
                                if 'Mala Carga' in df_terminal.columns:
                                    malas_cargas = df_terminal[df_terminal['Mala Carga'] == True]
                                    
                                    # Mostrar estadísticas
                                    porcentaje_malas = (len(malas_cargas) / len(df_terminal)) * 100 if len(df_terminal) > 0 else 0
                                    
                                    # Comparar con el promedio global
                                    global_pct = (df['Mala Carga'].sum() / len(df)) * 100 if len(df) > 0 else 0
                                    
                                    if porcentaje_malas > global_pct:
                                        st.markdown(show_alert(
                                            f"La terminal tiene un {porcentaje_malas:.2f}% de malas cargas, por encima del promedio global ({global_pct:.2f}%).",
                                            "warning",
                                            "Tasa de Malas Cargas Superior al Promedio",
                                            "⚠️"
                                        ), unsafe_allow_html=True)
                                    else:
                                        st.markdown(show_alert(
                                            f"La terminal tiene un {porcentaje_malas:.2f}% de malas cargas, por debajo del promedio global ({global_pct:.2f}%).",
                                            "success",
                                            "Buen Desempeño en Registro",
                                            "✅"
                                        ), unsafe_allow_html=True)
                                    
                                    # Mostrar tabla si hay malas cargas
                                    if not malas_cargas.empty:
                                        st.markdown("##### Últimas Malas Cargas")
                                        
                                        cols_relevantes = [col for col in [
                                            'Fecha', 'Hora', 'Número interno', 'Patente', 
                                            'Cantidad litros', 'Nombre Planillero'
                                        ] if col in malas_cargas.columns]
                                        
                                        st.dataframe(
                                            malas_cargas[cols_relevantes].sort_values('Fecha', ascending=False).head(5),
                                            use_container_width=True
                                        )
                            
                            with col2:
                                st.markdown("<h4>Sobreconsumos</h4>", unsafe_allow_html=True)
                                
                                if 'Sobreconsumo' in df_terminal.columns:
                                    sobreconsumos = df_terminal[df_terminal['Sobreconsumo'] == True]
                                    
                                    # Mostrar estadísticas
                                    porcentaje_sobre = (len(sobreconsumos) / len(df_terminal)) * 100 if len(df_terminal) > 0 else 0
                                    
                                    # Comparar con el promedio global
                                    global_pct = (df['Sobreconsumo'].sum() / len(df)) * 100 if len(df) > 0 else 0
                                    
                                    if porcentaje_sobre > global_pct:
                                        st.markdown(show_alert(
                                            f"La terminal tiene un {porcentaje_sobre:.2f}% de sobreconsumos, por encima del promedio global ({global_pct:.2f}%).",
                                            "warning",
                                            "Tasa de Sobreconsumo Superior al Promedio",
                                            "⚠️"
                                        ), unsafe_allow_html=True)
                                    else:
                                        st.markdown(show_alert(
                                            f"La terminal tiene un {porcentaje_sobre:.2f}% de sobreconsumos, por debajo del promedio global ({global_pct:.2f}%).",
                                            "success",
                                            "Buen Desempeño en Eficiencia",
                                            "✅"
                                        ), unsafe_allow_html=True)
                                    
                                    # Mostrar tabla si hay sobreconsumos
                                    if not sobreconsumos.empty:
                                        st.markdown("##### Últimos Sobreconsumos")
                                        
                                        cols_relevantes = [col for col in [
                                            'Fecha', 'Hora', 'Número interno', 'Modelo chasis',
                                            'Cantidad litros', 'Z-Score'
                                        ] if col in sobreconsumos.columns]
                                        
                                        st.dataframe(
                                            sobreconsumos[cols_relevantes].sort_values('Z-Score', ascending=False).head(5),
                                            use_container_width=True
                                        )
                            
                            # Tabla detallada de datos
                            with st.expander("Datos Detallados", expanded=False):
                                # Columnas relevantes para mostrar
                                cols_relevantes = [col for col in [
                                    'Fecha', 'Hora', 'Número interno', 'Patente', 'Modelo chasis',
                                    'Cantidad litros', 'Nombre conductor', 'Nombre Planillero',
                                    'Mala Carga', 'Sobreconsumo'
                                ] if col in df_terminal.columns]
                                
                                # Mostrar datos
                                st.dataframe(
                                    df_terminal[cols_relevantes].sort_values('Fecha', ascending=False),
                                    use_container_width=True,
                                    height=400
                                )
                    else:
                        st.warning("No se encontró la columna 'Terminal' en los datos.")
                
                # === ANÁLISIS POR BUSES ===
                elif section == "Análisis por Buses":
                    st.markdown("<h2 class='subtitle'>Análisis por Buses</h2>", unsafe_allow_html=True)
                    
                    if 'Número interno' in df.columns:
                        # Selector de bus
                        buses = sorted(df['Número interno'].unique())
                        
                        col1, col2, col3 = st.columns([3, 1, 1])
                        
                        with col1:
                            bus_seleccionado = st.selectbox("Seleccionar Bus", ['Todos'] + list(buses))
                        
                        with col2:
                            if 'Fecha' in df.columns:
                                fecha_min = df['Fecha'].min().date()
                                fecha_max = df['Fecha'].max().date()
                            else:
                                fecha_min = datetime.date.today() - datetime.timedelta(days=30)
                                fecha_max = datetime.date.today()
                            
                            fecha_inicio = st.date_input("Fecha inicio", fecha_min, key="bus_fecha_inicio")
                        
                        with col3:
                            fecha_fin = st.date_input("Fecha fin", fecha_max, key="bus_fecha_fin")
                        
                        # Filtrar datos por bus seleccionado y fechas
                        if bus_seleccionado != 'Todos':
                            df_bus = df[df['Número interno'] == bus_seleccionado]
                        else:
                            df_bus = df.copy()
                        
                        if 'Fecha' in df.columns:
                            df_bus = df_bus[
                                (df_bus['Fecha'].dt.date >= fecha_inicio) & 
                                (df_bus['Fecha'].dt.date <= fecha_fin)
                            ]
                        
                        # Si seleccionamos un bus específico
                        if bus_seleccionado != 'Todos':
                            # Encabezado e info del bus
                            st.markdown(f"<h3 class='section-title'>Análisis Detallado: Bus {bus_seleccionado}</h3>", unsafe_allow_html=True)
                            
                            # Info general del bus
                            info_bus = {}
                            
                            if 'Patente' in df_bus.columns and not df_bus['Patente'].empty:
                                info_bus["Patente"] = df_bus['Patente'].iloc[0]
                            
                            if 'Modelo chasis' in df_bus.columns and not df_bus['Modelo chasis'].empty:
                                info_bus["Modelo"] = df_bus['Modelo chasis'].iloc[0]
                            
                            # Mostrar información básica
                            st.markdown("<h4>Información General</h4>", unsafe_allow_html=True)
                            
                            col1, col2, col3 = st.columns(3)
                            
                            with col1:
                                st.markdown(f"""
                                <div class="card">
                                    <h4>Identificación</h4>
                                    <p><strong>Número Interno:</strong> {bus_seleccionado}</p>
                                    <p><strong>Patente:</strong> {info_bus.get("Patente", "N/A")}</p>
                                    <p><strong>Modelo:</strong> {info_bus.get("Modelo", "N/A")}</p>
                                </div>
                                """, unsafe_allow_html=True)
                            
                            with col2:
                                st.markdown(f"""
                                <div class="card">
                                    <h4>Estadísticas de Consumo</h4>
                                    <p><strong>Total Cargas:</strong> {len(df_bus)}</p>
                                    <p><strong>Total Litros:</strong> {df_bus['Cantidad litros'].sum():.2f} L</p>
                                    <p><strong>Promedio por Carga:</strong> {df_bus['Cantidad litros'].mean():.2f} L</p>
                                </div>
                                """, unsafe_allow_html=True)
                            
                            with col3:
                                # Calcular métricas adicionales
                                malas_cargas = df_bus['Mala Carga'].sum() if 'Mala Carga' in df_bus.columns else 0
                                sobreconsumos = df_bus['Sobreconsumo'].sum() if 'Sobreconsumo' in df_bus.columns else 0
                                
                                st.markdown(f"""
                                <div class="card">
                                    <h4>Métricas de Calidad</h4>
                                    <p><strong>Malas Cargas:</strong> {malas_cargas} ({(malas_cargas/len(df_bus)*100 if len(df_bus) > 0 else 0):.2f}%)</p>
                                    <p><strong>Sobreconsumos:</strong> {sobreconsumos} ({(sobreconsumos/len(df_bus)*100 if len(df_bus) > 0 else 0):.2f}%)</p>
                                    <p><strong>Rendimiento:</strong> {df_bus['Rendimiento'].mean():.2f} km/L</p>
                                </div>
                                """, unsafe_allow_html=True)
                            
                            # Evolución de consumo
                            st.markdown("<h4>Evolución del Consumo</h4>", unsafe_allow_html=True)
                            
                            if 'Fecha' in df_bus.columns and 'Cantidad litros' in df_bus.columns:
                                tab1, tab2 = st.tabs(["📊 Consumo por Carga", "📈 Odómetro y Recorrido"])
                                
                                with tab1:
                                    # Gráfico de consumo por carga
                                    df_bus_sorted = df_bus.sort_values('Fecha')
                                    
                                    fig = px.line(
                                        df_bus_sorted,
                                        x='Fecha',
                                        y='Cantidad litros',
                                        title=f'Evolución del Consumo - Bus {bus_seleccionado}',
                                        markers=True
                                    )
                                    
                                    # Añadir media móvil si hay suficientes datos
                                    if len(df_bus_sorted) > 5:
                                        df_bus_sorted['Media Móvil'] = df_bus_sorted['Cantidad litros'].rolling(window=3, min_periods=1).mean()
                                        
                                        fig.add_scatter(
                                            x=df_bus_sorted['Fecha'],
                                            y=df_bus_sorted['Media Móvil'],
                                            mode='lines',
                                            name='Media Móvil (3 cargas)',
                                            line=dict(color='red', width=2, dash='dash')
                                        )
                                    
                                    fig.update_layout(
                                        xaxis_title="Fecha",
                                        yaxis_title="Litros por Carga",
                                        plot_bgcolor='white',
                                        hovermode='x unified',
                                        height=400
                                    )
                                    
                                    st.plotly_chart(fig, use_container_width=True)
                                
                                with tab2:
                                    if 'Odómetro' in df_bus.columns and 'Kilómetros Recorridos' in df_bus.columns:
                                        # Gráfico de odómetro y kilómetros recorridos
                                        df_bus_sorted = df_bus.sort_values('Fecha')
                                        
                                        # Crear figura con dos ejes Y
                                        fig = make_subplots(specs=[[{"secondary_y": True}]])
                                        
                                        # Añadir odómetro al eje primario
                                        fig.add_trace(
                                            go.Scatter(
                                                x=df_bus_sorted['Fecha'],
                                                y=df_bus_sorted['Odómetro'],
                                                name='Odómetro',
                                                line=dict(color=COLORS['primary'], width=2),
                                                mode='lines+markers'
                                            ),
                                            secondary_y=False
                                        )
                                        
                                        # Añadir kilómetros recorridos al eje secundario
                                        fig.add_trace(
                                            go.Bar(
                                                x=df_bus_sorted['Fecha'],
                                                y=df_bus_sorted['Kilómetros Recorridos'],
                                                name='Km Recorridos',
                                                marker_color=COLORS['secondary']
                                            ),
                                            secondary_y=True
                                        )
                                        
                                        # Actualizar ejes
                                        fig.update_layout(
                                            title=f'Evolución del Odómetro y Kilómetros Recorridos - Bus {bus_seleccionado}',
                                            plot_bgcolor='white',
                                            hovermode='x unified',
                                            height=400,
                                            legend=dict(
                                                orientation="h",
                                                yanchor="bottom",
                                                y=1.02,
                                                xanchor="right",
                                                x=1
                                            )
                                        )
                                        
                                        fig.update_xaxes(title_text="Fecha")
                                        fig.update_yaxes(title_text="Odómetro (km)", secondary_y=False)
                                        fig.update_yaxes(title_text="Kilómetros Recorridos", secondary_y=True)
                                        
                                        st.plotly_chart(fig, use_container_width=True)
                                    else:
                                        st.warning("No se encontraron datos de odómetro o kilómetros recorridos.")
                            
                            # Análisis de rendimiento
                            st.markdown("<h4>Análisis de Rendimiento</h4>", unsafe_allow_html=True)
                            
                            if 'Rendimiento' in df_bus.columns:
                                col1, col2 = st.columns(2)
                                
                                with col1:
                                    # Estadísticas de rendimiento
                                    rendimiento_medio = df_bus['Rendimiento'].mean()
                                    rendimiento_mediana = df_bus['Rendimiento'].median()
                                    rendimiento_min = df_bus['Rendimiento'].min()
                                    rendimiento_max = df_bus['Rendimiento'].max()
                                    
                                    # Comparar con el promedio del modelo
                                    rendimiento_modelo = None
                                    if 'Modelo chasis' in df_bus.columns and 'Rendimiento Promedio Modelo' in df_bus.columns:
                                        rendimiento_modelo = df_bus['Rendimiento Promedio Modelo'].mean()
                                    
                                    st.markdown(f"""
                                    <div class="card">
                                        <h4>Estadísticas de Rendimiento</h4>
                                        <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 10px;">
                                            {create_mini_stat("📊", f"{rendimiento_medio:.2f} km/L", "Promedio")}
                                            {create_mini_stat("📏", f"{rendimiento_mediana:.2f} km/L", "Mediana")}
                                            {create_mini_stat("📉", f"{rendimiento_min:.2f} km/L", "Mínimo")}
                                            {create_mini_stat("📈", f"{rendimiento_max:.2f} km/L", "Máximo")}
                                        </div>
                                        {f'<p><strong>Comparación con el Modelo:</strong> {rendimiento_medio:.2f} km/L vs {rendimiento_modelo:.2f} km/L ({((rendimiento_medio/rendimiento_modelo)-1)*100:.2f}%)</p>' if rendimiento_modelo else ''}
                                    </div>
                                    """, unsafe_allow_html=True)
                                
                                with col2:
                                    # Gráfico de distribución de rendimiento
                                    fig = px.histogram(
                                        df_bus,
                                        x='Rendimiento',
                                        title='Distribución del Rendimiento',
                                        color_discrete_sequence=[COLORS['primary']]
                                    )
                                    
                                    fig.update_layout(
                                        xaxis_title="Rendimiento (km/L)",
                                        yaxis_title="Frecuencia",
                                        plot_bgcolor='white'
                                    )
                                    
                                    # Añadir línea para la media
                                    fig.add_vline(
                                        x=rendimiento_medio,
                                        line_dash="dash",
                                        line_color="red",
                                        annotation_text=f"Media: {rendimiento_medio:.2f} km/L"
                                    )
                                    
                                    # Añadir línea para el modelo si está disponible
                                    if rendimiento_modelo:
                                        fig.add_vline(
                                            x=rendimiento_modelo,
                                            line_dash="dash",
                                            line_color="green",
                                            annotation_text=f"Modelo: {rendimiento_modelo:.2f} km/L"
                                        )
                                    
                                    st.plotly_chart(fig, use_container_width=True)
                                
                                # Evolución del rendimiento en el tiempo
                                if 'Fecha' in df_bus.columns:
                                    df_bus_sorted = df_bus.sort_values('Fecha')
                                    
                                    fig = px.line(
                                        df_bus_sorted,
                                        x='Fecha',
                                        y='Rendimiento',
                                        title='Evolución del Rendimiento en el Tiempo',
                                        markers=True
                                    )
                                    
                                    # Añadir media móvil si hay suficientes datos
                                    if len(df_bus_sorted) > 3:
                                        df_bus_sorted['Media Móvil'] = df_bus_sorted['Rendimiento'].rolling(window=3, min_periods=1).mean()
                                        
                                        fig.add_scatter(
                                            x=df_bus_sorted['Fecha'],
                                            y=df_bus_sorted['Media Móvil'],
                                            mode='lines',
                                            name='Media Móvil (3 cargas)',
                                            line=dict(color='red', width=2, dash='dash')
                                        )
                                    
                                    fig.update_layout(
                                        xaxis_title="Fecha",
                                        yaxis_title="Rendimiento (km/L)",
                                        plot_bgcolor='white',
                                        hovermode='x unified',
                                        height=400
                                    )
                                    
                                    st.plotly_chart(fig, use_container_width=True)
                            
                            # Patrones operativos
                            st.markdown("<h4>Patrones Operativos</h4>", unsafe_allow_html=True)
                            
                            if 'Hora Numérica' in df_bus.columns:
                                col1, col2 = st.columns(2)
                                
                                with col1:
                                    # Distribución por hora del día
                                    hora_stats = df_bus.groupby(np.round(df_bus['Hora Numérica']).astype(int)).size().reset_index()
                                    hora_stats.columns = ['Hora', 'Cargas']
                                    
                                    fig = px.bar(
                                        hora_stats,
                                        x='Hora',
                                        y='Cargas',
                                        title='Cargas por Hora del Día',
                                        color='Cargas',
                                        color_continuous_scale='Viridis'
                                    )
                                    
                                    fig.update_layout(
                                        xaxis_title="Hora del Día",
                                        yaxis_title="Cantidad de Cargas",
                                        plot_bgcolor='white'
                                    )
                                    
                                    st.plotly_chart(fig, use_container_width=True)
                                
                                with col2:
                                    if 'Terminal' in df_bus.columns:
                                        # Distribución por terminal
                                        terminal_stats = df_bus.groupby('Terminal').size().reset_index()
                                        terminal_stats.columns = ['Terminal', 'Cargas']
                                        
                                        fig = px.pie(
                                            terminal_stats,
                                            names='Terminal',
                                            values='Cargas',
                                            title='Distribución por Terminal',
                                            color_discrete_sequence=px.colors.sequential.Viridis
                                        )
                                        
                                        fig.update_layout(
                                            plot_bgcolor='white'
                                        )
                                        
                                        st.plotly_chart(fig, use_container_width=True)
                            
                            # Tabla de datos históricos
                            with st.expander("Historial Detallado de Cargas", expanded=False):
                                # Columnas relevantes para mostrar
                                cols_relevantes = [col for col in [
                                    'Fecha', 'Hora', 'Terminal', 'Cantidad litros', 
                                    'Odómetro', 'Kilómetros Recorridos', 'Rendimiento',
                                    'Nombre conductor', 'Mala Carga', 'Sobreconsumo'
                                ] if col in df_bus.columns]
                                
                                # Mostrar datos
                                st.dataframe(
                                    df_bus[cols_relevantes].sort_values('Fecha', ascending=False),
                                    use_container_width=True,
                                    height=400
                                )
                        
                        # Si es vista de todos los buses
                        else:
                            st.markdown("<h3 class='section-title'>Resumen de la Flota</h3>", unsafe_allow_html=True)
                            
                            # Estadísticas generales
                            col1, col2, col3, col4 = st.columns(4)
                            
                            with col1:
                                total_buses = df_bus['Número interno'].nunique()
                                st.markdown(create_kpi_card(
                                    "Total de Buses", 
                                    f"{total_buses}",
                                    None,
                                    icon="🚍",
                                    color_class="primary",
                                    footnote=f"En operación durante el período"
                                ), unsafe_allow_html=True)
                            
                            with col2:
                                total_modelos = df_bus['Modelo chasis'].nunique() if 'Modelo chasis' in df_bus.columns else "N/A"
                                st.markdown(create_kpi_card(
                                    "Modelos Diferentes", 
                                    f"{total_modelos}",
                                    None,
                                    icon="🔧",
                                    color_class="info",
                                    footnote=f"Variedad en la flota"
                                ), unsafe_allow_html=True)
                            
                            with col3:
                                promedio_cargas = len(df_bus) / total_buses if total_buses > 0 else 0
                                st.markdown(create_kpi_card(
                                    "Cargas por Bus", 
                                    f"{promedio_cargas:.2f}",
                                    None,
                                    icon="⛽",
                                    color_class="success",
                                    footnote=f"Promedio durante el período"
                                ), unsafe_allow_html=True)
                            
                            with col4:
                                if 'Rendimiento' in df_bus.columns:
                                    rendimiento_promedio = df_bus['Rendimiento'].mean()
                                    st.markdown(create_kpi_card(
                                        "Rendimiento Promedio", 
                                        f"{rendimiento_promedio:.2f} km/L",
                                        None,
                                        icon="🔄",
                                        color_class="warning",
                                        footnote=f"De toda la flota"
                                    ), unsafe_allow_html=True)
                            
                            # Análisis por uso
                            st.markdown("<h4>Análisis por Frecuencia de Uso</h4>", unsafe_allow_html=True)
                            
                            # Contar cargas por bus
                            cargas_por_bus = df_bus.groupby('Número interno').size().reset_index()
                            cargas_por_bus.columns = ['Bus', 'Cargas']
                            cargas_por_bus = cargas_por_bus.sort_values('Cargas', ascending=False)
                            
                            # Clasificar por uso
                            def clasificar_uso(cargas):
                                if cargas >= 30:
                                    return 'Alto Uso'
                                elif cargas >= 15:
                                    return 'Uso Medio'
                                else:
                                    return 'Bajo Uso'
                            
                            cargas_por_bus['Categoría'] = cargas_por_bus['Cargas'].apply(clasificar_uso)
                            
                            # Mostrar distribución
                            col1, col2 = st.columns(2)
                            
                            with col1:
                                # Gráfico de categorías
                                categoria_counts = cargas_por_bus['Categoría'].value_counts().reset_index()
                                categoria_counts.columns = ['Categoría', 'Buses']
                                
                                # Ordenar categorías
                                orden_categoria = {'Alto Uso': 0, 'Uso Medio': 1, 'Bajo Uso': 2}
                                categoria_counts['Orden'] = categoria_counts['Categoría'].map(orden_categoria)
                                categoria_counts = categoria_counts.sort_values('Orden')
                                
                                # Gráfico
                                fig = px.pie(
                                    categoria_counts,
                                    names='Categoría',
                                    values='Buses',
                                    title='Distribución de Buses por Intensidad de Uso',
                                    color='Categoría',
                                    color_discrete_map={
                                        'Alto Uso': COLORS['danger'],
                                        'Uso Medio': COLORS['warning'],
                                        'Bajo Uso': COLORS['success']
                                    }
                                )
                                
                                fig.update_layout(
                                    plot_bgcolor='white'
                                )
                                
                                st.plotly_chart(fig, use_container_width=True)
                            
                            with col2:
                                # Top 10 buses por uso
                                fig = px.bar(
                                    cargas_por_bus.head(10),
                                    x='Bus',
                                    y='Cargas',
                                    title='Top 10 Buses con Mayor Frecuencia de Cargas',
                                    color='Categoría',
                                    color_discrete_map={
                                        'Alto Uso': COLORS['danger'],
                                        'Uso Medio': COLORS['warning'],
                                        'Bajo Uso': COLORS['success']
                                    },
                                    text='Cargas'
                                )
                                
                                fig.update_traces(
                                    texttemplate='%{text}',
                                    textposition='outside'
                                )
                                
                                fig.update_layout(
                                    xaxis_title="Bus",
                                    yaxis_title="Cantidad de Cargas",
                                    plot_bgcolor='white'
                                )
                                
                                st.plotly_chart(fig, use_container_width=True)
                            
                            # Análisis por consumo
                            st.markdown("<h4>Análisis por Consumo Total</h4>", unsafe_allow_html=True)
                            
                            if 'Cantidad litros' in df_bus.columns:
                                # Calcular consumo total por bus
                                consumo_por_bus = df_bus.groupby('Número interno')['Cantidad litros'].sum().reset_index()
                                consumo_por_bus.columns = ['Bus', 'Total Litros']
                                consumo_por_bus = consumo_por_bus.sort_values('Total Litros', ascending=False)
                                
                                # Top 10 por consumo
                                fig = px.bar(
                                    consumo_por_bus.head(10),
                                    x='Bus',
                                    y='Total Litros',
                                    title='Top 10 Buses por Consumo Total',
                                    color='Total Litros',
                                    color_continuous_scale='Viridis',
                                    text='Total Litros'
                                )
                                
                                fig.update_traces(
                                    texttemplate='%{text:.0f}',
                                    textposition='outside'
                                )
                                
                                fig.update_layout(
                                    xaxis_title="Bus",
                                    yaxis_title="Total Litros",
                                    plot_bgcolor='white'
                                )
                                
                                st.plotly_chart(fig, use_container_width=True)
                            
                            # Análisis por rendimiento
                            st.markdown("<h4>Análisis por Rendimiento</h4>", unsafe_allow_html=True)
                            
                            if 'Rendimiento' in df_bus.columns:
                                # Calcular rendimiento promedio por bus
                                rendimiento_por_bus = df_bus.groupby('Número interno')['Rendimiento'].mean().reset_index()
                                rendimiento_por_bus.columns = ['Bus', 'Rendimiento']
                                
                                # Filtrar buses con suficientes datos
                                conteo_por_bus = df_bus.groupby('Número interno').size().reset_index()
                                conteo_por_bus.columns = ['Bus', 'Cargas']
                                
                                rendimiento_por_bus = pd.merge(rendimiento_por_bus, conteo_por_bus, on='Bus')
                                rendimiento_por_bus = rendimiento_por_bus[rendimiento_por_bus['Cargas'] >= 5]  # Al menos 5 cargas
                                
                                tab1, tab2 = st.tabs(["📊 Mejores Buses", "📉 Peores Buses"])
                                
                                with tab1:
                                    # Top 10 mejores rendimientos
                                    mejores = rendimiento_por_bus.sort_values('Rendimiento', ascending=False).head(10)
                                    
                                    fig = px.bar(
                                        mejores,
                                        x='Bus',
                                        y='Rendimiento',
                                        title='Top 10 Buses con Mejor Rendimiento',
                                        color='Rendimiento',
                                        color_continuous_scale='Viridis',
                                        text='Rendimiento'
                                    )
                                    
                                    fig.update_traces(
                                        texttemplate='%{text:.2f}',
                                        textposition='outside'
                                    )
                                    
                                    fig.update_layout(
                                        xaxis_title="Bus",
                                        yaxis_title="Rendimiento (km/L)",
                                        plot_bgcolor='white'
                                    )
                                    
                                    st.plotly_chart(fig, use_container_width=True)
                                
                                with tab2:
                                    # Top 10 peores rendimientos
                                    peores = rendimiento_por_bus.sort_values('Rendimiento').head(10)
                                    
                                    fig = px.bar(
                                        peores,
                                        x='Bus',
                                        y='Rendimiento',
                                        title='Top 10 Buses con Peor Rendimiento',
                                        color='Rendimiento',
                                        color_continuous_scale='Viridis_r',  # Escala invertida
                                        text='Rendimiento'
                                    )
                                    
                                    fig.update_traces(
                                        texttemplate='%{text:.2f}',
                                        textposition='outside'
                                    )
                                    
                                    fig.update_layout(
                                        xaxis_title="Bus",
                                        yaxis_title="Rendimiento (km/L)",
                                        plot_bgcolor='white'
                                    )
                                    
                                    st.plotly_chart(fig, use_container_width=True)
                            
                            # Tabla completa
                            st.markdown("<h4>Resumen por Bus</h4>", unsafe_allow_html=True)
                            
                            # Crear tabla de resumen completa
                            resumen_buses = df_bus.groupby('Número interno').agg({
                                'Cantidad litros': ['sum', 'mean', 'count'] if 'Cantidad litros' in df_bus.columns else 'count',
                                'Mala Carga': 'sum' if 'Mala Carga' in df_bus.columns else 'count',
                                'Sobreconsumo': 'sum' if 'Sobreconsumo' in df_bus.columns else 'count',
                                'Rendimiento': 'mean' if 'Rendimiento' in df_bus.columns else 'count'
                            }).reset_index()
                            
                            # Aplanar columnas
                            if 'Cantidad litros' in df_bus.columns:
                                resumen_buses.columns = ['Bus', 'Total Litros', 'Promedio Litros', 'Cantidad Cargas', 
                                                      'Malas Cargas', 'Sobreconsumos', 'Rendimiento Promedio']
                                
                                # Añadir porcentajes
                                resumen_buses['% Malas Cargas'] = (resumen_buses['Malas Cargas'] / resumen_buses['Cantidad Cargas'] * 100).round(2)
                                resumen_buses['% Sobreconsumos'] = (resumen_buses['Sobreconsumos'] / resumen_buses['Cantidad Cargas'] * 100).round(2)
                                
                                # Ordenar por total litros (descendente)
                                resumen_buses = resumen_buses.sort_values('Total Litros', ascending=False)
                            else:
                                resumen_buses.columns = ['Bus', 'Cantidad Cargas', 'Malas Cargas', 'Sobreconsumos', 'Rendimiento Count']
                                
                                # Ordenar por cantidad de cargas
                                resumen_buses = resumen_buses.sort_values('Cantidad Cargas', ascending=False)
                            
                            # Añadir modelo si está disponible
                            if 'Modelo chasis' in df_bus.columns:
                                # Obtener el modelo más común para cada bus
                                modelos_por_bus = df_bus.groupby('Número interno')['Modelo chasis'].agg(lambda x: x.mode().iloc[0] if not x.mode().empty else 'Desconocido')
                                
                                # Añadir a la tabla de resumen
                                resumen_buses = pd.merge(
                                    resumen_buses,
                                    modelos_por_bus.reset_index().rename(columns={'Número interno': 'Bus', 'Modelo chasis': 'Modelo'}),
                                    on='Bus'
                                )
                            
                            # Mostrar tabla
                            st.dataframe(
                                resumen_buses,
                                column_config={
                                    "Bus": st.column_config.TextColumn("Bus"),
                                    "Total Litros": st.column_config.NumberColumn("Total Litros", format="%,.2f"),
                                    "Promedio Litros": st.column_config.NumberColumn("Prom. L/Carga", format="%.2f"),
                                    "Cantidad Cargas": st.column_config.NumberColumn("# Cargas", format="%d"),
                                    "Malas Cargas": st.column_config.NumberColumn("Malas", format="%d"),
                                    "% Malas Cargas": st.column_config.NumberColumn("% Malas", format="%.2f%%"),
                                    "Sobreconsumos": st.column_config.NumberColumn("Sobre", format="%d"),
                                    "% Sobreconsumos": st.column_config.NumberColumn("% Sobre", format="%.2f%%"),
                                    "Rendimiento Promedio": st.column_config.NumberColumn("Rendimiento (km/L)", format="%.2f"),
                                    "Modelo": st.column_config.TextColumn("Modelo")
                                },
                                use_container_width=True
                            )
                    else:
                        st.warning("No se encontró la columna 'Número interno' en los datos.")
                
                # === ANÁLISIS POR PERSONAL ===
                elif section == "Análisis por Personal":
                    st.markdown("<h2 class='subtitle'>Análisis por Personal</h2>", unsafe_allow_html=True)
                    
                    # Seleccionar tipo de personal
                    tabs = []
                    
                    # Verificar qué tipos de personal existen en los datos
                    if 'Nombre conductor' in df.columns:
                        tabs.append("Conductores")
                    
                    if 'Nombre Planillero' in df.columns:
                        tabs.append("Planilleros")
                    
                    if 'Nombre supervisor' in df.columns:
                        tabs.append("Supervisores")
                    
                    if not tabs:
                        st.warning("No se encontraron columnas de personal en los datos.")
                    else:
                        tipo_personal = st.radio("Tipo de personal:", tabs, horizontal=True)
                        
                        # Análisis por Conductores
                        if tipo_personal == "Conductores" and 'Nombre conductor' in df.columns:
                            # Selector de conductor
                            conductores = sorted(df['Nombre conductor'].dropna().unique())
                            
                            col1, col2, col3 = st.columns([3, 1, 1])
                            
                            with col1:
                                conductor_seleccionado = st.selectbox("Seleccionar Conductor", ['Todos'] + list(conductores))
                            
                            with col2:
                                if 'Fecha' in df.columns:
                                    fecha_min = df['Fecha'].min().date()
                                    fecha_max = df['Fecha'].max().date()
                                else:
                                    fecha_min = datetime.date.today() - datetime.timedelta(days=30)
                                    fecha_max = datetime.date.today()
                                
                                fecha_inicio = st.date_input("Fecha inicio", fecha_min, key="conductor_fecha_inicio")
                            
                            with col3:
                                fecha_fin = st.date_input("Fecha fin", fecha_max, key="conductor_fecha_fin")
                            
                            # Filtrar datos por conductor y fechas
                            if conductor_seleccionado != 'Todos':
                                df_conductor = df[df['Nombre conductor'] == conductor_seleccionado]
                            else:
                                df_conductor = df.copy()
                            
                            if 'Fecha' in df.columns:
                                df_conductor = df_conductor[
                                    (df_conductor['Fecha'].dt.date >= fecha_inicio) & 
                                    (df_conductor['Fecha'].dt.date <= fecha_fin)
                                ]
                            
                            # Si seleccionamos un conductor específico
                            if conductor_seleccionado != 'Todos':
                                st.markdown(f"<h3 class='section-title'>Análisis Detallado: {conductor_seleccionado}</h3>", unsafe_allow_html=True)
                                
                                # Métricas principales
                                col1, col2, col3, col4 = st.columns(4)
                                
                                with col1:
                                    cargas_conductor = len(df_conductor)
                                    st.markdown(create_kpi_card(
                                        "Total Cargas", 
                                        f"{cargas_conductor}",
                                        None,
                                        icon="🚌",
                                        color_class="primary",
                                        footnote=f"En el período analizado"
                                    ), unsafe_allow_html=True)
                                
                                with col2:
                                    if 'Cantidad litros' in df_conductor.columns:
                                        litros_conductor = df_conductor['Cantidad litros'].sum()
                                        st.markdown(create_kpi_card(
                                            "Total Litros", 
                                            f"{litros_conductor:,.0f}",
                                            None,
                                            icon="⛽",
                                            color_class="info",
                                            footnote=f"Consumo total"
                                        ), unsafe_allow_html=True)
                                
                                with col3:
                                    if 'Sobreconsumo' in df_conductor.columns:
                                        sobreconsumos = df_conductor['Sobreconsumo'].sum()
                                        porcentaje_sobre = (sobreconsumos / len(df_conductor)) * 100 if len(df_conductor) > 0 else 0
                                        
                                        # Comparar con el promedio global
                                        porcentaje_global = (df['Sobreconsumo'].sum() / len(df)) * 100 if len(df) > 0 else 0
                                        diferencia = porcentaje_sobre - porcentaje_global
                                        
                                        color_class = "success" if diferencia < -2 else "warning" if diferencia < 2 else "danger"
                                        
                                        st.markdown(create_kpi_card(
                                            "Sobreconsumos", 
                                            f"{sobreconsumos} ({porcentaje_sobre:.2f}%)",
                                            -diferencia,  # Negativo porque menos es mejor
                                            icon="⚠️",
                                            color_class=color_class,
                                            footnote=f"vs {porcentaje_global:.2f}% global"
                                        ), unsafe_allow_html=True)
                                
                                with col4:
                                    if 'Rendimiento' in df_conductor.columns:
                                        rendimiento_conductor = df_conductor['Rendimiento'].mean()
                                        rendimiento_global = df['Rendimiento'].mean()
                                        diferencia_pct = ((rendimiento_conductor / rendimiento_global) - 1) * 100 if rendimiento_global > 0 else 0
                                        
                                        color_class = "success" if diferencia_pct > 2 else "warning" if diferencia_pct > -2 else "danger"
                                        
                                        st.markdown(create_kpi_card(
                                            "Rendimiento", 
                                            f"{rendimiento_conductor:.2f} km/L",
                                            diferencia_pct,  # Positivo porque más es mejor
                                            icon="🔄",
                                            color_class=color_class,
                                            footnote=f"vs {rendimiento_global:.2f} global"
                                        ), unsafe_allow_html=True)
                                
                                # Análisis de eficiencia
                                st.markdown("<h4>Análisis de Eficiencia</h4>", unsafe_allow_html=True)
                                
                                # Calcular índice de eficiencia
                                eficiencia = 100
                                
                                if 'Sobreconsumo' in df_conductor.columns:
                                    porcentaje_sobre = (df_conductor['Sobreconsumo'].sum() / len(df_conductor)) * 100 if len(df_conductor) > 0 else 0
                                    eficiencia -= porcentaje_sobre * 0.7  # Penalización por sobreconsumo
                                
                                if 'Rendimiento' in df_conductor.columns:
                                    rendimiento_conductor = df_conductor['Rendimiento'].mean()
                                    rendimiento_global = df['Rendimiento'].mean()
                                    
                                    if rendimiento_global > 0:
                                        # Ajustar eficiencia basado en rendimiento
                                        diferencia_pct = ((rendimiento_conductor / rendimiento_global) - 1) * 100
                                        if diferencia_pct < -10:
                                            eficiencia -= 10  # Penalización por rendimiento muy bajo
                                        elif diferencia_pct < -5:
                                            eficiencia -= 5   # Penalización por rendimiento bajo
                                        elif diferencia_pct > 10:
                                            eficiencia += 5   # Bonificación por rendimiento muy alto
                                        elif diferencia_pct > 5:
                                            eficiencia += 2.5  # Bonificación por rendimiento alto
                                
                                eficiencia = max(0, min(100, eficiencia))  # Asegurar rango 0-100
                                
                                # Mostrar índice de eficiencia
                                col1, col2 = st.columns([1, 3])
                                
                                with col1:
                                    # Determinar color según eficiencia
                                    if eficiencia >= 90:
                                        color = COLORS['success']
                                        categoria = "Excelente"
                                    elif eficiencia >= 80:
                                        color = COLORS['info']
                                        categoria = "Bueno"
                                    elif eficiencia >= 70:
                                        color = COLORS['warning']
                                        categoria = "Regular"
                                    else:
                                        color = COLORS['danger']
                                        categoria = "Necesita Mejora"
                                    
                                    st.markdown(f"""
                                    <div class="card" style="text-align: center;">
                                        <h4>Índice de Eficiencia</h4>
                                        <div style="font-size: 3rem; font-weight: 700; color: {color};">{eficiencia:.1f}</div>
                                        <div style="font-size: 1.2rem; color: {color};">{categoria}</div>
                                    </div>
                                    """, unsafe_allow_html=True)
                                
                                with col2:
                                    # Mostrar factores que influyen en la eficiencia
                                    st.markdown(f"""
                                    <div class="card">
                                        <h4>Factores de Eficiencia</h4>
                                        <div class="progress" style="margin-bottom: 15px;">
                                            <div class="progress-bar progress-bar-{'success' if porcentaje_sobre < 5 else 'warning' if porcentaje_sobre < 10 else 'danger'}" style="width: {min(100, porcentaje_sobre * 10)}%"></div>
                                        </div>
                                        <p><strong>Tasa de Sobreconsumo:</strong> {porcentaje_sobre:.2f}% (vs {(df['Sobreconsumo'].sum() / len(df)) * 100:.2f}% global)</p>
                                        
                                        <div class="progress" style="margin-bottom: 15px;">
                                            <div class="progress-bar progress-bar-{'danger' if diferencia_pct < -10 else 'warning' if diferencia_pct < 0 else 'success'}" style="width: {min(100, (rendimiento_conductor / (rendimiento_global * 1.5)) * 100)}%"></div>
                                        </div>
                                        <p><strong>Rendimiento:</strong> {rendimiento_conductor:.2f} km/L (vs {rendimiento_global:.2f} km/L global)</p>
                                    </div>
                                    """, unsafe_allow_html=True)
                                
                                # Análisis de buses conducidos
                                if 'Número interno' in df_conductor.columns:
                                    st.markdown("<h4>Buses Conducidos</h4>", unsafe_allow_html=True)
                                    
                                    # Contar cargas por bus
                                    buses_conductor = df_conductor.groupby('Número interno').size().reset_index()
                                    buses_conductor.columns = ['Bus', 'Cargas']
                                    buses_conductor = buses_conductor.sort_values('Cargas', ascending=False)
                                    
                                    # Calcular porcentaje
                                    buses_conductor['% del Total'] = (buses_conductor['Cargas'] / buses_conductor['Cargas'].sum() * 100).round(2)
                                    
                                    col1, col2 = st.columns(2)
                                    
                                    with col1:
                                        # Gráfico de buses conducidos
                                        fig = px.bar(
                                            buses_conductor.head(10),
                                            x='Bus',
                                            y='Cargas',
                                            title='Top 10 Buses Conducidos',
                                            color='Cargas',
                                            color_continuous_scale='Viridis',
                                            text='Cargas'
                                        )
                                        
                                        fig.update_traces(
                                            texttemplate='%{text}',
                                            textposition='outside'
                                        )
                                        
                                        fig.update_layout(
                                            xaxis_title="Bus",
                                            yaxis_title="Cantidad de Cargas",
                                            plot_bgcolor='white'
                                        )
                                        
                                        st.plotly_chart(fig, use_container_width=True)
                                    
                                    with col2:
                                        # Distribución del rendimiento por bus
                                        if 'Rendimiento' in df_conductor.columns:
                                            rendimiento_por_bus = df_conductor.groupby('Número interno')['Rendimiento'].mean().reset_index()
                                            rendimiento_por_bus.columns = ['Bus', 'Rendimiento']
                                            
                                            # Fusionar con buses_conductor para mostrar solo los más usados
                                            rendimiento_por_bus = pd.merge(
                                                rendimiento_por_bus,
                                                buses_conductor[['Bus', 'Cargas']],
                                                on='Bus'
                                            )
                                            
                                            # Filtrar buses con al menos 3 cargas
                                            rendimiento_por_bus = rendimiento_por_bus[rendimiento_por_bus['Cargas'] >= 3]
                                            rendimiento_por_bus = rendimiento_por_bus.sort_values('Rendimiento', ascending=False)
                                            
                                            # Mostrar gráfico
                                            fig = px.bar(
                                                rendimiento_por_bus.head(10),
                                                x='Bus',
                                                y='Rendimiento',
                                                title='Rendimiento por Bus (Top 10)',
                                                color='Rendimiento',
                                                color_continuous_scale='Viridis',
                                                text='Rendimiento'
                                            )
                                            
                                            fig.update_traces(
                                                texttemplate='%{text:.2f}',
                                                textposition='outside'
                                            )
                                            
                                            fig.update_layout(
                                                xaxis_title="Bus",
                                                yaxis_title="Rendimiento (km/L)",
                                                plot_bgcolor='white'
                                            )
                                            
                                            st.plotly_chart(fig, use_container_width=True)
                                
                                # Historial detallado
                                with st.expander("Historial Detallado de Cargas", expanded=False):
                                    # Columnas relevantes para mostrar
                                    cols_relevantes = [col for col in [
                                        'Fecha', 'Hora', 'Terminal', 'Número interno',
                                        'Cantidad litros', 'Rendimiento', 'Sobreconsumo'
                                    ] if col in df_conductor.columns]
                                    
                                    # Mostrar datos
                                    st.dataframe(
                                        df_conductor[cols_relevantes].sort_values('Fecha', ascending=False),
                                        use_container_width=True,
                                        height=400
                                    )
                            
                            # Si es vista de todos los conductores
                            else:
                                st.markdown("<h3 class='section-title'>Resumen de Conductores</h3>", unsafe_allow_html=True)
                                
                                # Crear tabla de resumen
                                resumen_conductores = df_conductor.groupby('Nombre conductor').agg({
                                    'Cantidad litros': ['sum', 'mean', 'count'] if 'Cantidad litros' in df_conductor.columns else 'count',
                                    'Sobreconsumo': 'sum' if 'Sobreconsumo' in df_conductor.columns else 'count',
                                    'Rendimiento': 'mean' if 'Rendimiento' in df_conductor.columns else 'count'
                                }).reset_index()
                                
                                # Aplanar columnas
                                if 'Cantidad litros' in df_conductor.columns:
                                    resumen_conductores.columns = ['Conductor', 'Total Litros', 'Promedio Litros', 'Cantidad Cargas', 
                                                              'Sobreconsumos', 'Rendimiento Promedio']
                                    
                                    # Añadir porcentajes
                                    resumen_conductores['% Sobreconsumos'] = (resumen_conductores['Sobreconsumos'] / resumen_conductores['Cantidad Cargas'] * 100).round(2)
                                    
                                    # Calcular índice de eficiencia
                                    resumen_conductores['Índice Eficiencia'] = 100 - resumen_conductores['% Sobreconsumos'] * 0.7
                                    
                                    # Ajustar por rendimiento si está disponible
                                    if 'Rendimiento' in df_conductor.columns:
                                        rendimiento_global = df['Rendimiento'].mean()
                                        
                                        if rendimiento_global > 0:
                                            resumen_conductores['% vs Rendimiento Global'] = ((resumen_conductores['Rendimiento Promedio'] / rendimiento_global) - 1) * 100
                                            
                                            # Ajustar índice de eficiencia
                                            ajuste_rendimiento = np.where(
                                                resumen_conductores['% vs Rendimiento Global'] > 5, 5,
                                                np.where(
                                                    resumen_conductores['% vs Rendimiento Global'] > 0, 2.5,
                                                    np.where(
                                                        resumen_conductores['% vs Rendimiento Global'] < -10, -10,
                                                        np.where(
                                                            resumen_conductores['% vs Rendimiento Global'] < -5, -5, 0
                                                        )
                                                    )
                                                )
                                            )
                                            
                                            resumen_conductores['Índice Eficiencia'] += ajuste_rendimiento
                                    
                                    # Asegurar rango 0-100
                                    resumen_conductores['Índice Eficiencia'] = resumen_conductores['Índice Eficiencia'].clip(0, 100).round(1)
                                    
                                    # Ordenar por índice de eficiencia (descendente)
                                    resumen_conductores = resumen_conductores.sort_values('Índice Eficiencia', ascending=False)
                                else:
                                    resumen_conductores.columns = ['Conductor', 'Cantidad Cargas', 'Sobreconsumos', 'Rendimiento Count']
                                    
                                    # Ordenar por cantidad de cargas
                                    resumen_conductores = resumen_conductores.sort_values('Cantidad Cargas', ascending=False)
                                
                                # Mostrar gráficos de ranking
                                if 'Índice Eficiencia' in resumen_conductores.columns:
                                    # Mostrar top 10 más eficientes
                                    fig = px.bar(
                                        resumen_conductores.head(10),
                                        x='Conductor',
                                        y='Índice Eficiencia',
                                        title='Top 10 Conductores por Eficiencia',
                                        color='Índice Eficiencia',
                                        color_continuous_scale='RdYlGn',
                                        text='Índice Eficiencia'
                                    )
                                    
                                    fig.update_traces(
                                        texttemplate='%{text:.1f}',
                                        textposition='outside'
                                    )
                                    
                                    fig.update_layout(
                                        xaxis_title="Conductor",
                                        yaxis_title="Índice de Eficiencia",
                                        plot_bgcolor='white'
                                    )
                                    
                                    st.plotly_chart(fig, use_container_width=True)
                                
                                # Mostrar tabla completa
                                st.markdown("<h4>Tabla Completa de Conductores</h4>", unsafe_allow_html=True)
                                
                                # Mostrar tabla
                                st.dataframe(
                                    resumen_conductores,
                                    column_config={
                                        "Conductor": st.column_config.TextColumn("Conductor"),
                                        "Total Litros": st.column_config.NumberColumn("Total Litros", format="%,.2f"),
                                        "Promedio Litros": st.column_config.NumberColumn("Prom. L/Carga", format="%.2f"),
                                        "Cantidad Cargas": st.column_config.NumberColumn("# Cargas", format="%d"),
                                        "Sobreconsumos": st.column_config.NumberColumn("Sobre", format="%d"),
                                        "% Sobreconsumos": st.column_config.NumberColumn("% Sobre", format="%.2f%%"),
                                        "Rendimiento Promedio": st.column_config.NumberColumn("Rendimiento (km/L)", format="%.2f"),
                                        "% vs Rendimiento Global": st.column_config.NumberColumn("% vs Global", format="%.2f%%"),
                                        "Índice Eficiencia": st.column_config.ProgressColumn("Eficiencia", format="%.1f", min_value=0, max_value=100)
                                    },
                                    use_container_width=True
                                )
                                
                                # Análisis de sobreconsumos
                                if 'Sobreconsumo' in df_conductor.columns:
                                    st.markdown("<h4>Análisis de Conductores con Mayor Tasa de Sobreconsumo</h4>", unsafe_allow_html=True)
                                    
                                    # Filtrar conductores con al menos 5 cargas
                                    conductores_sobreconsumo = resumen_conductores[resumen_conductores['Cantidad Cargas'] >= 5].copy()
                                    
                                    # Ordenar por porcentaje de sobreconsumos (descendente)
                                    conductores_sobreconsumo = conductores_sobreconsumo.sort_values('% Sobreconsumos', ascending=False)
                                    
                                    # Mostrar gráfico
                                    fig = px.bar(
                                        conductores_sobreconsumo.head(10),
                                        x='Conductor',
                                        y='% Sobreconsumos',
                                        title='Top 10 Conductores con Mayor % de Sobreconsumo',
                                        color='% Sobreconsumos',
                                        color_continuous_scale='Reds',
                                        text='% Sobreconsumos'
                                    )
                                    
                                    fig.update_traces(
                                        texttemplate='%{text:.2f}%',
                                        textposition='outside'
                                    )
                                    
                                    fig.update_layout(
                                        xaxis_title="Conductor",
                                        yaxis_title="% de Sobreconsumo",
                                        plot_bgcolor='white'
                                    )
                                    
                                    # Añadir línea para el promedio
                                    promedio_sobreconsumo = (df_conductor['Sobreconsumo'].sum() / len(df_conductor)) * 100
                                    
                                    fig.add_hline(
                                        y=promedio_sobreconsumo,
                                        line_dash="dash",
                                        line_color="red",
                                        annotation_text=f"Promedio: {promedio_sobreconsumo:.2f}%"
                                    )
                                    
                                    st.plotly_chart(fig, use_container_width=True)
                                    
                                    # Mostrar recomendaciones si hay conductores con alta tasa
                                    if not conductores_sobreconsumo.empty and conductores_sobreconsumo['% Sobreconsumos'].iloc[0] > 15:
                                        st.markdown(show_alert(
                                            f"Los conductores con mayor tasa de sobreconsumo deberían recibir capacitación adicional sobre técnicas de conducción eficiente.",
                                            "warning",
                                            "Oportunidad de Mejora Detectada",
                                            "⚠️"
                                        ), unsafe_allow_html=True)
                        
                        # Análisis por Planilleros
                        elif tipo_personal == "Planilleros" and 'Nombre Planillero' in df.columns:
                            # Selector de planillero
                            planilleros = sorted(df['Nombre Planillero'].dropna().unique())
                            
                            col1, col2, col3 = st.columns([3, 1, 1])
                            
                            with col1:
                                planillero_seleccionado = st.selectbox("Seleccionar Planillero", ['Todos'] + list(planilleros))
                            
                            with col2:
                                if 'Fecha' in df.columns:
                                    fecha_min = df['Fecha'].min().date()
                                    fecha_max = df['Fecha'].max().date()
                                else:
                                    fecha_min = datetime.date.today() - datetime.timedelta(days=30)
                                    fecha_max = datetime.date.today()
                                
                                fecha_inicio = st.date_input("Fecha inicio", fecha_min, key="planillero_fecha_inicio")
                            
                            with col3:
                                fecha_fin = st.date_input("Fecha fin", fecha_max, key="planillero_fecha_fin")
                            
                            # Filtrar datos por planillero y fechas
                            if planillero_seleccionado != 'Todos':
                                df_planillero = df[df['Nombre Planillero'] == planillero_seleccionado]
                            else:
                                df_planillero = df.copy()
                            
                            if 'Fecha' in df.columns:
                                df_planillero = df_planillero[
                                    (df_planillero['Fecha'].dt.date >= fecha_inicio) & 
                                    (df_planillero['Fecha'].dt.date <= fecha_fin)
                                ]
                            
                            # Si seleccionamos un planillero específico
                            if planillero_seleccionado != 'Todos':
                                st.markdown(f"<h3 class='section-title'>Análisis Detallado: {planillero_seleccionado}</h3>", unsafe_allow_html=True)
                                
                                # Métricas principales
                                col1, col2, col3 = st.columns(3)
                                
                                with col1:
                                    cargas_planillero = len(df_planillero)
                                    st.markdown(create_kpi_card(
                                        "Total Cargas", 
                                        f"{cargas_planillero}",
                                        None,
                                        icon="📝",
                                        color_class="primary",
                                        footnote=f"En el período analizado"
                                    ), unsafe_allow_html=True)
                                
                                with col2:
                                    if 'Mala Carga' in df_planillero.columns:
                                        malas_cargas = df_planillero['Mala Carga'].sum()
                                        porcentaje_malas = (malas_cargas / len(df_planillero)) * 100 if len(df_planillero) > 0 else 0
                                        
                                        # Comparar con el promedio global
                                        porcentaje_global = (df['Mala Carga'].sum() / len(df)) * 100 if len(df) > 0 else 0
                                        diferencia = porcentaje_malas - porcentaje_global
                                        
                                        color_class = "success" if diferencia < -2 else "warning" if diferencia < 2 else "danger"
                                        
                                        st.markdown(create_kpi_card(
                                            "Malas Cargas", 
                                            f"{malas_cargas} ({porcentaje_malas:.2f}%)",
                                            -diferencia,  # Negativo porque menos es mejor
                                            icon="❌",
                                            color_class=color_class,
                                            footnote=f"vs {porcentaje_global:.2f}% global"
                                        ), unsafe_allow_html=True)
                                
                                with col3:
                                    if 'Terminal' in df_planillero.columns:
                                        terminales = df_planillero['Terminal'].nunique()
                                        st.markdown(create_kpi_card(
                                            "Terminales", 
                                            f"{terminales}",
                                            None,
                                            icon="🏢",
                                            color_class="info",
                                            footnote=f"Diferentes terminales"
                                        ), unsafe_allow_html=True)
                                
                                # Evolución temporal
                                if 'Fecha' in df_planillero.columns and len(df_planillero) > 1:
                                    st.markdown("<h4>Evolución Temporal</h4>", unsafe_allow_html=True)
                                    
                                    # Agrupar por fecha
                                    cargas_por_dia = df_planillero.groupby(df_planillero['Fecha'].dt.date).size().reset_index()
                                    cargas_por_dia.columns = ['Fecha', 'Cargas']
                                    
                                    # Calcular malas cargas por día
                                    if 'Mala Carga' in df_planillero.columns:
                                        malas_por_dia = df_planillero.groupby(df_planillero['Fecha'].dt.date)['Mala Carga'].sum().reset_index()
                                        malas_por_dia.columns = ['Fecha', 'Malas Cargas']
                                        
                                        # Unir dataframes
                                        cargas_por_dia = pd.merge(cargas_por_dia, malas_por_dia, on='Fecha')
                                        
                                        # Calcular porcentaje
                                        cargas_por_dia['% Malas'] = (cargas_por_dia['Malas Cargas'] / cargas_por_dia['Cargas'] * 100).round(2)
                                    
                                    # Crear gráfico
                                    fig = make_subplots(specs=[[{"secondary_y": True}]])
                                    
                                    # Añadir barras para cargas totales
                                    fig.add_trace(
                                        go.Bar(
                                            x=cargas_por_dia['Fecha'],
                                            y=cargas_por_dia['Cargas'],
                                            name='Total Cargas',
                                            marker_color=COLORS['primary']
                                        ),
                                        secondary_y=False
                                    )
                                    
                                    # Añadir línea para porcentaje de malas cargas si está disponible
                                    if 'Mala Carga' in df_planillero.columns:
                                        fig.add_trace(
                                            go.Scatter(
                                                x=cargas_por_dia['Fecha'],
                                                y=cargas_por_dia['% Malas'],
                                                name='% Malas Cargas',
                                                line=dict(color=COLORS['danger'], width=2),
                                                mode='lines+markers'
                                            ),
                                            secondary_y=True
                                        )
                                    
                                    # Actualizar diseño
                                    fig.update_layout(
                                        title='Evolución de Cargas y % de Malas Cargas',
                                        plot_bgcolor='white',
                                        hovermode='x unified',
                                        height=400,
                                        legend=dict(
                                            orientation="h",
                                            yanchor="bottom",
                                            y=1.02,
                                            xanchor="right",
                                            x=1
                                        )
                                    )
                                    
                                    fig.update_xaxes(title_text="Fecha")
                                    fig.update_yaxes(title_text="Cantidad de Cargas", secondary_y=False)
                                    
                                    if 'Mala Carga' in df_planillero.columns:
                                        fig.update_yaxes(title_text="% de Malas Cargas", secondary_y=True)
                                    
                                    st.plotly_chart(fig, use_container_width=True)
                                
                                # Análisis por terminal
                                if 'Terminal' in df_planillero.columns:
                                    st.markdown("<h4>Análisis por Terminal</h4>", unsafe_allow_html=True)
                                    
                                    # Agrupar por terminal
                                    terminal_stats = df_planillero.groupby('Terminal').size().reset_index()
                                    terminal_stats.columns = ['Terminal', 'Cargas']
                                    
                                    # Calcular porcentaje
                                    terminal_stats['% del Total'] = (terminal_stats['Cargas'] / terminal_stats['Cargas'].sum() * 100).round(2)
                                    
                                    # Calcular malas cargas por terminal
                                    if 'Mala Carga' in df_planillero.columns:
                                        malas_por_terminal = df_planillero.groupby('Terminal')['Mala Carga'].sum().reset_index()
                                        malas_por_terminal.columns = ['Terminal', 'Malas Cargas']
                                        
                                        # Unir dataframes
                                        terminal_stats = pd.merge(terminal_stats, malas_por_terminal, on='Terminal')
                                        
                                        # Calcular porcentaje de malas cargas
                                        terminal_stats['% Malas'] = (terminal_stats['Malas Cargas'] / terminal_stats['Cargas'] * 100).round(2)
                                    
                                    # Ordenar por cantidad de cargas (descendente)
                                    terminal_stats = terminal_stats.sort_values('Cargas', ascending=False)
                                    
                                    # Mostrar en pestañas
                                    tab1, tab2 = st.tabs(["📊 Gráfico", "📋 Tabla"])
                                    
                                    with tab1:
                                        # Crear gráfico
                                        if 'Mala Carga' in df_planillero.columns:
                                            # Gráfico con cargas totales y % de malas cargas
                                            fig = make_subplots(specs=[[{"secondary_y": True}]])
                                            
                                            # Añadir barras para cargas totales
                                            fig.add_trace(
                                                go.Bar(
                                                    x=terminal_stats['Terminal'],
                                                    y=terminal_stats['Cargas'],
                                                    name='Total Cargas',
                                                    marker_color=COLORS['primary']
                                                ),
                                                secondary_y=False
                                            )
                                            
                                            # Añadir línea para porcentaje de malas cargas
                                            fig.add_trace(
                                                go.Scatter(
                                                    x=terminal_stats['Terminal'],
                                                    y=terminal_stats['% Malas'],
                                                    name='% Malas Cargas',
                                                    line=dict(color=COLORS['danger'], width=2),
                                                    mode='lines+markers'
                                                ),
                                                secondary_y=True
                                            )
                                            
                                            # Actualizar diseño
                                            fig.update_layout(
                                                title='Cargas y % de Malas Cargas por Terminal',
                                                plot_bgcolor='white',
                                                hovermode='x unified',
                                                height=400,
                                                legend=dict(
                                                    orientation="h",
                                                    yanchor="bottom",
                                                    y=1.02,
                                                    xanchor="right",
                                                    x=1
                                                )
                                            )
                                            
                                            fig.update_xaxes(title_text="Terminal")
                                            fig.update_yaxes(title_text="Cantidad de Cargas", secondary_y=False)
                                            fig.update_yaxes(title_text="% de Malas Cargas", secondary_y=True)
                                        else:
                                            # Gráfico simple de cargas por terminal
                                            fig = px.bar(
                                                terminal_stats,
                                                x='Terminal',
                                                y='Cargas',
                                                title='Cargas por Terminal',
                                                color='Cargas',
                                                color_continuous_scale='Viridis',
                                                text='Cargas'
                                            )
                                            
                                            fig.update_traces(
                                                texttemplate='%{text}',
                                                textposition='outside'
                                            )
                                            
                                            fig.update_layout(
                                                xaxis_title="Terminal",
                                                yaxis_title="Cantidad de Cargas",
                                                plot_bgcolor='white'
                                            )
                                        
                                        st.plotly_chart(fig, use_container_width=True)
                                    
                                    with tab2:
                                        # Mostrar tabla
                                        st.dataframe(
                                            terminal_stats,
                                            column_config={
                                                "Terminal": st.column_config.TextColumn("Terminal"),
                                                "Cargas": st.column_config.NumberColumn("Cargas", format="%d"),
                                                "% del Total": st.column_config.NumberColumn("% del Total", format="%.2f%%"),
                                                "Malas Cargas": st.column_config.NumberColumn("Malas", format="%d"),
                                                "% Malas": st.column_config.NumberColumn("% Malas", format="%.2f%%")
                                            },
                                            use_container_width=True
                                        )
                                
                                # Historial detallado
                                with st.expander("Historial Detallado de Cargas", expanded=False):
                                    # Columnas relevantes para mostrar
                                    cols_relevantes = [col for col in [
                                        'Fecha', 'Hora', 'Terminal', 'Número interno',
                                        'Cantidad litros', 'Mala Carga'
                                    ] if col in df_planillero.columns]
                                    
                                    # Mostrar datos
                                    st.dataframe(
                                        df_planillero[cols_relevantes].sort_values('Fecha', ascending=False),
                                        use_container_width=True,
                                        height=400
                                    )
                            
                            # Si es vista de todos los planilleros
                            else:
                                st.markdown("<h3 class='section-title'>Resumen de Planilleros</h3>", unsafe_allow_html=True)
                                
                                # Crear tabla de resumen
                                resumen_planilleros = df_planillero.groupby('Nombre Planillero').agg({
                                    'Mala Carga': ['sum', 'count'] if 'Mala Carga' in df_planillero.columns else 'count'
                                }).reset_index()
                                
                                # Aplanar columnas
                                if 'Mala Carga' in df_planillero.columns:
                                    resumen_planilleros.columns = ['Planillero', 'Malas Cargas', 'Total Cargas']
                                    
                                    # Añadir porcentajes
                                    resumen_planilleros['% Malas Cargas'] = (resumen_planilleros['Malas Cargas'] / resumen_planilleros['Total Cargas'] * 100).round(2)
                                    
                                    # Ordenar por porcentaje de malas cargas (ascendente - mejores primero)
                                    resumen_planilleros = resumen_planilleros.sort_values('% Malas Cargas')
                                else:
                                    resumen_planilleros.columns = ['Planillero', 'Total Cargas']
                                    
                                    # Ordenar por cantidad de cargas (descendente)
                                    resumen_planilleros = resumen_planilleros.sort_values('Total Cargas', ascending=False)
                                
                                # Mostrar gráficos de ranking
                                if 'Mala Carga' in df_planillero.columns:
                                    # Filtrar planilleros con al menos 10 cargas
                                    planilleros_filtrados = resumen_planilleros[resumen_planilleros['Total Cargas'] >= 10].copy()
                                    
                                    # Mostrar mejores planilleros
                                    fig = px.bar(
                                        planilleros_filtrados.head(10),  # Top 10 mejores
                                        x='Planillero',
                                        y='% Malas Cargas',
                                        title='Top 10 Planilleros con Menor % de Malas Cargas',
                                        color='% Malas Cargas',
                                        color_continuous_scale='RdYlGn_r',  # Escala invertida
                                        text='% Malas Cargas'
                                    )
                                    
                                    fig.update_traces(
                                        texttemplate='%{text:.2f}%',
                                        textposition='outside'
                                    )
                                    
                                    fig.update_layout(
                                        xaxis_title="Planillero",
                                        yaxis_title="% de Malas Cargas",
                                        plot_bgcolor='white'
                                    )
                                    
                                    st.plotly_chart(fig, use_container_width=True)
                                    
                                    # Mostrar peores planilleros
                                    fig = px.bar(
                                        planilleros_filtrados.sort_values('% Malas Cargas', ascending=False).head(10),  # Top 10 peores
                                        x='Planillero',
                                        y='% Malas Cargas',
                                        title='Top 10 Planilleros con Mayor % de Malas Cargas',
                                        color='% Malas Cargas',
                                        color_continuous_scale='RdYlGn_r',  # Escala invertida
                                        text='% Malas Cargas'
                                    )
                                    
                                    fig.update_traces(
                                        texttemplate='%{text:.2f}%',
                                        textposition='outside'
                                    )
                                    
                                    fig.update_layout(
                                        xaxis_title="Planillero",
                                        yaxis_title="% de Malas Cargas",
                                        plot_bgcolor='white'
                                    )
                                    
                                    # Añadir línea para el promedio
                                    promedio_malas = (df_planillero['Mala Carga'].sum() / len(df_planillero)) * 100
                                    
                                    fig.add_hline(
                                        y=promedio_malas,
                                        line_dash="dash",
                                        line_color="red",
                                        annotation_text=f"Promedio: {promedio_malas:.2f}%"
                                    )
                                    
                                    st.plotly_chart(fig, use_container_width=True)
                                    
                                    # Mostrar recomendaciones si hay planilleros con alta tasa
                                    if not planilleros_filtrados.empty and planilleros_filtrados['% Malas Cargas'].max() > 10:
                                        st.markdown(show_alert(
                                            f"Los planilleros con mayor tasa de malas cargas deberían recibir capacitación adicional sobre los procedimientos correctos de registro.",
                                            "warning",
                                            "Oportunidad de Mejora Detectada",
                                            "⚠️"
                                        ), unsafe_allow_html=True)
                                
                                # Mostrar tabla completa
                                st.markdown("<h4>Tabla Completa de Planilleros</h4>", unsafe_allow_html=True)
                                
                                # Mostrar tabla
                                if 'Mala Carga' in df_planillero.columns:
                                    st.dataframe(
                                        resumen_planilleros,
                                        column_config={
                                            "Planillero": st.column_config.TextColumn("Planillero"),
                                            "Malas Cargas": st.column_config.NumberColumn("Malas", format="%d"),
                                            "Total Cargas": st.column_config.NumberColumn("Total", format="%d"),
                                            "% Malas Cargas": st.column_config.NumberColumn("% Malas", format="%.2f%%")
                                        },
                                        use_container_width=True
                                    )
                                else:
                                    st.dataframe(
                                        resumen_planilleros,
                                        column_config={
                                          "Planillero": st.column_config.TextColumn("Planillero"),
                                            "Total Cargas": st.column_config.NumberColumn("Total", format="%d")
                                        },
                                        use_container_width=True
                                    )
                        
                        # Análisis por Supervisores
                        elif tipo_personal == "Supervisores" and 'Nombre supervisor' in df.columns:
                            # Selector de supervisor
                            supervisores = sorted(df['Nombre supervisor'].dropna().unique())
                            
                            col1, col2, col3 = st.columns([3, 1, 1])
                            
                            with col1:
                                supervisor_seleccionado = st.selectbox("Seleccionar Supervisor", ['Todos'] + list(supervisores))
                            
                            with col2:
                                if 'Fecha' in df.columns:
                                    fecha_min = df['Fecha'].min().date()
                                    fecha_max = df['Fecha'].max().date()
                                else:
                                    fecha_min = datetime.date.today() - datetime.timedelta(days=30)
                                    fecha_max = datetime.date.today()
                                
                                fecha_inicio = st.date_input("Fecha inicio", fecha_min, key="supervisor_fecha_inicio")
                            
                            with col3:
                                fecha_fin = st.date_input("Fecha fin", fecha_max, key="supervisor_fecha_fin")
                            
                            # Filtrar datos por supervisor y fechas
                            if supervisor_seleccionado != 'Todos':
                                df_supervisor = df[df['Nombre supervisor'] == supervisor_seleccionado]
                            else:
                                df_supervisor = df.copy()
                            
                            if 'Fecha' in df.columns:
                                df_supervisor = df_supervisor[
                                    (df_supervisor['Fecha'].dt.date >= fecha_inicio) & 
                                    (df_supervisor['Fecha'].dt.date <= fecha_fin)
                                ]
                            
                            # Si seleccionamos un supervisor específico
                            if supervisor_seleccionado != 'Todos':
                                st.markdown(f"<h3 class='section-title'>Análisis Detallado: {supervisor_seleccionado}</h3>", unsafe_allow_html=True)
                                
                                # Métricas principales
                                col1, col2, col3, col4 = st.columns(4)
                                
                                with col1:
                                    cargas_supervisor = len(df_supervisor)
                                    st.markdown(create_kpi_card(
                                        "Total Cargas", 
                                        f"{cargas_supervisor}",
                                        None,
                                        icon="📋",
                                        color_class="primary",
                                        footnote=f"En el período analizado"
                                    ), unsafe_allow_html=True)
                                
                                with col2:
                                    if 'Mala Carga' in df_supervisor.columns:
                                        malas_cargas = df_supervisor['Mala Carga'].sum()
                                        porcentaje_malas = (malas_cargas / len(df_supervisor)) * 100 if len(df_supervisor) > 0 else 0
                                        
                                        # Comparar con el promedio global
                                        porcentaje_global = (df['Mala Carga'].sum() / len(df)) * 100 if len(df) > 0 else 0
                                        diferencia = porcentaje_malas - porcentaje_global
                                        
                                        color_class = "success" if diferencia < -2 else "warning" if diferencia < 2 else "danger"
                                        
                                        st.markdown(create_kpi_card(
                                            "Malas Cargas", 
                                            f"{malas_cargas} ({porcentaje_malas:.2f}%)",
                                            -diferencia,  # Negativo porque menos es mejor
                                            icon="❌",
                                            color_class=color_class,
                                            footnote=f"vs {porcentaje_global:.2f}% global"
                                        ), unsafe_allow_html=True)
                                
                                with col3:
                                    if 'Sobreconsumo' in df_supervisor.columns:
                                        sobreconsumos = df_supervisor['Sobreconsumo'].sum()
                                        porcentaje_sobre = (sobreconsumos / len(df_supervisor)) * 100 if len(df_supervisor) > 0 else 0
                                        
                                        # Comparar con el promedio global
                                        porcentaje_global = (df['Sobreconsumo'].sum() / len(df)) * 100 if len(df) > 0 else 0
                                        diferencia = porcentaje_sobre - porcentaje_global
                                        
                                        color_class = "success" if diferencia < -2 else "warning" if diferencia < 2 else "danger"
                                        
                                        st.markdown(create_kpi_card(
                                            "Sobreconsumos", 
                                            f"{sobreconsumos} ({porcentaje_sobre:.2f}%)",
                                            -diferencia,  # Negativo porque menos es mejor
                                            icon="⚠️",
                                            color_class=color_class,
                                            footnote=f"vs {porcentaje_global:.2f}% global"
                                        ), unsafe_allow_html=True)
                                
                                with col4:
                                    if 'Terminal' in df_supervisor.columns:
                                        terminales = df_supervisor['Terminal'].nunique()
                                        st.markdown(create_kpi_card(
                                            "Terminales", 
                                            f"{terminales}",
                                            None,
                                            icon="🏢",
                                            color_class="info",
                                            footnote=f"Diferentes terminales"
                                        ), unsafe_allow_html=True)
                                
                                # Análisis por terminal
                                if 'Terminal' in df_supervisor.columns:
                                    st.markdown("<h4>Análisis por Terminal</h4>", unsafe_allow_html=True)
                                    
                                    # Agrupar por terminal
                                    terminal_stats = df_supervisor.groupby('Terminal').agg({
                                        'Mala Carga': ['sum', 'count'] if 'Mala Carga' in df_supervisor.columns else 'count',
                                        'Sobreconsumo': ['sum'] if 'Sobreconsumo' in df_supervisor.columns else 'count'
                                    }).reset_index()
                                    
                                    # Aplanar columnas
                                    if 'Mala Carga' in df_supervisor.columns and 'Sobreconsumo' in df_supervisor.columns:
                                        terminal_stats.columns = ['Terminal', 'Malas Cargas', 'Total Cargas', 'Sobreconsumos']
                                        
                                        # Añadir porcentajes
                                        terminal_stats['% Malas'] = (terminal_stats['Malas Cargas'] / terminal_stats['Total Cargas'] * 100).round(2)
                                        terminal_stats['% Sobre'] = (terminal_stats['Sobreconsumos'] / terminal_stats['Total Cargas'] * 100).round(2)
                                        
                                        # Añadir índice de calidad
                                        terminal_stats['Índice Calidad'] = 100 - (terminal_stats['% Malas'] * 0.5 + terminal_stats['% Sobre'] * 0.5)
                                        terminal_stats['Índice Calidad'] = terminal_stats['Índice Calidad'].clip(0, 100).round(1)
                                    elif 'Mala Carga' in df_supervisor.columns:
                                        terminal_stats.columns = ['Terminal', 'Malas Cargas', 'Total Cargas']
                                        
                                        # Añadir porcentajes
                                        terminal_stats['% Malas'] = (terminal_stats['Malas Cargas'] / terminal_stats['Total Cargas'] * 100).round(2)
                                    else:
                                        terminal_stats.columns = ['Terminal', 'Total Cargas']
                                    
                                    # Ordenar por total cargas (descendente)
                                    terminal_stats = terminal_stats.sort_values('Total Cargas', ascending=False)
                                    
                                    # Mostrar gráfico y tabla en pestañas
                                    tab1, tab2 = st.tabs(["📊 Gráfico", "📋 Tabla"])
                                    
                                    with tab1:
                                        if 'Mala Carga' in df_supervisor.columns:
                                            # Gráfico con índice de calidad si está disponible
                                            if 'Índice Calidad' in terminal_stats.columns:
                                                fig = px.bar(
                                                    terminal_stats,
                                                    x='Terminal',
                                                    y='Índice Calidad',
                                                    title='Índice de Calidad por Terminal',
                                                    color='Índice Calidad',
                                                    color_continuous_scale='RdYlGn',
                                                    text='Índice Calidad'
                                                )
                                                
                                                fig.update_traces(
                                                    texttemplate='%{text:.1f}',
                                                    textposition='outside'
                                                )
                                                
                                                fig.update_layout(
                                                    xaxis_title="Terminal",
                                                    yaxis_title="Índice de Calidad",
                                                    plot_bgcolor='white'
                                                )
                                            else:
                                                # Gráfico con porcentaje de malas cargas
                                                fig = px.bar(
                                                    terminal_stats,
                                                    x='Terminal',
                                                    y='% Malas',
                                                    title='% de Malas Cargas por Terminal',
                                                    color='% Malas',
                                                    color_continuous_scale='RdYlGn_r',  # Escala invertida
                                                    text='% Malas'
                                                )
                                                
                                                fig.update_traces(
                                                    texttemplate='%{text:.2f}%',
                                                    textposition='outside'
                                                )
                                                
                                                fig.update_layout(
                                                    xaxis_title="Terminal",
                                                    yaxis_title="% de Malas Cargas",
                                                    plot_bgcolor='white'
                                                )
                                        else:
                                            # Gráfico simple de cargas por terminal
                                            fig = px.bar(
                                                terminal_stats,
                                                x='Terminal',
                                                y='Total Cargas',
                                                title='Cargas por Terminal',
                                                color='Total Cargas',
                                                color_continuous_scale='Viridis',
                                                text='Total Cargas'
                                            )
                                            
                                            fig.update_traces(
                                                texttemplate='%{text}',
                                                textposition='outside'
                                            )
                                            
                                            fig.update_layout(
                                                xaxis_title="Terminal",
                                                yaxis_title="Cantidad de Cargas",
                                                plot_bgcolor='white'
                                            )
                                        
                                        st.plotly_chart(fig, use_container_width=True)
                                    
                                    with tab2:
                                        # Mostrar tabla
                                        if 'Índice Calidad' in terminal_stats.columns:
                                            st.dataframe(
                                                terminal_stats,
                                                column_config={
                                                    "Terminal": st.column_config.TextColumn("Terminal"),
                                                    "Malas Cargas": st.column_config.NumberColumn("Malas", format="%d"),
                                                    "Total Cargas": st.column_config.NumberColumn("Total", format="%d"),
                                                    "% Malas": st.column_config.NumberColumn("% Malas", format="%.2f%%"),
                                                    "Sobreconsumos": st.column_config.NumberColumn("Sobre", format="%d"),
                                                    "% Sobre": st.column_config.NumberColumn("% Sobre", format="%.2f%%"),
                                                    "Índice Calidad": st.column_config.ProgressColumn("Calidad", format="%.1f", min_value=0, max_value=100)
                                                },
                                                use_container_width=True
                                            )
                                        elif 'Mala Carga' in df_supervisor.columns:
                                            st.dataframe(
                                                terminal_stats,
                                                column_config={
                                                    "Terminal": st.column_config.TextColumn("Terminal"),
                                                    "Malas Cargas": st.column_config.NumberColumn("Malas", format="%d"),
                                                    "Total Cargas": st.column_config.NumberColumn("Total", format="%d"),
                                                    "% Malas": st.column_config.NumberColumn("% Malas", format="%.2f%%")
                                                },
                                                use_container_width=True
                                            )
                                        else:
                                            st.dataframe(
                                                terminal_stats,
                                                column_config={
                                                    "Terminal": st.column_config.TextColumn("Terminal"),
                                                    "Total Cargas": st.column_config.NumberColumn("Total", format="%d")
                                                },
                                                use_container_width=True
                                            )
                                
                                # Historial detallado
                                with st.expander("Historial Detallado de Cargas", expanded=False):
                                    # Columnas relevantes para mostrar
                                    cols_relevantes = [col for col in [
                                        'Fecha', 'Hora', 'Terminal', 'Número interno',
                                        'Cantidad litros', 'Mala Carga', 'Sobreconsumo',
                                        'Nombre conductor', 'Nombre Planillero'
                                    ] if col in df_supervisor.columns]
                                    
                                    # Mostrar datos
                                    st.dataframe(
                                        df_supervisor[cols_relevantes].sort_values('Fecha', ascending=False),
                                        use_container_width=True,
                                        height=400
                                    )
                            
                            # Si es vista de todos los supervisores
                            else:
                                st.markdown("<h3 class='section-title'>Resumen de Supervisores</h3>", unsafe_allow_html=True)
                                
                                # Crear tabla de resumen
                                resumen_supervisores = df_supervisor.groupby('Nombre supervisor').agg({
                                    'Mala Carga': ['sum', 'count'] if 'Mala Carga' in df_supervisor.columns else 'count',
                                    'Sobreconsumo': ['sum'] if 'Sobreconsumo' in df_supervisor.columns else 'count'
                                }).reset_index()
                                
                                # Aplanar columnas
                                if 'Mala Carga' in df_supervisor.columns and 'Sobreconsumo' in df_supervisor.columns:
                                    resumen_supervisores.columns = ['Supervisor', 'Malas Cargas', 'Total Cargas', 'Sobreconsumos']
                                    
                                    # Añadir porcentajes
                                    resumen_supervisores['% Malas'] = (resumen_supervisores['Malas Cargas'] / resumen_supervisores['Total Cargas'] * 100).round(2)
                                    resumen_supervisores['% Sobre'] = (resumen_supervisores['Sobreconsumos'] / resumen_supervisores['Total Cargas'] * 100).round(2)
                                    
                                    # Añadir índice de calidad
                                    resumen_supervisores['Índice Calidad'] = 100 - (resumen_supervisores['% Malas'] * 0.5 + resumen_supervisores['% Sobre'] * 0.5)
                                    resumen_supervisores['Índice Calidad'] = resumen_supervisores['Índice Calidad'].clip(0, 100).round(1)
                                    
                                    # Ordenar por índice de calidad (descendente - mejores primero)
                                    resumen_supervisores = resumen_supervisores.sort_values('Índice Calidad', ascending=False)
                                elif 'Mala Carga' in df_supervisor.columns:
                                    resumen_supervisores.columns = ['Supervisor', 'Malas Cargas', 'Total Cargas']
                                    
                                    # Añadir porcentajes
                                    resumen_supervisores['% Malas'] = (resumen_supervisores['Malas Cargas'] / resumen_supervisores['Total Cargas'] * 100).round(2)
                                    
                                    # Ordenar por porcentaje de malas cargas (ascendente - mejores primero)
                                    resumen_supervisores = resumen_supervisores.sort_values('% Malas')
                                else:
                                    resumen_supervisores.columns = ['Supervisor', 'Total Cargas']
                                    
                                    # Ordenar por cantidad de cargas (descendente)
                                    resumen_supervisores = resumen_supervisores.sort_values('Total Cargas', ascending=False)
                                
                                # Mostrar gráficos de ranking
                                if 'Índice Calidad' in resumen_supervisores.columns:
                                    # Filtrar supervisores con al menos 10 cargas
                                    supervisores_filtrados = resumen_supervisores[resumen_supervisores['Total Cargas'] >= 10].copy()
                                    
                                    # Mostrar mejores supervisores
                                    fig = px.bar(
                                        supervisores_filtrados.head(10),  # Top 10 mejores
                                        x='Supervisor',
                                        y='Índice Calidad',
                                        title='Top 10 Supervisores por Índice de Calidad',
                                        color='Índice Calidad',
                                        color_continuous_scale='RdYlGn',
                                        text='Índice Calidad'
                                    )
                                    
                                    fig.update_traces(
                                        texttemplate='%{text:.1f}',
                                        textposition='outside'
                                    )
                                    
                                    fig.update_layout(
                                        xaxis_title="Supervisor",
                                        yaxis_title="Índice de Calidad",
                                        plot_bgcolor='white'
                                    )
                                    
                                    st.plotly_chart(fig, use_container_width=True)
                                
                                # Mostrar tabla completa
                                st.markdown("<h4>Tabla Completa de Supervisores</h4>", unsafe_allow_html=True)
                                
                                # Mostrar tabla
                                if 'Índice Calidad' in resumen_supervisores.columns:
                                    st.dataframe(
                                        resumen_supervisores,
                                        column_config={
                                            "Supervisor": st.column_config.TextColumn("Supervisor"),
                                            "Malas Cargas": st.column_config.NumberColumn("Malas", format="%d"),
                                            "Total Cargas": st.column_config.NumberColumn("Total", format="%d"),
                                            "% Malas": st.column_config.NumberColumn("% Malas", format="%.2f%%"),
                                            "Sobreconsumos": st.column_config.NumberColumn("Sobre", format="%d"),
                                            "% Sobre": st.column_config.NumberColumn("% Sobre", format="%.2f%%"),
                                            "Índice Calidad": st.column_config.ProgressColumn("Calidad", format="%.1f", min_value=0, max_value=100)
                                        },
                                        use_container_width=True
                                    )
                                elif 'Mala Carga' in df_supervisor.columns:
                                    st.dataframe(
                                        resumen_supervisores,
                                        column_config={
                                            "Supervisor": st.column_config.TextColumn("Supervisor"),
                                            "Malas Cargas": st.column_config.NumberColumn("Malas", format="%d"),
                                            "Total Cargas": st.column_config.NumberColumn("Total", format="%d"),
                                            "% Malas": st.column_config.NumberColumn("% Malas", format="%.2f%%")
                                        },
                                        use_container_width=True
                                    )
                                else:
                                    st.dataframe(
                                        resumen_supervisores,
                                        column_config={
                                            "Supervisor": st.column_config.TextColumn("Supervisor"),
                                            "Total Cargas": st.column_config.NumberColumn("Total", format="%d")
                                        },
                                        use_container_width=True
                                    )
                
                # === MALAS CARGAS ===
                elif section == "Malas Cargas":
                    st.markdown("<h2 class='subtitle'>Análisis de Malas Cargas</h2>", unsafe_allow_html=True)
                    
                    if 'Mala Carga' in df.columns:
                        # Fecha de análisis
                        col1, col2 = st.columns([1, 1])
                        
                        with col1:
                            if 'Fecha' in df.columns:
                                fecha_min = df['Fecha'].min().date()
                                fecha_max = df['Fecha'].max().date()
                            else:
                                fecha_min = datetime.date.today() - datetime.timedelta(days=30)
                                fecha_max = datetime.date.today()
                            
                            fecha_inicio = st.date_input("Fecha inicio", fecha_min, key="malas_fecha_inicio")
                        
                        with col2:
                            fecha_fin = st.date_input("Fecha fin", fecha_max, key="malas_fecha_fin")
                        
                        # Filtrar por fecha
                        if 'Fecha' in df.columns:
                            df_filtrado = df[
                                (df['Fecha'].dt.date >= fecha_inicio) & 
                                (df['Fecha'].dt.date <= fecha_fin)
                            ]
                        else:
                            df_filtrado = df.copy()
                        
                        # Filtrar malas cargas
                        df_malas = df_filtrado[df_filtrado['Mala Carga'] == True]
                        total_malas = len(df_malas)
                        
                        # Métricas principales
                        col1, col2, col3 = st.columns(3)
                        
                        with col1:
                            # Total malas cargas
                            porcentaje = (total_malas / len(df_filtrado) * 100) if len(df_filtrado) > 0 else 0
                            
                            # Color según porcentaje
                            color = "success" if porcentaje < 2 else "warning" if porcentaje < 5 else "danger"
                            
                            st.markdown(create_kpi_card(
                                "Total Malas Cargas", 
                                f"{total_malas}",
                                None,
                                icon="❌",
                                color_class=color,
                                footnote=f"{porcentaje:.2f}% del total"
                            ), unsafe_allow_html=True)
                        
                        with col2:
                            # Litros en malas cargas
                            if 'Cantidad litros' in df_malas.columns:
                                litros_malas = df_malas['Cantidad litros'].sum()
                                porcentaje_litros = (litros_malas / df_filtrado['Cantidad litros'].sum() * 100) if df_filtrado['Cantidad litros'].sum() > 0 else 0
                                
                                st.markdown(create_kpi_card(
                                    "Litros en Malas Cargas", 
                                    f"{litros_malas:,.0f}",
                                    None,
                                    icon="⛽",
                                    color_class="info",
                                    footnote=f"{porcentaje_litros:.2f}% del total"
                                ), unsafe_allow_html=True)
                        
                        with col3:
                            # Terminales afectadas
                            if 'Terminal' in df_malas.columns:
                                terminales_afectadas = df_malas['Terminal'].nunique()
                                porcentaje_terminales = (terminales_afectadas / df_filtrado['Terminal'].nunique() * 100) if df_filtrado['Terminal'].nunique() > 0 else 0
                                
                                st.markdown(create_kpi_card(
                                    "Terminales Afectadas", 
                                    f"{terminales_afectadas}",
                                    None,
                                    icon="🏢",
                                    color_class="primary",
                                    footnote=f"{porcentaje_terminales:.2f}% del total"
                                ), unsafe_allow_html=True)
                        
                        # Alerta general
                        if total_malas > 0:
                            if porcentaje >= 5:
                                st.markdown(show_alert(
                                    f"Se han detectado {total_malas} cargas registradas como malas, lo que representa el {porcentaje:.2f}% del total. Esto indica un problema significativo en los procedimientos de registro que requiere atención inmediata.",
                                    "danger",
                                    "ALERTA CRÍTICA: Alto porcentaje de malas cargas",
                                    "🚨"
                                ), unsafe_allow_html=True)
                            elif porcentaje >= 2:
                                st.markdown(show_alert(
                                    f"Se han detectado {total_malas} cargas registradas como malas, lo que representa el {porcentaje:.2f}% del total. Se recomienda revisar los procedimientos de registro.",
                                    "warning",
                                    "ALERTA: Malas cargas detectadas",
                                    "⚠️"
                                ), unsafe_allow_html=True)
                            else:
                                st.markdown(show_alert(
                                    f"Se han detectado {total_malas} cargas registradas como malas, lo que representa el {porcentaje:.2f}% del total. Aunque el porcentaje es bajo, se recomienda supervisar la situación.",
                                    "info",
                                    "AVISO: Malas cargas detectadas",
                                    "ℹ️"
                                ), unsafe_allow_html=True)
                        else:
                            st.markdown(show_alert(
                                f"No se han encontrado cargas registradas como malas en los datos analizados.",
                                "success",
                                "No se detectaron malas cargas",
                                "✅"
                            ), unsafe_allow_html=True)
                        
                        # Análisis por terminal
                        if total_malas > 0 and 'Terminal' in df_malas.columns:
                            st.markdown("<h3 class='section-title'>Malas Cargas por Terminal</h3>", unsafe_allow_html=True)
                            
                            # Tabs para diferentes vistas
                            tab1, tab2 = st.tabs(["📊 Gráfico", "📋 Tabla"])
                            
                            with tab1:
                                # Contar malas cargas por terminal
                                terminal_counts = df_malas['Terminal'].value_counts().reset_index()
                                terminal_counts.columns = ['Terminal', 'Malas Cargas']
                                
                                # Calcular porcentaje respecto al total de cargas en cada terminal
                                total_por_terminal = df_filtrado.groupby('Terminal').size().reset_index()
                                total_por_terminal.columns = ['Terminal', 'Total Cargas']
                                
                                # Unir dataframes
                                terminal_analysis = pd.merge(terminal_counts, total_por_terminal, on='Terminal')
                                terminal_analysis['Porcentaje'] = (terminal_analysis['Malas Cargas'] / terminal_analysis['Total Cargas'] * 100).round(2)
                                
                                # Ordenar por porcentaje (descendente)
                                terminal_analysis = terminal_analysis.sort_values('Porcentaje', ascending=False)
                                
                                # Gráfico
                                fig = px.bar(
                                    terminal_analysis,
                                    x='Terminal',
                                    y='Porcentaje',
                                    title='Porcentaje de Malas Cargas por Terminal',
                                    color='Porcentaje',
                                    color_continuous_scale='Reds',
                                    text='Porcentaje'
                                )
                                
                                fig.update_traces(
                                    texttemplate='%{text:.2f}%',
                                    textposition='outside'
                                )
                                
                                fig.update_layout(
                                    xaxis_title="Terminal",
                                    yaxis_title="% de Malas Cargas",
                                    plot_bgcolor='white'
                                )
                                
                                # Añadir línea para el promedio
                                fig.add_hline(
                                    y=porcentaje,
                                    line_dash="dash",
                                    line_color="red",
                                    annotation_text=f"Promedio: {porcentaje:.2f}%"
                                )
                                
                                st.plotly_chart(fig, use_container_width=True)
                                
                                # Gráfico de cantidades absolutas
                                fig = px.bar(
                                    terminal_analysis.sort_values('Malas Cargas', ascending=False),
                                    x='Terminal',
                                    y='Malas Cargas',
                                    title='Cantidad de Malas Cargas por Terminal',
                                    color='Porcentaje',
                                    color_continuous_scale='Reds',
                                    text='Malas Cargas'
                                )
                                
                                fig.update_traces(
                                    texttemplate='%{text}',
                                    textposition='outside'
                                )
                                
                                fig.update_layout(
                                    xaxis_title="Terminal",
                                    yaxis_title="Cantidad de Malas Cargas",
                                    plot_bgcolor='white'
                                )
                                
                                st.plotly_chart(fig, use_container_width=True)
                            
                            with tab2:
                                # Mostrar tabla
                                st.dataframe(
                                    terminal_analysis,
                                    column_config={
                                        "Terminal": st.column_config.TextColumn("Terminal"),
                                        "Malas Cargas": st.column_config.NumberColumn("Malas Cargas", format="%d"),
                                        "Total Cargas": st.column_config.NumberColumn("Total Cargas", format="%d"),
                                        "Porcentaje": st.column_config.NumberColumn("% Malas Cargas", format="%.2f%%")
                                    },
                                    use_container_width=True
                                )
                            
                            # Terminal con más problemas
                            if not terminal_analysis.empty:
                                peor_terminal = terminal_analysis.iloc[0]['Terminal']
                                porcentaje_peor = terminal_analysis.iloc[0]['Porcentaje']
                                
                                if porcentaje_peor > 5:
                                    st.markdown(show_alert(
                                        f"Esta terminal muestra un {porcentaje_peor:.2f}% de malas cargas, muy por encima del promedio. Se recomienda una revisión urgente de sus procedimientos.",
                                        "danger",
                                        f"Terminal Crítica: {peor_terminal}",
                                        "🚨"
                                    ), unsafe_allow_html=True)
                        
                        # Análisis por planillero
                        if total_malas > 0 and 'Nombre Planillero' in df_malas.columns:
                            st.markdown("<h3 class='section-title'>Malas Cargas por Planillero</h3>", unsafe_allow_html=True)
                            
                            # Contar malas cargas por planillero
                            planillero_counts = df_malas['Nombre Planillero'].value_counts().reset_index()
                            planillero_counts.columns = ['Planillero', 'Malas Cargas']
                            
                            # Calcular porcentaje respecto al total de cargas de cada planillero
                            total_por_planillero = df_filtrado.groupby('Nombre Planillero').size().reset_index()
                            total_por_planillero.columns = ['Planillero', 'Total Cargas']
                            
                            # Unir dataframes
                            planillero_analysis = pd.merge(planillero_counts, total_por_planillero, on='Planillero')
                            planillero_analysis['Porcentaje'] = (planillero_analysis['Malas Cargas'] / planillero_analysis['Total Cargas'] * 100).round(2)
                            
                            # Filtrar planilleros con al menos 5 cargas para evitar valores atípicos
                            planillero_analysis = planillero_analysis[planillero_analysis['Total Cargas'] >= 5]
                            
                            # Ordenar por porcentaje (descendente)
                            planillero_analysis = planillero_analysis.sort_values('Porcentaje', ascending=False)
                            
                            # Mostrar top 10
                            top10 = planillero_analysis.head(10)
                            
                            # Gráfico
                            fig = px.bar(
                                top10,
                                x='Planillero',
                                y='Porcentaje',
                                title='Top 10 Planilleros con Mayor % de Malas Cargas',
                                color='Porcentaje',
                                color_continuous_scale='Reds',
                                text='Porcentaje'
                            )
                            
                            fig.update_traces(
                                texttemplate='%{text:.2f}%',
                                textposition='outside'
                            )
                            
                            fig.update_layout(
                                xaxis_title="Planillero",
                                yaxis_title="% de Malas Cargas",
                                plot_bgcolor='white'
                            )
                            
                            # Añadir línea para el promedio
                            fig.add_hline(
                                y=porcentaje,
                                line_dash="dash",
                                line_color="red",
                                annotation_text=f"Promedio: {porcentaje:.2f}%"
                            )
                            
                            st.plotly_chart(fig, use_container_width=True)
                            
                            # Mostrar tabla
                            st.markdown("<h4>Top 10 Planilleros con Mayor % de Malas Cargas</h4>", unsafe_allow_html=True)
                            st.dataframe(
                                top10,
                                column_config={
                                    "Planillero": st.column_config.TextColumn("Planillero"),
                                    "Malas Cargas": st.column_config.NumberColumn("Malas Cargas", format="%d"),
                                    "Total Cargas": st.column_config.NumberColumn("Total Cargas", format="%d"),
                                    "Porcentaje": st.column_config.NumberColumn("% Malas Cargas", format="%.2f%%")
                                },
                                use_container_width=True
                            )
                            
                            # Planillero con más problemas
                            if not planillero_analysis.empty:
                                peor_planillero = planillero_analysis.iloc[0]['Planillero']
                                porcentaje_peor = planillero_analysis.iloc[0]['Porcentaje']
                                
                                if porcentaje_peor > 10:
                                    st.markdown(show_alert(
                                        f"Este planillero muestra un {porcentaje_peor:.2f}% de malas cargas, muy por encima del promedio. Se recomienda capacitación adicional o supervisión directa.",
                                        "danger",
                                        f"Planillero Crítico: {peor_planillero}",
                                        "🚨"
                                    ), unsafe_allow_html=True)
                        
                        # Tendencia temporal
                        if total_malas > 0 and 'Fecha' in df_malas.columns:
                            st.markdown("<h3 class='section-title'>Evolución Temporal de Malas Cargas</h3>", unsafe_allow_html=True)
                            
                            # Agrupar por fecha
                            malas_por_dia = df_malas.groupby(df_malas['Fecha'].dt.date).size().reset_index()
                            malas_por_dia.columns = ['Fecha', 'Malas Cargas']
                            
                            # Total por día
                            total_por_dia = df_filtrado.groupby(df_filtrado['Fecha'].dt.date).size().reset_index()
                            total_por_dia.columns = ['Fecha', 'Total Cargas']
                            
                            # Unir dataframes
                            tendencia = pd.merge(malas_por_dia, total_por_dia, on='Fecha')
                            tendencia['Porcentaje'] = (tendencia['Malas Cargas'] / tendencia['Total Cargas'] * 100).round(2)
                            
                            # Gráfico
                            fig = make_subplots(specs=[[{"secondary_y": True}]])
                            
                            # Añadir línea de malas cargas
                            fig.add_trace(
                                go.Bar(
                                    x=tendencia['Fecha'],
                                    y=tendencia['Malas Cargas'],
                                    name='Malas Cargas',
                                    marker_color=COLORS['danger']
                                ),
                                secondary_y=False
                            )
                            
                            # Añadir línea de porcentaje
                            fig.add_trace(
                                go.Scatter(
                                    x=tendencia['Fecha'],
                                    y=tendencia['Porcentaje'],
                                    name='% Malas Cargas',
                                    mode='lines+markers',
                                    marker=dict(color=COLORS['warning']),
                                    line=dict(color=COLORS['warning'], width=3)
                                ),
                                secondary_y=True
                            )
                            
                            # Actualizar ejes
                            fig.update_layout(
                                title='Evolución de Malas Cargas a lo Largo del Tiempo',
                                plot_bgcolor='white',
                                hovermode='x unified',
                                height=450,
                                legend=dict(
                                    orientation="h",
                                    yanchor="bottom",
                                    y=1.02,
                                    xanchor="right",
                                    x=1
                                )
                            )
                            
                            fig.update_xaxes(title_text='Fecha')
                            fig.update_yaxes(title_text='Cantidad de Malas Cargas', secondary_y=False)
                            fig.update_yaxes(title_text='% de Malas Cargas', secondary_y=True)
                            
                            # Añadir línea de promedio
                            fig.add_hline(
                                y=porcentaje,
                                line_dash="dash",
                                line_color="red",
                                secondary_y=True,
                                annotation_text=f"Promedio: {porcentaje:.2f}%"
                            )
                            
                            st.plotly_chart(fig, use_container_width=True)
                            
                            # Análisis de tendencia
                            if len(tendencia) > 3:
                                # Calcular tendencia lineal
                                X = np.array(range(len(tendencia))).reshape(-1, 1)
                                y = tendencia['Porcentaje'].values
                                
                                modelo = LinearRegression()
                                modelo.fit(X, y)
                                tendencia_valor = modelo.coef_[0]
                                
                                # Generar predicción para graficar línea de tendencia
                                y_pred = modelo.predict(X)
                                
                                # Determinar si es creciente o decreciente
                                if tendencia_valor > 0.1:
                                    st.markdown(show_alert(
                                        f"La tendencia de malas cargas está aumentando en un {tendencia_valor:.2f}% por día. Se recomienda revisar urgentemente los procedimientos de registro.",
                                        "danger",
                                        "Tendencia Creciente de Malas Cargas",
                                        "📈"
                                    ), unsafe_allow_html=True)
                                elif tendencia_valor < -0.1:
                                    st.markdown(show_alert(
                                        f"La tendencia de malas cargas está disminuyendo en un {abs(tendencia_valor):.2f}% por día. Las medidas correctivas parecen estar funcionando.",
                                        "success",
                                        "Tendencia Decreciente de Malas Cargas",
                                        "📉"
                                    ), unsafe_allow_html=True)
                        
                        # Recomendaciones específicas
                        st.markdown("<h3 class='section-title'>Recomendaciones para Reducir Malas Cargas</h3>", unsafe_allow_html=True)
                        
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.markdown(f"""
                            <div class="card">
                                <h4>Recomendaciones Operativas</h4>
                                <ul>
                                    <li><strong>Capacitación Específica:</strong> Realizar sesiones de capacitación focalizadas para planilleros con altas tasas de malas cargas.</li>
                                    <li><strong>Estandarización de Procesos:</strong> Implementar protocolos claros y estandarizados para el registro de cargas.</li>
                                    <li><strong>Supervisión Reforzada:</strong> Incrementar la supervisión en terminales con mayor incidencia de malas cargas.</li>
                                    <li><strong>Verificación Doble:</strong> Implementar un sistema de verificación doble para cargas grandes o inusuales.</li>
                                </ul>
                            </div>
                            """, unsafe_allow_html=True)
                        
                        with col2:
                            st.markdown(f"""
                            <div class="card">
                                <h4>Mejoras Tecnológicas</h4>
                                <ul>
                                    <li><strong>Automatización:</strong> Implementar sistemas de medición automática cuando sea posible.</li>
                                    <li><strong>Validación en Tiempo Real:</strong> Desarrollar controles de validación que detecten valores anómalos durante el ingreso.</li>
                                    <li><strong>Feedback Inmediato:</strong> Proporcionar retroalimentación inmediata a planilleros cuando se detecta un patrón de malas cargas.</li>
                                    <li><strong>Monitoreo Continuo:</strong> Establecer un sistema de monitoreo continuo con alertas para tendencias negativas.</li>
                                </ul>
                            </div>
                            """, unsafe_allow_html=True)
                        
                        # Listado detallado
                        st.markdown("<h3 class='section-title'>Listado Detallado de Malas Cargas</h3>", unsafe_allow_html=True)
                        
                        # Columnas relevantes
                        cols_relevantes = [col for col in [
                            'Fecha', 'Hora', 'Terminal', 'Número interno', 'Patente', 
                            'Cantidad litros', 'Tipo', 'Nombre Planillero', 'Nombre supervisor'
                        ] if col in df_malas.columns]
                        
                        # Mostrar datos con paginación
                        with st.expander("Ver listado completo", expanded=False):
                            # Crear selector de página
                            if len(df_malas) > 50:
                                # Calcular número de páginas (50 registros por página)
                                num_paginas = (len(df_malas) - 1) // 50 + 1
                                pagina_actual = st.selectbox("Página", range(1, num_paginas + 1), key="pagina_malas_cargas")
                                
                                # Índices de inicio y fin para la página seleccionada
                                inicio = (pagina_actual - 1) * 50
                                fin = min(inicio + 50, len(df_malas))
                                
                                # Mostrar datos de la página actual
                                st.dataframe(
                                    df_malas[cols_relevantes].sort_values('Fecha', ascending=False).iloc[inicio:fin],
                                    use_container_width=True,
                                    height=400
                                )
                                
                                st.info(f"Mostrando registros {inicio+1} a {fin} de {len(df_malas)}")
                            else:
                                # Si hay pocos registros, mostrar todos
                                st.dataframe(
                                    df_malas[cols_relevantes].sort_values('Fecha', ascending=False),
                                    use_container_width=True,
                                    height=400
                                )
                        
                        # Opción para exportar
                        if len(df_malas) > 0:
                            st.download_button(
                                label="📥 Descargar listado de malas cargas",
                                data=df_malas[cols_relevantes].sort_values('Fecha', ascending=False).to_csv(index=False).encode('utf-8'),
                                file_name="malas_cargas.csv",
                                mime="text/csv",
                                help="Descargar un archivo CSV con el listado completo de malas cargas",
                                use_container_width=True
                            )
                    else:
                        st.warning("No se encontró información sobre malas cargas en los datos.")
                
                # === DETECCIÓN DE SOBRECONSUMO ===
                elif section == "Detección de Sobreconsumo":
                    st.markdown("<h2 class='subtitle'>Análisis de Sobreconsumo</h2>", unsafe_allow_html=True)
                    
                    if 'Sobreconsumo' in df.columns:
                        # Fecha de análisis
                        col1, col2 = st.columns([1, 1])
                        
                        with col1:
                            if 'Fecha' in df.columns:
                                fecha_min = df['Fecha'].min().date()
                                fecha_max = df['Fecha'].max().date()
                            else:
                                fecha_min = datetime.date.today() - datetime.timedelta(days=30)
                                fecha_max = datetime.date.today()
                            
                            fecha_inicio = st.date_input("Fecha inicio", fecha_min, key="sobre_fecha_inicio")
                        
                        with col2:
                            fecha_fin = st.date_input("Fecha fin", fecha_max, key="sobre_fecha_fin")
                        
                        # Filtrar por fecha
                        if 'Fecha' in df.columns:
                            df_filtrado = df[
                                (df['Fecha'].dt.date >= fecha_inicio) & 
                                (df['Fecha'].dt.date <= fecha_fin)
                            ]
                        else:
                            df_filtrado = df.copy()
                        
                        # Filtrar sobreconsumos
                        df_sobre = df_filtrado[df_filtrado['Sobreconsumo'] == True]
                        total_sobre = len(df_sobre)
                        
                        # Métricas principales
                        col1, col2, col3 = st.columns(3)
                        
                        with col1:
                            # Total sobreconsumos
                            porcentaje = (total_sobre / len(df_filtrado) * 100) if len(df_filtrado) > 0 else 0
                            
                            # Color según porcentaje
                            color = "success" if porcentaje < 5 else "warning" if porcentaje < 10 else "danger"
                            
                            st.markdown(create_kpi_card(
                                "Total Sobreconsumos", 
                                f"{total_sobre}",
                                None,
                                icon="⚠️",
                                color_class=color,
                                footnote=f"{porcentaje:.2f}% del total"
                            ), unsafe_allow_html=True)
                        
                        with col2:
                            # Litros en sobreconsumo
                            if 'Cantidad litros' in df_sobre.columns:
                                litros_sobre = df_sobre['Cantidad litros'].sum()
                                porcentaje_litros = (litros_sobre / df_filtrado['Cantidad litros'].sum() * 100) if df_filtrado['Cantidad litros'].sum() > 0 else 0
                                
                                st.markdown(create_kpi_card(
                                    "Litros en Sobreconsumo", 
                                    f"{litros_sobre:,.0f}",
                                    None,
                                    icon="⛽",
                                    color_class="info",
                                    footnote=f"{porcentaje_litros:.2f}% del total"
                                ), unsafe_allow_html=True)
                        
                        with col3:
                            # Litros excedentes
                            if 'Cantidad litros' in df_sobre.columns and 'Promedio Modelo' in df_sobre.columns:
                                litros_excedentes = (df_sobre['Cantidad litros'] - df_sobre['Promedio Modelo']).sum()
                                porcentaje_excedente = (litros_excedentes / litros_sobre * 100) if litros_sobre > 0 else 0
                                
                                st.markdown(create_kpi_card(
                                    "Litros Excedentes", 
                                    f"{litros_excedentes:,.0f}",
                                    None,
                                    icon="💧",
                                    color_class="danger",
                                    footnote=f"{porcentaje_excedente:.2f}% del consumo"
                                ), unsafe_allow_html=True)
                        
                        # Alerta general
                        if total_sobre > 0:
                            if porcentaje >= 10:
                                st.markdown(show_alert(
                                    f"Se han detectado {total_sobre} cargas con sobreconsumo, lo que representa el {porcentaje:.2f}% del total. Esto indica un problema significativo en la eficiencia de consumo que requiere atención inmediata.",
                                    "danger",
                                    "ALERTA CRÍTICA: Alto porcentaje de sobreconsumo",
                                    "🚨"
                                ), unsafe_allow_html=True)
                            elif porcentaje >= 5:
                                st.markdown(show_alert(
                                    f"Se han detectado {total_sobre} cargas con sobreconsumo, lo que representa el {porcentaje:.2f}% del total. Se recomienda revisar la eficiencia de los buses y los procedimientos de carga.",
                                    "warning",
                                    "ALERTA: Sobreconsumo detectado",
                                    "⚠️"
                                ), unsafe_allow_html=True)
                            else:
                                st.markdown(show_alert(
                                    f"Se han detectado {total_sobre} cargas con sobreconsumo, lo que representa el {porcentaje:.2f}% del total. Aunque el porcentaje es bajo, se recomienda supervisar la situación.",
                                    "info",
                                    "AVISO: Sobreconsumo detectado",
                                    "ℹ️"
                                ), unsafe_allow_html=True)
                        else:
                            st.markdown(show_alert(
                                f"No se han encontrado cargas con sobreconsumo en los datos analizados.",
                                "success",
                                "No se detectaron casos de sobreconsumo",
                                "✅"
                            ), unsafe_allow_html=True)
                        
                        # Análisis por modelo
                        if total_sobre > 0 and 'Modelo chasis' in df_sobre.columns:
                            st.markdown("<h3 class='section-title'>Sobreconsumo por Modelo de Chasis</h3>", unsafe_allow_html=True)
                            
                            # Tabs para diferentes vistas
                            tab1, tab2 = st.tabs(["📊 Gráfico", "📋 Tabla"])
                            
                            with tab1:
                                # Contar sobreconsumos por modelo
                                modelo_counts = df_sobre['Modelo chasis'].value_counts().reset_index()
                                modelo_counts.columns = ['Modelo', 'Sobreconsumos']
                                
                                # Calcular porcentaje respecto al total de cargas en cada modelo
                                total_por_modelo = df_filtrado.groupby('Modelo chasis').size().reset_index()
                                total_por_modelo.columns = ['Modelo', 'Total Cargas']
                                
                                # Unir dataframes
                                modelo_analysis = pd.merge(modelo_counts, total_por_modelo, on='Modelo')
                                modelo_analysis['Porcentaje'] = (modelo_analysis['Sobreconsumos'] / modelo_analysis['Total Cargas'] * 100).round(2)
                                
                                # Añadir información de consumo
                                if 'Cantidad litros' in df_filtrado.columns:
                                    # Promedio por modelo
                                    promedio_por_modelo = df_filtrado.groupby('Modelo chasis')['Cantidad litros'].mean().reset_index()
                                    promedio_por_modelo.columns = ['Modelo', 'Promedio Litros']
                                    modelo_analysis = pd.merge(modelo_analysis, promedio_por_modelo, on='Modelo')
                                    
                                    # Promedio en sobreconsumos
                                    promedio_sobre = df_sobre.groupby('Modelo chasis')['Cantidad litros'].mean().reset_index()
                                    promedio_sobre.columns = ['Modelo', 'Promedio Sobreconsumo']
                                    modelo_analysis = pd.merge(modelo_analysis, promedio_sobre, on='Modelo')
                                    
                                    # Calcular exceso
                                    modelo_analysis['Exceso Litros'] = (modelo_analysis['Promedio Sobreconsumo'] - modelo_analysis['Promedio Litros']).round(2)
                                    modelo_analysis['Exceso %'] = ((modelo_analysis['Promedio Sobreconsumo'] / modelo_analysis['Promedio Litros'] - 1) * 100).round(2)
                                
                                # Ordenar por porcentaje
                                modelo_analysis = modelo_analysis.sort_values('Porcentaje', ascending=False)
                                
                                # Gráfico
                                fig = px.bar(
                                    modelo_analysis,
                                    x='Modelo',
                                    y='Porcentaje',
                                    title='Porcentaje de Sobreconsumo por Modelo',
                                    color='Porcentaje',
                                    color_continuous_scale='Reds',
                                    text='Porcentaje'
                                )
                                
                                fig.update_traces(
                                    texttemplate='%{text:.2f}%',
                                    textposition='outside'
                                )
                                
                                fig.update_layout(
                                    xaxis_title="Modelo de Chasis",
                                    yaxis_title="% de Sobreconsumo",
                                    plot_bgcolor='white'
                                )
                                
                                # Añadir línea para el promedio
                                fig.add_hline(
                                    y=porcentaje,
                                    line_dash="dash",
                                    line_color="red",
                                    annotation_text=f"Promedio: {porcentaje:.2f}%"
                                )
                                
                                st.plotly_chart(fig, use_container_width=True)
                                
                                # Gráfico de exceso si está disponible
                                if 'Exceso %' in modelo_analysis.columns:
                                    fig = px.bar(
                                        modelo_analysis.sort_values('Exceso %', ascending=False),
                                        x='Modelo',
                                        y='Exceso %',
                                        title='Porcentaje de Exceso en Sobreconsumos por Modelo',
                                        color='Exceso %',
                                        color_continuous_scale='Reds',
                                        text='Exceso %'
                                    )
                                    
                                    fig.update_traces(
                                        texttemplate='%{text:.2f}%',
                                        textposition='outside'
                                    )
                                    
                                    fig.update_layout(
                                        xaxis_title="Modelo de Chasis",
                                        yaxis_title="% de Exceso",
                                        plot_bgcolor='white'
                                    )
                                    
                                    st.plotly_chart(fig, use_container_width=True)
                            
                            with tab2:
                                # Mostrar tabla
                                if 'Exceso %' in modelo_analysis.columns:
                                    st.dataframe(
                                        modelo_analysis,
                                        column_config={
                                            "Modelo": st.column_config.TextColumn("Modelo de Chasis"),
                                            "Sobreconsumos": st.column_config.NumberColumn("Sobreconsumos", format="%d"),
                                            "Total Cargas": st.column_config.NumberColumn("Total Cargas", format="%d"),
                                            "Porcentaje": st.column_config.NumberColumn("% Sobreconsumo", format="%.2f%%"),
                                            "Promedio Litros": st.column_config.NumberColumn("Promedio Normal (L)", format="%.2f"),
                                            "Promedio Sobreconsumo": st.column_config.NumberColumn("Promedio en Sobreconsumo (L)", format="%.2f"),
                                            "Exceso Litros": st.column_config.NumberColumn("Exceso (L)", format="%.2f"),
                                            "Exceso %": st.column_config.NumberColumn("Exceso (%)", format="%.2f%%")
                                        },
                                        use_container_width=True
                                    )
                                else:
                                    st.dataframe(
                                        modelo_analysis,
                                        column_config={
                                            "Modelo": st.column_config.TextColumn("Modelo de Chasis"),
                                            "Sobreconsumos": st.column_config.NumberColumn("Sobreconsumos", format="%d"),
                                            "Total Cargas": st.column_config.NumberColumn("Total Cargas", format="%d"),
                                            "Porcentaje": st.column_config.NumberColumn("% Sobreconsumo", format="%.2f%%")
                                        },
                                        use_container_width=True
                                    )
                                
                                # Modelo con más problemas
                                if not modelo_analysis.empty:
                                    peor_modelo = modelo_analysis.iloc[0]['Modelo']
                                    porcentaje_peor = modelo_analysis.iloc[0]['Porcentaje']
                                    
                                    if porcentaje_peor > 15:
                                        exceso_info = ""
                                        if 'Exceso %' in modelo_analysis.columns:
                                            exceso_info = f" con un exceso promedio del {modelo_analysis.iloc[0]['Exceso %']:.2f}%"
                                        
                                        st.markdown(show_alert(
                                            f"Este modelo muestra un {porcentaje_peor:.2f}% de sobreconsumos{exceso_info}. Se recomienda una revisión técnica de este modelo y capacitación especializada para sus conductores.",
                                            "danger",
                                            f"Modelo Crítico: {peor_modelo}",
                                            "🚨"
                                        ), unsafe_allow_html=True)
                        
                        # Análisis por conductor
                        if total_sobre > 0 and 'Nombre conductor' in df_sobre.columns:
                            st.markdown("<h3 class='section-title'>Sobreconsumo por Conductor</h3>", unsafe_allow_html=True)
                            
                            # Contar sobreconsumos por conductor
                            conductor_counts = df_sobre['Nombre conductor'].value_counts().reset_index()
                            conductor_counts.columns = ['Conductor', 'Sobreconsumos']
                            
                            # Calcular porcentaje respecto al total de cargas de cada conductor
                            total_por_conductor = df_filtrado.groupby('Nombre conductor').size().reset_index()
                            total_por_conductor.columns = ['Conductor', 'Total Cargas']
                            
                            # Unir dataframes
                            conductor_analysis = pd.merge(conductor_counts, total_por_conductor, on='Conductor')
                            conductor_analysis['Porcentaje'] = (conductor_analysis['Sobreconsumos'] / conductor_analysis['Total Cargas'] * 100).round(2)
                            
                            # Filtrar conductores con al menos 5 cargas para evitar valores atípicos
                            conductor_analysis = conductor_analysis[conductor_analysis['Total Cargas'] >= 5]
                            
                            # Ordenar por porcentaje (descendente)
                            conductor_analysis = conductor_analysis.sort_values('Porcentaje', ascending=False)
                            
                            # Mostrar top 15
                            top15 = conductor_analysis.head(15)
                            
                            # Gráfico
                            fig = px.bar(
                                top15,
                                x='Conductor',
                                y='Porcentaje',
                                title='Top 15 Conductores con Mayor % de Sobreconsumo',
                                color='Porcentaje',
                                color_continuous_scale='Reds',
                                text='Porcentaje'
                            )
                            
                            fig.update_traces(
                                texttemplate='%{text:.2f}%',
                                textposition='outside'
                            )
                            
                            fig.update_layout(
                                xaxis_title="Conductor",
                                yaxis_title="% de Sobreconsumo",
                                plot_bgcolor='white'
                            )
                            
                            # Añadir línea para el promedio
                            fig.add_hline(
                                y=porcentaje,
                                line_dash="dash",
                                line_color="red",
                                annotation_text=f"Promedio: {porcentaje:.2f}%"
                            )
                            
                            st.plotly_chart(fig, use_container_width=True)
                            
                            # Mostrar tabla
                            st.markdown("<h4>Top 15 Conductores con Mayor % de Sobreconsumo</h4>", unsafe_allow_html=True)
                            st.dataframe(
                                top15,
                                column_config={
                                    "Conductor": st.column_config.TextColumn("Conductor"),
                                    "Sobreconsumos": st.column_config.NumberColumn("Sobreconsumos", format="%d"),
                                    "Total Cargas": st.column_config.NumberColumn("Total Cargas", format="%d"),
                                    "Porcentaje": st.column_config.NumberColumn("% Sobreconsumo", format="%.2f%%")
                                },
                                use_container_width=True
                            )
                            
                            # Conductor con más problemas
                            if not conductor_analysis.empty:
                                peor_conductor = conductor_analysis.iloc[0]['Conductor']
                                porcentaje_peor = conductor_analysis.iloc[0]['Porcentaje']
                                
                                if porcentaje_peor > 20:
                                    st.markdown(show_alert(
                                        f"Este conductor muestra un {porcentaje_peor:.2f}% de sobreconsumos, muy por encima del promedio. Se recomienda capacitación especializada en técnicas de conducción eficiente.",
                                        "danger",
                                        f"Conductor Crítico: {peor_conductor}",
                                        "🚨"
                                    ), unsafe_allow_html=True)
                        
                        # Distribución por severidad
                        if total_sobre > 0 and 'Z-Score' in df_sobre.columns:
                            st.markdown("<h3 class='section-title'>Distribución por Severidad</h3>", unsafe_allow_html=True)
                            
                            # Crear categorías de severidad
                            def categorizar_severidad(z):
                                if z <= 2.5:
                                    return "Leve (2-2.5σ)"
                                elif z <= 3:
                                    return "Moderado (2.5-3σ)"
                                elif z <= 4:
                                    return "Alto (3-4σ)"
                                else:
                                    return "Extremo (>4σ)"
                            
                            df_sobre['Severidad'] = df_sobre['Z-Score'].apply(categorizar_severidad)
                            
                            # Contar por severidad
                            severidad_counts = df_sobre['Severidad'].value_counts().reset_index()
                            severidad_counts.columns = ['Severidad', 'Cantidad']
                            
                            # Ordenar por severidad
                            orden_severidad = {
                                "Leve (2-2.5σ)": 1,
                                "Moderado (2.5-3σ)": 2,
                                "Alto (3-4σ)": 3,
                                "Extremo (>4σ)": 4
                            }
                            severidad_counts['Orden'] = severidad_counts['Severidad'].map(orden_severidad)
                            severidad_counts = severidad_counts.sort_values('Orden')
                            
                            # Tabs para diferentes vistas
                            tab1, tab2 = st.tabs(["📊 Gráfico", "📋 Tabla"])
                            
                            with tab1:
                                # Gráfico
                                colores_severidad = {
                                    "Leve (2-2.5σ)": "#FFD700",
                                    "Moderado (2.5-3σ)": "#FFA500",
                                    "Alto (3-4σ)": "#FF4500",
                                    "Extremo (>4σ)": "#FF0000"
                                }
                                
                                fig = px.pie(
                                    severidad_counts,
                                    names='Severidad',
                                    values='Cantidad',
                                    title='Distribución de Sobreconsumos por Severidad',
                                    color='Severidad',
                                    color_discrete_map=colores_severidad
                                )
                                
                                fig.update_traces(
                                    textposition='inside',
                                    textinfo='percent+label'
                                )
                                
                                fig.update_layout(
                                    plot_bgcolor='white'
                                )
                                
                                st.plotly_chart(fig, use_container_width=True)
                            
                            with tab2:
                                # Mostrar tabla
                                st.dataframe(
                                    severidad_counts[['Severidad', 'Cantidad']],
                                    column_config={
                                        "Severidad": st.column_config.TextColumn("Nivel de Severidad"),
                                        "Cantidad": st.column_config.NumberColumn("Cantidad", format="%d")
                                    },
                                    use_container_width=True
                                )
                                
                                # Añadir porcentajes
                                severidad_counts['Porcentaje'] = (severidad_counts['Cantidad'] / severidad_counts['Cantidad'].sum() * 100).round(2)
                                
                                # Mostrar estadísticas
                                st.markdown(f"""
                                <div class="card">
                                    <h4>Análisis de Severidad</h4>
                                    <ul>
                                        <li><strong>Leves (2-2.5σ):</strong> {severidad_counts[severidad_counts['Severidad']=='Leve (2-2.5σ)']['Porcentaje'].values[0] if 'Leve (2-2.5σ)' in severidad_counts['Severidad'].values else 0:.2f}%</li>
                                        <li><strong>Moderados (2.5-3σ):</strong> {severidad_counts[severidad_counts['Severidad']=='Moderado (2.5-3σ)']['Porcentaje'].values[0] if 'Moderado (2.5-3σ)' in severidad_counts['Severidad'].values else 0:.2f}%</li>
                                        <li><strong>Altos (3-4σ):</strong> {severidad_counts[severidad_counts['Severidad']=='Alto (3-4σ)']['Porcentaje'].values[0] if 'Alto (3-4σ)' in severidad_counts['Severidad'].values else 0:.2f}%</li>
                                        <li><strong>Extremos (>4σ):</strong> {severidad_counts[severidad_counts['Severidad']=='Extremo (>4σ)']['Porcentaje'].values[0] if 'Extremo (>4σ)' in severidad_counts['Severidad'].values else 0:.2f}%</li>
                                    </ul>
                                </div>
                                """, unsafe_allow_html=True)
                            
                            # Casos extremos
                            extremos = df_sobre[df_sobre['Severidad'] == "Extremo (>4σ)"]
                            if not extremos.empty:
                                st.markdown(show_alert(
                                    f"Se han identificado casos con desviaciones superiores a 4 sigmas del promedio. Estos casos requieren atención inmediata ya que pueden indicar fugas, manipulación o problemas mecánicos graves.",
                                    "danger",
                                    f"⚠️ {len(extremos)} casos de sobreconsumo extremo detectados",
                                    "🚨"
                                ), unsafe_allow_html=True)
                        
                        # Tendencia temporal
                        if total_sobre > 0 and 'Fecha' in df_sobre.columns:
                            st.markdown("<h3 class='section-title'>Evolución Temporal de Sobreconsumos</h3>", unsafe_allow_html=True)
                            
                            # Agrupar por fecha
                            sobre_por_dia = df_sobre.groupby(df_sobre['Fecha'].dt.date).size().reset_index()
                            sobre_por_dia.columns = ['Fecha', 'Sobreconsumos']
                            
                            # Total por día
                            total_por_dia = df_filtrado.groupby(df_filtrado['Fecha'].dt.date).size().reset_index()
                            total_por_dia.columns = ['Fecha', 'Total Cargas']
                            
                            # Unir dataframes
                            tendencia = pd.merge(sobre_por_dia, total_por_dia, on='Fecha')
                            tendencia['Porcentaje'] = (tendencia['Sobreconsumos'] / tendencia['Total Cargas'] * 100).round(2)
                            
                            # Gráfico
                            fig = make_subplots(specs=[[{"secondary_y": True}]])
                            
                            # Añadir barras de sobreconsumos
                            fig.add_trace(
                                go.Bar(
                                    x=tendencia['Fecha'],
                                    y=tendencia['Sobreconsumos'],
                                    name='Sobreconsumos',
                                    marker_color=COLORS['danger']
                                ),
                                secondary_y=False
                            )
                            
                            # Añadir línea de porcentaje
                            fig.add_trace(
                                go.Scatter(
                                    x=tendencia['Fecha'],
                                    y=tendencia['Porcentaje'],
                                    name='% Sobreconsumos',
                                    mode='lines+markers',
                                    marker=dict(color=COLORS['warning']),
                                    line=dict(color=COLORS['warning'], width=3)
                                ),
                                secondary_y=True
                            )
                            
                            # Actualizar ejes
                            fig.update_layout(
                                title='Evolución de Sobreconsumos a lo Largo del Tiempo',
                                plot_bgcolor='white',
                                hovermode='x unified',
                                height=450,
                                legend=dict(
                                    orientation="h",
                                    yanchor="bottom",
                                    y=1.02,
                                    xanchor="right",
                                    x=1
                                )
                            )
                            
                            fig.update_xaxes(title_text='Fecha')
                            fig.update_yaxes(title_text='Cantidad de Sobreconsumos', secondary_y=False)
                            fig.update_yaxes(title_text='% de Sobreconsumos', secondary_y=True)
                            
                            # Añadir línea de promedio
                            fig.add_hline(
                                y=porcentaje,
                                line_dash="dash",
                                line_color="red",
                                secondary_y=True,
                                annotation_text=f"Promedio: {porcentaje:.2f}%"
                            )
                            
                            st.plotly_chart(fig, use_container_width=True)
                            
                            # Análisis de tendencia
                            if len(tendencia) > 3:
                                # Calcular tendencia lineal
                                X = np.array(range(len(tendencia))).reshape(-1, 1)
                                y = tendencia['Porcentaje'].values
                                
                                modelo = LinearRegression()
                                modelo.fit(X, y)
                                tendencia_valor = modelo.coef_[0]
                                
                                # Determinar si es creciente o decreciente
                                if tendencia_valor > 0.1:
                                    st.markdown(show_alert(
                                        f"La tendencia de sobreconsumos está aumentando en un {tendencia_valor:.2f}% por día. Se recomienda revisar urgentemente la eficiencia de la flota.",
                                        "danger",
                                        "Tendencia Creciente de Sobreconsumos",
                                        "📈"
                                    ), unsafe_allow_html=True)
                                elif tendencia_valor < -0.1:
                                    st.markdown(show_alert(
                                        f"La tendencia de sobreconsumos está disminuyendo en un {abs(tendencia_valor):.2f}% por día. Las medidas de eficiencia parecen estar funcionando.",
                                        "success",
                                        "Tendencia Decreciente de Sobreconsumos",
                                        "📉"
                                    ), unsafe_allow_html=True)
                        
                        # Recomendaciones específicas
                        st.markdown("<h3 class='section-title'>Recomendaciones para Reducir Sobreconsumos</h3>", unsafe_allow_html=True)
                        
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.markdown(f"""
                            <div class="card">
                                <h4>Recomendaciones Operativas</h4>
                                <ul>
                                    <li><strong>Mantenimiento Preventivo:</strong> Implementar un programa riguroso de mantenimiento preventivo, con especial atención a los modelos con mayor tasa de sobreconsumo.</li>
                                    <li><strong>Capacitación en Conducción Eficiente:</strong> Proporcionar capacitación especializada a conductores con altas tasas de sobreconsumo.</li>
                                    <li><strong>Optimización de Rutas:</strong> Analizar y optimizar las rutas para minimizar tiempos de ralentí y maximizar la eficiencia.</li>
                                    <li><strong>Presión de Neumáticos:</strong> Establecer un protocolo de verificación regular de la presión de neumáticos.</li>
                                </ul>
                            </div>
                            """, unsafe_allow_html=True)
                        
                        with col2:
                            st.markdown(f"""
                            <div class="card">
                                <h4>Mejoras Tecnológicas</h4>
                                <ul>
                                    <li><strong>Sistemas de Monitoreo:</strong> Implementar sistemas de monitoreo en tiempo real para detectar patrones de sobreconsumo inmediatamente.</li>
                                    <li><strong>Limitadores de Velocidad:</strong> Considerar la instalación de limitadores de velocidad en vehículos con tendencia al sobreconsumo.</li>
                                    <li><strong>Análisis de Telemetría:</strong> Utilizar datos de telemetría para identificar comportamientos de conducción que aumentan el consumo.</li>
                                    <li><strong>Actualización de Flotas:</strong> Considerar la renovación o actualización de los modelos con mayores tasas de sobreconsumo.</li>
                                </ul>
                            </div>
                            """, unsafe_allow_html=True)
                        
                        # Listado detallado
                        st.markdown("<h3 class='section-title'>Listado Detallado de Sobreconsumos</h3>", unsafe_allow_html=True)
                        
                        # Columnas relevantes
                        cols_relevantes = [col for col in [
                            'Fecha', 'Hora', 'Terminal', 'Número interno', 'Patente', 'Modelo chasis',
                            'Cantidad litros', 'Z-Score', 'Promedio Modelo', 'Nombre conductor', 'Nombre supervisor'
                        ] if col in df_sobre.columns]
                        
                        # Mostrar datos con paginación
                        with st.expander("Ver listado completo", expanded=False):
                            # Crear selector de página
                            if len(df_sobre) > 50:
                                # Calcular número de páginas (50 registros por página)
                                num_paginas = (len(df_sobre) - 1) // 50 + 1
                                pagina_actual = st.selectbox("Página", range(1, num_paginas + 1), key="pagina_sobreconsumos")
                                
                                # Índices de inicio y fin para la página seleccionada
                                inicio = (pagina_actual - 1) * 50
                                fin = min(inicio + 50, len(df_sobre))
                                
                                # Mostrar datos de la página actual
                                st.dataframe(
                                    df_sobre[cols_relevantes].sort_values('Z-Score', ascending=False).iloc[inicio:fin],
                                    use_container_width=True,
                                    height=400
                                )
                                
                                st.info(f"Mostrando registros {inicio+1} a {fin} de {len(df_sobre)}")
                            else:
                                # Si hay pocos registros, mostrar todos
                                st.dataframe(
                                    df_sobre[cols_relevantes].sort_values('Z-Score', ascending=False),
                                    use_container_width=True,
                                    height=400
                                )
                        
                        # Opción para exportar
                        if len(df_sobre) > 0:
                            st.download_button(
                                label="📥 Descargar listado de sobreconsumos",
                                data=df_sobre[cols_relevantes].sort_values('Z-Score', ascending=False).to_csv(index=False).encode('utf-8'),
                                file_name="sobreconsumos.csv",
                                mime="text/csv",
                                help="Descargar un archivo CSV con el listado completo de sobreconsumos",
                                use_container_width=True
                            )
                    else:
                        st.warning("No se encontró información sobre sobreconsumo en los datos.")
                
                # === ANÁLISIS TEMPORAL ===
                elif section == "Análisis Temporal":
                    st.markdown("<h2 class='subtitle'>Análisis Temporal</h2>", unsafe_allow_html=True)
                    
                    if 'Fecha' in df.columns:
                        # Selección de vista y métrica
                        col1, col2, col3 = st.columns([2, 2, 1])
                        
                        with col1:
                            vista = st.selectbox(
                                "Seleccionar vista temporal",
                                ["Diaria", "Semanal", "Mensual", "Trimestral"]
                            )
                        
                        with col2:
                            if 'Cantidad litros' in df.columns:
                                metrica = st.selectbox(
                                    "Seleccionar métrica",
                                    ["Total Litros", "Promedio Litros/Carga", "Cantidad de Cargas"]
                                )
                            else:
                                metrica = "Cantidad de Cargas"
                        
                        with col3:
                            # Opción para mostrar tendencia
                            mostrar_tendencia = st.checkbox("Mostrar tendencia", value=True)
                        
                        # Preparar datos según vista seleccionada
                        if vista == "Diaria":
                            fecha_group = df['Fecha'].dt.date
                            titulo = "Evolución Diaria"
                            fecha_format = "%d/%m/%Y"
                        elif vista == "Semanal":
                            # Crear columna de semana
                            df['Semana'] = df['Fecha'].dt.isocalendar().week
                            df['Año'] = df['Fecha'].dt.isocalendar().year
                            fecha_group = df['Año'].astype(str) + "-W" + df['Semana'].astype(str).str.zfill(2)
                            titulo = "Evolución Semanal"
                            fecha_format = "%Y-W%W"
                        elif vista == "Mensual":
                            fecha_group = df['Fecha'].dt.strftime('%Y-%m')
                            titulo = "Evolución Mensual"
                            fecha_format = "%Y-%m"
                        else:  # Trimestral
                            df['Trimestre'] = df['Fecha'].dt.quarter
                            df['Año'] = df['Fecha'].dt.year
                            fecha_group = df['Año'].astype(str) + "-Q" + df['Trimestre'].astype(str)
                            titulo = "Evolución Trimestral"
                            fecha_format = "%Y-Q%q"
                        
                        # Agrupar datos
                        if metrica == "Total Litros" and 'Cantidad litros' in df.columns:
                            evolución = df.groupby(fecha_group)['Cantidad litros'].sum().reset_index()
                            evolución.columns = ['Período', 'Total Litros']
                            y_column = 'Total Litros'
                            titulo += " - Total de Litros"
                        elif metrica == "Promedio Litros/Carga" and 'Cantidad litros' in df.columns:
                            evolución = df.groupby(fecha_group)['Cantidad litros'].mean().reset_index()
                            evolución.columns = ['Período', 'Promedio Litros']
                            y_column = 'Promedio Litros'
                            titulo += " - Promedio de Litros por Carga"
                        else:  # Cantidad de Cargas
                            evolución = df.groupby(fecha_group).size().reset_index()
                            evolución.columns = ['Período', 'Cantidad Cargas']
                            y_column = 'Cantidad Cargas'
                            titulo += " - Cantidad de Cargas"
                        
                        # Ordenar por período
                        evolución = evolución.sort_values('Período')
                        
                        # Añadir columnas adicionales si están disponibles
                        if 'Mala Carga' in df.columns:
                            malas_cargas = df.groupby(fecha_group)['Mala Carga'].sum().reset_index()
                            malas_cargas.columns = ['Período', 'Malas Cargas']
                            evolución = pd.merge(evolución, malas_cargas, on='Período')
                            evolución['% Malas Cargas'] = (evolución['Malas Cargas'] / df.groupby(fecha_group).size().reset_index()[0] * 100).round(2)
                        
                        if 'Sobreconsumo' in df.columns:
                            sobreconsumos = df.groupby(fecha_group)['Sobreconsumo'].sum().reset_index()
                            sobreconsumos.columns = ['Período', 'Sobreconsumos']
                            evolución = pd.merge(evolución, sobreconsumos, on='Período')
                            evolución['% Sobreconsumos'] = (evolución['Sobreconsumos'] / df.groupby(fecha_group).size().reset_index()[0] * 100).round(2)
                        
                        # Visualizar
                        tab1, tab2 = st.tabs(["📊 Gráfico", "📋 Tabla"])
                        
                        with tab1:
                            # Gráfico de evolución
                            st.markdown("<h3 class='section-title'>Evolución Temporal</h3>", unsafe_allow_html=True)
                            
                            fig = px.line(
                                evolución,
                                x='Período',
                                y=y_column,
                                title=titulo,
                                markers=True
                            )
                            
                            fig.update_traces(
                                line=dict(color=COLORS['primary'], width=3),
                                marker=dict(color=COLORS['primary'], size=8)
                            )
                            
                            fig.update_layout(
                                xaxis_title="Período",
                                yaxis_title=y_column,
                                plot_bgcolor='white',
                                hovermode='x unified',
                                height=450
                            )
                            
                            # Calcular y añadir tendencia
                            if mostrar_tendencia and len(evolución) > 2:
                                X = np.array(range(len(evolución))).reshape(-1, 1)
                                y = evolución[y_column].values
                                
                                try:
                                    # Intentar con regresión robusta primero
                                    modelo = HuberRegressor()
                                    modelo.fit(X, y)
                                    tendencia = modelo.coef_[0]
                                except:
                                    # Fallback a regresión lineal estándar
                                    modelo = LinearRegression()
                                    modelo.fit(X, y)
                                    tendencia = modelo.coef_[0]
                                
                                # Añadir línea de tendencia
                                tendencia_y = modelo.predict(X)
                                
                                fig.add_trace(
                                    go.Scatter(
                                        x=evolución['Período'],
                                        y=tendencia_y,
                                        mode='lines',
                                        name='Tendencia',
                                        line=dict(color='red', width=2, dash='dash')
                                    )
                                )
                                
                                # Calcular porcentaje de cambio
                                promedio = np.mean(y)
                                porcentaje_cambio = (tendencia / promedio) * 100
                                
                                # Añadir anotación con la tendencia
                                if vista == "Diaria":
                                    unidad = "diario"
                                elif vista == "Semanal":
                                    unidad = "semanal"
                                elif vista == "Mensual":
                                    unidad = "mensual"
                                else:
                                    unidad = "trimestral"
                                
                                tendencia_texto = f"Tendencia: {'+' if tendencia > 0 else ''}{tendencia:.2f} ({'+' if tendencia > 0 else ''}{porcentaje_cambio:.2f}% {unidad})"
                                
                                fig.add_annotation(
                                    x=evolución['Período'].iloc[-1],
                                    y=tendencia_y[-1],
                                    text=tendencia_texto,
                                    showarrow=True,
                                    arrowhead=2,
                                    arrowsize=1,
                                    arrowwidth=2,
                                    arrowcolor='red',
                                    ax=50,
                                    ay=-30
                                )
                            
                            st.plotly_chart(fig, use_container_width=True)
                            
                            # Gráfico adicional para malas cargas y sobreconsumos si están disponibles
                            if 'Mala Carga' in df.columns and 'Sobreconsumo' in df.columns:
                                st.markdown("<h3 class='section-title'>Evolución de Indicadores de Calidad</h3>", unsafe_allow_html=True)
                                
                                fig = go.Figure()
                                
                                # Añadir línea para % de malas cargas
                                fig.add_trace(
                                    go.Scatter(
                                        x=evolución['Período'],
                                        y=evolución['% Malas Cargas'],
                                        mode='lines+markers',
                                        name='% Malas Cargas',
                                        line=dict(color=COLORS['danger'], width=3),
                                        marker=dict(color=COLORS['danger'], size=8)
                                    )
                                )
                                
                                # Añadir línea para % de sobreconsumos
                                fig.add_trace(
                                    go.Scatter(
                                        x=evolución['Período'],
                                        y=evolución['% Sobreconsumos'],
                                        mode='lines+markers',
                                        name='% Sobreconsumos',
                                        line=dict(color=COLORS['warning'], width=3),
                                        marker=dict(color=COLORS['warning'], size=8)
                                    )
                                )
                                
                                # Actualizar diseño
                                fig.update_layout(
                                    title='Evolución de Malas Cargas y Sobreconsumos',
                                    xaxis_title='Período',
                                    yaxis_title='Porcentaje (%)',
                                    plot_bgcolor='white',
                                    hovermode='x unified',
                                    height=450,
                                    legend=dict(
                                        orientation="h",
                                        yanchor="bottom",
                                        y=1.02,
                                        xanchor="right",
                                        x=1
                                    )
                                )
                                
                                st.plotly_chart(fig, use_container_width=True)
                        
                        with tab2:
                            # Mostrar tabla de evolución
                            st.dataframe(
                                evolución,
                                use_container_width=True
                            )
                        
                        # Análisis por día de semana
                        if 'Día Semana' in df.columns:
                            st.markdown("<h3 class='section-title'>Análisis por Día de Semana</h3>", unsafe_allow_html=True)
                            
                            # Agrupar por día de la semana
                            dias_orden = {
                                'Monday': 0, 'Tuesday': 1, 'Wednesday': 2, 'Thursday': 3, 
                                'Friday': 4, 'Saturday': 5, 'Sunday': 6,
                                'Lunes': 0, 'Martes': 1, 'Miércoles': 2, 'Jueves': 3, 
                                'Viernes': 4, 'Sábado': 5, 'Domingo': 6
                            }
                            
                            # Verificar si los días están en español o inglés
                            if df['Día Semana'].iloc[0] in list(dias_orden.keys())[:7]:  # Inglés
                                dict_dias = {0: 'Monday', 1: 'Tuesday', 2: 'Wednesday', 3: 'Thursday', 
                                           4: 'Friday', 5: 'Saturday', 6: 'Sunday'}
                            else:  # Español
                                dict_dias = {0: 'Lunes', 1: 'Martes', 2: 'Miércoles', 3: 'Jueves', 
                                           4: 'Viernes', 5: 'Sábado', 6: 'Domingo'}
                            
                            df['Día Orden'] = df['Día Semana'].map(dias_orden)
                            
                            # Datos para gráfico
                            if 'Cantidad litros' in df.columns:
                                cols = ['Día Semana', 'Día Orden', 'Cantidad litros']
                                metrics = ['sum', 'mean', 'count']
                                dia_agg = df[cols].groupby(['Día Semana', 'Día Orden']).agg({'Cantidad litros': metrics})
                                dia_agg.columns = ['Total Litros', 'Promedio Litros', 'Cantidad Cargas']
                                dia_agg = dia_agg.reset_index()
                            else:
                                cols = ['Día Semana', 'Día Orden']
                                dia_agg = df[cols].groupby(['Día Semana', 'Día Orden']).size().reset_index()
                                dia_agg.columns = ['Día Semana', 'Día Orden', 'Cantidad Cargas']
                            
                            # Ordenar por día de la semana
                            dia_agg = dia_agg.sort_values('Día Orden')
                            
                            # Mapear nombres de días
                            dia_agg['Día'] = dia_agg['Día Orden'].map(dict_dias)
                            
                            # Pestañas para diferentes métricas
                            tab1, tab2, tab3 = st.tabs(["📊 Cantidad de Cargas", "⛽ Total de Litros", "📏 Promedio de Litros"])
                            
                            with tab1:
                                # Gráfico de cargas por día
                                fig = px.bar(
                                    dia_agg,
                                    x='Día',
                                    y='Cantidad Cargas',
                                    title='Cantidad de Cargas por Día de Semana',
                                    color='Cantidad Cargas',
                                    color_continuous_scale='Viridis',
                                    text='Cantidad Cargas'
                                )
                                
                                fig.update_traces(
                                    texttemplate='%{text}',
                                    textposition='outside'
                                )
                                
                                fig.update_layout(
                                    xaxis_title="Día de Semana",
                                    yaxis_title="Cantidad de Cargas",
                                    plot_bgcolor='white'
                                )
                                
                                st.plotly_chart(fig, use_container_width=True)
                            
                            if 'Cantidad litros' in df.columns:
                                with tab2:
                                    # Gráfico de total de litros por día
                                    fig = px.bar(
                                        dia_agg,
                                        x='Día',
                                        y='Total Litros',
                                        title='Total de Litros por Día de Semana',
                                        color='Total Litros',
                                        color_continuous_scale='Viridis',
                                        text='Total Litros'
                                    )
                                    
                                    fig.update_traces(
                                        texttemplate='%{text:.1f}',
                                        textposition='outside'
                                    )
                                    
                                    fig.update_layout(
                                        xaxis_title="Día de Semana",
                                        yaxis_title="Total de Litros",
                                        plot_bgcolor='white'
                                    )
                                    
                                    st.plotly_chart(fig, use_container_width=True)
                                
                                with tab3:
                                    # Gráfico de promedio de litros por día
                                    fig = px.bar(
                                        dia_agg,
                                        x='Día',
                                        y='Promedio Litros',
                                        title='Promedio de Litros por Día de Semana',
                                        color='Promedio Litros',
                                        color_continuous_scale='Viridis',
                                        text='Promedio Litros'
                                    )
                                    
                                    fig.update_traces(
                                        texttemplate='%{text:.1f}',
                                        textposition='outside'
                                    )
                                    
                                    fig.update_layout(
                                        xaxis_title="Día de Semana",
                                        yaxis_title="Promedio de Litros",
                                        plot_bgcolor='white'
                                    )
                                    
                                    st.plotly_chart(fig, use_container_width=True)
                            
                            # Análisis comparativo de días laborales vs fin de semana
                            st.markdown("<h3 class='section-title'>Comparativa Laboral vs Fin de Semana</h3>", unsafe_allow_html=True)
                            
                            # Definir días laborales y fin de semana
                            dias_laborales = [0, 1, 2, 3, 4]  # Lunes a Viernes
                            fin_semana = [5, 6]  # Sábado y Domingo
                            
                            # Filtrar datos
                            df_laboral = df[df['Día Orden'].isin(dias_laborales)]
                            df_finde = df[df['Día Orden'].isin(fin_semana)]
                            
                            # Calcular métricas
                            metricas_comparacion = {
                                "Cantidad Cargas": [len(df_laboral), len(df_finde)],
                                "% del Total": [len(df_laboral) / len(df) * 100, len(df_finde) / len(df) * 100]
                            }
                            
                            if 'Cantidad litros' in df.columns:
                                metricas_comparacion["Total Litros"] = [df_laboral['Cantidad litros'].sum(), df_finde['Cantidad litros'].sum()]
                                metricas_comparacion["Promedio Litros"] = [df_laboral['Cantidad litros'].mean(), df_finde['Cantidad litros'].mean()]
                            
                            if 'Mala Carga' in df.columns:
                                metricas_comparacion["% Malas Cargas"] = [df_laboral['Mala Carga'].mean() * 100, df_finde['Mala Carga'].mean() * 100]
                            
                            if 'Sobreconsumo' in df.columns:
                                metricas_comparacion["% Sobreconsumos"] = [df_laboral['Sobreconsumo'].mean() * 100, df_finde['Sobreconsumo'].mean() * 100]
                            
                            # Crear DataFrame para visualización
                            df_comparacion = pd.DataFrame({
                                "Métrica": list(metricas_comparacion.keys()),
                                "Laboral (L-V)": [metricas_comparacion[k][0] for k in metricas_comparacion.keys()],
                                "Fin de Semana (S-D)": [metricas_comparacion[k][1] for k in metricas_comparacion.keys()],
                                "Diferencia %": [(metricas_comparacion[k][1] / metricas_comparacion[k][0] - 1) * 100 if metricas_comparacion[k][0] > 0 else 0 for k in metricas_comparacion.keys()]
                            })
                            
                            # Mostrar tabla
                            st.dataframe(
                                df_comparacion,
                                column_config={
                                    "Métrica": st.column_config.TextColumn("Métrica"),
                                    "Laboral (L-V)": st.column_config.NumberColumn("Laboral (L-V)", format="%.2f"),
                                    "Fin de Semana (S-D)": st.column_config.NumberColumn("Fin de Semana (S-D)", format="%.2f"),
                                    "Diferencia %": st.column_config.NumberColumn("Diferencia %", format="%.2f%%")
                                },
                                use_container_width=True
                            )
                            
                            # Gráfico comparativo
                            fig = go.Figure()
                            
                            # Barras para días laborales
                            fig.add_trace(
                                go.Bar(
                                    x=df_comparacion['Métrica'],
                                    y=df_comparacion['Laboral (L-V)'],
                                    name='Laboral (L-V)',
                                    marker_color=COLORS['primary']
                                )
                            )
                            
                            # Barras para fin de semana
                            fig.add_trace(
                                go.Bar(
                                    x=df_comparacion['Métrica'],
                                    y=df_comparacion['Fin de Semana (S-D)'],
                                    name='Fin de Semana (S-D)',
                                    marker_color=COLORS['secondary']
                                )
                            )
                            
                            fig.update_layout(
                                title='Comparativa Días Laborales vs Fin de Semana',
                                xaxis_title='Métrica',
                                yaxis_title='Valor',
                                barmode='group',
                                plot_bgcolor='white'
                            )
                            
                            st.plotly_chart(fig, use_container_width=True)
                            
                            # Insights y recomendaciones
                            if 'Promedio Litros' in metricas_comparacion:
                                diff_litros = (metricas_comparacion['Promedio Litros'][1] / metricas_comparacion['Promedio Litros'][0] - 1) * 100
                                
                                if abs(diff_litros) > 10:
                                    if diff_litros > 0:
                                        st.markdown(show_alert(
                                            f"El promedio de litros por carga es un {diff_litros:.2f}% mayor durante los fines de semana que en días laborales. Se recomienda revisar los patrones operativos durante los fines de semana para identificar posibles oportunidades de mejora.",
                                            "warning",
                                            "Diferencia Significativa en Consumo de Fin de Semana",
                                            "⚠️"
                                        ), unsafe_allow_html=True)
                                    else:
                                        st.markdown(show_alert(
                                            f"El promedio de litros por carga es un {abs(diff_litros):.2f}% menor durante los fines de semana que en días laborales. Esto podría indicar que las rutas o la operación durante el fin de semana son más eficientes.",
                                            "info",
                                            "Consumo Menor en Fin de Semana",
                                            "ℹ️"
                                        ), unsafe_allow_html=True)
                        
                        # Análisis por hora del día
                        if 'Hora Numérica' in df.columns:
                            st.markdown("<h3 class='section-title'>Análisis por Hora del Día</h3>", unsafe_allow_html=True)
                            
                            # Redondear horas para agrupar
                            df['Hora Redondeada'] = np.round(df['Hora Numérica']).astype(int)
                            
                            # Agrupar por hora
                            if 'Cantidad litros' in df.columns:
                                hora_agg = df.groupby('Hora Redondeada').agg({
                                    'Cantidad litros': ['sum', 'mean', 'count']
                                }).reset_index()
                                hora_agg.columns = ['Hora', 'Total Litros', 'Promedio Litros', 'Cantidad Cargas']
                            else:
                                hora_agg = df.groupby('Hora Redondeada').size().reset_index()
                                hora_agg.columns = ['Hora', 'Cantidad Cargas']
                            
                            # Ordenar por hora
                            hora_agg = hora_agg.sort_values('Hora')
                            
                            # Pestañas para diferentes métricas
                            if 'Cantidad litros' in df.columns:
                                tab1, tab2, tab3 = st.tabs(["📊 Cantidad de Cargas", "⛽ Total de Litros", "📏 Promedio de Litros"])
                            else:
                                tab1 = st.container()  # Solo una pestaña si no hay datos de litros
                            
                            # Gráfico de cargas por hora
                            if 'Cantidad litros' in df.columns:
                                with tab1:
                                    fig = px.bar(
                                        hora_agg,
                                        x='Hora',
                                        y='Cantidad Cargas',
                                        title='Cantidad de Cargas por Hora del Día',
                                        color='Cantidad Cargas',
                                        color_continuous_scale='Viridis',
                                        text='Cantidad Cargas'
                                    )
                                    
                                    fig.update_traces(
                                        texttemplate='%{text}',
                                        textposition='outside'
                                    )
                                    
                                    fig.update_layout(
                                        xaxis_title="Hora del Día",
                                        yaxis_title="Cantidad de Cargas",
                                        plot_bgcolor='white',
                                        xaxis=dict(
                                            tickmode='linear',
                                            tick0=0,
                                            dtick=2  # Mostrar etiquetas cada 2 horas
                                        )
                                    )
                                    
                                    st.plotly_chart(fig, use_container_width=True)
                                
                                with tab2:
                                    # Gráfico de total de litros por hora
                                    fig = px.bar(
                                        hora_agg,
                                        x='Hora',
                                        y='Total Litros',
                                        title='Total de Litros por Hora del Día',
                                        color='Total Litros',
                                        color_continuous_scale='Viridis',
                                        text='Total Litros'
                                    )
                                    
                                    fig.update_traces(
                                        texttemplate='%{text:.1f}',
                                        textposition='outside'
                                    )
                                    
                                    fig.update_layout(
                                        xaxis_title="Hora del Día",
                                        yaxis_title="Total de Litros",
                                        plot_bgcolor='white',
                                        xaxis=dict(
                                            tickmode='linear',
                                            tick0=0,
                                            dtick=2  # Mostrar etiquetas cada 2 horas
                                        )
                                    )
                                    
                                    st.plotly_chart(fig, use_container_width=True)
                                
                                with tab3:
                                    # Gráfico de promedio de litros por hora
                                    fig = px.bar(
                                        hora_agg,
                                        x='Hora',
                                        y='Promedio Litros',
                                        title='Promedio de Litros por Hora del Día',
                                        color='Promedio Litros',
                                        color_continuous_scale='Viridis',
                                        text='Promedio Litros'
                                    )
                                    
                                    fig.update_traces(
                                        texttemplate='%{text:.1f}',
                                        textposition='outside'
                                    )
                                    
                                    fig.update_layout(
                                        xaxis_title="Hora del Día",
                                        yaxis_title="Promedio de Litros",
                                        plot_bgcolor='white',
                                        xaxis=dict(
                                            tickmode='linear',
                                            tick0=0,
                                            dtick=2  # Mostrar etiquetas cada 2 horas
                                        )
                                    )
                                    
                                    st.plotly_chart(fig, use_container_width=True)
                            else:
                                # Gráfico simple si solo hay datos de cantidad
                                fig = px.bar(
                                    hora_agg,
                                    x='Hora',
                                    y='Cantidad Cargas',
                                    title='Cantidad de Cargas por Hora del Día',
                                    color='Cantidad Cargas',
                                    color_continuous_scale='Viridis',
                                    text='Cantidad Cargas'
                                )
                                
                                fig.update_traces(
                                    texttemplate='%{text}',
                                    textposition='outside'
                                )
                                
                                fig.update_layout(
                                    xaxis_title="Hora del Día",
                                    yaxis_title="Cantidad de Cargas",
                                    plot_bgcolor='white',
                                    xaxis=dict(
                                        tickmode='linear',
                                        tick0=0,
                                        dtick=2  # Mostrar etiquetas cada 2 horas
                                    )
                                )
                                
                                st.plotly_chart(fig, use_container_width=True)
                            
                            # Análisis por período del día
                            if 'Período' in df.columns:
                                st.markdown("<h3 class='section-title'>Análisis por Período del Día</h3>", unsafe_allow_html=True)
                                
                                # Agrupar por período
                                if 'Cantidad litros' in df.columns:
                                    periodo_agg = df.groupby('Período').agg({
                                        'Cantidad litros': ['sum', 'mean', 'count']
                                    }).reset_index()
                                    periodo_agg.columns = ['Período', 'Total Litros', 'Promedio Litros', 'Cantidad Cargas']
                                else:
                                    periodo_agg = df.groupby('Período').size().reset_index()
                                    periodo_agg.columns = ['Período', 'Cantidad Cargas']
                                
                                # Ordenar períodos
                                orden_periodo = {'Madrugada': 0, 'Mañana': 1, 'Tarde': 2, 'Noche': 3}
                                periodo_agg['Orden'] = periodo_agg['Período'].map(orden_periodo)
                                periodo_agg = periodo_agg.sort_values('Orden')
                                
                                # Crear gráfico
                                col1, col2 = st.columns(2)
                                
                                with col1:
                                    # Gráfico de cargas por período
                                    fig = px.pie(
                                        periodo_agg,
                                        names='Período',
                                        values='Cantidad Cargas',
                                        title='Distribución de Cargas por Período',
                                        color_discrete_sequence=px.colors.sequential.Viridis
                                    )
                                    
                                    fig.update_traces(
                                        textposition='inside',
                                        textinfo='percent+label'
                                    )
                                    
                                    fig.update_layout(
                                        plot_bgcolor='white'
                                    )
                                    
                                    st.plotly_chart(fig, use_container_width=True)
                                
                                if 'Cantidad litros' in df.columns:
                                    with col2:
                                        # Gráfico de litros por período
                                        fig = px.pie(
                                            periodo_agg,
                                            names='Período',
                                            values='Total Litros',
                                            title='Distribución de Litros por Período',
                                            color_discrete_sequence=px.colors.sequential.Viridis
                                        )
                                        
                                        fig.update_traces(
                                            textposition='inside',
                                            textinfo='percent+label'
                                        )
                                        
                                        fig.update_layout(
                                            plot_bgcolor='white'
                                        )
                                        
                                        st.plotly_chart(fig, use_container_width=True)
                                    
                                    # Mostrar tabla de datos
                                    st.dataframe(
                                        periodo_agg[['Período', 'Cantidad Cargas', 'Total Litros', 'Promedio Litros']],
                                        column_config={
                                            "Período": st.column_config.TextColumn("Período del Día"),
                                            "Cantidad Cargas": st.column_config.NumberColumn("Cantidad Cargas", format="%d"),
                                            "Total Litros": st.column_config.NumberColumn("Total Litros", format="%.2f"),
                                            "Promedio Litros": st.column_config.NumberColumn("Promedio Litros", format="%.2f")
                                        },
                                        use_container_width=True
                                    )
                        
                        # Proporcionar conclusiones y recomendaciones
                        st.markdown("<h3 class='section-title'>Conclusiones y Recomendaciones</h3>", unsafe_allow_html=True)
                        
                        # Calcular métricas para insights
                        insights = []
                        
                        # Tendencia
                        if 'tendencia_valor' in locals():
                            if tendencia_valor > 0.05 * promedio:
                                insights.append({
                                    "tipo": "warning",
                                    "titulo": f"Tendencia al alza en {metrica}",
                                    "mensaje": f"Se observa una tendencia creciente de {tendencia_valor:.2f} ({porcentaje_cambio:.2f}% {unidad}). Considere revisar la eficiencia operativa para controlar el aumento."
                                })
                            elif tendencia_valor < -0.05 * promedio:
                                insights.append({
                                    "tipo": "success",
                                    "titulo": f"Tendencia a la baja en {metrica}",
                                    "mensaje": f"Se observa una tendencia decreciente de {tendencia_valor:.2f} ({porcentaje_cambio:.2f}% {unidad}). Las medidas de eficiencia parecen estar funcionando correctamente."
                                })
                        
                        # Diferencias entre días de semana
                        if 'dia_agg' in locals() and 'Cantidad Cargas' in dia_agg.columns:
                            max_dia = dia_agg.loc[dia_agg['Cantidad Cargas'].idxmax()]
                            min_dia = dia_agg.loc[dia_agg['Cantidad Cargas'].idxmin()]
                            
                            diff_dias = (max_dia['Cantidad Cargas'] / min_dia['Cantidad Cargas'] - 1) * 100
                            
                            if diff_dias > 50:
                                insights.append({
                                    "tipo": "info",
                                    "titulo": "Alta variabilidad entre días de la semana",
                                    "mensaje": f"El día con mayor actividad ({max_dia['Día']}) tiene un {diff_dias:.2f}% más de cargas que el día con menor actividad ({min_dia['Día']}). Considere redistribuir las cargas de manera más uniforme para optimizar los recursos."
                                })
                        
                        # Diferencia entre laboral y fin de semana
                        if 'df_comparacion' in locals() and len(df_comparacion) > 0:
                            diff_litros_finde = df_comparacion.loc[df_comparacion['Métrica'] == 'Promedio Litros', 'Diferencia %'].values[0] if 'Promedio Litros' in df_comparacion['Métrica'].values else 0
                            
                            if abs(diff_litros_finde) > 15:
                                tipo = "warning" if diff_litros_finde > 0 else "info"
                                mensaje = f"El promedio de litros es un {abs(diff_litros_finde):.2f}% {'mayor' if diff_litros_finde > 0 else 'menor'} en fin de semana que en días laborales."
                                
                                if diff_litros_finde > 0:
                                    mensaje += " Se recomienda revisar las operaciones de fin de semana para identificar posibles ineficiencias."
                                else:
                                    mensaje += " Las operaciones de fin de semana parecen ser más eficientes que en días laborales."
                                
                                insights.append({
                                    "tipo": tipo,
                                    "titulo": "Diferencia significativa entre días laborales y fin de semana",
                                    "mensaje": mensaje
                                })
                        
                        # Concentración de cargas por hora
                        if 'hora_agg' in locals() and 'Cantidad Cargas' in hora_agg.columns:
                            max_hora = hora_agg.loc[hora_agg['Cantidad Cargas'].idxmax()]
                            
                            # Calcular el porcentaje de cargas en la hora pico
                            pct_hora_pico = (max_hora['Cantidad Cargas'] / hora_agg['Cantidad Cargas'].sum()) * 100
                            
                            if pct_hora_pico > 15:
                                insights.append({
                                    "tipo": "warning",
                                    "titulo": "Alta concentración de cargas en hora pico",
                                    "mensaje": f"El {pct_hora_pico:.2f}% de las cargas se realizan a las {max_hora['Hora']}:00 horas. Considere distribuir las cargas a lo largo del día para evitar congestionamiento y mejorar la eficiencia."
                                })
                        
                        # Mostrar insights
                        if insights:
                            for insight in insights:
                                st.markdown(show_alert(
                                    insight["mensaje"],
                                    insight["tipo"],
                                    insight["titulo"],
                                    "📊" if insight["tipo"] == "info" else "⚠️" if insight["tipo"] == "warning" else "✅"
                                ), unsafe_allow_html=True)
                        else:
                            st.markdown(show_alert(
                                "No se han detectado patrones temporales significativos en los datos analizados.",
                                "info",
                                "Sin patrones temporales destacables",
                                "ℹ️"
                            ), unsafe_allow_html=True)
                    else:
                        st.warning("No se encontró información temporal en los datos.")
                
                # === RENDIMIENTO ===
                elif section == "Rendimiento":
                    st.markdown("<h2 class='subtitle'>Análisis de Rendimiento</h2>", unsafe_allow_html=True)
                    
                    if 'Rendimiento' in df.columns:
                        # Fecha de análisis
                        col1, col2 = st.columns([1, 1])
                        
                        with col1:
                            if 'Fecha' in df.columns:
                                fecha_min = df['Fecha'].min().date()
                                fecha_max = df['Fecha'].max().date()
                            else:
                                fecha_min = datetime.date.today() - datetime.timedelta(days=30)
                                fecha_max = datetime.date.today()
                            
                            fecha_inicio = st.date_input("Fecha inicio", fecha_min, key="rendimiento_fecha_inicio")
                        
                        with col2:
                            fecha_fin = st.date_input("Fecha fin", fecha_max, key="rendimiento_fecha_fin")
                        
                        # Filtrar por fecha
                        if 'Fecha' in df.columns:
                            df_filtrado = df[
                                (df['Fecha'].dt.date >= fecha_inicio) & 
                                (df['Fecha'].dt.date <= fecha_fin)
                            ]
                        else:
                            df_filtrado = df.copy()
                        
                        # Filtrar valores válidos de rendimiento
                        df_rendimiento = df_filtrado.dropna(subset=['Rendimiento'])
                        
                        # Métricas principales
                        col1, col2, col3, col4 = st.columns(4)
                        
                        with col1:
                            # Rendimiento promedio
                            rendimiento_promedio = df_rendimiento['Rendimiento'].mean()
                            
                            st.markdown(create_kpi_card(
                                "Rendimiento Promedio", 
                                f"{rendimiento_promedio:.2f} km/L",
                                None,
                                icon="🔄",
                                color_class="primary",
                                footnote="De todas las cargas"
                            ), unsafe_allow_html=True)
                        
                        with col2:
                            # Rendimiento mediana
                            rendimiento_mediana = df_rendimiento['Rendimiento'].median()
                            
                            st.markdown(create_kpi_card(
                                "Rendimiento Mediana", 
                                f"{rendimiento_mediana:.2f} km/L",
                                None,
                                icon="📊",
                                color_class="info",
                                footnote="Valor central"
                            ), unsafe_allow_html=True)
                        
                        with col3:
                            # Mejor rendimiento
                            if len(df_rendimiento) > 0:
                                mejor_rendimiento = df_rendimiento['Rendimiento'].max()
                                
                                st.markdown(create_kpi_card(
                                    "Mejor Rendimiento", 
                                    f"{mejor_rendimiento:.2f} km/L",
                                    None,
                                    icon="🏆",
                                    color_class="success",
                                    footnote="Valor máximo"
                                ), unsafe_allow_html=True)
                        
                        with col4:
                            # Peor rendimiento
                            if len(df_rendimiento) > 0:
                                peor_rendimiento = df_rendimiento['Rendimiento'].min()
                                
                                st.markdown(create_kpi_card(
                                    "Peor Rendimiento", 
                                    f"{peor_rendimiento:.2f} km/L",
                                    None,
                                    icon="⚠️",
                                    color_class="danger",
                                    footnote="Valor mínimo"
                                ), unsafe_allow_html=True)
                        
                        # Distribución de rendimiento
                        st.markdown("<h3 class='section-title'>Distribución del Rendimiento</h3>", unsafe_allow_html=True)
                        
                        # Gráfico de histograma
                        fig = px.histogram(
                            df_rendimiento,
                            x='Rendimiento',
                            nbins=30,
                            marginal='box',
                            title='Distribución del Rendimiento (km/L)',
                            color_discrete_sequence=[COLORS['primary']]
                        )
                        
                        fig.update_layout(
                            xaxis_title="Rendimiento (km/L)",
                            yaxis_title="Frecuencia",
                            plot_bgcolor='white'
                        )
                        
                        # Añadir líneas para estadísticas clave
                        fig.add_vline(
                            x=rendimiento_promedio,
                            line_dash="dash",
                            line_color="red",
                            annotation_text=f"Media: {rendimiento_promedio:.2f}"
                        )
                        
                        fig.add_vline(
                            x=rendimiento_mediana,
                            line_dash="dash",
                            line_color="green",
                            annotation_text=f"Mediana: {rendimiento_mediana:.2f}"
                        )
                        
                        st.plotly_chart(fig, use_container_width=True)
                        
                        # Rendimiento por modelo
                        if 'Modelo chasis' in df_rendimiento.columns:
                            st.markdown("<h3 class='section-title'>Rendimiento por Modelo</h3>", unsafe_allow_html=True)
                            
                            # Agrupar por modelo
                            rendimiento_modelo = df_rendimiento.groupby('Modelo chasis')['Rendimiento'].agg(['mean', 'median', 'std', 'count']).reset_index()
                            rendimiento_modelo.columns = ['Modelo', 'Promedio', 'Mediana', 'Desviación', 'Cantidad']
                            
                            # Filtrar modelos con suficientes datos
                            rendimiento_modelo = rendimiento_modelo[rendimiento_modelo['Cantidad'] >= 5]
                            
                            # Ordenar por promedio (descendente)
                            rendimiento_modelo = rendimiento_modelo.sort_values('Promedio', ascending=False)
                            
                            # Tabs para diferentes vistas
                            tab1, tab2 = st.tabs(["📊 Gráfico", "📋 Tabla"])
                            
                            with tab1:
                                # Gráfico de barras
                                fig = px.bar(
                                    rendimiento_modelo,
                                    x='Modelo',
                                    y='Promedio',
                                    title='Rendimiento Promedio por Modelo',
                                    color='Promedio',
                                    color_continuous_scale='Viridis',
                                    text='Promedio',
                                    error_y='Desviación'  # Barras de error
                                )
                                
                                fig.update_traces(
                                    texttemplate='%{text:.2f}',
                                    textposition='outside'
                                )
                                
                                fig.update_layout(
                                    xaxis_title="Modelo",
                                    yaxis_title="Rendimiento Promedio (km/L)",
                                    plot_bgcolor='white'
                                )
                                
                                # Añadir línea para el promedio global
                                fig.add_hline(
                                    y=rendimiento_promedio,
                                    line_dash="dash",
                                    line_color="red",
                                    annotation_text=f"Promedio Global: {rendimiento_promedio:.2f}"
                                )
                                
                                st.plotly_chart(fig, use_container_width=True)
                            
                            with tab2:
                                # Mostrar tabla
                                st.dataframe(
                                    rendimiento_modelo,
                                    column_config={
                                        "Modelo": st.column_config.TextColumn("Modelo"),
                                        "Promedio": st.column_config.NumberColumn("Promedio (km/L)", format="%.2f"),
                                        "Mediana": st.column_config.NumberColumn("Mediana (km/L)", format="%.2f"),
                                        "Desviación": st.column_config.NumberColumn("Desviación Estándar", format="%.2f"),
                                        "Cantidad": st.column_config.NumberColumn("Cantidad de Cargas", format="%d")
                                    },
                                    use_container_width=True
                                )
                                
                                # Identificar modelos destacados
                                mejor_modelo = rendimiento_modelo.iloc[0]
                                peor_modelo = rendimiento_modelo.iloc[-1]
                                
                                # Mostrar insights
                                col1, col2 = st.columns(2)
                                
                                with col1:
                                    st.markdown(f"""
                                    <div class="card">
                                        <h4>🏆 Modelo más Eficiente</h4>
                                        <p><strong>Modelo:</strong> {mejor_modelo['Modelo']}</p>
                                        <p><strong>Rendimiento:</strong> {mejor_modelo['Promedio']:.2f} km/L</p>
                                        <p><strong>Diferencia vs Promedio:</strong> {((mejor_modelo['Promedio'] / rendimiento_promedio) - 1) * 100:.2f}%</p>
                                        <p><strong>Cargas Analizadas:</strong> {mejor_modelo['Cantidad']}</p>
                                    </div>
                                    """, unsafe_allow_html=True)
                                
                                with col2:
                                    st.markdown(f"""
                                    <div class="card">
                                        <h4>⚠️ Modelo menos Eficiente</h4>
                                        <p><strong>Modelo:</strong> {peor_modelo['Modelo']}</p>
                                        <p><strong>Rendimiento:</strong> {peor_modelo['Promedio']:.2f} km/L</p>
                                        <p><strong>Diferencia vs Promedio:</strong> {((peor_modelo['Promedio'] / rendimiento_promedio) - 1) * 100:.2f}%</p>
                                        <p><strong>Cargas Analizadas:</strong> {peor_modelo['Cantidad']}</p>
                                    </div>
                                    """, unsafe_allow_html=True)
                                
                                # Recomendaciones basadas en diferencias
                                diferencia_pct = ((mejor_modelo['Promedio'] / peor_modelo['Promedio']) - 1) * 100
                                
                                if diferencia_pct > 25:
                                    st.markdown(show_alert(
                                        f"La diferencia de rendimiento entre el modelo más eficiente y el menos eficiente es del {diferencia_pct:.2f}%. Se recomienda realizar un análisis técnico detallado de ambos modelos y evaluar la posibilidad de renovar la flota con modelos más eficientes.",
                                        "warning",
                                        "Diferencia Significativa entre Modelos",
                                        "⚠️"
                                    ), unsafe_allow_html=True)
                        
                        # Rendimiento por bus
                        if 'Número interno' in df_rendimiento.columns:
                            st.markdown("<h3 class='section-title'>Rendimiento por Bus</h3>", unsafe_allow_html=True)
                            
                            # Agrupar por bus
                            rendimiento_bus = df_rendimiento.groupby('Número interno')['Rendimiento'].agg(['mean', 'count']).reset_index()
                            rendimiento_bus.columns = ['Bus', 'Rendimiento', 'Cargas']
                            
                            # Filtrar buses con suficientes datos
                            rendimiento_bus = rendimiento_bus[rendimiento_bus['Cargas'] >= 5]
                            
                            # Tabs para mejores y peores buses
                            tab1, tab2 = st.tabs(["🏆 Top 10 Mejores", "⚠️ Top 10 Peores"])
                            
                            with tab1:
                                # Mejores buses
                                mejores_buses = rendimiento_bus.sort_values('Rendimiento', ascending=False).head(10)
                                
                                # Gráfico
                                fig = px.bar(
                                    mejores_buses,
                                    x='Bus',
                                    y='Rendimiento',
                                    title='Top 10 Buses con Mejor Rendimiento',
                                    color='Rendimiento',
                                    color_continuous_scale='Viridis',
                                    text='Rendimiento'
                                )
                                
                                fig.update_traces(
                                    texttemplate='%{text:.2f}',
                                    textposition='outside'
                                )
                                
                                fig.update_layout(
                                    xaxis_title="Bus",
                                    yaxis_title="Rendimiento (km/L)",
                                    plot_bgcolor='white'
                                )
                                
                                # Añadir línea para el promedio global
                                fig.add_hline(
                                    y=rendimiento_promedio,
                                    line_dash="dash",
                                    line_color="red",
                                    annotation_text=f"Promedio Global: {rendimiento_promedio:.2f}"
                                )
                                
                                st.plotly_chart(fig, use_container_width=True)
                                
                                # Tabla de mejores buses
                                st.dataframe(
                                    mejores_buses,
                                    column_config={
                                        "Bus": st.column_config.TextColumn("Bus"),
                                        "Rendimiento": st.column_config.NumberColumn("Rendimiento (km/L)", format="%.2f"),
                                        "Cargas": st.column_config.NumberColumn("Cargas", format="%d")
                                    },
                                    use_container_width=True
                                )
                            
                            with tab2:
                                # Peores buses
                                peores_buses = rendimiento_bus.sort_values('Rendimiento').head(10)
                                
                                # Gráfico
                                fig = px.bar(
                                    peores_buses,
                                    x='Bus',
                                    y='Rendimiento',
                                    title='Top 10 Buses con Peor Rendimiento',
                                    color='Rendimiento',
                                    color_continuous_scale='Viridis_r',  # Escala invertida
                                    text='Rendimiento'
                                )
                                
                                fig.update_traces(
                                    texttemplate='%{text:.2f}',
                                    textposition='outside'
                                )
                                
                                fig.update_layout(
                                    xaxis_title="Bus",
                                    yaxis_title="Rendimiento (km/L)",
                                    plot_bgcolor='white'
                                )
                                
                                # Añadir línea para el promedio global
                                fig.add_hline(
                                    y=rendimiento_promedio,
                                    line_dash="dash",
                                    line_color="red",
                                    annotation_text=f"Promedio Global: {rendimiento_promedio:.2f}"
                                )
                                
                                st.plotly_chart(fig, use_container_width=True)
                                
                                # Tabla de peores buses
                                st.dataframe(
                                    peores_buses,
                                    column_config={
                                        "Bus": st.column_config.TextColumn("Bus"),
                                        "Rendimiento": st.column_config.NumberColumn("Rendimiento (km/L)", format="%.2f"),
                                        "Cargas": st.column_config.NumberColumn("Cargas", format="%d")
                                    },
                                    use_container_width=True
                                )
                                
                                # Recomendaciones para buses de bajo rendimiento
                                peor_bus = peores_buses.iloc[0]
                                diferencia_peor = ((rendimiento_promedio / peor_bus['Rendimiento']) - 1) * 100
                                
                                if diferencia_peor > 25:
                                    st.markdown(show_alert(
                                        f"El bus {peor_bus['Bus']} tiene un rendimiento {diferencia_peor:.2f}% inferior al promedio. Se recomienda una revisión técnica detallada y mantenimiento preventivo para mejorar su eficiencia.",
                                        "warning",
                                        "Bus con Rendimiento Crítico",
                                        "⚠️"
                                    ), unsafe_allow_html=True)
                        
                        # Evolución del rendimiento en el tiempo
                        if 'Fecha' in df_rendimiento.columns:
                            st.markdown("<h3 class='section-title'>Evolución del Rendimiento</h3>", unsafe_allow_html=True)
                            
                            # Agrupar por fecha
                            rendimiento_tiempo = df_rendimiento.groupby(df_rendimiento['Fecha'].dt.date)['Rendimiento'].mean().reset_index()
                            rendimiento_tiempo.columns = ['Fecha', 'Rendimiento']
                            
                            # Ordenar por fecha
                            rendimiento_tiempo = rendimiento_tiempo.sort_values('Fecha')
                            
                            # Gráfico
                            fig = px.line(
                                rendimiento_tiempo,
                                x='Fecha',
                                y='Rendimiento',
                                title='Evolución del Rendimiento Promedio a lo Largo del Tiempo',
                                markers=True
                            )
                            
                            fig.update_traces(
                                line=dict(color=COLORS['primary'], width=3),
                                marker=dict(color=COLORS['primary'], size=8)
                            )
                            
                            fig.update_layout(
                                xaxis_title="Fecha",
                                yaxis_title="Rendimiento Promedio (km/L)",
                                plot_bgcolor='white',
                                hovermode='x unified',
                                height=450
                            )
                            
                            # Añadir línea para el promedio global
                            fig.add_hline(
                                y=rendimiento_promedio,
                                line_dash="dash",
                                line_color="red",
                                annotation_text=f"Promedio Global: {rendimiento_promedio:.2f}"
                            )
                            
                            # Calcular y añadir tendencia
                            if len(rendimiento_tiempo) > 3:
                                X = np.array(range(len(rendimiento_tiempo))).reshape(-1, 1)
                                y = rendimiento_tiempo['Rendimiento'].values
                                
                                try:
                                    # Intentar con regresión robusta primero
                                    modelo = HuberRegressor()
                                    modelo.fit(X, y)
                                    tendencia = modelo.coef_[0]
                                except:
                                    # Fallback a regresión lineal estándar
                                    modelo = LinearRegression()
                                    modelo.fit(X, y)
                                    tendencia = modelo.coef_[0]
                                
                                # Añadir línea de tendencia
                                tendencia_y = modelo.predict(X)
                                
                                fig.add_trace(
                                    go.Scatter(
                                        x=rendimiento_tiempo['Fecha'],
                                        y=tendencia_y,
                                        mode='lines',
                                        name='Tendencia',
                                        line=dict(color='red', width=2, dash='dash')
                                    )
                                )
                                
                                # Calcular porcentaje de cambio diario
                                porcentaje_cambio = (tendencia / rendimiento_promedio) * 100
                                
                                tendencia_texto = f"Tendencia: {'+' if tendencia > 0 else ''}{tendencia:.4f} ({'+' if tendencia > 0 else ''}{porcentaje_cambio:.2f}% diario)"
                                
                                fig.add_annotation(
                                    x=rendimiento_tiempo['Fecha'].iloc[-1],
                                    y=tendencia_y[-1],
                                    text=tendencia_texto,
                                    showarrow=True,
                                    arrowhead=2,
                                    arrowsize=1,
                                    arrowwidth=2,
                                    arrowcolor='red',
                                    ax=50,
                                    ay=-30
                                )
                                
                                # Comentarios sobre la tendencia
                                if tendencia > 0.01:
                                    st.markdown(show_alert(
                                        f"Se observa una tendencia positiva en el rendimiento, con un incremento de {tendencia:.4f} km/L ({porcentaje_cambio:.2f}%) por día. Las medidas de optimización parecen estar funcionando correctamente.",
                                        "success",
                                        "Tendencia Positiva en Rendimiento",
                                        "📈"
                                    ), unsafe_allow_html=True)
                                elif tendencia < -0.01:
                                    st.markdown(show_alert(
                                        f"Se observa una tendencia negativa en el rendimiento, con una disminución de {abs(tendencia):.4f} km/L ({abs(porcentaje_cambio):.2f}%) por día. Se recomienda revisar las condiciones operativas y el mantenimiento de la flota.",
                                        "warning",
                                        "Tendencia Negativa en Rendimiento",
                                        "📉"
                                    ), unsafe_allow_html=True)
                            
                            st.plotly_chart(fig, use_container_width=True)
                        
                        # Influencia de factores en el rendimiento
                        st.markdown("<h3 class='section-title'>Factores que Influyen en el Rendimiento</h3>", unsafe_allow_html=True)
                        
                        # Analizar correlaciones
                        factores_potenciales = []
                        
                        if 'Cantidad litros' in df_rendimiento.columns:
                            factores_potenciales.append('Cantidad litros')
                        
                        if 'Hora Numérica' in df_rendimiento.columns:
                            factores_potenciales.append('Hora Numérica')
                        
                        if 'Odómetro' in df_rendimiento.columns:
                            factores_potenciales.append('Odómetro')
                        
                        if 'Kilómetros Recorridos' in df_rendimiento.columns:
                            factores_potenciales.append('Kilómetros Recorridos')
                        
                        if factores_potenciales:
                            # Calcular correlaciones
                            correlaciones = []
                            
                            for factor in factores_potenciales:
                                corr = df_rendimiento['Rendimiento'].corr(df_rendimiento[factor])
                                correlaciones.append({
                                    "Factor": factor,
                                    "Correlación": corr,
                                    "Fuerza": abs(corr)
                                })
                            
                            # Ordenar por fuerza de correlación (descendente)
                            correlaciones = sorted(correlaciones, key=lambda x: x["Fuerza"], reverse=True)
                            
                            # Mostrar correlaciones
                            st.markdown("#### Correlaciones con el Rendimiento")
                            
                            df_corr = pd.DataFrame(correlaciones)
                            
                            st.dataframe(
                                df_corr,
                                column_config={
                                    "Factor": st.column_config.TextColumn("Factor"),
                                    "Correlación": st.column_config.NumberColumn("Correlación", format="%.3f"),
                                    "Fuerza": st.column_config.ProgressColumn("Fuerza de Correlación", format="%.2f", min_value=0, max_value=1)
                                },
                                use_container_width=True
                            )
                            
                            # Mostrar gráfico para el factor con mayor correlación
                            if correlaciones:
                                factor_principal = correlaciones[0]["Factor"]
                                
                                # Título descriptivo según el factor
                                titulos = {
                                    'Cantidad litros': 'Relación entre Tamaño de Carga y Rendimiento',
                                    'Hora Numérica': 'Relación entre Hora del Día y Rendimiento',
                                    'Odómetro': 'Relación entre Kilometraje Acumulado y Rendimiento',
                                    'Kilómetros Recorridos': 'Relación entre Distancia Recorrida y Rendimiento'
                                }
                                
                                titulo = titulos.get(factor_principal, f'Relación entre {factor_principal} y Rendimiento')
                                
                                # Gráfico de dispersión
                                fig = px.scatter(
                                    df_rendimiento,
                                    x=factor_principal,
                                    y='Rendimiento',
                                    title=titulo,
                                    opacity=0.7,
                                    color_discrete_sequence=[COLORS['primary']]
                                )
                                
                                fig.update_layout(
                                    xaxis_title=factor_principal,
                                    yaxis_title="Rendimiento (km/L)",
                                    plot_bgcolor='white'
                                )
                                
                                # Añadir línea de tendencia
                                fig.update_traces(
                                    marker=dict(size=8)
                                )
                                
                                fig.update_layout(
                                    shapes=[
                                        dict(
                                            type='line',
                                            xref='x',
                                            yref='y',
                                            x0=df_rendimiento[factor_principal].min(),
                                            y0=np.polyval(np.polyfit(df_rendimiento[factor_principal], df_rendimiento['Rendimiento'], 1), df_rendimiento[factor_principal].min()),
                                            x1=df_rendimiento[factor_principal].max(),
                                            y1=np.polyval(np.polyfit(df_rendimiento[factor_principal], df_rendimiento['Rendimiento'], 1), df_rendimiento[factor_principal].max()),
                                            line=dict(color='red', width=2, dash='dash'),
                                        )
                                    ]
                                )
                                
                                st.plotly_chart(fig, use_container_width=True)
                                
                                # Agregar interpretación según el factor
                                if factor_principal == 'Cantidad litros':
                                    corr = correlaciones[0]["Correlación"]
                                    if corr < -0.3:
                                        st.markdown(show_alert(
                                            f"Existe una correlación negativa significativa (r={corr:.3f}) entre el tamaño de la carga y el rendimiento. Esto sugiere que las cargas más grandes tienden a estar asociadas con menor rendimiento, posiblemente debido a un llenado excesivo o condiciones de operación diferentes.",
                                            "info",
                                            "Influencia del Tamaño de Carga",
                                            "ℹ️"
                                        ), unsafe_allow_html=True)
                                    elif corr > 0.3:
                                        st.markdown(show_alert(
                                            f"Existe una correlación positiva significativa (r={corr:.3f}) entre el tamaño de la carga y el rendimiento. Esto podría indicar que las cargas más grandes están asociadas con mejores condiciones operativas o vehículos más eficientes.",
                                            "info",
                                            "Influencia del Tamaño de Carga",
                                            "ℹ️"
                                        ), unsafe_allow_html=True)
                                
                                elif factor_principal == 'Hora Numérica':
                                    corr = correlaciones[0]["Correlación"]
                                    if abs(corr) > 0.2:
                                        periodo = "horas pico" if corr < 0 else "horas de baja congestión"
                                        st.markdown(show_alert(
                                            f"Se observa una correlación (r={corr:.3f}) entre la hora del día y el rendimiento. Esto sugiere que el rendimiento tiende a ser mejor durante {periodo}, posiblemente debido a diferencias en la congestión del tráfico.",
                                            "info",
                                            "Influencia del Horario",
                                            "ℹ️"
                                        ), unsafe_allow_html=True)
                                
                                elif factor_principal == 'Odómetro':
                                    corr = correlaciones[0]["Correlación"]
                                    if corr < -0.2:
                                        st.markdown(show_alert(
                                            f"Existe una correlación negativa (r={corr:.3f}) entre el kilometraje acumulado y el rendimiento. Esto indica que los vehículos con mayor kilometraje tienden a tener menor rendimiento, lo que sugiere un efecto del desgaste en la eficiencia.",
                                            "warning",
                                            "Influencia del Desgaste",
                                            "⚠️"
                                        ), unsafe_allow_html=True)
                        
                        # Recomendaciones generales
                        st.markdown("<h3 class='section-title'>Recomendaciones para Mejorar el Rendimiento</h3>", unsafe_allow_html=True)
                        
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.markdown(f"""
                            <div class="card">
                                <h4>Recomendaciones Operativas</h4>
                                <ul>
                                    <li><strong>Mantenimiento Preventivo:</strong> Implementar un programa riguroso de mantenimiento preventivo, especialmente para los buses con rendimiento por debajo del promedio.</li>
                                    <li><strong>Capacitación en Conducción Eficiente:</strong> Proporcionar capacitación a los conductores sobre técnicas de conducción eficiente, como arranques suaves y evitar frenados bruscos.</li>
                                    <li><strong>Optimización de Rutas:</strong> Analizar y optimizar las rutas para minimizar tiempo en congestión y maximizar la eficiencia.</li>
                                    <li><strong>Presión de Neumáticos:</strong> Establecer un protocolo de verificación regular de la presión de neumáticos.</li>
                                </ul>
                            </div>
                            """, unsafe_allow_html=True)
                        
                        with col2:
                            st.markdown(f"""
                            <div class="card">
                                <h4>Recomendaciones Técnicas</h4>
                                <ul>
                                    <li><strong>Evaluación Técnica:</strong> Realizar evaluaciones técnicas detalladas de los buses con rendimiento significativamente inferior al promedio.</li>
                                    <li><strong>Renovación de Flota:</strong> Considerar la renovación prioritaria de los modelos con peor rendimiento promedio.</li>
                                    <li><strong>Monitoreo Continuo:</strong> Implementar sistemas de monitoreo en tiempo real para detectar disminuciones en el rendimiento que puedan indicar problemas mecánicos.</li>
                                    <li><strong>Análisis de Combustible:</strong> Verificar la calidad del combustible utilizado y evaluar alternativas de mayor eficiencia.</li>
                                </ul>
                            </div>
                            """, unsafe_allow_html=True)
                    else:
                        st.warning("No se encontró información sobre rendimiento en los datos.")
                
                # === EXPORTAR RESULTADOS ===
                elif section == "Exportar Resultados":
                    st.markdown("<h2 class='subtitle'>Exportar Resultados</h2>", unsafe_allow_html=True)
                    
                    # Sección de opciones de exportación
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        st.markdown(f"""
                        <div class="card">
                            <h4>Opciones de Exportación</h4>
                            <p>Seleccione el formato y las secciones que desea incluir en el informe exportado.</p>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        # Selección de formato
                        formato = st.selectbox(
                            "Formato de exportación",
                            ["Excel (.xlsx)", "CSV (.csv)"]
                        )
                        
                        # Incluir gráficos (solo para Excel)
                        incluir_graficos = st.checkbox("Incluir gráficos", value=True, disabled=(formato != "Excel (.xlsx)"))
                        
                        # Secciones a incluir
                        st.markdown("##### Secciones a incluir")
                        
                        incluir_resumen = st.checkbox("Resumen general", value=True)
                        incluir_consumo = st.checkbox("Consumo diario", value=True)
                        incluir_terminales = st.checkbox("Análisis por terminal", value=True)
                        incluir_buses = st.checkbox("Análisis por buses", value=True)
                        incluir_modelos = st.checkbox("Análisis por modelo", value=True)
                        incluir_malas = st.checkbox("Malas cargas", value=True)
                        incluir_sobreconsumo = st.checkbox("Sobreconsumo", value=True)
                        incluir_conductores = st.checkbox("Análisis por conductor", value=True)
                        incluir_rendimiento = st.checkbox("Rendimiento", value=True)
                        incluir_datos = st.checkbox("Datos completos", value=True)
                        
                    with col2:
                        st.markdown(f"""
                        <div class="card">
                            <h4>Información del Informe</h4>
                            <p>Agregue información adicional para personalizar el informe exportado.</p>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        # Título personalizado
                        titulo_informe = st.text_input("Título del informe", "Análisis de Consumo de Combustible")
                        
                        # Información adicional
                        info_adicional = st.text_area("Notas o comentarios adicionales", placeholder="Ingrese aquí cualquier información adicional que desee incluir en el informe...")
                        
                        # Fecha de referencia
                        if 'Fecha' in df.columns:
                            fecha_min = df['Fecha'].min().date()
                            fecha_max = df['Fecha'].max().date()
                        else:
                            fecha_min = datetime.date.today() - datetime.timedelta(days=30)
                            fecha_max = datetime.date.today()
                        
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            fecha_inicio = st.date_input("Fecha inicio", fecha_min, key="export_fecha_inicio")
                        
                        with col2:
                            fecha_fin = st.date_input("Fecha fin", fecha_max, key="export_fecha_fin")
                    
                    # Botón para generar informe
                    if st.button("Generar Informe", type="primary", use_container_width=True):
                        # Mostrar mensaje de carga
                        with st.spinner("Generando informe... Por favor espere..."):
                            # Filtrar por fecha
                            if 'Fecha' in df.columns:
                                df_export = df[
                                    (df['Fecha'].dt.date >= fecha_inicio) & 
                                    (df['Fecha'].dt.date <= fecha_fin)
                                ]
                            else:
                                df_export = df.copy()
                            
                            # Generar informe según formato seleccionado
                            if formato == "Excel (.xlsx)":
                                # Generar Excel avanzado
                                output = export_to_excel(
                                    df_export, 
                                    filename=titulo_informe, 
                                    include_plots=incluir_graficos
                                )
                                
                                # Mostrar mensaje de éxito
                                st.success("Informe generado exitosamente. Haga clic en el botón para descargar.")
                                
                                # Botón de descarga
                                st.download_button(
                                    label="📥 Descargar Informe Excel",
                                    data=output,
                                    file_name=f"{titulo_informe.replace(' ', '_')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    help="Descargar el informe en formato Excel",
                                    use_container_width=True
                                )
                            else:
                                # Generar archivo CSV
                                csv = df_export.to_csv(index=False).encode('utf-8')
                                
                                # Mostrar mensaje de éxito
                                st.success("Datos exportados exitosamente. Haga clic en el botón para descargar.")
                                
                                # Botón de descarga
                                st.download_button(
                                    label="📥 Descargar Datos CSV",
                                    data=csv,
                                    file_name=f"{titulo_informe.replace(' ', '_')}.csv",
                                    mime="text/csv",
                                    help="Descargar los datos en formato CSV",
                                    use_container_width=True
                                )
                    
                    # Consejos y notas
                    st.markdown(f"""
                    <div class="card">
                        <h4>Notas sobre la Exportación</h4>
                        <ul>
                            <li><strong>Formato Excel:</strong> Incluye múltiples hojas con análisis específicos y gráficos integrados (si se selecciona la opción).</li>
                            <li><strong>Formato CSV:</strong> Exporta únicamente los datos en formato plano, sin análisis ni gráficos.</li>
                            <li><strong>Datos Incluidos:</strong> Solo se exportan los datos dentro del rango de fechas seleccionado.</li>
                            <li><strong>Tamaño del Archivo:</strong> Los archivos Excel con gráficos pueden ser de mayor tamaño, especialmente con grandes volúmenes de datos.</li>
                        </ul>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    # Opciones de informes programados
                    with st.expander("Informes Programados (Próximamente)", expanded=False):
                        st.markdown(f"""
                        <div style="text-align: center; padding: 20px;">
                            <h3>🔜 Función en Desarrollo</h3>
                            <p>La opción para programar informes automáticos estará disponible en una próxima actualización.</p>
                            <p>Esta funcionalidad permitirá generar y enviar informes periódicos de forma automática a los destinatarios que especifique.</p>
                        </div>
                        """, unsafe_allow_html=True)
            else:
                st.warning("Por favor, haga clic en 'Iniciar Análisis Avanzado' para procesar los datos.")
    else:
        # Mostrar banner de bienvenida
        show_welcome_banner()
        
        # Características principales
        st.markdown("<h2 class='subtitle'>Características principales</h2>", unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.markdown(f"""
            <div class="card">
                <h3 style="color: {COLORS['primary']};">📊 Análisis Preciso</h3>
                <p>Detección automática de patrones de consumo, identificación de anomalías y visualizaciones detalladas con tecnología de machine learning.</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
            <div class="card">
                <h3 style="color: {COLORS['primary']};">⚠️ Alertas Inteligentes</h3>
                <p>Sistema avanzado de detección de sobreconsumo, malas cargas y anomalías operativas con tableros de control en tiempo real.</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            st.markdown(f"""
            <div class="card">
                <h3 style="color: {COLORS['primary']};">📱 Informes Completos</h3>
                <p>Informes personalizables por terminal, buses, personal y períodos, exportables a Excel con formatos profesionales.</p>
            </div>
            """, unsafe_allow_html=True)
        
        # Cómo funciona
        st.markdown("<h2 class='subtitle'>¿Cómo funciona?</h2>", unsafe_allow_html=True)
        
        col1, col2 = st.columns([3, 2])
        
        with col1:
            st.markdown(f"""
            <div class="card">
                <h3 style="color: {COLORS['primary']};">Proceso de Análisis</h3>
                <p>Smart Fuel Analytics 3.0 utiliza algoritmos avanzados para procesar y analizar sus datos de combustible:</p>
                <ol>
                    <li><strong>Carga de datos:</strong> Suba sus archivos Excel o CSV con registros de cargas de combustible.</li>
                    <li><strong>Preprocesamiento automático:</strong> El sistema limpia, normaliza y enriquece los datos, detectando automáticamente formatos y columnas.</li>
                    <li><strong>Detección de anomalías:</strong> Utilizamos técnicas de IA como Isolation Forest y análisis estadístico para identificar patrones anómalos.</li>
                    <li><strong>Análisis multidimensional:</strong> Examinamos el consumo por terminal, bus, modelo, conductor y períodos de tiempo.</li>
                    <li><strong>Generación de informes:</strong> Visualizaciones interactivas, métricas clave y reportes exportables para una toma de decisiones informada.</li>
                </ol>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
            <div class="card">
                <h3 style="color: {COLORS['primary']};">Beneficios Clave</h3>
                <ul>
                    <li><strong>Ahorro de combustible</strong> mediante la identificación de ineficiencias</li>
                    <li><strong>Reducción de errores operativos</strong> a través de la detección de malas cargas</li>
                    <li><strong>Optimización de flota</strong> con análisis de rendimiento por modelo y bus</li>
                    <li><strong>Mejora de procesos</strong> en terminales y personal</li>
                    <li><strong>Toma de decisiones basada en datos</strong> con métricas claras y precisas</li>
                </ul>
                <div style="text-align: center; margin-top: 1.5rem;">
                    <p>👈 Comience cargando un archivo desde el panel lateral</p>
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        # Instrucciones
        st.markdown("<h2 class='subtitle'>Instrucciones de uso</h2>", unsafe_allow_html=True)
        
        st.markdown(f"""
        <div class="card">
            <ol>
                <li><strong>Sube tu archivo Excel o CSV</strong> con datos de carga de combustible desde el panel lateral</li>
                <li>Haz clic en <strong>Iniciar Análisis Avanzado</strong> para procesar los datos con nuestros algoritmos</li>
                <li>Navega entre las diferentes secciones para ver análisis específicos y visualizaciones detalladas</li>
                <li>Utiliza los filtros interactivos para profundizar en aspectos particulares de tus datos</li>
                <li>Exporta los resultados a Excel para compartir o guardar cuando sea necesario</li>
            </ol>
            <p style="margin-top: 1rem;"><strong>Formato esperado del archivo:</strong></p>
            <ul>
                <li>Título en fila A3</li>
                <li>Datos desde la fila A4</li>
                <li>Columnas principales: Turno, Fecha, Hora, Terminal, Número interno, Patente, Cantidad litros, etc.</li>
                <li>El sistema detectará automáticamente "Carga Masiva" en cualquier columna, incluida la última</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)

# Ejecutar la aplicación
if __name__ == "__main__":
    main()